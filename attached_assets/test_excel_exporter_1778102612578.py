"""Tests for the Excel exporter."""

from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timezone
from pathlib import Path
import unittest
from uuid import uuid4

from openpyxl import load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.cross_cut_engine import compute_cross_cuts
from src.excel_exporter import (
    export_cross_cuts_only,
    export_filtered_single_cuts,
    export_single_cuts,
)
from src.models import (
    AuditRecord,
    AnalysisType,
    CrossCutResult,
    CrossCutSpec,
    DataQualityReport,
    DenominatorPolicy,
    FilteredSingleCutResult,
    FilterSpec,
    GridSingleSelectResult,
    MultiSelectResult,
    NumericResult,
    QuestionSpec,
    QuestionType,
    SingleSelectResult,
    SkipRecord,
    SurveySchema,
)
from tests.conftest import CROSS_CUT_30_RESPONDENTS_PATH


UTC_NOW = datetime(2026, 5, 4, 12, 0, tzinfo=timezone.utc)
LONG_ID = "Q_VERY_LONG_SINGLE_SELECT_EXPORT_NAME"
FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"


def make_audit(
    metric_name: str,
    question_id: str,
    source_columns: tuple[str, ...],
    value_raw: float,
    valid_n: int,
    missing_n: int,
) -> AuditRecord:
    return AuditRecord(
        output_sheet=f"SC_{question_id}",
        metric_name=metric_name,
        source_question_id=question_id,
        source_columns=source_columns,
        filter_expr=None,
        numerator=None,
        denominator=valid_n,
        formula=f"{metric_name} formula",
        value_raw=value_raw,
        valid_n=valid_n,
        missing_n=missing_n,
        timestamp=UTC_NOW,
    )


def make_export_fixture() -> tuple[
    list,
    list[SkipRecord],
    SurveySchema,
    DataQualityReport,
    CalculationLog,
]:
    ss_audit = make_audit("rate_per_value", "Q_SS_EXPORT", ("Q_SS_EXPORT",), 10, 10, 0)
    ms_audit = make_audit(
        "selection_rate",
        "Q_MS_EXPORT",
        ("Q_MS_EXPORTr1", "Q_MS_EXPORTr2"),
        10,
        10,
        0,
    )
    num_audit = make_audit("numeric_summary", "Q_NUM_EXPORT", ("Q_NUM_EXPORT",), 5.5, 10, 0)
    grid_row_1_audit = make_audit(
        "rate_per_value", "Q_GRID_EXPORTr1", ("Q_GRID_EXPORTr1",), 10, 10, 0
    )
    grid_row_2_audit = make_audit(
        "rate_per_value", "Q_GRID_EXPORTr2", ("Q_GRID_EXPORTr2",), 10, 10, 0
    )
    grid_parent_audit = make_audit(
        "grid_overall",
        "Q_GRID_EXPORT",
        ("Q_GRID_EXPORTr1", "Q_GRID_EXPORTr2"),
        10,
        10,
        0,
    )
    long_audit = make_audit("rate_per_value", LONG_ID, (LONG_ID,), 10, 10, 0)

    single_result = SingleSelectResult(
        question_id="Q_SS_EXPORT",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Yes", "count": 6, "rate": 0.6},
            2: {"label": "No", "count": 4, "rate": 0.4},
        },
        audit_records=(ss_audit,),
    )
    multi_result = MultiSelectResult(
        question_id="Q_MS_EXPORT",
        question_type=QuestionType.MULTI_SELECT_BINARY,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        selections={
            "Q_MS_EXPORTr1": {"label": "First", "count": 5, "selection_rate": 0.5},
            "Q_MS_EXPORTr2": {"label": "Second", "count": 2, "selection_rate": 0.2},
        },
        respondents_who_answered_any=10,
        audit_records=(ms_audit,),
    )
    numeric_result = NumericResult(
        question_id="Q_NUM_EXPORT",
        question_type=QuestionType.DIRECT_NUMERIC,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        mean=5.5,
        median=5.5,
        std=2.9,
        min_val=1.0,
        max_val=10.0,
        percentiles={25: 3.0, 50: 5.5, 75: 8.0},
        audit_records=(num_audit,),
    )
    grid_row_1 = SingleSelectResult(
        question_id="Q_GRID_EXPORTr1",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Low", "count": 4, "rate": 0.4},
            2: {"label": "High", "count": 6, "rate": 0.6},
        },
        audit_records=(grid_row_1_audit,),
    )
    grid_row_2 = SingleSelectResult(
        question_id="Q_GRID_EXPORTr2",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Low", "count": 5, "rate": 0.5},
            2: {"label": "High", "count": 5, "rate": 0.5},
        },
        audit_records=(grid_row_2_audit,),
    )
    grid_result = GridSingleSelectResult(
        question_id="Q_GRID_EXPORT",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        rows={"Q_GRID_EXPORTr1": grid_row_1, "Q_GRID_EXPORTr2": grid_row_2},
        overall_valid_n=10,
        audit_records=(grid_parent_audit,),
    )
    long_result = SingleSelectResult(
        question_id=LONG_ID,
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "A", "count": 7, "rate": 0.7},
            2: {"label": "B", "count": 3, "rate": 0.3},
        },
        audit_records=(long_audit,),
        warnings=("long id warning",),
    )
    results = [single_result, multi_result, numeric_result, grid_result, long_result]

    questions = (
        QuestionSpec(
            question_id="[Q_SS_EXPORT]",
            canonical_id="Q_SS_EXPORT",
            question_text="Single select export question",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_SS_EXPORT",),
            option_map={1: "Yes", 2: "No"},
        ),
        QuestionSpec(
            question_id="[Q_MS_EXPORT]",
            canonical_id="Q_MS_EXPORT",
            question_text="Multi select export question",
            question_type=QuestionType.MULTI_SELECT_BINARY,
            raw_columns=("Q_MS_EXPORTr1", "Q_MS_EXPORTr2"),
            option_map={"Q_MS_EXPORTr1": "First", "Q_MS_EXPORTr2": "Second"},
            value_range=(0, 1),
        ),
        QuestionSpec(
            question_id="[Q_NUM_EXPORT]",
            canonical_id="Q_NUM_EXPORT",
            question_text="Numeric export question",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_NUM_EXPORT",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_GRID_EXPORT]",
            canonical_id="Q_GRID_EXPORT",
            question_text="Grid export question",
            question_type=QuestionType.GRID_SINGLE_SELECT,
            raw_columns=("Q_GRID_EXPORTr1", "Q_GRID_EXPORTr2"),
            option_map={1: "Low", 2: "High"},
            value_range=(1, 2),
            grid_row_labels={
                "Q_GRID_EXPORTr1": "Grid first row",
                "Q_GRID_EXPORTr2": "Grid second row",
            },
        ),
        QuestionSpec(
            question_id=f"[{LONG_ID}]",
            canonical_id=LONG_ID,
            question_text="Long sheet name export question",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=(LONG_ID,),
            option_map={1: "A", 2: "B"},
        ),
        QuestionSpec(
            question_id="[Q_TEXT_EXPORT]",
            canonical_id="Q_TEXT_EXPORT",
            question_text="Open text export question",
            question_type=QuestionType.OPEN_TEXT,
            raw_columns=("Q_TEXT_EXPORT",),
            option_map={},
        ),
    )
    schema = SurveySchema(
        questions=questions,
        respondent_id_column="respondent_id",
        total_respondents=10,
        source_datamap_path="datamap.xlsx",
        source_rawdata_path="raw.csv",
        parsed_at=UTC_NOW,
    )
    skips = [
        SkipRecord(
            question_id="[Q_TEXT_EXPORT]",
            canonical_id="Q_TEXT_EXPORT",
            question_type=QuestionType.OPEN_TEXT,
            skip_reason="unsupported_type: OPEN_TEXT",
        ),
        SkipRecord(
            question_id="[Q_FAIL_EXPORT]",
            canonical_id="Q_FAIL_EXPORT",
            question_type=QuestionType.SINGLE_SELECT,
            skip_reason="calculation_error",
            details="ValueError: raw column not found in data",
        ),
    ]
    quality_report = DataQualityReport(
        total_rows=10,
        total_columns=8,
        columns_in_datamap=7,
        columns_not_in_datamap=("extra_col",),
        per_column_missing_pct={"Q_SS_EXPORT": 0.1, "Q_NUM_EXPORT": 0.0, "extra_col": 0.5},
        per_column_out_of_range_pct={"Q_NUM_EXPORT": 0.2},
        coercion_log=(
            {
                "column": "Q_NUM_EXPORT",
                "from_type": "string",
                "to_type": "numeric",
                "values_coerced": ["abc"],
                "rows_affected": 1,
            },
        ),
        warnings=("column extra_col has 50.0% missing values",),
    )
    log = CalculationLog()
    for record in (
        ss_audit,
        ms_audit,
        num_audit,
        grid_row_1_audit,
        grid_row_2_audit,
        grid_parent_audit,
        long_audit,
    ):
        log.record(record)
    return results, skips, schema, quality_report, log


def make_cross_cut_schema() -> SurveySchema:
    questions = (
        QuestionSpec(
            question_id="[Q_SEG_1]",
            canonical_id="Q_SEG_1",
            question_text="Segment",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_SEG_1",),
            option_map={1: "Segment 1", 2: "Segment 2", 3: "Segment 3"},
        ),
        QuestionSpec(
            question_id="[Q_TGT_1]",
            canonical_id="Q_TGT_1",
            question_text="Target categorical",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_TGT_1",),
            option_map={1: "A", 2: "B", 3: "C", 4: "D"},
        ),
        QuestionSpec(
            question_id="[Q_NUM_3]",
            canonical_id="Q_NUM_3",
            question_text="Numeric metric",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_NUM_3",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_EXP_1]",
            canonical_id="Q_EXP_1",
            question_text="Expected",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_EXP_1",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_REAL_1]",
            canonical_id="Q_REAL_1",
            question_text="Realized",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_REAL_1",),
            option_map={},
        ),
    )
    return SurveySchema(
        questions=questions,
        respondent_id_column="respondent_id",
        total_respondents=30,
        source_datamap_path="cross_datamap.xlsx",
        source_rawdata_path="cross_raw.csv",
        parsed_at=UTC_NOW,
    )


def grid_segment_question() -> QuestionSpec:
    return QuestionSpec(
        question_id="[Q_GRID_SEG]",
        canonical_id="Q_GRID_SEG",
        question_text="Grid segment",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        raw_columns=("Q_GRID_SEGr1", "Q_GRID_SEGr2", "Q_GRID_SEGr3"),
        option_map={1: "Selected"},
        value_range=(0, 1),
        grid_row_labels={
            "Q_GRID_SEGr1": "Segment 1",
            "Q_GRID_SEGr2": "Segment 2",
            "Q_GRID_SEGr3": "Segment 3",
        },
    )


def grid_cross_tab_export_result(display_mode: str) -> tuple[CrossCutResult, SurveySchema, CalculationLog]:
    dataframe = pd.read_csv(CROSS_CUT_30_RESPONDENTS_PATH)
    dataframe["Q_GRID_SEGr1"] = [1] * 10 + [0] * 20
    dataframe["Q_GRID_SEGr2"] = [0] * 10 + [1] * 10 + [0] * 10
    dataframe["Q_GRID_SEGr3"] = [0] * 20 + [1] * 10
    base_schema = make_cross_cut_schema()
    schema = SurveySchema(
        questions=base_schema.questions + (grid_segment_question(),),
        respondent_id_column=base_schema.respondent_id_column,
        total_respondents=base_schema.total_respondents,
        source_datamap_path=base_schema.source_datamap_path,
        source_rawdata_path=base_schema.source_rawdata_path,
        parsed_at=base_schema.parsed_at,
    )
    log = CalculationLog()
    results, skips = compute_cross_cuts(
        [
            CrossCutSpec(
                cross_cut_id="CC_GRID_EXPORT",
                title="Grid segment by target",
                analysis_type=AnalysisType.CROSS_TAB,
                source_question_ids=("Q_GRID_SEG", "Q_TGT_1"),
                display_mode=display_mode,
            )
        ],
        schema,
        dataframe,
        log,
    )
    if skips:
        raise AssertionError(skips[0].details)
    return results[0], schema, log


def make_cross_cut_export_fixture():
    results, skips, schema, quality_report, log = make_export_fixture()
    dataframe = pd.read_csv(CROSS_CUT_30_RESPONDENTS_PATH)
    cross_schema = make_cross_cut_schema()
    export_schema = SurveySchema(
        questions=schema.questions + cross_schema.questions,
        respondent_id_column=schema.respondent_id_column,
        total_respondents=schema.total_respondents,
        source_datamap_path=schema.source_datamap_path,
        source_rawdata_path=schema.source_rawdata_path,
        parsed_at=schema.parsed_at,
    )
    cross_specs = [
        CrossCutSpec(
            cross_cut_id="CC_TAB_EXPORT",
            title="Segment by target",
            analysis_type=AnalysisType.CROSS_TAB,
            source_question_ids=("Q_SEG_1", "Q_TGT_1"),
        ),
        CrossCutSpec(
            cross_cut_id="CC_SEG_EXPORT",
            title="Segment profile",
            analysis_type=AnalysisType.SEGMENT_PROFILE,
            source_question_ids=("Q_SEG_1", "Q_TGT_1"),
            filter_expr="Q_SEG_1 == 1",
            filter_mask_description="Q_SEG_1 = Segment 1",
        ),
        CrossCutSpec(
            cross_cut_id="CC_GROUP_EXPORT",
            title="Group comparison",
            analysis_type=AnalysisType.GROUP_COMPARISON,
            source_question_ids=("Q_SEG_1", "Q_NUM_3"),
        ),
        CrossCutSpec(
            cross_cut_id="CC_EVR_EXPORT",
            title="Expected vs realized",
            analysis_type=AnalysisType.EXPECTED_VS_REALIZED,
            source_question_ids=("Q_EXP_1", "Q_REAL_1"),
        ),
    ]
    cross_results, _ = compute_cross_cuts(cross_specs, cross_schema, dataframe, log)
    cross_skips = [
        SkipRecord(
            question_id="CC_BAD_EXPORT",
            canonical_id="CC_BAD_EXPORT",
            question_type=QuestionType.UNKNOWN,
            skip_reason="cross_cut_error",
            details="ValueError: synthetic cross cut failure",
        )
    ]
    return results, skips, export_schema, quality_report, log, cross_results, cross_skips


def make_filtered_export_fixture():
    (
        results,
        _skips,
        schema,
        _quality_report,
        log,
        cross_results,
        _cross_skips,
    ) = make_cross_cut_export_fixture()
    unrelated_audit = make_audit(
        "rate_per_value",
        "Q_UNRELATED_EXPORT",
        ("Q_UNRELATED_EXPORT",),
        1.0,
        1,
        0,
    )
    log.record(unrelated_audit)
    filtered_results = [
        FilteredSingleCutResult(
            target_question_id="Q_SS_EXPORT",
            filters_applied=(FilterSpec("Q_SEG_1", 1),),
            dispatch_mode="single_cut_filtered",
            single_cut_result=results[0],
            cross_cut_result=None,
            filtered_n=6,
            audit_records=results[0].audit_records,
            warnings=("Filtered sample size 6 is below reliability threshold (30); results may not be reliable.",),
        ),
        FilteredSingleCutResult(
            target_question_id="Q_SS_EXPORT",
            filters_applied=(FilterSpec("Q_SEG_1", 2),),
            dispatch_mode="single_cut_filtered",
            single_cut_result=replace(results[0], valid_n=4),
            cross_cut_result=None,
            filtered_n=4,
            audit_records=results[0].audit_records,
        ),
        FilteredSingleCutResult(
            target_question_id="Q_TGT_1",
            filters_applied=(FilterSpec("Q_SEG_1"),),
            dispatch_mode="cross_cut_breakdown",
            single_cut_result=None,
            cross_cut_result=cross_results[0],
            filtered_n=30,
            audit_records=cross_results[0].audit_records,
        ),
    ]
    return filtered_results, schema, log


def sheet_values(workbook_path: Path, sheet_name: str) -> list[object]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return [
            cell
            for row in workbook[sheet_name].iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]
    finally:
        workbook.close()


class TestExcelExporter(unittest.TestCase):
    def export_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        export_single_cuts(results, skips, schema, quality_report, log, str(output_path))
        return output_path

    def export_workbook_with_cross_cuts(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        (
            results,
            skips,
            schema,
            quality_report,
            log,
            cross_results,
            cross_skips,
        ) = make_cross_cut_export_fixture()
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            cross_cut_results=cross_results,
            cross_cut_skips=cross_skips,
        )
        return output_path

    def export_cross_cut_workbook(self, cross_results: list) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        (
            _results,
            _skips,
            schema,
            _quality_report,
            log,
            _cross_results,
            _cross_skips,
        ) = make_cross_cut_export_fixture()
        export_cross_cuts_only(cross_results, schema, log, str(output_path))
        return output_path

    def export_filtered_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        filtered_results, schema, log = make_filtered_export_fixture()
        export_filtered_single_cuts(filtered_results, schema, log, str(output_path))
        return output_path

    def test_export_creates_file_at_path(self) -> None:
        output_path = self.export_workbook()

        self.assertTrue(output_path.exists())
        self.assertGreater(output_path.stat().st_size, 0)

    def test_export_writes_all_required_sheets(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(
            workbook.sheetnames[:4],
            ["Run_Summary", "Question_Metadata", "Single_Cut_Index", "Skip_Log"],
        )
        self.assertEqual(
            workbook.sheetnames[-4:],
            ["Calculation_Log", "Filter_Log", "Data_Quality", "Warnings"],
        )
        self.assertIn("SC_Q_SS_EXPORT", workbook.sheetnames)
        self.assertIn("SC_Q_MS_EXPORT", workbook.sheetnames)
        self.assertIn("SC_Q_NUM_EXPORT", workbook.sheetnames)
        self.assertIn("SC_Q_GRID_EXPORT", workbook.sheetnames)

    def test_run_summary_contains_correct_totals(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Run_Summary"]

        self.assertEqual(ws["B4"].value, 10)
        self.assertEqual(ws["B5"].value, 6)
        self.assertEqual(ws["B6"].value, 5)
        self.assertEqual(ws["B7"].value, 2)
        self.assertEqual(ws["B8"].value, 0)
        self.assertEqual(ws["B9"].value, 0)
        self.assertEqual(ws["B10"].value, 0)
        self.assertEqual(ws["B11"].value, 7)
        self.assertEqual(ws["B12"].value, 1)
        self.assertEqual(ws["B13"].value, 1)

    def test_question_metadata_one_row_per_question(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(workbook["Question_Metadata"].max_row, 7)

    def test_single_cut_index_links_to_sc_sheets(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Single_Cut_Index"]

        self.assertEqual(ws["D2"].value, "SC_Q_SS_EXPORT")
        self.assertIsNotNone(ws["D2"].hyperlink)
        self.assertIn("SC_Q_SS_EXPORT", workbook.sheetnames)

    def test_sc_sheet_for_single_select_has_distribution_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["SC_Q_SS_EXPORT"]

        self.assertEqual([ws["A7"].value, ws["B7"].value, ws["C7"].value, ws["D7"].value], ["Code", "Label", "Count", "%"])
        self.assertEqual(ws["A8"].value, 1)
        self.assertEqual(ws["B8"].value, "Yes")
        self.assertEqual(ws["C8"].value, 6)
        self.assertEqual(ws["D8"].value, 0.6)

    def test_sc_sheet_for_multi_select_has_selections_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["SC_Q_MS_EXPORT"]

        self.assertEqual(ws["A7"].value, "Sub-Column ID")
        self.assertEqual(ws["C7"].value, "Count selected")
        self.assertEqual(ws["D7"].value, "Selection %")
        self.assertEqual(ws["A8"].value, "Q_MS_EXPORTr1")
        self.assertEqual(ws["D8"].value, 0.5)
        self.assertEqual(ws["A11"].value, "Respondents who answered any: 10")

    def test_sc_sheet_for_numeric_has_stats_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["SC_Q_NUM_EXPORT"]

        self.assertEqual(ws["A7"].value, "Mean")
        self.assertEqual(ws["B7"].value, 5.5)
        self.assertEqual(ws["A8"].value, "Median")
        self.assertEqual(ws["B8"].value, 5.5)

    def test_sc_sheet_for_grid_has_per_row_blocks(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["SC_Q_GRID_EXPORT"]

        self.assertEqual(ws["A7"].value, "Per-row distributions")
        self.assertEqual(ws["A9"].value, "Grid first row (n=10)")
        self.assertEqual(ws["A10"].value, "Code")
        self.assertEqual(ws["A15"].value, "Grid second row (n=10)")

    def test_calculation_log_one_row_per_audit_record(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(workbook["Calculation_Log"].max_row, 8)

    def test_skip_log_one_row_per_skip(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(workbook["Skip_Log"].max_row, 3)

    def test_data_quality_sheet_has_three_sections(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        values = [
            cell
            for row in workbook["Data_Quality"].iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]

        self.assertIn("Per-column missing %", values)
        self.assertIn("Per-column out-of-range %", values)
        self.assertIn("Coercion log", values)

    def test_sheet_names_truncated_to_31_chars(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertTrue(all(len(sheet_name) <= 31 for sheet_name in workbook.sheetnames))
        long_sheets = [
            sheet_name
            for sheet_name in workbook.sheetnames
            if sheet_name.startswith("SC_Q_VERY_LONG")
        ]
        self.assertEqual(len(long_sheets), 1)
        self.assertEqual(len(long_sheets[0]), 31)

    def test_percentages_stored_as_raw_floats(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        value = workbook["SC_Q_SS_EXPORT"]["D8"].value

        self.assertEqual(value, 0.6)
        self.assertLessEqual(value, 1.0)

    def test_workbook_opens_without_corruption(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Run_Summary", workbook.sheetnames)

    def test_export_with_cross_cuts_writes_cross_cut_index(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Cross_Cut_Index"]

        self.assertEqual(ws["A1"].value, "Cross Cut ID")
        self.assertEqual(ws["A2"].value, "CC_TAB_EXPORT")
        self.assertEqual(ws["C2"].value, "CROSS_TAB")

    def test_export_with_cross_cuts_writes_cc_sheets(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("CC_CC_TAB_EXPORT", workbook.sheetnames)
        self.assertIn("CC_CC_SEG_EXPORT", workbook.sheetnames)
        self.assertIn("CC_CC_GROUP_EXPORT", workbook.sheetnames)
        self.assertIn("CC_CC_EVR_EXPORT", workbook.sheetnames)

    def test_export_writes_cross_tab_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_TAB_EXPORT"]

        self.assertEqual(ws["A13"].value, "Counts")
        self.assertEqual(ws["C14"].value, 1)
        self.assertEqual(ws["C16"].value, 5)
        self.assertEqual(ws["A20"].value, "Row %")

    def test_export_writes_segment_profile_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_SEG_EXPORT"]

        self.assertEqual(ws["A8"].value, "Filter applied:")
        self.assertEqual(ws["A10"].value, "= Segment 1")
        self.assertEqual(ws["A14"].value, "Target distribution")
        self.assertEqual(ws["A15"].value, "Code")

    def test_export_writes_group_comparison_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_GROUP_EXPORT"]

        self.assertEqual(ws["A8"].value, "Segments (rows):")
        self.assertEqual(ws["A13"].value, "Per-segment comparison")
        self.assertEqual(ws["A14"].value, "Segment")
        self.assertEqual(ws["D15"].value, 55)
        self.assertEqual(ws["A18"].value, "Overall")

    def test_export_writes_expected_vs_realized_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_EVR_EXPORT"]

        self.assertEqual(ws["A8"].value, "Expected:")
        self.assertEqual(ws["A13"].value, "Expected vs Realized")
        self.assertEqual(ws["A15"].value, "Mean")
        self.assertEqual(ws["D15"].value, -5)
        self.assertEqual(ws["A20"].value, "Paired N: 12")

    def test_export_filter_log_populated_for_segment_profile(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filter_Log"]

        self.assertEqual(ws["A1"].value, "Cross Cut ID")
        self.assertEqual(ws["A2"].value, "CC_SEG_EXPORT")
        self.assertEqual(ws["C2"].value, "Q_SEG_1 == 1")
        self.assertEqual(ws["E2"].value, "CC_CC_SEG_EXPORT")

    def test_export_skip_log_includes_cross_cut_skips(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        rows = list(workbook["Skip_Log"].iter_rows(values_only=True))

        self.assertEqual(rows[0][0], "Source")
        self.assertIn(
            ("cross_cut", "CC_BAD_EXPORT", "UNKNOWN", "cross_cut_error",
             "ValueError: synthetic cross cut failure"),
            rows,
        )

    def test_export_run_summary_has_cross_cut_totals(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Run_Summary"]

        self.assertEqual(ws["B8"].value, 4)
        self.assertEqual(ws["B9"].value, 1)
        self.assertEqual(ws["B10"].value, 1)

    def test_cross_tab_sheet_includes_axis_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_TAB_EXPORT"]

        self.assertEqual(ws["A8"].value, "Rows (vertical): Q_SEG_1")
        self.assertEqual(ws["A9"].value, "Segment")
        self.assertEqual(ws["A10"].value, "Columns (horizontal): Q_TGT_1")
        self.assertEqual(ws["A11"].value, "Target categorical")

    def test_segment_profile_sheet_includes_filter_and_target_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_SEG_EXPORT"]

        self.assertEqual(ws["A8"].value, "Filter applied:")
        self.assertEqual(ws["A9"].value, "Q_SEG_1: Segment")
        self.assertEqual(ws["A10"].value, "= Segment 1")
        self.assertEqual(ws["A11"].value, "Target question:")
        self.assertEqual(ws["A12"].value, "Q_TGT_1: Target categorical")

    def test_group_comparison_sheet_includes_segment_and_metric_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_GROUP_EXPORT"]

        self.assertEqual(ws["A8"].value, "Segments (rows):")
        self.assertEqual(ws["A9"].value, "Q_SEG_1: Segment")
        self.assertEqual(ws["A10"].value, "Metric (columns):")
        self.assertEqual(ws["A11"].value, "Q_NUM_3: Numeric metric")

    def test_expected_vs_realized_sheet_includes_both_question_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_EVR_EXPORT"]

        self.assertEqual(ws["A8"].value, "Expected:")
        self.assertEqual(ws["A9"].value, "Q_EXP_1: Expected")
        self.assertEqual(ws["A10"].value, "Realized:")
        self.assertEqual(ws["A11"].value, "Q_REAL_1: Realized")

    def test_cross_tab_sheet_has_corner_orientation_label(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_TAB_EXPORT"]

        self.assertEqual(ws["A14"].value, "↓ Q_SEG_1  →  Q_TGT_1")
        self.assertEqual(ws["A21"].value, "↓ Q_SEG_1  →  Q_TGT_1")
        self.assertEqual(ws["A28"].value, "↓ Q_SEG_1  →  Q_TGT_1")

    def test_cross_tab_sheet_renders_only_counts_in_counts_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="counts")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertNotIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_only_row_pct_in_row_pct_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="row_pct")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertNotIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_both_blocks_in_both_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="both")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_all_three_blocks_in_all_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="all")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertIn("Column %", values)

    def test_grid_cross_tab_respects_display_mode(self) -> None:
        result, schema, log = grid_cross_tab_export_result(display_mode="row_pct")
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        export_cross_cuts_only([result], schema, log, str(output_path))

        values = sheet_values(output_path, "CC_CC_GRID_EXPORT")

        self.assertNotIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)
        self.assertIn("Q_GRID_SEGr1", values)
        self.assertIn("Segment 1", values)

    def test_cross_tab_includes_copy_friendly_block(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Copy-friendly", values)
        self.assertIn("Row Code", values)
        self.assertIn("Row Label", values)
        self.assertIn("Column Code", values)
        self.assertIn("Column Label", values)
        self.assertIn("Count", values)

    def test_segment_profile_sheet_respects_display_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[1], display_mode="counts")]
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_SEG_EXPORT"]

        self.assertEqual(ws["A15"].value, "Code")
        self.assertEqual(ws["C15"].value, "Count")
        self.assertIsNone(ws["D15"].value)

    def test_group_comparison_sheet_ignores_display_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[2], display_mode="counts")]
        )
        values = sheet_values(output_path, "CC_CC_GROUP_EXPORT")

        self.assertIn("Per-segment comparison", values)
        self.assertIn("Mean", values)
        self.assertIn("Median", values)
        self.assertIn("Std", values)

    def test_export_cross_cuts_only_writes_selected_cc_sheets(self) -> None:
        (
            _results,
            _skips,
            schema,
            _quality_report,
            log,
            cross_results,
            _cross_skips,
        ) = make_cross_cut_export_fixture()
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        export_cross_cuts_only(cross_results[:2], schema, log, str(output_path))
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Cross_Cut_Index", workbook.sheetnames)
        self.assertIn("CC_CC_TAB_EXPORT", workbook.sheetnames)
        self.assertIn("CC_CC_SEG_EXPORT", workbook.sheetnames)
        self.assertNotIn("CC_CC_GROUP_EXPORT", workbook.sheetnames)

    def test_export_backward_compatible_without_cross_cuts(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertNotIn("Cross_Cut_Index", workbook.sheetnames)
        self.assertFalse(any(name.startswith("CC_") for name in workbook.sheetnames))

    def test_export_filtered_single_cuts_creates_file(self) -> None:
        output_path = self.export_filtered_workbook()

        self.assertTrue(output_path.exists())
        self.assertGreater(output_path.stat().st_size, 0)

    def test_filtered_cut_index_lists_all_results(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filtered_Cut_Index"]

        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws["A1"].value, "Sheet Name")
        self.assertEqual(ws["B2"].value, "Q_SS_EXPORT")
        self.assertEqual(ws["C2"].value, "Q_SEG_1 == 1 (Segment 1)")
        self.assertEqual(ws["D4"].value, "cross_cut_breakdown")
        self.assertEqual(ws["E4"].value, 30)

    def test_fsc_sheet_for_single_cut_filtered_renders_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["FSC_Q_SS_EXPORT_01"]
        values = sheet_values(output_path, "FSC_Q_SS_EXPORT_01")

        self.assertEqual(ws["A1"].value, "Target question:")
        self.assertEqual(ws["A4"].value, "Q_SS_EXPORT")
        self.assertIn("  Q_SEG_1 == 1 (Segment 1)", values)
        self.assertIn("Code", values)
        self.assertIn("Yes", values)
        self.assertIn(6, values)

    def test_fsc_sheet_for_cross_cut_breakdown_renders_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        values = sheet_values(output_path, "FSC_Q_TGT_1")

        self.assertIn("  Q_SEG_1 (breakdown - no specific value)", values)
        self.assertIn("Rows (vertical): Q_SEG_1", values)
        self.assertIn("Columns (horizontal): Q_TGT_1", values)
        self.assertIn("Counts", values)

    def test_fsc_sheet_naming_handles_duplicate_targets(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("FSC_Q_SS_EXPORT_01", workbook.sheetnames)
        self.assertIn("FSC_Q_SS_EXPORT_02", workbook.sheetnames)
        self.assertIn("FSC_Q_TGT_1", workbook.sheetnames)

    def test_export_filtered_single_cuts_filter_log_populated(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        rows = list(workbook["Filter_Log"].iter_rows(values_only=True))

        self.assertEqual(rows[0], (
            "Sheet Name",
            "Filter Question",
            "Filter Value",
            "Filter Description",
        ))
        self.assertIn(
            (
                "FSC_Q_SS_EXPORT_01",
                "Q_SEG_1",
                1,
                "Q_SEG_1 == 1 (Segment 1)",
            ),
            rows,
        )
        self.assertIn(
            (
                "FSC_Q_TGT_1",
                "Q_SEG_1",
                None,
                "Q_SEG_1 (breakdown - no specific value)",
            ),
            rows,
        )

    def test_filtered_export_calculation_log_filtered_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        values = sheet_values(output_path, "Calculation_Log")

        self.assertIn("Q_SS_EXPORT", values)
        self.assertIn("CC_CC_TAB_EXPORT", values)
        self.assertNotIn("Q_MS_EXPORT", values)
        self.assertNotIn("Q_UNRELATED_EXPORT", values)


if __name__ == "__main__":
    unittest.main()
