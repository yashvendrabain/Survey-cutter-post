"""Tests for CompactTwoColumnAdapter raw-data sanity check."""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook

from src.adapters.compact_two_column import (
    CompactTwoColumnAdapter,
    RAW_COL_MATCH_THRESHOLD,
    _qid_match_ratio,
    _sample_qids_from_compact_sheet,
)


def _build_workbook(
    raw_sheet_name: str,
    raw_columns: list[str],
    codebook_sheet_name: str,
    codebook_blocks: list[tuple[str, str, list[tuple[int, str]]]],
) -> Workbook:
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = raw_sheet_name
    ws_raw.append(raw_columns)
    ws_raw.append([None] * len(raw_columns))  # one filler row

    ws_cb = wb.create_sheet(codebook_sheet_name)
    for qid, qtext, options in codebook_blocks:
        ws_cb.append([qid, qtext])
        for code, label in options:
            ws_cb.append([code, label])
        ws_cb.append([None, None])
    return wb


# -----------------------------------------------------------------------------
# Whole-adapter behavior
# -----------------------------------------------------------------------------

def test_alien_format_with_compact_codebook_returns_zero_confidence():
    """The new test fixture: compact codebook but raw cols are 1q / 6q1a."""

    wb = _build_workbook(
        raw_sheet_name="responses",
        raw_columns=["id", "status", "1q", "2q", "3q", "6q1a", "6q2a", "6q3a"],
        codebook_sheet_name="codebook",
        codebook_blocks=[
            ("q1", "Revenue?", [(1, "Low"), (2, "High")]),
            ("q2", "AI maturity?", [(1, "None"), (2, "Some")]),
            ("q3", "Region?", [(1, "AMS"), (2, "EMEA")]),
            ("q6", "Channels?", [(1, "Field"), (2, "Inside"), (3, "Partner")]),
        ],
    )
    raw_df = pd.DataFrame(columns=["id", "status", "1q", "2q", "3q", "6q1a", "6q2a", "6q3a"])

    result = CompactTwoColumnAdapter().detect(wb, raw_df)
    assert result.confidence == 0.0
    assert "unknown survey format" in result.reason
    assert result.is_certain is False


def test_winvslag_style_compact_codebook_keeps_high_confidence():
    """winvslag2024: codebook QIDs match raw cols via exact or colon-prefix."""

    wb = _build_workbook(
        raw_sheet_name="Raw Data",
        raw_columns=[
            "Respondent", "status", "Q1", "Q2", "Q3",
            "Q6: Field sales", "Q6: Inside sales", "Q6: Partner",
        ],
        codebook_sheet_name="Data map",
        codebook_blocks=[
            ("Q1", "Revenue?", [(1, "Low"), (2, "High")]),
            ("Q2", "AI maturity?", [(1, "None"), (2, "Some")]),
            ("Q3", "Region?", [(1, "AMS"), (2, "EMEA")]),
            ("Q6", "Channels?", [(1, "Field"), (2, "Inside"), (3, "Partner")]),
        ],
    )
    raw_df = pd.DataFrame(columns=[
        "Respondent", "status", "Q1", "Q2", "Q3",
        "Q6: Field sales", "Q6: Inside sales", "Q6: Partner",
    ])

    result = CompactTwoColumnAdapter().detect(wb, raw_df)
    assert result.confidence == 0.9


def test_bcn_style_raw_columns_still_match_compact_codebook():
    """Raw cols Q1, Q3r1, Q3r2 — Q1 exact, Q3 matched via 'r' suffix.
    Compact must NOT falsely abdicate; BCN will out-score it elsewhere."""

    wb = _build_workbook(
        raw_sheet_name="Raw",
        raw_columns=["id", "Q1", "Q2", "Q3r1", "Q3r2", "Q3r3"],
        codebook_sheet_name="codebook",
        codebook_blocks=[
            ("Q1", "Revenue?", [(1, "Low"), (2, "High")]),
            ("Q2", "AI maturity?", [(1, "None"), (2, "Some")]),
            ("Q3", "Channels?", [(1, "Field"), (2, "Inside"), (3, "Partner")]),
        ],
    )
    raw_df = pd.DataFrame(columns=["id", "Q1", "Q2", "Q3r1", "Q3r2", "Q3r3"])

    result = CompactTwoColumnAdapter().detect(wb, raw_df)
    assert result.confidence == 0.9


def test_no_raw_df_preserves_compact_confidence():
    """Graceful degradation: if raw_df is None, historical behavior holds."""

    wb = _build_workbook(
        raw_sheet_name="responses",
        raw_columns=["id", "1q", "6q1a"],
        codebook_sheet_name="codebook",
        codebook_blocks=[
            ("q1", "Revenue?", [(1, "Low"), (2, "High")]),
            ("q2", "AI maturity?", [(1, "None"), (2, "Some")]),
        ],
    )
    result = CompactTwoColumnAdapter().detect(wb, raw_df=None)
    assert result.confidence == 0.9


def test_no_datamap_sheet_returns_zero():
    wb = Workbook()
    wb.active.title = "responses"
    wb.active.append(["id", "q1"])
    result = CompactTwoColumnAdapter().detect(wb, pd.DataFrame(columns=["id", "q1"]))
    assert result.confidence == 0.0
    assert "no data map sheet" in result.reason


# -----------------------------------------------------------------------------
# Helper unit tests
# -----------------------------------------------------------------------------

def test_qid_match_ratio_exact_and_prefix_matches():
    qids = ["Q1", "Q2", "Q6"]
    raw = ["Q1", "Q2", "Q6: Field sales", "Q6: Inside sales"]
    assert _qid_match_ratio(qids, raw) == 1.0


def test_qid_match_ratio_alien_cols_return_zero():
    qids = ["q1", "q2", "q3", "q6"]
    raw = ["id", "status", "1q", "2q", "6q1a", "6q2a"]
    assert _qid_match_ratio(qids, raw) == 0.0


def test_qid_match_ratio_rejects_q1_matching_q10():
    """Suffix requirement prevents 'q1' from spuriously matching 'q10'."""
    assert _qid_match_ratio(["q1"], ["q10", "q11"]) == 0.0


def test_qid_match_ratio_empty_qids_returns_zero():
    assert _qid_match_ratio([], ["Q1", "Q2"]) == 0.0


def test_qid_match_ratio_partial_match_is_fractional():
    qids = ["Q1", "Q2", "Q3", "Q4"]
    raw = ["Q1", "Q3"]
    assert _qid_match_ratio(qids, raw) == 0.5


def test_sample_qids_skips_option_rows_and_blanks():
    wb = Workbook()
    ws = wb.active
    ws.append(["status", "Participant status"])
    ws.append([1, "complete"])
    ws.append([2, "partial"])
    ws.append([None, None])
    ws.append(["q1", "Revenue?"])
    ws.append([1, "Low"])
    ws.append([2, "High"])
    ws.append([None, None])
    ws.append(["q2", "Maturity?"])
    ws.append([1, "None"])
    ws.append([2, "Some"])

    sampled = _sample_qids_from_compact_sheet(ws, limit=20)
    assert sampled == ["status", "q1", "q2"]


def test_sample_qids_respects_limit():
    wb = Workbook()
    ws = wb.active
    for i in range(1, 30):
        ws.append([f"q{i}", f"Question {i}"])
        ws.append([1, "Yes"])
        ws.append([2, "No"])
        ws.append([None, None])
    sampled = _sample_qids_from_compact_sheet(ws, limit=5)
    assert len(sampled) == 5
    assert sampled[0] == "q1"
    assert sampled[-1] == "q5"


def test_threshold_constant_is_30_percent():
    """Lock the threshold so future tuning is a deliberate change."""
    assert RAW_COL_MATCH_THRESHOLD == 0.30
