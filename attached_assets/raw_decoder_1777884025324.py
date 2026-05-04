"""Raw survey data decoder for the Survey Insight Engine."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

from src.datamap_parser import DataMap, ParsedQuestion
from src.models import DataQualityReport

try:
    from config import MISSING_VALUE_TOKENS, HIGH_MISSING_THRESHOLD
except ModuleNotFoundError as exc:
    if exc.name != "config":
        raise
    MISSING_VALUE_TOKENS = {"", "NA", "N/A", "NULL", "null", "None", "nan"}
    HIGH_MISSING_THRESHOLD = 0.5


RESPONDENT_ID_CANDIDATES = (
    "record",
    "uuid",
    "respondent_id",
    "id",
    "ID",
    "RespondentID",
)


def decode_raw_data(
    path: str,
    data_map: DataMap,
) -> tuple[pd.DataFrame, DataQualityReport]:
    """Load raw survey data and return a decoded DataFrame plus quality report."""

    expected_columns = _expected_columns_from_datamap(data_map)
    dataframe = _load_raw_file(path, expected_columns)
    dataframe = _strip_string_values(dataframe)
    dataframe = _replace_missing_tokens(dataframe)

    original_columns = tuple(str(column) for column in dataframe.columns)
    respondent_id_column = _find_respondent_id_column(dataframe.columns.tolist())
    warnings: list[str] = []
    generated_respondent_id = False
    sheet_warning = dataframe.attrs.get("_raw_decoder_sheet_warning")
    if sheet_warning:
        warnings.append(str(sheet_warning))

    if respondent_id_column is None:
        respondent_id_column = "respondent_id"
        dataframe[respondent_id_column] = range(1, len(dataframe) + 1)
        generated_respondent_id = True
        warnings.append("no respondent ID column found; generated sequential IDs")

    columns_in_datamap = len(set(original_columns).intersection(expected_columns))
    columns_not_in_datamap = tuple(
        column for column in original_columns if column not in expected_columns
    )
    value_ranges = _value_ranges_by_column(data_map)
    no_numeric_coercion_columns = _string_preserved_columns(data_map)

    coercion_log: list[dict] = []
    for column in dataframe.columns:
        if column == respondent_id_column:
            continue
        if column in no_numeric_coercion_columns:
            continue
        dataframe[column] = _coerce_column_to_numeric(
            dataframe[column], str(column), coercion_log
        )

    missing_pct = {
        str(column): float(dataframe[column].isna().mean())
        for column in dataframe.columns
    }
    out_of_range_pct = {
        str(column): _out_of_range_pct(dataframe[column], value_ranges.get(str(column)))
        for column in dataframe.columns
    }

    for column, pct in missing_pct.items():
        if pct > HIGH_MISSING_THRESHOLD:
            warnings.append(f"column {column} has {pct:.1%} missing values")
    for column, pct in out_of_range_pct.items():
        if pct > 0.05:
            warnings.append(f"column {column} has {pct:.1%} out-of-range values")

    report = DataQualityReport(
        total_rows=int(len(dataframe)),
        total_columns=int(len(dataframe.columns)),
        columns_in_datamap=int(columns_in_datamap),
        columns_not_in_datamap=columns_not_in_datamap,
        per_column_missing_pct=missing_pct,
        per_column_out_of_range_pct=out_of_range_pct,
        coercion_log=tuple(coercion_log),
        warnings=tuple(warnings),
    )

    if generated_respondent_id:
        dataframe[respondent_id_column] = dataframe[respondent_id_column].astype("Int64")

    return dataframe, report


def _load_raw_file(
    path: str,
    expected_columns: set[str] | None = None,
) -> pd.DataFrame:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, dtype=str)
    if suffix == ".xlsx":
        sheet_name: int | str = 0
        first_sheet_name = _first_sheet_name(path)
        fallback_warning = None
        if expected_columns:
            detected, fallback_warning = _find_data_sheet(path, expected_columns)
            sheet_name = detected
        dataframe = pd.read_excel(
            path,
            dtype=str,
            engine="openpyxl",
            sheet_name=sheet_name,
        )
        if fallback_warning:
            dataframe.attrs["_raw_decoder_sheet_warning"] = fallback_warning
        elif isinstance(sheet_name, str) and sheet_name != first_sheet_name:
            dataframe.attrs["_raw_decoder_sheet_warning"] = (
                f"data loaded from sheet '{sheet_name}' "
                "(not the first sheet)"
            )
        return dataframe
    raise ValueError(f"unsupported raw data file extension: {suffix or '<none>'}")


def _find_data_sheet(path: str, expected_columns: set[str]) -> tuple[str, str | None]:
    """Return the sheet name whose headers best match expected_columns."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) == 1:
            return workbook.sheetnames[0], None

        first_sheet = workbook.sheetnames[0]
        best_sheet = None
        best_score = 0
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            first_row = next(
                worksheet.iter_rows(max_row=1, values_only=True),
                (),
            )
            headers = {
                str(cell).strip()
                for cell in first_row
                if cell is not None
            }
            score = len(headers.intersection(expected_columns))
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
        if best_score == 0:
            return (
                first_sheet,
                "no sheet matched expected columns; loaded first sheet "
                f"'{first_sheet}' as fallback",
            )
        return str(best_sheet), None
    finally:
        workbook.close()


def _first_sheet_name(path: str) -> str | None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames[0] if workbook.sheetnames else None
    finally:
        workbook.close()


def _strip_string_values(dataframe: pd.DataFrame) -> pd.DataFrame:
    stripped = dataframe.copy()
    for column in stripped.columns:
        stripped[column] = stripped[column].map(
            lambda value: value.strip() if isinstance(value, str) else value
        )
    return stripped


def _replace_missing_tokens(dataframe: pd.DataFrame) -> pd.DataFrame:
    replaced = dataframe.copy()
    for column in replaced.columns:
        replaced[column] = replaced[column].map(_replace_missing_value)
    return replaced


def _replace_missing_value(value: Any) -> Any:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, str) and value in MISSING_VALUE_TOKENS:
        return pd.NA
    return value


def _find_respondent_id_column(columns: list[str]) -> str | None:
    for candidate in RESPONDENT_ID_CANDIDATES:
        if candidate in columns:
            return candidate
    return None


def _expected_columns_from_datamap(data_map: DataMap) -> set[str]:
    expected: set[str] = set()
    for question in data_map["questions"]:
        expected.add(question["canonical_id"])
        sub_columns = question["sub_columns"]
        if sub_columns:
            expected.update(sub_column_id for sub_column_id, _ in sub_columns)
    return expected


def _value_ranges_by_column(data_map: DataMap) -> dict[str, tuple[int, int]]:
    value_ranges: dict[str, tuple[int, int]] = {}
    for question in data_map["questions"]:
        value_range = question["value_range"]
        if value_range is None:
            continue
        for column in _question_expected_columns(question):
            value_ranges[column] = value_range
    return value_ranges


def _string_preserved_columns(data_map: DataMap) -> set[str]:
    preserved: set[str] = set()
    for question in data_map["questions"]:
        if question["type_hint"] in {"open_text", "open_numeric"}:
            preserved.update(_question_expected_columns(question))
    return preserved


def _question_expected_columns(question: ParsedQuestion) -> tuple[str, ...]:
    if question["sub_columns"]:
        return tuple(sub_column_id for sub_column_id, _ in question["sub_columns"])
    return (question["canonical_id"],)


def _coerce_column_to_numeric(
    series: pd.Series, column_name: str, coercion_log: list[dict]
) -> pd.Series:
    original = series.copy()
    numeric_series = pd.to_numeric(original, errors="coerce")
    affected_mask = original.notna() & numeric_series.isna()
    affected_count = int(affected_mask.sum())

    if affected_count > 0:
        values_coerced = sorted(
            {str(value) for value in original[affected_mask].dropna().tolist()}
        )
        coercion_log.append(
            {
                "column": column_name,
                "from_type": "string",
                "to_type": "numeric",
                "values_coerced": values_coerced,
                "rows_affected": affected_count,
            }
        )

    if int(numeric_series.notna().sum()) == 0:
        return original
    return numeric_series


def _out_of_range_pct(
    series: pd.Series, value_range: tuple[int, int] | None
) -> float:
    if value_range is None or not pd.api.types.is_numeric_dtype(series):
        return 0.0

    valid_values = series.dropna()
    valid_n = int(len(valid_values))
    if valid_n == 0:
        return 0.0

    low, high = value_range
    out_of_range_count = int((~valid_values.between(low, high)).sum())
    return float(out_of_range_count / valid_n)
