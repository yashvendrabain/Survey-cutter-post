"""Tests for the compact two-column data-map adapter."""

from __future__ import annotations

import unittest

from openpyxl import Workbook, load_workbook

from src.adapters.compact_two_column import CompactTwoColumnAdapter
from tests.conftest import DATAMAP_FIXTURE_PATH


def make_compact_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["Q1", "What is your current employment status?"])
    worksheet.append([1, "Full-time"])
    worksheet.append([2, "Part-time"])
    worksheet.append(["a", "Alphabetic option"])
    worksheet.append([None, None])
    worksheet.append(["Q2", "What was your most recent industry?"])
    worksheet.append([1, "Technology"])
    worksheet.append([2, "Retail"])
    return workbook


class TestCompactTwoColumnAdapter(unittest.TestCase):
    def test_detects_compact_workbook(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook())

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("compact", result.reason)

    def test_low_confidence_for_bcn_fixture(self) -> None:
        workbook = load_workbook(DATAMAP_FIXTURE_PATH, read_only=True, data_only=True)
        try:
            result = CompactTwoColumnAdapter().detect(workbook)
        finally:
            workbook.close()

        self.assertEqual(result.confidence, 0.0)

    def test_parse_compact_workbook_preserves_codes_and_options(self) -> None:
        workbook = make_compact_workbook()
        setattr(workbook, "_survey_source_path", "compact.xlsx")

        parsed = CompactTwoColumnAdapter().parse(workbook)

        questions = {question["canonical_id"]: question for question in parsed["questions"]}
        self.assertEqual(parsed["source_path"], "compact.xlsx")
        self.assertEqual(
            questions["Q1"]["options"],
            [(1, "Full-time"), (2, "Part-time"), ("a", "Alphabetic option")],
        )
        self.assertEqual(questions["Q2"]["question_text"], "What was your most recent industry?")


if __name__ == "__main__":
    unittest.main()
