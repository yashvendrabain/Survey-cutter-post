"""Tests for example-driven wizard parsing (integer-key linking)."""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import Workbook

from src.adapters.wizard_configured import (
    WizardConfig,
    WizardConfiguredAdapter,
    infer_multi_select_spec,
    infer_question_id_regex,
    _parse_questions_by_example,
)


def _cfg(qid_example: str, multi_example: str, **over) -> WizardConfig:
    base = dict(
        raw_data_sheet_name="responses",
        data_map_sheet_name="codebook",
        respondent_id_column="id",
        question_id_pattern="",
        sub_column_separator="",
        option_code_position="column_b",
        section_prefixes=(),
        question_id_example=qid_example,
        multi_select_example=multi_example,
    )
    base.update(over)
    return WizardConfig(**base)


def _codebook(blocks) -> "Workbook":
    wb = Workbook()
    ws = wb.active
    ws.title = "codebook"
    for block in blocks:
        for row in block:
            ws.append(list(row))
    return wb


# --- inference ---

def test_infer_question_id_regex_forms():
    assert infer_question_id_regex("q1").match("q17").group("key") == "17"
    assert infer_question_id_regex("1q").match("11q").group("key") == "11"
    assert infer_question_id_regex("q-1").match("q-9").group("key") == "9"
    assert infer_question_id_regex("Q15").match("Q3").group("key") == "3"
    assert infer_question_id_regex("Item_3").match("Item_42").group("key") == "42"
    # single-answer form must NOT match a longer multi-answer column
    assert infer_question_id_regex("1q").match("6q1a") is None


def test_infer_question_id_regex_no_integer_raises():
    with pytest.raises(ValueError):
        infer_question_id_regex("status")


def test_infer_multi_index_mode():
    for ex, probe, key, opt in [
        ("6q1a", "14q3a", "14", "3"),
        ("Q6r1", "Q6r4", "6", "4"),
        ("q6_1", "q6_2", "6", "2"),
    ]:
        mode, multi, _single = infer_multi_select_spec(ex)
        assert mode == "index"
        m = multi.match(probe)
        assert m.group("key") == key and m.group("opt") == opt


def test_infer_multi_label_mode():
    mode, multi, single = infer_multi_select_spec("Q6: Field sales")
    assert mode == "label"
    m = multi.match("Q6: Inside sales")
    assert m.group("key") == "6" and m.group("label") == "Inside sales"
    assert single.match("Q6").group("key") == "6"


def test_infer_multi_no_integer_raises():
    with pytest.raises(ValueError):
        infer_multi_select_spec("channels")


# --- integer-key parsing (the alien format) ---

def test_parse_by_example_alien_index_mode():
    wb = _codebook([
        [("q1", "Revenue?"), (1, "Low"), (2, "High"), (None, None)],
        [("q6", "Channels?"), (1, "Field sales"), (2, "Inside sales"),
         (3, "Partner"), (None, None)],
        # grid with a SECOND code block that must be ignored for labels:
        [("q14", "Region motion?"), (1, "Americas"), (2, "EMEA"), (3, "APAC"),
         (None, "(cell values for the columns above)"),
         (1, "Direct"), (2, "Channel"), (3, "Digital"), (None, None)],
        [("q10", "Pct new revenue?")],  # numeric, no options
    ])
    raw_columns = (
        "id", "status", "1q", "6q1a", "6q2a", "6q3a",
        "14q1a", "14q2a", "14q3a", "10q",
    )
    qs = {q["canonical_id"]: q
          for q in _parse_questions_by_example(wb["codebook"], raw_columns,
                                               _cfg("q1", "6q1a"))}

    assert qs["q1"]["raw_id"] == "1q"
    assert qs["q1"]["value_range"] == (1, 2)

    assert qs["q6"]["sub_columns"] == [
        ("6q1a", "Field sales"), ("6q2a", "Inside sales"), ("6q3a", "Partner"),
    ]

    # note row closed the option list -> region labels, NOT Direct/Channel/Digital
    assert qs["q14"]["sub_columns"] == [
        ("14q1a", "Americas"), ("14q2a", "EMEA"), ("14q3a", "APAC"),
    ]

    assert qs["q10"]["raw_id"] == "10q"
    assert qs["q10"]["type_hint"] == "open_numeric"
    assert qs["q10"]["value_range"] is None


def test_parse_by_example_label_mode():
    wb = _codebook([
        [("Q6", "Channels?"), (1, "Field sales"), (2, "Inside sales"), (None, None)],
    ])
    raw_columns = ("Respondent", "Q6: Field sales", "Q6: Inside sales")
    qs = {q["canonical_id"]: q
          for q in _parse_questions_by_example(wb["codebook"], raw_columns,
                                               _cfg("Q1", "Q6: Field sales"))}
    assert qs["Q6"]["sub_columns"] == [
        ("Q6: Field sales", "Field sales"),
        ("Q6: Inside sales", "Inside sales"),
    ]


# --- config ---

def test_post_init_derives_pattern_from_example():
    cfg = _cfg("q1", "6q1a")
    assert cfg.question_id_pattern == r"^q(?P<key>\d+)$"
    assert cfg.sub_column_separator == "none"


def test_post_init_example_only_config_is_valid():
    # construction alone must not raise
    _cfg("1q", "6q1a")


# --- adapter end-to-end through parse() branch ---

def test_parse_routes_to_example_engine():
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "responses"
    ws_raw.append(["id", "1q", "6q1a", "6q2a"])
    ws_raw.append(["r1", 1, 1, 0])
    ws_cb = wb.create_sheet("codebook")
    for row in [("q1", "Revenue?"), (1, "Low"), (2, "High"), (None, None),
                ("q6", "Channels?"), (1, "Field sales"), (2, "Inside sales"),
                (None, None)]:
        ws_cb.append(list(row))
    raw_df = pd.DataFrame(columns=["id", "1q", "6q1a", "6q2a"])

    parsed = WizardConfiguredAdapter(_cfg("q1", "6q1a")).parse(wb, raw_df)
    by_id = {q["canonical_id"]: q for q in parsed["questions"]}
    assert by_id["q1"]["raw_id"] == "1q"
    assert by_id["q6"]["sub_columns"] == [
        ("6q1a", "Field sales"), ("6q2a", "Inside sales"),
    ]
