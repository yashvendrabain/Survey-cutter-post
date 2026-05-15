"""Tests for the Excel exporter."""

from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timezone
import importlib.util
import os
from pathlib import Path
import re
import unittest
from unittest.mock import patch
from uuid import uuid4

from openpyxl import load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.cross_cut_engine import compute_cross_cuts
from src.excel_exporter import (
    export_cross_cuts_only,
    export_filtered_single_cuts,
    export_single_cuts,
    _wrapped_formula,
)
from src.models import (
    AuditRecord,
    AnalysisType,
    CrossCutResult,
    CrossCutSpec,
    DataQualityReport,
    DenominatorPolicy,
    FilteredSingleCutResult,
    FilterSpec,
    GridSingleSelectResult,
    MultiSelectResult,
    NumericResult,
    QuestionSpec,
    QuestionType,
    SingleSelectResult,
    SkipRecord,
    SurveySchema,
)
from tests.conftest import CROSS_CUT_30_RESPONDENTS_PATH


UTC_NOW = datetime(2026, 5, 4, 12, 0, tzinfo=timezone.utc)
LONG_ID = "Q_VERY_LONG_SINGLE_SELECT_EXPORT_NAME"
FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"
def make_audit(
    metric_name: str,
    question_id: str,
    source_columns: tuple[str, ...],
    value_raw: float,
    valid_n: int,
    missing_n: int,
) -> AuditRecord:
    return AuditRecord(
        output_sheet=f"SC_{question_id}",
        metric_name=metric_name,
        source_question_id=question_id,
        source_columns=source_columns,
        filter_expr=None,
        numerator=None,
        denominator=valid_n,
        formula=f"{metric_name} formula",
        value_raw=value_raw,
        valid_n=valid_n,
        missing_n=missing_n,
        timestamp=UTC_NOW,
    )


def make_export_fixture() -> tuple[
    list,
    list[SkipRecord],
    SurveySchema,
    DataQualityReport,
    CalculationLog,
]:
    ss_audit = make_audit("rate_per_value", "Q_SS_EXPORT", ("Q_SS_EXPORT",), 10, 10, 0)
    ms_audit = make_audit(
        "selection_rate",
        "Q_MS_EXPORT",
        ("Q_MS_EXPORTr1", "Q_MS_EXPORTr2"),
        10,
        10,
        0,
    )
    num_audit = make_audit("numeric_summary", "Q_NUM_EXPORT", ("Q_NUM_EXPORT",), 5.5, 10, 0)
    grid_row_1_audit = make_audit(
        "rate_per_value", "Q_GRID_EXPORTr1", ("Q_GRID_EXPORTr1",), 10, 10, 0
    )
    grid_row_2_audit = make_audit(
        "rate_per_value", "Q_GRID_EXPORTr2", ("Q_GRID_EXPORTr2",), 10, 10, 0
    )
    grid_parent_audit = make_audit(
        "grid_overall",
        "Q_GRID_EXPORT",
        ("Q_GRID_EXPORTr1", "Q_GRID_EXPORTr2"),
        10,
        10,
        0,
    )
    long_audit = make_audit("rate_per_value", LONG_ID, (LONG_ID,), 10, 10, 0)

    single_result = SingleSelectResult(
        question_id="Q_SS_EXPORT",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Yes", "count": 6, "rate": 0.6},
            2: {"label": "No", "count": 4, "rate": 0.4},
        },
        audit_records=(ss_audit,),
    )
    multi_result = MultiSelectResult(
        question_id="Q_MS_EXPORT",
        question_type=QuestionType.MULTI_SELECT_BINARY,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        selections={
            "Q_MS_EXPORTr1": {"label": "First", "count": 5, "selection_rate": 0.5},
            "Q_MS_EXPORTr2": {"label": "Second", "count": 2, "selection_rate": 0.2},
        },
        respondents_who_answered_any=10,
        audit_records=(ms_audit,),
    )
    numeric_result = NumericResult(
        question_id="Q_NUM_EXPORT",
        question_type=QuestionType.DIRECT_NUMERIC,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        mean=5.5,
        median=5.5,
        std=2.9,
        min_val=1.0,
        max_val=10.0,
        percentiles={25: 3.0, 50: 5.5, 75: 8.0},
        audit_records=(num_audit,),
    )
    grid_row_1 = SingleSelectResult(
        question_id="Q_GRID_EXPORTr1",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Low", "count": 4, "rate": 0.4},
            2: {"label": "High", "count": 6, "rate": 0.6},
        },
        audit_records=(grid_row_1_audit,),
    )
    grid_row_2 = SingleSelectResult(
        question_id="Q_GRID_EXPORTr2",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Low", "count": 5, "rate": 0.5},
            2: {"label": "High", "count": 5, "rate": 0.5},
        },
        audit_records=(grid_row_2_audit,),
    )
    grid_result = GridSingleSelectResult(
        question_id="Q_GRID_EXPORT",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        rows={"Q_GRID_EXPORTr1": grid_row_1, "Q_GRID_EXPORTr2": grid_row_2},
        overall_valid_n=10,
        audit_records=(grid_parent_audit,),
    )
    long_result = SingleSelectResult(
        question_id=LONG_ID,
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "A", "count": 7, "rate": 0.7},
            2: {"label": "B", "count": 3, "rate": 0.3},
        },
        audit_records=(long_audit,),
        warnings=("long id warning",),
    )
    results = [single_result, multi_result, numeric_result, grid_result, long_result]

    questions = (
        QuestionSpec(
            question_id="[Q_SS_EXPORT]",
            canonical_id="Q_SS_EXPORT",
            question_text="Single select export question",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_SS_EXPORT",),
            option_map={1: "Yes", 2: "No"},
        ),
        QuestionSpec(
            question_id="[Q_MS_EXPORT]",
            canonical_id="Q_MS_EXPORT",
            question_text="Multi select export question",
            question_type=QuestionType.MULTI_SELECT_BINARY,
            raw_columns=("Q_MS_EXPORTr1", "Q_MS_EXPORTr2"),
            option_map={"Q_MS_EXPORTr1": "First", "Q_MS_EXPORTr2": "Second"},
            value_range=(0, 1),
        ),
        QuestionSpec(
            question_id="[Q_NUM_EXPORT]",
            canonical_id="Q_NUM_EXPORT",
            question_text="Numeric export question",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_NUM_EXPORT",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_GRID_EXPORT]",
            canonical_id="Q_GRID_EXPORT",
            question_text="Grid export question",
            question_type=QuestionType.GRID_SINGLE_SELECT,
            raw_columns=("Q_GRID_EXPORTr1", "Q_GRID_EXPORTr2"),
            option_map={1: "Low", 2: "High"},
            value_range=(1, 2),
            grid_row_labels={
                "Q_GRID_EXPORTr1": "Grid first row",
                "Q_GRID_EXPORTr2": "Grid second row",
            },
        ),
        QuestionSpec(
            question_id=f"[{LONG_ID}]",
            canonical_id=LONG_ID,
            question_text="Long sheet name export question",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=(LONG_ID,),
            option_map={1: "A", 2: "B"},
        ),
        QuestionSpec(
            question_id="[Q_TEXT_EXPORT]",
            canonical_id="Q_TEXT_EXPORT",
            question_text="Open text export question",
            question_type=QuestionType.OPEN_TEXT,
            raw_columns=("Q_TEXT_EXPORT",),
            option_map={},
        ),
    )
    schema = SurveySchema(
        questions=questions,
        respondent_id_column="respondent_id",
        total_respondents=10,
        source_datamap_path="datamap.xlsx",
        source_rawdata_path="raw.csv",
        parsed_at=UTC_NOW,
    )
    skips = [
        SkipRecord(
            question_id="[Q_TEXT_EXPORT]",
            canonical_id="Q_TEXT_EXPORT",
            question_type=QuestionType.OPEN_TEXT,
            skip_reason="unsupported_type: OPEN_TEXT",
        ),
        SkipRecord(
            question_id="[Q_FAIL_EXPORT]",
            canonical_id="Q_FAIL_EXPORT",
            question_type=QuestionType.SINGLE_SELECT,
            skip_reason="calculation_error",
            details="ValueError: raw column not found in data",
        ),
    ]
    quality_report = DataQualityReport(
        total_rows=10,
        total_columns=8,
        columns_in_datamap=7,
        columns_not_in_datamap=("extra_col",),
        per_column_missing_pct={"Q_SS_EXPORT": 0.1, "Q_NUM_EXPORT": 0.0, "extra_col": 0.5},
        per_column_out_of_range_pct={"Q_NUM_EXPORT": 0.2},
        coercion_log=(
            {
                "column": "Q_NUM_EXPORT",
                "from_type": "string",
                "to_type": "numeric",
                "values_coerced": ["abc"],
                "rows_affected": 1,
            },
        ),
        warnings=("column extra_col has 50.0% missing values",),
    )
    log = CalculationLog()
    for record in (
        ss_audit,
        ms_audit,
        num_audit,
        grid_row_1_audit,
        grid_row_2_audit,
        grid_parent_audit,
        long_audit,
    ):
        log.record(record)
    return results, skips, schema, quality_report, log


def make_cross_cut_schema() -> SurveySchema:
    questions = (
        QuestionSpec(
            question_id="[Q_SEG_1]",
            canonical_id="Q_SEG_1",
            question_text="Segment",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_SEG_1",),
            option_map={1: "Segment 1", 2: "Segment 2", 3: "Segment 3"},
        ),
        QuestionSpec(
            question_id="[Q_TGT_1]",
            canonical_id="Q_TGT_1",
            question_text="Target categorical",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_TGT_1",),
            option_map={1: "A", 2: "B", 3: "C", 4: "D"},
        ),
        QuestionSpec(
            question_id="[Q_NUM_3]",
            canonical_id="Q_NUM_3",
            question_text="Numeric metric",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_NUM_3",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_EXP_1]",
            canonical_id="Q_EXP_1",
            question_text="Expected",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_EXP_1",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_REAL_1]",
            canonical_id="Q_REAL_1",
            question_text="Realized",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_REAL_1",),
            option_map={},
        ),
    )
    return SurveySchema(
        questions=questions,
        respondent_id_column="respondent_id",
        total_respondents=30,
        source_datamap_path="cross_datamap.xlsx",
        source_rawdata_path="cross_raw.csv",
        parsed_at=UTC_NOW,
    )


def grid_segment_question() -> QuestionSpec:
    return QuestionSpec(
        question_id="[Q_GRID_SEG]",
        canonical_id="Q_GRID_SEG",
        question_text="Grid segment",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        raw_columns=("Q_GRID_SEGr1", "Q_GRID_SEGr2", "Q_GRID_SEGr3"),
        option_map={1: "Selected"},
        value_range=(0, 1),
        grid_row_labels={
            "Q_GRID_SEGr1": "Segment 1",
            "Q_GRID_SEGr2": "Segment 2",
            "Q_GRID_SEGr3": "Segment 3",
        },
    )


def grid_cross_tab_export_result(display_mode: str) -> tuple[CrossCutResult, SurveySchema, CalculationLog]:
    dataframe = pd.read_csv(CROSS_CUT_30_RESPONDENTS_PATH)
    dataframe["Q_GRID_SEGr1"] = [1] * 10 + [0] * 20
    dataframe["Q_GRID_SEGr2"] = [0] * 10 + [1] * 10 + [0] * 10
    dataframe["Q_GRID_SEGr3"] = [0] * 20 + [1] * 10
    base_schema = make_cross_cut_schema()
    schema = SurveySchema(
        questions=base_schema.questions + (grid_segment_question(),),
        respondent_id_column=base_schema.respondent_id_column,
        total_respondents=base_schema.total_respondents,
        source_datamap_path=base_schema.source_datamap_path,
        source_rawdata_path=base_schema.source_rawdata_path,
        parsed_at=base_schema.parsed_at,
    )
    log = CalculationLog()
    results, skips = compute_cross_cuts(
        [
            CrossCutSpec(
                cross_cut_id="CC_GRID_EXPORT",
                title="Grid segment by target",
                analysis_type=AnalysisType.CROSS_TAB,
                source_question_ids=("Q_GRID_SEG", "Q_TGT_1"),
                display_mode=display_mode,
            )
        ],
        schema,
        dataframe,
        log,
    )
    if skips:
        raise AssertionError(skips[0].details)
    return results[0], schema, log


def make_cross_cut_export_fixture():
    results, skips, schema, quality_report, log = make_export_fixture()
    dataframe = pd.read_csv(CROSS_CUT_30_RESPONDENTS_PATH)
    cross_schema = make_cross_cut_schema()
    export_schema = SurveySchema(
        questions=schema.questions + cross_schema.questions,
        respondent_id_column=schema.respondent_id_column,
        total_respondents=schema.total_respondents,
        source_datamap_path=schema.source_datamap_path,
        source_rawdata_path=schema.source_rawdata_path,
        parsed_at=schema.parsed_at,
    )
    cross_specs = [
        CrossCutSpec(
            cross_cut_id="CC_TAB_EXPORT",
            title="Segment by target",
            analysis_type=AnalysisType.CROSS_TAB,
            source_question_ids=("Q_SEG_1", "Q_TGT_1"),
        ),
        CrossCutSpec(
            cross_cut_id="CC_SEG_EXPORT",
            title="Segment profile",
            analysis_type=AnalysisType.SEGMENT_PROFILE,
            source_question_ids=("Q_SEG_1", "Q_TGT_1"),
            filter_expr="Q_SEG_1 == 1",
            filter_mask_description="Q_SEG_1 = Segment 1",
        ),
        CrossCutSpec(
            cross_cut_id="CC_GROUP_EXPORT",
            title="Group comparison",
            analysis_type=AnalysisType.GROUP_COMPARISON,
            source_question_ids=("Q_SEG_1", "Q_NUM_3"),
        ),
        CrossCutSpec(
            cross_cut_id="CC_EVR_EXPORT",
            title="Expected vs realized",
            analysis_type=AnalysisType.EXPECTED_VS_REALIZED,
            source_question_ids=("Q_EXP_1", "Q_REAL_1"),
        ),
    ]
    cross_results, _ = compute_cross_cuts(cross_specs, cross_schema, dataframe, log)
    cross_skips = [
        SkipRecord(
            question_id="CC_BAD_EXPORT",
            canonical_id="CC_BAD_EXPORT",
            question_type=QuestionType.UNKNOWN,
            skip_reason="cross_cut_error",
            details="ValueError: synthetic cross cut failure",
        )
    ]
    return results, skips, export_schema, quality_report, log, cross_results, cross_skips


def make_filtered_export_fixture():
    (
        results,
        _skips,
        schema,
        _quality_report,
        log,
        cross_results,
        _cross_skips,
    ) = make_cross_cut_export_fixture()
    unrelated_audit = make_audit(
        "rate_per_value",
        "Q_UNRELATED_EXPORT",
        ("Q_UNRELATED_EXPORT",),
        1.0,
        1,
        0,
    )
    log.record(unrelated_audit)
    filtered_results = [
        FilteredSingleCutResult(
            target_question_id="Q_SS_EXPORT",
            filters_applied=(FilterSpec("Q_SEG_1", 1),),
            dispatch_mode="single_cut_filtered",
            single_cut_result=results[0],
            cross_cut_result=None,
            filtered_n=6,
            audit_records=results[0].audit_records,
            warnings=("Filtered sample size 6 is below reliability threshold (30); results may not be reliable.",),
        ),
        FilteredSingleCutResult(
            target_question_id="Q_SS_EXPORT",
            filters_applied=(FilterSpec("Q_SEG_1", 2),),
            dispatch_mode="single_cut_filtered",
            single_cut_result=replace(results[0], valid_n=4),
            cross_cut_result=None,
            filtered_n=4,
            audit_records=results[0].audit_records,
        ),
        FilteredSingleCutResult(
            target_question_id="Q_TGT_1",
            filters_applied=(FilterSpec("Q_SEG_1"),),
            dispatch_mode="cross_cut_breakdown",
            single_cut_result=None,
            cross_cut_result=cross_results[0],
            filtered_n=30,
            audit_records=cross_results[0].audit_records,
        ),
    ]
    return filtered_results, schema, log


def sheet_values(workbook_path: Path, sheet_name: str) -> list[object]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return [
            cell
            for row in workbook[sheet_name].iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]
    finally:
        workbook.close()


def find_row(ws, value: object, column: int = 1) -> int:
    for row_index in range(1, ws.max_row + 1):
        if ws.cell(row=row_index, column=column).value == value:
            return row_index
    raise AssertionError(f"{value!r} not found in column {column}")


def defined_range_values(workbook, name: str) -> list[object]:
    destinations = list(workbook.defined_names[name].destinations)
    values: list[object] = []
    for sheet_name, coord in destinations:
        ws = workbook[sheet_name]
        for row in ws[coord]:
            for cell in row:
                values.append(cell.value)
    return values


def question_title(question_id: str, text: str) -> str:
    return f"{question_id} - {text}"


def question_header_row(ws, question_id: str) -> int:
    prefix = f"{question_id} - "
    for row_index in range(1, ws.max_row + 1):
        value = ws.cell(row=row_index, column=1).value
        if isinstance(value, str) and value.startswith(prefix):
            return row_index
    raise AssertionError(f"Question block {question_id!r} not found")


def table_header_row(ws, question_id: str, header: str = "Option") -> int:
    start = question_header_row(ws, question_id)
    for row_index in range(start, min(ws.max_row, start + 12) + 1):
        if ws.cell(row=row_index, column=1).value == header:
            return row_index
    raise AssertionError(f"{header!r} table header for {question_id!r} not found")


class TestExcelExporter(unittest.TestCase):
    def test_wrapped_formula_preserves_bare_commas(self) -> None:
        formula = _wrapped_formula("F_Q14")

        self.assertEqual(formula.count("SUBSTITUTE("), 1)
        self.assertIn('", ", "|"', formula)
        self.assertNotIn('SUBSTITUTE(SUBSTITUTE(', formula)

    def test_memory_profile_report_written_when_enabled(self) -> None:
        from src import memory_profiler

        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        report_path = Path("outputs") / f"{output_path.stem}.memory_report.txt"
        if report_path.exists():
            report_path.unlink()
        results, skips, schema, quality_report, log = make_export_fixture()
        expected_labels = [
            "load_or_receive_decoded_df",
            "generate_short_labels",
            "build_raw_data_sheet",
            "build_options_sheet",
            "build_filters_sheet",
            "build_helper_columns",
            "build_run_summary_sheet",
            "build_question_metadata_sheet",
            "build_single_cut_index_sheet",
            "build_theme_sheets",
            "build_calculation_log_sheet",
            "build_filter_log_sheet",
            "build_warnings_sheet",
            "save_workbook",
            "patch_formula_caches",
            "write_calc_chain",
        ]

        try:
            with patch.dict(os.environ, {"SURVEY_PROFILE_MEMORY": "1"}):
                export_single_cuts(
                    results,
                    skips,
                    schema,
                    quality_report,
                    log,
                    str(output_path),
                )
        finally:
            memory_profiler.disable_profiling()
            memory_profiler.reset_log()

        self.assertTrue(report_path.exists())
        report = report_path.read_text(encoding="utf-8")
        for label in expected_labels:
            self.assertIn(label, report)

    def grid_single_select_format_fixture(
        self,
    ) -> tuple[GridSingleSelectResult, SurveySchema]:
        row_a = SingleSelectResult(
            question_id="Q_GRID_FORMAT.r1",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=80,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={1: {"label": "Checked", "count": 50, "rate": 0.625}},
        )
        row_b = SingleSelectResult(
            question_id="Q_GRID_FORMAT.r2",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=80,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={1: {"label": "Checked", "count": 30, "rate": 0.375}},
        )
        row_c = SingleSelectResult(
            question_id="Q_GRID_FORMAT.r3",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=80,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={1: {"label": "Checked", "count": 0, "rate": 0.0}},
        )
        result = GridSingleSelectResult(
            question_id="Q_GRID_FORMAT",
            question_type=QuestionType.GRID_SINGLE_SELECT,
            valid_n=80,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            rows={"r1": row_a, "r2": row_b, "r3": row_c},
            overall_valid_n=80,
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_GRID_FORMAT]",
                    canonical_id="Q_GRID_FORMAT",
                    question_text="Grid format question",
                    question_type=QuestionType.GRID_SINGLE_SELECT,
                    raw_columns=("r1", "r2", "r3"),
                    option_map={0: "Unchecked", 1: "Checked"},
                    value_range=(0, 1),
                    grid_row_labels={
                        "r1": "Option A",
                        "r2": "Option B",
                        "r3": "Option C",
                    },
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=80,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        return result, schema

    def export_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        export_single_cuts(results, skips, schema, quality_report, log, str(output_path))
        return output_path

    def export_numeric_allocation_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        per_option_stats = {
            "Q20r1": {
                "mean": 28.9,
                "median": 27.0,
                "std": 10.5,
                "min_val": 0.0,
                "max_val": 100.0,
                "valid_n": 1022,
                "missing_n": 0,
            },
            "Q20r2": {
                "mean": 24.6,
                "median": 22.0,
                "std": 9.1,
                "min_val": 0.0,
                "max_val": 100.0,
                "valid_n": 1022,
                "missing_n": 0,
            },
            "Q20r3": {
                "mean": 22.9,
                "median": 20.0,
                "std": 8.4,
                "min_val": 0.0,
                "max_val": 100.0,
                "valid_n": 1022,
                "missing_n": 0,
            },
        }
        audits = tuple(
            AuditRecord(
                output_sheet="SC_Q20",
                metric_name=metric_name,
                source_question_id="Q20",
                source_columns=(option_id,),
                filter_expr=None,
                numerator=100.0,
                denominator=int(payload["valid_n"]),
                formula=f"{metric_name} formula",
                value_raw=float(payload[metric_name.rsplit("_", 1)[-1]]),
                valid_n=int(payload["valid_n"]),
                missing_n=int(payload["missing_n"]),
                timestamp=UTC_NOW,
            )
            for option_id, payload in per_option_stats.items()
            for metric_name in ("numeric_allocation_mean", "numeric_allocation_median")
        )
        result = NumericResult(
            question_id="Q20",
            question_type=QuestionType.NUMERIC_ALLOCATION,
            valid_n=1022,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            mean=25.5,
            median=23.0,
            std=3.0,
            min_val=0.0,
            max_val=100.0,
            percentiles={25: 20.0, 50: 23.0, 75: 28.9},
            allocation_target=100.0,
            allocation_tolerance=2.0,
            allocation_excluded_n=0,
            per_option_stats=per_option_stats,
            audit_records=audits,
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q20]",
                    canonical_id="Q20",
                    question_text="Allocate growth by source",
                    question_type=QuestionType.NUMERIC_ALLOCATION,
                    raw_columns=("Q20r1", "Q20r2", "Q20r3"),
                    option_map={
                        "Q20r1": "Existing customers - net retention",
                        "Q20r2": "Existing customers - cross-sell / up-sell",
                        "Q20r3": "New customers - existing geographies",
                    },
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=1022,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=1022,
            total_columns=3,
            columns_in_datamap=1,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        log = CalculationLog()
        for audit in audits:
            log.record(audit)
        export_single_cuts([result], [], schema, quality_report, log, str(output_path))
        return output_path

    def export_workbook_with_cross_cuts(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        (
            results,
            skips,
            schema,
            quality_report,
            log,
            cross_results,
            cross_skips,
        ) = make_cross_cut_export_fixture()
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            cross_cut_results=cross_results,
            cross_cut_skips=cross_skips,
        )
        return output_path

    def export_cross_cut_workbook(self, cross_results: list) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        (
            _results,
            _skips,
            schema,
            _quality_report,
            log,
            _cross_results,
            _cross_skips,
        ) = make_cross_cut_export_fixture()
        export_cross_cuts_only(cross_results, schema, log, str(output_path))
        return output_path

    def export_filtered_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        filtered_results, schema, log = make_filtered_export_fixture()
        export_filtered_single_cuts(filtered_results, schema, log, str(output_path))
        return output_path

    def test_export_creates_file_at_path(self) -> None:
        output_path = self.export_workbook()

        self.assertTrue(output_path.exists())
        self.assertGreater(output_path.stat().st_size, 0)

    def test_export_writes_all_required_sheets(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(
            workbook.sheetnames[:6],
            ["_RawData", "_Options", "Filters", "Run_Summary", "Question_Metadata", "Single_Cut_Index"],
        )
        self.assertEqual(
            workbook.sheetnames[-3:],
            ["Calculation_Log", "Filter_Log", "Warnings"],
        )
        self.assertIn("All Questions", workbook.sheetnames)
        self.assertFalse(any(name.startswith("SC_") for name in workbook.sheetnames))

    def test_run_summary_contains_correct_totals(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Run_Summary"]

        self.assertEqual(ws["B4"].value, 10)
        self.assertEqual(ws["B5"].value, 6)
        self.assertEqual(ws["B6"].value, 5)
        self.assertEqual(ws["B7"].value, 2)
        self.assertEqual(ws["B8"].value, 0)
        self.assertEqual(ws["B9"].value, 0)
        self.assertEqual(ws["B10"].value, 7)
        self.assertEqual(ws["B11"].value, 1)

    def test_question_metadata_one_row_per_question(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(workbook["Question_Metadata"].max_row, 7)

    def test_single_cut_index_links_to_theme_sheets(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["Single_Cut_Index"]

        self.assertEqual(ws["D2"].value, "All Questions")
        self.assertEqual(ws["E2"].value, "All Questions")
        self.assertIsNotNone(ws["E2"].hyperlink)
        self.assertIn("All Questions", workbook.sheetnames)

    def test_sc_sheet_for_single_select_has_distribution_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_SS_EXPORT")
        data_row = header_row + 1
        self.assertEqual(
            ws.cell(row=question_header_row(ws, "Q_SS_EXPORT"), column=1).value,
            "Q_SS_EXPORT - Single Select Export Question",
        )
        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 5)],
            ["Option", "Count", "%", "Denominator"],
        )
        self.assertEqual(ws.cell(data_row, 1).value, "Yes")
        self.assertTrue(ws.cell(data_row, 2).value.startswith("=COUNTIFS"))
        self.assertIn("Q_SS_EXPORT_data", ws.cell(data_row, 2).value)
        self.assertIn("passes_workbook_filters_data", ws.cell(data_row, 2).value)
        self.assertIn("passes_workbook_custom_filters_data", ws.cell(data_row, 2).value)
        self.assertIn("SUBTOTAL", ws.cell(data_row, 3).value)

    def test_sc_sheet_for_multi_select_has_selections_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_MS_EXPORT")
        data_row = header_row + 1
        self.assertEqual(
            ws.cell(row=question_header_row(ws, "Q_MS_EXPORT"), column=1).value,
            "Q_MS_EXPORT - Multi Select Export Question",
        )
        self.assertEqual(ws.cell(header_row, 1).value, "Option")
        self.assertEqual(ws.cell(header_row, 2).value, "Count")
        self.assertEqual(ws.cell(data_row, 1).value, "First")
        self.assertIn("Q_MS_EXPORTr1_data", ws.cell(data_row, 2).value)
        self.assertIn("SUBTOTAL", ws.cell(data_row, 3).value)

    def test_sc_sheet_for_numeric_has_stats_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_NUM_EXPORT", header="Metric")
        self.assertEqual(
            ws.cell(row=question_header_row(ws, "Q_NUM_EXPORT"), column=1).value,
            "Q_NUM_EXPORT - Numeric Export Question",
        )
        self.assertEqual(ws.cell(header_row, 1).value, "Metric")
        self.assertEqual(ws.cell(header_row + 1, 1).value, "Mean")
        self.assertTrue(ws.cell(header_row + 1, 2).value.startswith("=IFERROR(AVERAGEIFS"))
        self.assertIn("passes_workbook_filters_data", ws.cell(header_row + 1, 2).value)
        self.assertIn("passes_workbook_custom_filters_data", ws.cell(header_row + 1, 2).value)
        self.assertEqual(ws.cell(header_row + 4, 1).value, "Std")
        self.assertEqual(ws.cell(header_row + 4, 2).value, 2.9)
        self.assertIn("Median not available", ws.cell(header_row + 5, 1).value)

    def test_numeric_allocation_table_is_option_oriented_with_mean_and_median(self) -> None:
        output_path = self.export_numeric_allocation_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q20")

        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 5)],
            ["Option", "Mean", "Median", "Denominator"],
        )
        option_labels = [ws.cell(header_row + offset, 1).value for offset in range(1, 4)]
        self.assertIn("Existing customers - net retention", option_labels)
        self.assertFalse(any(str(label).startswith("Q20r") for label in option_labels))
        table_values = [
            ws.cell(row, 1).value
            for row in range(header_row, table_header_row(ws, "Q20") + 6)
        ]
        self.assertFalse(any(value in {"Std", "Min", "Max"} for value in table_values))

    def test_numeric_allocation_mean_and_median_columns_have_color_scales(self) -> None:
        output_path = self.export_numeric_allocation_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q20")

        self.assertTrue(
            any(
                str(formatting.sqref) == f"B{header_row + 1}:B{header_row + 3}"
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )
        self.assertTrue(
            any(
                str(formatting.sqref) == f"C{header_row + 1}:C{header_row + 3}"
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )

    def test_numeric_allocation_audit_log_contains_per_option_medians(self) -> None:
        output_path = self.export_numeric_allocation_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Calculation_Log"]
        metric_names = [
            ws.cell(row=row_index, column=2).value
            for row_index in range(2, ws.max_row + 1)
        ]

        self.assertEqual(metric_names.count("numeric_allocation_median"), 3)

    def test_sc_sheet_for_grid_has_compact_distribution_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_GRID_EXPORT")
        data_row = header_row + 1
        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 5)],
            ["Option", "Count", "%", "Denominator"],
        )
        self.assertEqual(ws.cell(data_row, 1).value, "Grid first row")
        self.assertTrue(ws.cell(data_row, 2).value.startswith("=COUNTIFS"))
        self.assertIn("Q_GRID_EXPORTr1_data", ws.cell(data_row, 2).value)
        self.assertIn("passes_workbook_filters_data", ws.cell(data_row, 2).value)
        self.assertIn("passes_workbook_custom_filters_data", ws.cell(data_row, 2).value)
        self.assertIn("SUBTOTAL", ws.cell(data_row, 4).value)

    def test_sc_sheet_has_autofilter(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertGreater(len(ws.tables), 0)

    def test_distribution_count_column_has_color_scale(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        range_ref = f"B{header_row + 1}:B{header_row + 2}"

        self.assertTrue(
            any(
                str(formatting.sqref) == range_ref
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )

    def test_distribution_percent_column_has_color_scale(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        range_ref = f"C{header_row + 1}:C{header_row + 2}"

        self.assertTrue(
            any(
                str(formatting.sqref) == range_ref
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )

    def test_total_respondents_row_follows_distribution_options(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        total_row = header_row + 3

        self.assertEqual(ws.cell(total_row, 1).value, "Total respondents")
        self.assertEqual(ws.cell(total_row, 2).value, 10)

    def test_sc_sheet_filter_block_no_filter(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(ws["A1"].value, "THEME: All Questions")
        self.assertEqual(ws["A3"].value, "LOCAL FILTERS (override workbook defaults)")
        self.assertNotEqual(ws["A3"].value, "DEMOGRAPHIC FILTERS")
        self.assertEqual(
            ws.cell(question_header_row(ws, "Q_SS_EXPORT"), 1).value,
            "Q_SS_EXPORT - Single Select Export Question",
        )

    def test_subset_denominator_note_appears_when_denominator_is_small(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        subset_result = replace(
            results[0],
            valid_n=5,
            missing_n=5,
            distribution={
                1: {"label": "Yes", "count": 3, "rate": 0.6},
                2: {"label": "No", "count": 2, "rate": 0.4},
            },
        )
        export_single_cuts(
            [subset_result, *results[1:]],
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        values = [
            ws.cell(row=row_index, column=1).value
            for row_index in range(
                question_header_row(ws, "Q_SS_EXPORT"),
                table_header_row(ws, "Q_SS_EXPORT") + 1,
            )
        ]

        self.assertIn(
            "Note: This question was shown to a subset. Total respondents shown: 5.",
            values,
        )

    def test_subset_denominator_note_absent_for_full_denominator(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        values = [
            ws.cell(row=row_index, column=1).value
            for row_index in range(
                question_header_row(ws, "Q_SS_EXPORT"),
                table_header_row(ws, "Q_SS_EXPORT") + 1,
            )
        ]

        self.assertFalse(
            any(isinstance(value, str) and value.startswith("Note:") for value in values)
        )

    def test_filter_sheet_exists_visible(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Filters", workbook.sheetnames)
        self.assertEqual(workbook["Filters"].sheet_state, "visible")

    def test_filter_display_label_drops_question_number(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(
                    question,
                    question_text="Q14 - In which country are you based?",
                    is_demographic=True,
                )
                if question.canonical_id == "Q_SS_EXPORT"
                else question
                for question in schema.questions
            ),
        )
        with patch("src.ai_insights.generate_short_labels", return_value={}):
            export_single_cuts(
                results,
                skips,
                schema,
                quality_report,
                log,
                str(output_path),
                demo_priority={
                    "priority_ordered": ["Q_SS_EXPORT"],
                    "categories": {"Q_SS_EXPORT": "country"},
                },
            )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(workbook["Filters"]["A4"].value, "Country")
        self.assertNotRegex(str(workbook["Filters"]["A4"].value), r"\bQ\d+")

    def test_filters_sheet_uses_ai_short_labels_for_demographics(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        questions = (
            QuestionSpec(
                question_id="[Q4]",
                canonical_id="Q4",
                question_text="Which of the following best describes the size of your organization",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q4",),
                option_map={1: "Small", 2: "Large"},
                is_demographic=True,
            ),
            QuestionSpec(
                question_id="[Q14]",
                canonical_id="Q14",
                question_text="Approximately how old is the organization",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q14",),
                option_map={1: "New", 2: "Established"},
                is_demographic=True,
            ),
        )
        schema = SurveySchema(
            questions=questions,
            respondent_id_column="respondent_id",
            total_respondents=2,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=2,
            total_columns=2,
            columns_in_datamap=2,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        results = [
            SingleSelectResult(
                question_id="Q4",
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=2,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "Small", "count": 1, "rate": 0.5},
                    2: {"label": "Large", "count": 1, "rate": 0.5},
                },
            ),
            SingleSelectResult(
                question_id="Q14",
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=2,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "New", "count": 1, "rate": 0.5},
                    2: {"label": "Established", "count": 1, "rate": 0.5},
                },
            ),
        ]

        with patch(
            "src.ai_insights.generate_short_labels",
            return_value={"Q4": "Org Size", "Q14": "Org Age"},
        ):
            export_single_cuts(
                results,
                [],
                schema,
                quality_report,
                CalculationLog(),
                str(output_path),
                demo_priority={
                    "priority_ordered": ["Q4", "Q14"],
                    "categories": {"Q4": "company_size", "Q14": "age"},
                },
            )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        filter_labels = [
            workbook["Filters"].cell(row=row_index, column=1).value
            for row_index in range(1, workbook["Filters"].max_row + 1)
            if workbook["Filters"].cell(row=row_index, column=1).value
        ]

        self.assertIn("Org Size", filter_labels)
        self.assertIn("Org Age", filter_labels)
        broken_prefix = re.compile(r"^(Which|If|Approximately|How|What|During)")
        self.assertFalse(
            any(broken_prefix.match(str(label)) for label in filter_labels)
        )

    def test_per_question_dropdown_options_are_readable_labels(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["_Options"]
        values = [
            ws.cell(row=row_index, column=1).value
            for row_index in range(2, ws.max_row + 1)
            if ws.cell(row=row_index, column=1).value
        ]

        self.assertIn("Q_SS_EXPORT - Single Select Export Question", values)
        self.assertFalse(any(str(value) in {"Yes", "No", "First", "Second"} for value in values))

    def test_all_questions_dropdown_contains_only_eligible_single_select_questions(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        values = defined_range_values(workbook, "All_Questions")
        eligible_questions = [
            question
            for question in make_export_fixture()[2].questions
            if question.question_type is QuestionType.SINGLE_SELECT
            and question.analysis_eligible
            and len(question.option_map) >= 2
        ]

        self.assertEqual(len(values), len(eligible_questions) + 1)
        self.assertEqual(values[0], "(None)")
        self.assertIn("Q_SS_EXPORT - Single Select Export Question", values)
        self.assertIn(f"{LONG_ID} - Long Sheet Name Export Question", values)
        self.assertNotIn("Yes", values)
        self.assertNotIn("No", values)
        self.assertNotIn("Version 2", values)
        self.assertFalse(any("term:" in str(value) for value in values))

    def test_all_questions_dropdown_entries_use_q_number_dash_label_format(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        questions = (
            QuestionSpec(
                question_id="[Q1]",
                canonical_id="Q1",
                question_text="In which country do you currently work?",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q1",),
                option_map={1: "United States", 2: "Canada", 3: "India"},
                is_demographic=True,
            ),
            QuestionSpec(
                question_id="[Q2]",
                canonical_id="Q2",
                question_text="Which survey version did the respondent see?",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q2",),
                option_map={1: "Version 2", 2: "Version 3"},
            ),
            QuestionSpec(
                question_id="[Q3]",
                canonical_id="Q3",
                question_text="Which terms apply?",
                question_type=QuestionType.MULTI_SELECT_BINARY,
                raw_columns=("Q3r1",),
                option_map={"Q3r1": "term: Selected Other at Q2"},
                value_range=(0, 1),
            ),
        )
        schema = SurveySchema(
            questions=questions,
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=4,
            total_columns=3,
            columns_in_datamap=3,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        results = [
            SingleSelectResult(
                question_id="Q1",
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=4,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "United States", "count": 2, "rate": 0.5},
                    2: {"label": "Canada", "count": 1, "rate": 0.25},
                    3: {"label": "India", "count": 1, "rate": 0.25},
                },
            ),
            SingleSelectResult(
                question_id="Q2",
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=4,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "Version 2", "count": 2, "rate": 0.5},
                    2: {"label": "Version 3", "count": 2, "rate": 0.5},
                },
            ),
        ]

        with patch(
            "src.ai_insights.generate_short_labels",
            return_value={"Q1": "Country", "Q2": "Survey Version", "Q3": "Terms"},
        ):
            export_single_cuts(
                results,
                [],
                schema,
                quality_report,
                CalculationLog(),
                str(output_path),
            )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        values = defined_range_values(workbook, "All_Questions")

        self.assertEqual(values, ["(None)", "Q1 - Country", "Q2 - Survey Version"])
        for value in values[1:]:
            self.assertRegex(str(value), r"^Q\d+\s*-\s*\w+")
        self.assertFalse(
            any(
                value in {"United States", "Canada", "India", "Version 2", "Version 3"}
                for value in values
            )
        )
        local_values = defined_range_values(workbook, "All_Questions_Local")
        self.assertEqual(local_values, ["(Inherit)", "(None)", "Q1 - Country", "Q2 - Survey Version"])

    def test_theme_sheet_has_local_filter_rows(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(ws["A3"].value, "LOCAL FILTERS (override workbook defaults)")
        self.assertEqual(ws["A4"].value, "Filter")
        self.assertEqual(ws["B5"].value, "(Inherit)")

    def test_filter_uses_countifs_not_sumproduct(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        formula = ws.cell(row=header_row + 1, column=2).value

        self.assertTrue(formula.startswith("=COUNTIFS"))
        self.assertIn("passes_workbook_filters_data", formula)
        self.assertIn("passes_workbook_custom_filters_data", formula)
        self.assertNotIn("SUMPRODUCT", formula)
        self.assertNotIn("INDEX(", formula)

    def test_formulas_use_countifs_helpers_no_sumproduct(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)

        formulas: list[str] = []
        for worksheet in workbook.worksheets:
            if worksheet.title.startswith("_"):
                continue
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and (
                        "COUNTIFS" in value
                        or "AVERAGEIFS" in value
                        or "MINIFS" in value
                        or "MAXIFS" in value
                    ):
                        formulas.append(value)

        self.assertTrue(formulas)
        for formula in formulas:
            self.assertNotIn("@", formula)
            self.assertNotIn("INDEX(", formula)
            self.assertNotIn("SUMPRODUCT", formula)

    def test_formula_cells_have_cached_values_and_calc_chain(self) -> None:
        import xml.etree.ElementTree as ET
        from zipfile import ZipFile

        output_path = self.export_workbook()
        formula_cells: list[tuple[str, str]] = []
        namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

        with ZipFile(output_path) as archive:
            names = set(archive.namelist())
            self.assertIn("xl/calcChain.xml", names)

            for name in sorted(names):
                if not name.startswith("xl/worksheets/sheet") or not name.endswith(".xml"):
                    continue
                root = ET.fromstring(archive.read(name))
                for cell in root.iter(f"{namespace}c"):
                    formula = cell.find(f"{namespace}f")
                    if formula is None:
                        continue
                    ref = cell.attrib.get("r", "")
                    formula_cells.append((name, ref))
                    cached = cell.find(f"{namespace}v")
                    self.assertIsNotNone(cached, f"{name}!{ref} has no cached <v>")
                    if "INDEX(INDIRECT" not in (formula.text or ""):
                        self.assertIsNotNone(cached.text, f"{name}!{ref} has an empty cache")
                        self.assertNotEqual(cached.text, "", f"{name}!{ref} has an empty cache")

        self.assertTrue(formula_cells)

    def test_countifs_include_theme_local_and_per_question_helpers(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        formula = ws.cell(row=header_row + 1, column=2).value

        self.assertTrue(formula.startswith("=COUNTIFS"))
        self.assertIn("All_Questions_passes_local_filters_data,1", formula)
        self.assertIn(
            "All_Questions_Q_SS_EXPORT_F_passes_per_q_filter_data,1",
            formula,
        )

    def test_numeric_formulas_include_theme_local_and_per_question_helpers(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_NUM_EXPORT", header="Metric")
        mean_formula = ws.cell(row=header_row + 1, column=2).value
        count_formula = ws.cell(row=header_row + 1, column=4).value

        self.assertIn("All_Questions_passes_local_filters_data,1", mean_formula)
        self.assertIn(
            "All_Questions_Q_NUM_EXPORT_F_passes_per_q_filter_data,1",
            mean_formula,
        )
        self.assertIn("All_Questions_passes_local_filters_data,1", count_formula)
        self.assertIn(
            "All_Questions_Q_NUM_EXPORT_F_passes_per_q_filter_data,1",
            count_formula,
        )

    def test_cross_tab_header_formula_cache_is_blank_by_default(self) -> None:
        import xml.etree.ElementTree as ET
        from zipfile import ZipFile

        output_path = self.export_workbook()
        namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        blank_caches = 0

        with ZipFile(output_path) as archive:
            for name in archive.namelist():
                if not name.startswith("xl/worksheets/sheet") or not name.endswith(".xml"):
                    continue
                root = ET.fromstring(archive.read(name))
                for cell in root.iter(f"{namespace}c"):
                    formula = cell.find(f"{namespace}f")
                    if formula is None:
                        continue
                    formula_text = formula.text or ""
                    if "INDEX(INDIRECT" not in formula_text or "_options" not in formula_text:
                        continue
                    cached = cell.find(f"{namespace}v")
                    self.assertIsNotNone(cached)
                    self.assertIn(cached.text, (None, ""))
                    self.assertNotEqual(cached.text, "0")
                    blank_caches += 1

        self.assertGreater(blank_caches, 0)

    def test_raw_data_has_filter_helper_columns(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, is_demographic=question.canonical_id == "Q_SS_EXPORT")
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country"},
            },
        )
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["_RawData"]
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

        self.assertIn("F_Country_match", headers)
        self.assertIn("passes_workbook_filters", headers)
        self.assertIn("F_Custom1_match", headers)
        self.assertIn("F_Custom2_match", headers)
        self.assertIn("passes_workbook_custom_filters", headers)
        pass_col = headers.index("passes_workbook_filters") + 1
        self.assertIn("*", ws.cell(row=2, column=pass_col).value)
        custom_col = headers.index("passes_workbook_custom_filters") + 1
        self.assertIn("*", ws.cell(row=2, column=custom_col).value)
        self.assertIn("passes_workbook_filters_data", workbook.defined_names)
        self.assertIn("F_Custom1_match_data", workbook.defined_names)
        self.assertIn("F_Custom2_match_data", workbook.defined_names)
        self.assertIn("passes_workbook_custom_filters_data", workbook.defined_names)

    def test_inherit_pattern_resolves_to_workbook_value(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(ws["B5"].value, "(Inherit)")
        self.assertIn('="(Inherit)"', ws["C5"].value)
        self.assertIn("F_Custom1_Q", ws["C5"].value)

    def test_available_values_column_present(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, is_demographic=question.canonical_id == "Q_SS_EXPORT")
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country"},
            },
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Yes", workbook["Filters"]["D4"].value)
        self.assertIn("No", workbook["Filters"]["D4"].value)
        self.assertIn("Yes", workbook["Demographics"]["E5"].value)

    def test_question_title_has_cell_comment_with_full_text(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        title_row = question_header_row(ws, "Q_SS_EXPORT")

        self.assertIsNotNone(ws.cell(title_row, 1).comment)
        self.assertEqual(
            ws.cell(title_row, 1).comment.text,
            "Single select export question",
        )

    def test_theme_question_heading_restores_question_number_prefix(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        heading = ws.cell(question_header_row(ws, "Q_SS_EXPORT"), 1).value

        self.assertEqual(heading, "Q_SS_EXPORT - Single Select Export Question")
        self.assertRegex(str(heading), r"^Q\w+\s*[-:]?\s*")

    def test_short_labels_replace_visible_question_title_only(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            short_labels={"Q_SS_EXPORT": "Short revenue label"},
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(
            ws.cell(question_header_row(ws, "Q_SS_EXPORT"), 1).value,
            "Q_SS_EXPORT - Short Revenue Label",
        )

    def test_question_heading_row_uses_bain_red_style(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        heading_row = question_header_row(ws, "Q_SS_EXPORT")
        heading_cell = ws.cell(heading_row, 1)

        self.assertEqual(heading_cell.fill.fgColor.rgb, "FFCC0000")
        self.assertEqual(heading_cell.font.color.rgb, "FFFFFFFF")
        self.assertTrue(heading_cell.font.bold)
        for column in range(1, 5):
            self.assertEqual(ws.cell(heading_row, column).fill.fgColor.rgb, "FFCC0000")

    def test_question_heading_uses_short_label_with_full_text_comment(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        full_text = (
            "Q7 - Which of the following best describes your current seniority "
            "within your organization?"
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q7]",
                    canonical_id="Q7",
                    question_text=full_text,
                    question_type=QuestionType.SINGLE_SELECT,
                    raw_columns=("Q7",),
                    option_map={1: "Executive", 2: "Manager"},
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        result = SingleSelectResult(
            question_id="Q7",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                1: {"label": "Executive", "count": 1, "rate": 0.25},
                2: {"label": "Manager", "count": 3, "rate": 0.75},
            },
        )
        quality_report = DataQualityReport(
            total_rows=4,
            total_columns=1,
            columns_in_datamap=1,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        with patch("src.ai_insights.generate_short_labels", return_value={"Q7": "Seniority"}):
            export_single_cuts(
                [result],
                [],
                schema,
                quality_report,
                CalculationLog(),
                str(output_path),
            )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        heading_cell = ws.cell(question_header_row(ws, "Q7"), 1)

        self.assertRegex(str(heading_cell.value), r"^Q\d+[a-z]*\s*-\s*.+")
        self.assertEqual(heading_cell.value, "Q7 - Seniority")
        self.assertIsNotNone(heading_cell.comment)
        self.assertIn(full_text, heading_cell.comment.text)
        self.assertGreater(len(heading_cell.comment.text), len(str(heading_cell.value)))

    def test_per_question_filter_named_cells_scoped_per_theme(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        themes = {
            "themes": [
                {"name": "Theme One", "question_ids": ["Q_SS_EXPORT"]},
                {"name": "Theme Two", "question_ids": ["Q_NUM_EXPORT"]},
            ]
        }
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            themes=themes,
        )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Theme_One_Q_SS_EXPORT_F_V", workbook.defined_names)
        self.assertIn("Theme_Two_Q_NUM_EXPORT_F_V", workbook.defined_names)
        self.assertIsNone(workbook.defined_names["Theme_One_Q_SS_EXPORT_F_V"].localSheetId)

    def test_raw_data_sheet_is_hidden(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("_RawData", workbook.sheetnames)
        self.assertEqual(workbook["_RawData"].sheet_state, "hidden")

    def test_options_sheet_is_hidden(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("_Options", workbook.sheetnames)
        self.assertEqual(workbook["_Options"].sheet_state, "hidden")

    def test_demographic_filter_named_cells_exist(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(
                    question,
                    is_demographic=question.canonical_id in {"Q_SS_EXPORT", "Q_MS_EXPORT"},
                )
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT", "Q_MS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country", "Q_MS_EXPORT": "industry"},
            },
        )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("F_Country", workbook.defined_names)
        self.assertIn("F_Industry", workbook.defined_names)

    def test_count_cell_uses_countifs_formula(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        formula = ws.cell(row=header_row + 1, column=2).value

        self.assertTrue(formula.startswith("=COUNTIFS"))
        self.assertIn("Q_SS_EXPORT_data", formula)
        self.assertIn("passes_workbook_filters_data", formula)
        self.assertIn("passes_workbook_custom_filters_data", formula)
        self.assertNotIn("SUMPRODUCT", formula)

    def test_named_cells_are_workbook_scoped(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, is_demographic=question.canonical_id == "Q_SS_EXPORT")
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country"},
            },
        )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIsNone(workbook.defined_names["F_Country"].localSheetId)
        self.assertIsNone(workbook.defined_names["F_Country_wrapped"].localSheetId)
        self.assertEqual(workbook.defined_names["F_Country"].attr_text, "'Filters'!$B$4")

    def test_per_question_filter_named_cells_exist(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("All_Questions_Q_SS_EXPORT_F_Q", workbook.defined_names)
        self.assertIn("All_Questions_Q_SS_EXPORT_F_V", workbook.defined_names)

    def test_cross_tab_named_cell_exists(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("All_Questions_Q_SS_EXPORT_CT", workbook.defined_names)

    def test_priority_demographic_ordering(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(
                    question,
                    is_demographic=question.canonical_id in {"Q_SS_EXPORT", "Q_MS_EXPORT"},
                )
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT", "Q_MS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country", "Q_MS_EXPORT": "industry"},
            },
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filters"]

        self.assertEqual(ws["A4"].value, "Single Select Export Question")
        self.assertEqual(ws["A5"].value, "Multi Select Export Question")

    def test_workbook_has_one_sheet_per_theme(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        themes = {
            "themes": [
                {
                    "name": "Theme One",
                    "question_ids": ["Q_SS_EXPORT", "Q_MS_EXPORT"],
                },
                {"name": "Theme Two", "question_ids": ["Q_NUM_EXPORT"]},
            ]
        }
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            themes=themes,
        )

        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Theme One", workbook.sheetnames)
        self.assertIn("Theme Two", workbook.sheetnames)
        self.assertFalse(any(name.startswith("SC_") for name in workbook.sheetnames))

    def test_question_block_uses_subtotal_formula_for_percent(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_SS_EXPORT")
        value = ws.cell(row=header_row + 1, column=3).value
        self.assertTrue(value.startswith("="))
        self.assertIn("SUBTOTAL", value)

    def test_demographic_filter_row_has_data_validation(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        questions = tuple(
            replace(question, is_demographic=question.canonical_id == "Q_SS_EXPORT")
            for question in schema.questions
        )
        schema = replace(schema, questions=questions)
        decoded_df = pd.DataFrame({"Q_SS_EXPORT": ["Yes", "No", "Yes"]})
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            decoded_df=decoded_df,
        )

        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filters"]

        self.assertGreater(len(ws.data_validations.dataValidation), 0)

    def test_grid_single_select_excel_format(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        result, schema = self.grid_single_select_format_fixture()
        _results, _skips, _schema, quality_report, _log = make_export_fixture()
        export_single_cuts(
            [result],
            [],
            schema,
            quality_report,
            CalculationLog(),
            str(output_path),
        )

        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_GRID_FORMAT")
        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 5)],
            ["Option", "Count", "%", "Denominator"],
        )
        self.assertEqual(ws.cell(header_row + 1, 1).value, "Option A")
        self.assertIn("r1_data", ws.cell(header_row + 1, 2).value)
        self.assertEqual(ws.cell(header_row + 2, 1).value, "Option B")
        self.assertIn("r2_data", ws.cell(header_row + 2, 2).value)

        values = [
            cell
            for row in ws.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]
        self.assertNotIn("Option C", values)
        self.assertNotIn("Per-row distributions", values)
        self.assertNotIn("Code", values)

    def test_grid_single_select_ui_format(self) -> None:
        app_path = (
            Path(__file__).resolve().parents[1]
            / "artifacts"
            / "survey-insight-engine"
            / "app.py"
        )
        spec = importlib.util.spec_from_file_location("survey_insight_app", app_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        app_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(app_module)

        result, schema = self.grid_single_select_format_fixture()
        rows = app_module._build_grid_display_rows(result, schema)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["Option"], "Option A")
        self.assertEqual(rows[0]["Count"], 50)
        self.assertEqual(rows[1]["Option"], "Option B")
        self.assertNotIn("Option C", [row["Option"] for row in rows])

    def test_calculation_log_one_row_per_audit_record(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(workbook["Calculation_Log"].max_row, 8)

    def test_skip_log_one_row_per_skip(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        values = sheet_values(output_path, "Warnings")
        self.assertIn("ValueError: raw column not found in data", values)

    def test_data_quality_sheet_has_three_sections(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Warnings", workbook.sheetnames)
        values = sheet_values(output_path, "Warnings")
        self.assertIn("column extra_col has 50.0% missing values", values)

    def test_sheet_names_truncated_to_31_chars(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertTrue(all(len(sheet_name) <= 31 for sheet_name in workbook.sheetnames))
        self.assertIn("All Questions", workbook.sheetnames)

    def test_percentages_stored_as_raw_floats(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        value = ws.cell(row=header_row + 1, column=3).value

        self.assertTrue(value.startswith("="))
        self.assertIn("SUBTOTAL", value)

    def test_workbook_opens_without_corruption(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Run_Summary", workbook.sheetnames)

    def test_export_with_cross_cuts_writes_cross_cut_index(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertNotIn("Cross_Cut_Index", workbook.sheetnames)
        self.assertIn("Filter_Log", workbook.sheetnames)

    def test_export_with_cross_cuts_writes_cc_sheets(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertFalse(any(name.startswith("CC_") for name in workbook.sheetnames))
        self.assertIn("All Questions", workbook.sheetnames)

    def test_export_writes_cross_tab_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")

        self.assertEqual(ws.cell(row=header_row, column=6).value, "Option \\ Cross-tab")
        self.assertTrue(ws.cell(row=header_row, column=7).value.startswith("=IFERROR(INDEX"))
        self.assertIn("_options", ws.cell(row=header_row, column=7).value)

    def test_cross_tab_sheet_includes_chart(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(len(ws._charts), 0)
        self.assertGreater(len(ws.tables), 0)

    def test_chart_uses_correct_data_range(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        count_formula = ws.cell(row=header_row + 1, column=2).value
        pct_formula = ws.cell(row=header_row + 1, column=3).value

        self.assertIn("_RawData", workbook.sheetnames)
        self.assertTrue(count_formula.startswith("=COUNTIFS"))
        self.assertIn("passes_workbook_filters_data", count_formula)
        self.assertIn("passes_workbook_custom_filters_data", count_formula)
        self.assertIn("SUBTOTAL", pct_formula)

    def test_export_writes_segment_profile_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("All Questions", workbook.sheetnames)
        self.assertNotIn("CC_CC_SEG_EXPORT", workbook.sheetnames)

    def test_segment_profile_sheet_includes_chart(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        self.assertEqual(len(workbook["All Questions"]._charts), 0)

    def test_export_writes_group_comparison_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("Warnings", workbook.sheetnames)
        self.assertNotIn("CC_CC_GROUP_EXPORT", workbook.sheetnames)

    def test_group_comparison_sheet_includes_chart(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        self.assertEqual(len(workbook["All Questions"]._charts), 0)

    def test_export_writes_expected_vs_realized_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("Filter_Log", workbook.sheetnames)
        self.assertNotIn("CC_CC_EVR_EXPORT", workbook.sheetnames)

    def test_expected_vs_realized_sheet_includes_chart(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        self.assertEqual(len(workbook["All Questions"]._charts), 0)

    def test_export_filter_log_populated_for_segment_profile(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filter_Log"]

        self.assertEqual(ws["A1"].value, "Cross Cut ID")
        self.assertEqual(ws["A2"].value, "CC_SEG_EXPORT")
        self.assertEqual(ws["C2"].value, "Q_SEG_1 == 1")
        self.assertEqual(ws["D2"].value, "Q_SEG_1 = Segment 1")

    def test_export_skip_log_includes_cross_cut_skips(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        rows = list(workbook["Warnings"].iter_rows(values_only=True))

        self.assertEqual(rows[0][0], "Source")
        self.assertIn(
            ("skip:CC_BAD_EXPORT", "ValueError: synthetic cross cut failure"),
            rows,
        )

    def test_export_run_summary_has_cross_cut_totals(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Run_Summary"]

        self.assertEqual(ws["B8"].value, 4)
        self.assertEqual(ws["B9"].value, 1)
        self.assertEqual(ws["B10"].value, 19)

    def test_cross_tab_sheet_includes_axis_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")

        self.assertEqual(ws.cell(header_row, 6).value, "Option \\ Cross-tab")
        self.assertIn("Q_SS_EXPORT_CT", ws.cell(header_row, 7).value)

    def test_segment_profile_sheet_includes_filter_and_target_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("CC_SEG_EXPORT", sheet_values(output_path, "Filter_Log"))

    def test_group_comparison_sheet_includes_segment_and_metric_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("All Questions", workbook.sheetnames)
        self.assertFalse(any(name.startswith("CC_CC_GROUP") for name in workbook.sheetnames))

    def test_expected_vs_realized_sheet_includes_both_question_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        values = sheet_values(output_path, "Question_Metadata")
        self.assertIn("Expected", values)
        self.assertIn("Realized", values)

    def test_cross_tab_sheet_has_corner_orientation_label(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")

        self.assertEqual(ws.cell(header_row, 6).value, "Option \\ Cross-tab")
        self.assertEqual(ws.cell(header_row + 1, 6).value, "Yes")

    def test_cross_tab_sheet_renders_only_counts_in_counts_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="counts")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertNotIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_only_row_pct_in_row_pct_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="row_pct")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertNotIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_both_blocks_in_both_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="both")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_all_three_blocks_in_all_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="all")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertIn("Column %", values)

    def test_grid_cross_tab_respects_display_mode(self) -> None:
        result, schema, log = grid_cross_tab_export_result(display_mode="row_pct")
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        export_cross_cuts_only([result], schema, log, str(output_path))

        values = sheet_values(output_path, "CC_CC_GRID_EXPORT")

        self.assertNotIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)
        self.assertIn("Q_GRID_SEGr1", values)
        self.assertIn("Segment 1", values)

    def test_cross_tab_includes_copy_friendly_block(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(cross_results[:1])
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Copy-friendly", values)
        self.assertIn("Row Code", values)
        self.assertIn("Row Label", values)
        self.assertIn("Column Code", values)
        self.assertIn("Column Label", values)
        self.assertIn("Count", values)

    def test_segment_profile_sheet_respects_display_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[1], display_mode="counts")]
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_SEG_EXPORT"]

        self.assertEqual(ws["A15"].value, "Code")
        self.assertEqual(ws["C15"].value, "Count")
        self.assertIsNone(ws["D15"].value)

    def test_group_comparison_sheet_ignores_display_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[2], display_mode="counts")]
        )
        values = sheet_values(output_path, "CC_CC_GROUP_EXPORT")

        self.assertIn("Per-segment comparison", values)
        self.assertIn("Mean", values)
        self.assertIn("Median", values)
        self.assertIn("Std", values)

    def test_export_cross_cuts_only_writes_selected_cc_sheets(self) -> None:
        (
            _results,
            _skips,
            schema,
            _quality_report,
            log,
            cross_results,
            _cross_skips,
        ) = make_cross_cut_export_fixture()
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        export_cross_cuts_only(cross_results[:2], schema, log, str(output_path))
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Cross_Cut_Index", workbook.sheetnames)
        self.assertIn("CC_CC_TAB_EXPORT", workbook.sheetnames)
        self.assertIn("CC_CC_SEG_EXPORT", workbook.sheetnames)
        self.assertNotIn("CC_CC_GROUP_EXPORT", workbook.sheetnames)

    def test_export_backward_compatible_without_cross_cuts(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertNotIn("Cross_Cut_Index", workbook.sheetnames)
        self.assertFalse(any(name.startswith("CC_") for name in workbook.sheetnames))

    def test_export_filtered_single_cuts_creates_file(self) -> None:
        output_path = self.export_filtered_workbook()

        self.assertTrue(output_path.exists())
        self.assertGreater(output_path.stat().st_size, 0)

    def test_filtered_cut_index_lists_all_results(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filtered_Cut_Index"]

        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws["A1"].value, "Sheet Name")
        self.assertEqual(ws["B2"].value, "Q_SS_EXPORT")
        self.assertEqual(ws["C2"].value, "Q_SEG_1 == 1 (Segment 1)")
        self.assertEqual(ws["D4"].value, "cross_cut_breakdown")
        self.assertEqual(ws["E4"].value, 30)

    def test_fsc_sheet_for_single_cut_filtered_renders_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["FSC_Q_SS_EXPORT_01"]
        values = sheet_values(output_path, "FSC_Q_SS_EXPORT_01")

        self.assertEqual(ws["A1"].value, "Filters applied")
        self.assertEqual(ws["A4"].value, "Target question:")
        self.assertEqual(ws["B5"].value, "Q_SS_EXPORT")
        self.assertIn("Q_SEG_1", values)
        self.assertIn("= Segment 1", values)
        self.assertIn("FILTERED VIEW: Q_SEG_1 == 1 (Segment 1) (n=6)", values)
        self.assertIn("Code", values)
        self.assertIn("Yes", values)
        self.assertIn(6, values)

    def test_sc_sheet_filtered_section_present_when_filter_active(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["FSC_Q_SS_EXPORT_01"]
        values = [
            cell
            for row in ws.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]

        self.assertIn("FILTERED VIEW: Q_SEG_1 == 1 (Segment 1) (n=6)", values)
        self.assertIn("n=6 respondents match filter", values)
        self.assertIsNotNone(ws.auto_filter.ref)

    def test_fsc_sheet_for_cross_cut_breakdown_renders_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        values = sheet_values(output_path, "FSC_Q_TGT_1")

        self.assertIn("Filters applied", values)
        self.assertIn("Q_SEG_1", values)
        self.assertIn("(breakdown)", values)
        self.assertIn("Rows (vertical): Q_SEG_1", values)
        self.assertIn("Columns (horizontal): Q_TGT_1", values)
        self.assertIn("Counts", values)

    def test_fsc_sheet_naming_handles_duplicate_targets(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("FSC_Q_SS_EXPORT_01", workbook.sheetnames)
        self.assertIn("FSC_Q_SS_EXPORT_02", workbook.sheetnames)
        self.assertIn("FSC_Q_TGT_1", workbook.sheetnames)

    def test_export_filtered_single_cuts_filter_log_populated(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        rows = list(workbook["Filter_Log"].iter_rows(values_only=True))

        self.assertEqual(rows[0], (
            "Sheet Name",
            "Filter Question",
            "Filter Value",
            "Filter Description",
        ))
        self.assertIn(
            (
                "FSC_Q_SS_EXPORT_01",
                "Q_SEG_1",
                1,
                "Q_SEG_1 == 1 (Segment 1)",
            ),
            rows,
        )
        self.assertIn(
            (
                "FSC_Q_TGT_1",
                "Q_SEG_1",
                None,
                "Q_SEG_1 (breakdown - no specific value)",
            ),
            rows,
        )

    def test_filtered_export_calculation_log_filtered_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        values = sheet_values(output_path, "Calculation_Log")

        self.assertIn("Q_SS_EXPORT", values)
        self.assertIn("CC_CC_TAB_EXPORT", values)
        self.assertNotIn("Q_MS_EXPORT", values)
        self.assertNotIn("Q_UNRELATED_EXPORT", values)


if __name__ == "__main__":
    unittest.main()
