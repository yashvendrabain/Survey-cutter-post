"""Unified survey file intake layer."""

from __future__ import annotations

from dataclasses import dataclass
import io as stdlib_io
import os
from pathlib import Path
import re
import tempfile
from typing import Any, Iterable, Protocol

from openpyxl import load_workbook
import pandas as pd

from src.datamap_parser import DataMap, parse_datamap
from src.models import LoadReport
from src.raw_decoder import decode_raw_data
from src.word_survey_parser import parse_word_survey


class UploadedFile(Protocol):
    """Small protocol matching Streamlit UploadedFile behavior."""

    name: str

    def read(self) -> bytes:
        ...

    def seek(self, offset: int) -> object:
        ...


DATA_SHEET_KEYWORDS = {
    "data",
    "raw",
    "responses",
    "results",
    "sheet1",
    "survey data",
    "rawdata",
}
MAP_SHEET_KEYWORDS = {
    "map",
    "datamap",
    "data map",
    "codebook",
    "variables",
    "questions",
    "questionnaire",
    "coding",
    "legend",
    "metadata",
}
DATAMAP_NAME_KEYWORDS = {
    "map",
    "datamap",
    "data map",
    "codebook",
    "question",
    "questionnaire",
    "variable",
    "coding",
    "legend",
    "metadata",
}
RAW_NAME_KEYWORDS = {
    "raw",
    "data",
    "response",
    "responses",
    "result",
    "results",
}

MANUAL_COHORT_SHEET_KEYS = {
    "winnerslaggards",
    "winvslag",
    "winnerlaggard",
}


@dataclass(frozen=True, slots=True)
class ManualCohortInput:
    """Validated manual winner/laggard respondent-id lists from an upload."""

    winner_uuids: tuple[str, ...]
    laggard_uuids: tuple[str, ...]
    invalid_uuids: tuple[str, ...] = ()
    overlap_uuids: tuple[str, ...] = ()
    source: str = ""
    id_column: str | None = None

    @property
    def winner_count(self) -> int:
        return len(self.winner_uuids)

    @property
    def laggard_count(self) -> int:
        return len(self.laggard_uuids)


def _normalise_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pandas 3.x convert_dtypes() produces nullable dtypes (Int64, StringDtype,
    boolean) that use pd.NA for missing values. openpyxl cannot serialize pd.NA.
    Coerce all nullable dtypes back to numpy-backed equivalents.
    """

    result = df.copy()
    for col in result.columns:
        dtype = str(result[col].dtype)
        if dtype in ("str", "string"):
            result[col] = result[col].astype(object)
        elif dtype.startswith("Int") or dtype.startswith("UInt"):
            result[col] = result[col].astype(float)
        elif dtype == "boolean":
            result[col] = result[col].astype(object)
        elif dtype in ("Float32", "Float64"):
            result[col] = result[col].astype(float)

        if str(result[col].dtype) == "object":
            result[col] = result[col].where(~result[col].isna(), None)
    return result


def load_survey_inputs(
    uploaded_files: list[UploadedFile],
) -> tuple[DataMap, pd.DataFrame, LoadReport]:
    """Load uploaded survey inputs into a DataMap and decoded DataFrame."""

    if not uploaded_files:
        raise ValueError("at least one uploaded file is required")

    scenario = _detect_scenario(uploaded_files)
    if scenario == "C_word_datamap":
        return _load_scenario_c(uploaded_files)
    if scenario == "B_combined_xlsx":
        xlsx_files = [
            file for file in uploaded_files if file.name.lower().endswith(".xlsx")
        ]
        if len(xlsx_files) != 1:
            raise ValueError("combined xlsx scenario requires exactly one .xlsx file")
        return _load_scenario_b(xlsx_files[0])
    return _load_scenario_a(uploaded_files)


def _detect_scenario(files: list[UploadedFile]) -> str:
    names = [file.name.lower() for file in files]
    extensions = {Path(name).suffix for name in names}

    if ".docx" in extensions:
        return "C_word_datamap"

    xlsx_files = [file for file in files if file.name.lower().endswith(".xlsx")]
    csv_files = [file for file in files if file.name.lower().endswith(".csv")]

    if len(xlsx_files) == 1 and not csv_files:
        return _probe_xlsx_for_scenario(xlsx_files[0])

    return "A_separate_files"


def _probe_xlsx_for_scenario(file: UploadedFile) -> str:
    content = _read_upload(file)
    workbook = load_workbook(stdlib_io.BytesIO(content), read_only=True)
    try:
        sheets = [sheet_name.lower() for sheet_name in workbook.sheetnames]
    finally:
        workbook.close()

    has_data = any(
        any(keyword in sheet_name for keyword in DATA_SHEET_KEYWORDS)
        for sheet_name in sheets
    )
    has_map = any(
        any(keyword in sheet_name for keyword in MAP_SHEET_KEYWORDS)
        for sheet_name in sheets
    )

    if has_data and has_map:
        return "B_combined_xlsx"
    if len(sheets) >= 2:
        return "B_combined_xlsx"
    return "A_separate_files"


def _load_scenario_a(
    files: list[UploadedFile],
) -> tuple[DataMap, pd.DataFrame, LoadReport]:
    datamap_file, raw_file = _identify_file_roles(files)
    dm_path = _write_upload_to_temp(datamap_file, Path(datamap_file.name).suffix)
    raw_path = _write_upload_to_temp(raw_file, Path(raw_file.name).suffix)

    try:
        data_map = parse_datamap(dm_path)
        raw_df, _quality_report = decode_raw_data(raw_path, data_map)
        raw_df = _normalise_dataframe(raw_df)
    finally:
        _safe_unlink(dm_path)
        _safe_unlink(raw_path)

    return data_map, raw_df, LoadReport(
        scenario="A_separate_files",
        raw_data_source=raw_file.name,
        datamap_source=datamap_file.name,
        raw_rows=int(len(raw_df)),
        raw_columns=int(len(raw_df.columns)),
        questions_parsed=int(len(data_map["questions"])),
        parser_warnings=data_map["parser_warnings"],
        detection_notes=["Two files detected: separate raw data + data map"],
    )


def _load_scenario_b(
    combined_file: UploadedFile,
) -> tuple[DataMap, pd.DataFrame, LoadReport]:
    content = _read_upload(combined_file)
    sheets = _workbook_sheet_names(content)
    data_sheet, map_sheet = _identify_combined_sheets(sheets)
    detection_notes = [
        "Combined xlsx detected. "
        f"Data sheet: {data_sheet!r}, Map sheet: {map_sheet!r}",
    ]

    dm_path = _write_datamap_sheet_as_sheet1(content, map_sheet)
    combined_path = _write_bytes_to_temp(content, ".xlsx")

    try:
        data_map = parse_datamap(dm_path)
        raw_df, _quality_report = decode_raw_data(combined_path, data_map)
        raw_df = _normalise_dataframe(raw_df)
        manual_cohort_input = detect_manual_cohort_sheet(
            content,
            raw_df,
            respondent_id_column="record",
        )
    finally:
        _safe_unlink(dm_path)
        _safe_unlink(combined_path)

    if manual_cohort_input is not None:
        detection_notes.append(
            "Manual cohort sheet detected: "
            f"{manual_cohort_input.winner_count} winners, "
            f"{manual_cohort_input.laggard_count} laggards"
        )

    return data_map, raw_df, LoadReport(
        scenario="B_combined_xlsx",
        raw_data_source=f"sheet:{data_sheet}",
        datamap_source=f"sheet:{map_sheet}",
        raw_rows=int(len(raw_df)),
        raw_columns=int(len(raw_df.columns)),
        questions_parsed=int(len(data_map["questions"])),
        parser_warnings=data_map["parser_warnings"],
        detection_notes=detection_notes,
        manual_cohort_input=manual_cohort_input,
    )


def detect_manual_cohort_sheet(
    workbook_content: bytes,
    raw_df: pd.DataFrame,
    respondent_id_column: str | None = None,
) -> ManualCohortInput | None:
    """Return a parsed manual cohort sheet from a combined workbook, if present."""

    workbook = load_workbook(stdlib_io.BytesIO(workbook_content), read_only=True, data_only=True)
    try:
        sheet_name = _find_manual_cohort_sheet_name(workbook.sheetnames)
        if sheet_name is None:
            return None
        primary_column = _respondent_id_column(raw_df, respondent_id_column)
        respondent_ids = _respondent_id_values(raw_df, primary_column)
        parsed = _parse_manual_cohort_worksheet(
            workbook[sheet_name],
            valid_uuids=respondent_ids,
            source=f"embedded sheet:{sheet_name}",
            id_column=primary_column,
        )
        fallback_column = _manual_cohort_fallback_column(raw_df, primary_column, parsed)
        if fallback_column is None:
            return parsed
        return _parse_manual_cohort_worksheet(
            workbook[sheet_name],
            valid_uuids=_respondent_id_values(raw_df, fallback_column),
            source=f"embedded sheet:{sheet_name}",
            id_column=fallback_column,
        )
    finally:
        workbook.close()


def parse_manual_cohort_workbook(
    workbook_source: str | Path | bytes,
    valid_uuids: Iterable[Any] | None = None,
    source: str = "separate upload",
    id_column: str | None = None,
) -> ManualCohortInput:
    """Parse a standalone or combined Winners & Laggards workbook."""

    if isinstance(workbook_source, bytes):
        source_obj: str | Path | stdlib_io.BytesIO = stdlib_io.BytesIO(workbook_source)
    else:
        source_obj = workbook_source
    workbook = load_workbook(source_obj, read_only=True, data_only=True)
    try:
        sheet_name = _find_manual_cohort_sheet_name(workbook.sheetnames)
        if sheet_name is None:
            if len(workbook.sheetnames) != 1:
                raise ValueError("manual cohort workbook must contain a Winners & Laggards sheet")
            sheet_name = workbook.sheetnames[0]
        return _parse_manual_cohort_worksheet(
            workbook[sheet_name],
            valid_uuids=valid_uuids,
            source=source,
            id_column=id_column,
        )
    finally:
        workbook.close()


def _find_manual_cohort_sheet_name(sheet_names: list[str]) -> str | None:
    for sheet_name in sheet_names:
        if _manual_key(sheet_name) in MANUAL_COHORT_SHEET_KEYS:
            return sheet_name
    return None


def _manual_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _respondent_id_values(
    raw_df: pd.DataFrame,
    respondent_id_column: str | None,
) -> set[str]:
    candidate = _respondent_id_column(raw_df, respondent_id_column)
    if candidate is None or candidate not in raw_df.columns:
        return set()
    return {
        cleaned
        for cleaned in (_clean_uuid(value) for value in raw_df[candidate].tolist())
        if cleaned is not None
    }


def _respondent_id_column(
    raw_df: pd.DataFrame,
    respondent_id_column: str | None,
) -> str | None:
    candidate = respondent_id_column if respondent_id_column in raw_df.columns else None
    if candidate is None and "record" in raw_df.columns:
        candidate = "record"
    if candidate is None and len(raw_df.columns) > 0:
        candidate = str(raw_df.columns[0])
    return candidate


def _manual_cohort_fallback_column(
    raw_df: pd.DataFrame,
    primary_column: str | None,
    parsed: ManualCohortInput,
) -> str | None:
    if parsed.winner_uuids or parsed.laggard_uuids or not parsed.invalid_uuids:
        return None
    primary_lower = str(primary_column or "").lower()
    for column in raw_df.columns:
        column_text = str(column)
        if column_text.lower() == primary_lower:
            continue
        if column_text.lower() in {"uuid", "respondent_id", "id"}:
            values = _respondent_id_values(raw_df, column_text)
            reparsed_valid = [value for value in parsed.invalid_uuids if value in values]
            if reparsed_valid:
                return column_text
    return None


def _parse_manual_cohort_worksheet(
    worksheet: Any,
    valid_uuids: Iterable[Any] | None,
    source: str,
    id_column: str | None,
) -> ManualCohortInput:
    header_row = next(
        worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
        (),
    )
    winner_col = _find_manual_cohort_header(header_row, "winner")
    laggard_col = _find_manual_cohort_header(header_row, "laggard")
    if winner_col is None or laggard_col is None:
        raise ValueError("manual cohort sheet must have Winners and Laggards columns")

    winners: list[str] = []
    laggards: list[str] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        winner_value = _clean_uuid(row[winner_col] if winner_col < len(row) else None)
        laggard_value = _clean_uuid(row[laggard_col] if laggard_col < len(row) else None)
        if winner_value is not None:
            winners.append(winner_value)
        if laggard_value is not None:
            laggards.append(laggard_value)

    valid_set = None
    if valid_uuids is not None:
        valid_set = {
            cleaned
            for cleaned in (_clean_uuid(value) for value in valid_uuids)
            if cleaned is not None
        }
    invalid: list[str] = []
    if valid_set is not None:
        valid_winners = []
        for value in winners:
            if value in valid_set:
                valid_winners.append(value)
            else:
                invalid.append(value)
        valid_laggards = []
        for value in laggards:
            if value in valid_set:
                valid_laggards.append(value)
            else:
                invalid.append(value)
        winners = valid_winners
        laggards = valid_laggards

    overlap = _preserve_order_intersection(winners, laggards)
    return ManualCohortInput(
        winner_uuids=tuple(dict.fromkeys(winners)),
        laggard_uuids=tuple(dict.fromkeys(laggards)),
        invalid_uuids=tuple(dict.fromkeys(invalid)),
        overlap_uuids=tuple(overlap),
        source=source,
        id_column=id_column,
    )


def _find_manual_cohort_header(header_row: tuple[Any, ...], target: str) -> int | None:
    for index, value in enumerate(header_row):
        key = _manual_key(value)
        if target == "winner" and key in {"winner", "winners", "winneruuid", "winnersuuid"}:
            return index
        if target == "laggard" and key in {
            "laggard",
            "laggards",
            "laggarduuid",
            "laggardsuuid",
            "loser",
            "losers",
        }:
            return index
    return None


def _clean_uuid(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if not text or text == "0":
        return None
    return text


def _preserve_order_intersection(left: list[str], right: list[str]) -> list[str]:
    right_set = set(right)
    seen: set[str] = set()
    result: list[str] = []
    for value in left:
        if value in right_set and value not in seen:
            result.append(value)
            seen.add(value)
    return result


def _load_scenario_c(
    files: list[UploadedFile],
) -> tuple[DataMap, pd.DataFrame, LoadReport]:
    docx_files = [file for file in files if file.name.lower().endswith(".docx")]
    raw_files = [file for file in files if not file.name.lower().endswith(".docx")]
    if len(docx_files) != 1 or len(raw_files) != 1:
        raise ValueError("word datamap scenario requires one .docx and one raw file")

    docx_file = docx_files[0]
    raw_file = raw_files[0]
    doc_path = _write_upload_to_temp(docx_file, ".docx")
    raw_path = _write_upload_to_temp(raw_file, Path(raw_file.name).suffix)

    try:
        data_map = parse_word_survey(doc_path)
        raw_df, _quality_report = decode_raw_data(raw_path, data_map)
        raw_df = _normalise_dataframe(raw_df)
    finally:
        _safe_unlink(doc_path)
        _safe_unlink(raw_path)

    return data_map, raw_df, LoadReport(
        scenario="C_word_datamap",
        raw_data_source=raw_file.name,
        datamap_source=docx_file.name,
        raw_rows=int(len(raw_df)),
        raw_columns=int(len(raw_df.columns)),
        questions_parsed=int(len(data_map["questions"])),
        parser_warnings=data_map["parser_warnings"],
        detection_notes=[
            "Word document detected as survey/datamap. "
            "Questions auto-parsed from document structure."
        ],
    )


def _identify_file_roles(files: list[UploadedFile]) -> tuple[UploadedFile, UploadedFile]:
    candidates = [
        file
        for file in files
        if Path(file.name.lower()).suffix in {".csv", ".xlsx"}
    ]
    if len(candidates) < 2:
        raise ValueError("separate-file scenario requires raw data and data map files")

    datamap_by_name = [
        file for file in candidates if _name_contains(file.name, DATAMAP_NAME_KEYWORDS)
    ]
    raw_by_name = [
        file
        for file in candidates
        if _name_contains(file.name, RAW_NAME_KEYWORDS)
        and file not in datamap_by_name
    ]
    if len(datamap_by_name) == 1:
        datamap_file = datamap_by_name[0]
        raw_file = raw_by_name[0] if raw_by_name else _first_other(candidates, datamap_file)
        return datamap_file, raw_file

    parsed_scores: list[tuple[int, UploadedFile]] = []
    for file in candidates:
        if not file.name.lower().endswith(".xlsx"):
            continue
        score = _try_parse_datamap_score(file)
        parsed_scores.append((score, file))

    parsed_scores.sort(key=lambda item: item[0], reverse=True)
    if parsed_scores and parsed_scores[0][0] > 0:
        datamap_file = parsed_scores[0][1]
        return datamap_file, _first_other(candidates, datamap_file)

    raise ValueError("could not identify data map file")


def _identify_combined_sheets(sheets: list[str]) -> tuple[str, str]:
    if not sheets:
        raise ValueError("combined workbook has no sheets")

    data_sheet = None
    map_sheet = None
    for sheet_name in sheets:
        lowered = sheet_name.lower()
        if data_sheet is None and any(
            keyword in lowered for keyword in DATA_SHEET_KEYWORDS
        ):
            data_sheet = sheet_name
        if map_sheet is None and any(
            keyword in lowered for keyword in MAP_SHEET_KEYWORDS
        ):
            map_sheet = sheet_name

    if data_sheet is None:
        data_sheet = sheets[0]
    if map_sheet is None:
        map_sheet = sheets[1] if len(sheets) > 1 else sheets[0]
    return data_sheet, map_sheet


def _try_parse_datamap_score(file: UploadedFile) -> int:
    temp_path = _write_upload_to_temp(file, Path(file.name).suffix)
    try:
        try:
            return len(parse_datamap(temp_path)["questions"])
        except Exception:
            return 0
    finally:
        _safe_unlink(temp_path)


def _write_datamap_sheet_as_sheet1(content: bytes, map_sheet: str) -> str:
    workbook = load_workbook(stdlib_io.BytesIO(content))
    try:
        if map_sheet not in workbook.sheetnames:
            raise ValueError(f"map sheet {map_sheet!r} not found")
        if map_sheet != "Sheet1":
            if "Sheet1" in workbook.sheetnames:
                workbook["Sheet1"].title = "__original_Sheet1__"
            workbook[map_sheet].title = "Sheet1"
        path = _empty_temp_path(".xlsx")
        workbook.save(path)
        return path
    finally:
        workbook.close()


def _workbook_sheet_names(content: bytes) -> list[str]:
    workbook = load_workbook(stdlib_io.BytesIO(content), read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _name_contains(name: str, keywords: set[str]) -> bool:
    lowered = name.lower()
    return any(keyword in lowered for keyword in keywords)


def _first_other(files: list[UploadedFile], excluded: UploadedFile) -> UploadedFile:
    for file in files:
        if file is not excluded:
            return file
    raise ValueError("raw data file not found")


def _read_upload(file: UploadedFile) -> bytes:
    file.seek(0)
    content = file.read()
    file.seek(0)
    return content


def _write_upload_to_temp(file: UploadedFile, suffix: str) -> str:
    return _write_bytes_to_temp(_read_upload(file), suffix)


def _write_bytes_to_temp(content: bytes, suffix: str) -> str:
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temp_file:
        temp_file.write(content)
        return temp_file.name


def _empty_temp_path(suffix: str) -> str:
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temp_file:
        return temp_file.name


def _safe_unlink(path: str) -> None:
    try:
        os.unlink(path)
    except FileNotFoundError:
        pass
