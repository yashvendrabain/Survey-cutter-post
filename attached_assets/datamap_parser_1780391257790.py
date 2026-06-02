"""State-machine parser for Survey Insight Engine data map workbooks."""

from __future__ import annotations

from enum import Enum
import re
from typing import Any, Literal, NotRequired, TypedDict

from openpyxl import load_workbook

try:
    from config import (
        DATAMAP_SHEET_NAME,
        QUESTION_HEADER_PATTERN,
        VALUES_LINE_PATTERN,
        OPEN_NUMERIC_LINE,
        OPEN_TEXT_LINE,
        SUB_COLUMN_PATTERN,
    )
except ModuleNotFoundError as exc:
    if exc.name != "config":
        raise
    DATAMAP_SHEET_NAME = "Sheet1"
    QUESTION_HEADER_PATTERN = r"^\[?([A-Za-z][A-Za-z0-9_]*)\]?:\s*(.+)$"
    VALUES_LINE_PATTERN = r"^Values:\s*(-?\d+)\s*-\s*(-?\d+)$"
    OPEN_NUMERIC_LINE = "Open numeric response"
    OPEN_TEXT_LINE = "Open text response"
    SUB_COLUMN_PATTERN = r"^[A-Za-z][A-Za-z0-9_]*$"


QUESTION_HEADER_RE = re.compile(QUESTION_HEADER_PATTERN)
VALUES_LINE_RE = re.compile(VALUES_LINE_PATTERN)
SUB_COLUMN_RE = re.compile(SUB_COLUMN_PATTERN)
PARENT_ROW_OE_RE = re.compile(r"^([A-Za-z][A-Za-z0-9_]*?)r\d+oe$")
PARENT_OE_RE = re.compile(r"^([A-Za-z][A-Za-z0-9_]*?)oe$")
PIPE_TOKEN_RE = re.compile(r"\[(?:pipe|pn)(?::|\s)+([^\]\s]+)\s*\]", re.IGNORECASE)
PIPE_QUESTION_RE = re.compile(r"\[pipe[:\s]+(Q\d+[A-Za-z0-9_]*)\s*\]", re.IGNORECASE)
NUMERIC_LABEL_RE = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*$")


TypeHint = Literal["values_range", "open_numeric", "open_text"]


class ParsedQuestion(TypedDict):
    canonical_id: str
    raw_id: str
    question_text: str
    type_hint: TypeHint | None
    value_range: tuple[int, int] | None
    options: list[tuple[int | str, str]]
    sub_columns: list[tuple[str, str]]
    parent_canonical_id: str | None
    source_row: int
    warnings: list[str]
    children: NotRequired[list["ParsedQuestion"]]
    conditional_on: NotRequired[str | None]
    label_to_numeric_value: NotRequired[dict[str, float]]
    na_label_set: NotRequired[frozenset[str]]
    allowed_numeric_range: NotRequired[tuple[float, float] | None]


class DataMap(TypedDict):
    questions: list[ParsedQuestion]
    source_path: str
    sheet_name: str
    total_rows_in_sheet: int
    parser_warnings: list[str]


class _State(Enum):
    BETWEEN_BLOCKS = "BETWEEN_BLOCKS"
    IN_HEADER = "IN_HEADER"
    IN_TYPE_HINT = "IN_TYPE_HINT"
    IN_OPTIONS = "IN_OPTIONS"


class _Block(TypedDict):
    canonical_id: str
    raw_id: str
    question_text: str
    type_hint: TypeHint | None
    value_range: tuple[int, int] | None
    options: list[tuple[int | str, str]]
    sub_columns: list[tuple[str, str]]
    source_row: int
    warnings: list[str]


def parse_datamap(
    path: str,
    raw_df: Any | None = None,
    *,
    min_questions: int = 2,
) -> DataMap:
    """Parse a data map workbook using the adapter registry."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if DATAMAP_SHEET_NAME not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"{DATAMAP_SHEET_NAME!r} sheet not found. "
                f"Available sheets: {available}"
            )

        setattr(workbook, "_survey_source_path", path)
        from src.adapters.registry import NoAdapterError, get_default_registry

        try:
            return get_default_registry().parse(
                workbook,
                raw_df,
                min_questions=min_questions,
            )
        except NoAdapterError as exc:
            raise ValueError(str(exc)) from exc
    finally:
        workbook.close()


def _detect_datamap_format(worksheet: Any) -> str:
    """Return 'bcn_multicolumn' or 'compact_two_column' or 'unknown'."""

    rows = list(worksheet.iter_rows(min_row=1, max_row=30, values_only=True))
    if not rows:
        return "unknown"

    col_a_nonempty = sum(
        1
        for row in rows
        if len(row) > 0 and row[0] is not None and str(row[0]).strip()
    )
    col_c_plus_nonempty = sum(
        1
        for row in rows
        for index in range(2, min(len(row), 8))
        if row[index] is not None and str(row[index]).strip()
    )

    first_row_vals = [str(value).lower().strip() if value else "" for value in rows[0]]
    if any(
        header in first_row_vals
        for header in ("question id", "question_id", "qid", "question text")
    ):
        return "bcn_multicolumn"

    header_like_rows = sum(
        1
        for row in rows
        if len(row) > 0
        and isinstance(row[0], str)
        and _header_match(str(row[0]).strip()) is not None
    )
    if header_like_rows:
        return "bcn_multicolumn"

    col_a_qids = sum(
        1
        for row in rows
        if len(row) > 0 and row[0] and _is_question_id(str(row[0]).strip())
    )

    if (
        col_a_qids >= 1
        and col_a_nonempty > 1
        and col_c_plus_nonempty < max(1, col_a_nonempty * 0.3)
    ):
        return "compact_two_column"

    if col_a_qids >= 3:
        return "compact_two_column"

    if col_c_plus_nonempty >= max(3, col_a_nonempty * 0.3):
        return "bcn_multicolumn"

    return "unknown"


def _parse_compact_datamap(worksheet: Any) -> list[ParsedQuestion]:
    """Parse the two-column compact format."""

    questions: list[ParsedQuestion] = []
    current_q: ParsedQuestion | None = None

    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        col_a = (
            str(row[0]).strip()
            if len(row) > 0 and row[0] is not None and str(row[0]).strip()
            else ""
        )
        col_b = (
            str(row[1]).strip()
            if len(row) > 1 and row[1] is not None and str(row[1]).strip()
            else ""
        )

        if not col_a and not col_b:
            if current_q is not None:
                questions.append(current_q)
                current_q = None
            continue

        if _is_question_id(col_a):
            if current_q is not None:
                questions.append(current_q)
            current_q = {
                "canonical_id": col_a,
                "raw_id": col_a,
                "question_text": col_b,
                "type_hint": "values_range",
                "value_range": None,
                "options": [],
                "sub_columns": [],
                "parent_canonical_id": None,
                "source_row": row_number,
                "warnings": [],
            }
            conditional_on = _first_pipe_reference(col_b)
            if conditional_on:
                current_q["conditional_on"] = conditional_on
            continue

        if current_q is not None and col_a:
            code: int | str
            code = int(col_a) if col_a.isdigit() else col_a
            current_q["options"].append((code, col_b))

    if current_q is not None:
        questions.append(current_q)

    return questions


def _is_question_id(value: str) -> bool:
    """Detect Q1, Q33, Q33r1, Q33r1c2, S1, D1, and F1-style IDs."""

    if not value or len(value) < 2:
        return False
    normalized = value.strip()
    return bool(re.match(r"^[QSDF]\d+(?:[rc]\d+)*(?:\.\d+)?$", normalized, re.I))


def _normalise_row(row: tuple[Any, ...]) -> tuple[Any | None, ...]:
    return tuple(_normalise_cell(value) for value in row)


def _normalise_cell(value: Any) -> Any | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _cell(row: tuple[Any | None, ...], index: int) -> Any | None:
    if index >= len(row):
        return None
    return row[index]


def _is_blank_row(col_a: Any | None, col_b: Any | None) -> bool:
    return col_a is None and col_b is None


def _header_match(value: Any | None) -> re.Match[str] | None:
    if not isinstance(value, str):
        return None
    if _looks_like_type_hint(value):
        return None
    return QUESTION_HEADER_RE.match(value)


def _looks_like_type_hint(value: str) -> bool:
    return (
        value == OPEN_NUMERIC_LINE
        or value == OPEN_TEXT_LINE
        or value.startswith("Values:")
    )


def _start_block(
    header_value: Any | None, header_match: re.Match[str], row_number: int
) -> _Block:
    if not isinstance(header_value, str):
        raise ValueError("header value must be a string")

    raw_id = header_value.split(":", 1)[0].strip()
    return {
        "canonical_id": header_match.group(1).strip(),
        "raw_id": raw_id,
        "question_text": header_match.group(2).strip(),
        "type_hint": None,
        "value_range": None,
        "options": [],
        "sub_columns": [],
        "source_row": row_number,
        "warnings": [],
    }


def _capture_type_hint(
    block: _Block, value: Any | None, row_number: int
) -> None:
    if not isinstance(value, str):
        block["warnings"].append(f"unrecognised type hint: {value!r}")
        block["type_hint"] = None
        block["value_range"] = None
        return

    if value == OPEN_NUMERIC_LINE:
        block["type_hint"] = "open_numeric"
        block["value_range"] = None
        return

    if value == OPEN_TEXT_LINE:
        block["type_hint"] = "open_text"
        block["value_range"] = None
        return

    values_match = VALUES_LINE_RE.match(value)
    if values_match:
        low = int(values_match.group(1))
        high = int(values_match.group(2))
        block["type_hint"] = "values_range"
        block["value_range"] = (low, high)
        if low > high:
            block["warnings"].append(f"value range inverted: {low} > {high}")
        return

    block["warnings"].append(f"unrecognised type hint: {value!r}")
    block["type_hint"] = None
    block["value_range"] = None


def _capture_option_row(
    block: _Block, col_b: Any | None, col_c: Any | None, row_number: int
) -> None:
    label = _parse_label(col_c, block["warnings"], row_number)
    if label is None:
        return

    if isinstance(col_b, str) and col_b.startswith("[") and col_b.endswith("]"):
        inner_id = col_b[1:-1].strip()
        if not SUB_COLUMN_RE.match(inner_id):
            block["warnings"].append(
                f"sub-column id does not match pattern: {inner_id!r} "
                f"at row {row_number}"
            )
            return
        block["sub_columns"].append((inner_id, label))
        return

    try:
        code = int(str(col_b).strip())
    except (TypeError, ValueError):
        block["warnings"].append(
            f"option code in col B is not an integer: {col_b!r} "
            f"at row {row_number}"
        )
        return

    block["options"].append((code, label))


def _parse_label(
    value: Any | None, warnings: list[str], row_number: int
) -> str | None:
    if not isinstance(value, str) or not value:
        warnings.append(f"option label is empty at row {row_number}")
        return None
    return value


def _finalise_block(block: _Block) -> ParsedQuestion:
    type_hint = block["type_hint"]
    parent_canonical_id = (
        _derive_parent_canonical_id(block["canonical_id"])
        if type_hint == "open_text"
        else None
    )
    question: ParsedQuestion = {
        "canonical_id": block["canonical_id"],
        "raw_id": block["raw_id"],
        "question_text": block["question_text"],
        "type_hint": type_hint,
        "value_range": block["value_range"],
        "options": list(block["options"]),
        "sub_columns": list(block["sub_columns"]),
        "parent_canonical_id": parent_canonical_id,
        "source_row": block["source_row"],
        "warnings": list(block["warnings"]),
    }
    conditional_on = _first_pipe_reference(block["question_text"])
    if conditional_on:
        question["conditional_on"] = conditional_on
    return question


def _derive_parent_canonical_id(canonical_id: str) -> str | None:
    row_oe_match = PARENT_ROW_OE_RE.match(canonical_id)
    if row_oe_match:
        return row_oe_match.group(1)

    oe_match = PARENT_OE_RE.match(canonical_id)
    if oe_match:
        return oe_match.group(1)

    return None


def _first_pipe_reference(question_text: str) -> str | None:
    match = PIPE_TOKEN_RE.search(question_text or "")
    if match is None:
        return None
    return match.group(1).strip()


def _validate_conditional_on(questions: list[ParsedQuestion]) -> list[ParsedQuestion]:
    canonical_ids = {question["canonical_id"] for question in questions}
    for question in questions:
        conditional_on = question.get("conditional_on")
        if conditional_on and conditional_on not in canonical_ids:
            question.setdefault("warnings", []).append(
                f"conditional_on={conditional_on!r} does not resolve to a known question; cleared"
            )
            question["conditional_on"] = None

    for question in questions:
        _recover_conditional_on_from_text(question, canonical_ids)
    return questions


def _recover_conditional_on_from_text(
    question: ParsedQuestion,
    canonical_ids: set[str],
) -> None:
    if question.get("conditional_on"):
        return
    for match in PIPE_QUESTION_RE.finditer(question.get("question_text", "") or ""):
        candidate = match.group(1)
        if candidate in canonical_ids:
            question["conditional_on"] = candidate
            return


def _attach_numeric_label_metadata(
    questions: list[ParsedQuestion],
) -> list[ParsedQuestion]:
    for question in questions:
        label_to_numeric_value, na_label_set, allowed_range = derive_numeric_label_metadata(
            question
        )
        if label_to_numeric_value:
            question["label_to_numeric_value"] = label_to_numeric_value
            question["na_label_set"] = na_label_set
            question["allowed_numeric_range"] = allowed_range
        for child in question.get("children", []) or []:
            child_mapping, child_na_labels, child_range = derive_numeric_label_metadata(
                child
            )
            if child_mapping:
                child["label_to_numeric_value"] = child_mapping
                child["na_label_set"] = child_na_labels
                child["allowed_numeric_range"] = child_range
    return questions


def derive_numeric_label_metadata(
    question: ParsedQuestion,
) -> tuple[dict[str, float], frozenset[str], tuple[float, float] | None]:
    if question.get("type_hint") == "open_text":
        return {}, frozenset(), None
    if question.get("value_range") == (0, 1) and question.get("sub_columns"):
        return {}, frozenset(), None

    labels = _numeric_metadata_option_labels(question)
    if len(labels) < 3:
        return {}, frozenset(), None
    if not all(_label_is_pure_number(label) for label in labels):
        return {}, frozenset(), None

    label_to_numeric_value: dict[str, float] = {}
    for label in labels:
        label_to_numeric_value[label] = float(str(label).strip())

    if not label_to_numeric_value:
        return {}, frozenset(), None

    values = tuple(label_to_numeric_value.values())
    return (
        label_to_numeric_value,
        frozenset(),
        (float(min(values)), float(max(values))),
    )


def _label_is_pure_number(label: str) -> bool:
    if not isinstance(label, str):
        return False
    stripped = label.strip()
    if not stripped or NUMERIC_LABEL_RE.match(stripped) is None:
        return False
    try:
        float(stripped)
    except (TypeError, ValueError):
        return False
    return True


def _numeric_metadata_option_labels(question: ParsedQuestion) -> list[str]:
    labels = [str(label) for _code, label in question.get("options", []) or []]
    if labels:
        return labels

    for child in question.get("children", []) or []:
        for _code, label in child.get("options", []) or []:
            labels.append(str(label))
    return labels


PER_ROW_CHILD_RE = re.compile(r"^(?P<parent>[A-Za-z_]*Q?\d+)r(?P<row>\d+)$")
QUESTION_TEXT_SEPARATOR_RE = re.compile(r"\s+[-\u2013\u2014]\s+")


def _merge_per_row_children(
    questions: list[ParsedQuestion],
) -> list[ParsedQuestion]:
    """Collapse QNrM-style child blocks into one synthetic parent block.

    The rest of the codebase consumes a flat ParsedQuestion list, so the
    synthetic parent also carries the merged child raw columns in sub_columns.
    The original children are retained under a ``children`` key for consumers
    that need row labels and c-column labels separately.
    """

    by_id = {question["canonical_id"]: question for question in questions}
    child_groups: dict[str, list[ParsedQuestion]] = {}
    for question in questions:
        parent_id = _per_row_parent_id(question["canonical_id"])
        if parent_id is None:
            continue
        child_groups.setdefault(parent_id, []).append(question)

    merge_children: dict[str, list[ParsedQuestion]] = {}
    for parent_id, children in child_groups.items():
        if len(children) < 2:
            continue
        existing_parent = by_id.get(parent_id)
        if existing_parent is not None and existing_parent["options"]:
            continue
        merge_children[parent_id] = sorted(
            children,
            key=lambda child: _per_row_sort_key(child["canonical_id"]),
        )

    if not merge_children:
        return questions

    consumed_child_ids = {
        child["canonical_id"]
        for children in merge_children.values()
        for child in children
    }
    emitted_synthetic: set[str] = set()
    merged_questions: list[ParsedQuestion] = []

    for question in questions:
        canonical_id = question["canonical_id"]
        if canonical_id in consumed_child_ids:
            parent_id = _per_row_parent_id(canonical_id)
            if parent_id is not None and parent_id not in emitted_synthetic:
                merged_questions.append(
                    _synthetic_parent_question(
                        parent_id,
                        merge_children[parent_id],
                        by_id.get(parent_id),
                    )
                )
                emitted_synthetic.add(parent_id)
            continue
        if canonical_id in merge_children:
            if canonical_id not in emitted_synthetic:
                merged_questions.append(
                    _synthetic_parent_question(
                        canonical_id,
                        merge_children[canonical_id],
                        question,
                    )
                )
                emitted_synthetic.add(canonical_id)
            continue
        merged_questions.append(question)

    return merged_questions


def _per_row_parent_id(canonical_id: str) -> str | None:
    match = PER_ROW_CHILD_RE.match(canonical_id)
    if match is None:
        return None
    return match.group("parent")


def _per_row_sort_key(canonical_id: str) -> tuple[int, str]:
    match = PER_ROW_CHILD_RE.match(canonical_id)
    if match is None:
        return (10**9, canonical_id)
    return (int(match.group("row")), canonical_id)


def _synthetic_parent_question(
    parent_id: str,
    children: list[ParsedQuestion],
    existing_parent: ParsedQuestion | None = None,
) -> ParsedQuestion:
    first_child = children[0]
    row_label, root_text = _split_per_row_child_text(first_child)
    del row_label
    parent_text = (
        existing_parent["question_text"]
        if existing_parent is not None and existing_parent["question_text"]
        else root_text
    )
    value_range = (
        existing_parent["value_range"]
        if existing_parent is not None and existing_parent["value_range"] is not None
        else first_child["value_range"]
    )
    type_hint = (
        existing_parent["type_hint"]
        if existing_parent is not None and existing_parent["type_hint"] is not None
        else first_child["type_hint"]
    )
    warnings = list(existing_parent["warnings"] if existing_parent is not None else [])
    child_value_ranges = {child["value_range"] for child in children}
    if len(child_value_ranges) > 1:
        warnings.append("merged child blocks have mixed value ranges")

    sub_columns: list[tuple[str, str]] = []
    for child in children:
        child_label, _child_root = _split_per_row_child_text(child)
        if child["sub_columns"]:
            sub_columns.extend(child["sub_columns"])
        else:
            sub_columns.append((child["canonical_id"], child_label))

    parent: ParsedQuestion = {
        "canonical_id": parent_id,
        "raw_id": existing_parent["raw_id"] if existing_parent is not None else parent_id,
        "question_text": parent_text,
        "type_hint": type_hint,
        "value_range": value_range,
        "options": list(existing_parent["options"] if existing_parent is not None else []),
        "sub_columns": sub_columns,
        "parent_canonical_id": (
            existing_parent["parent_canonical_id"]
            if existing_parent is not None
            else None
        ),
        "source_row": (
            existing_parent["source_row"]
            if existing_parent is not None
            else first_child["source_row"]
        ),
        "warnings": warnings,
    }
    conditional_values = {
        child.get("conditional_on")
        for child in children
        if child.get("conditional_on")
    }
    if existing_parent is not None and existing_parent.get("conditional_on"):
        parent["conditional_on"] = existing_parent["conditional_on"]
    elif len(conditional_values) == 1:
        parent["conditional_on"] = conditional_values.pop()
    parent["children"] = children
    return parent


def _split_per_row_child_text(question: ParsedQuestion) -> tuple[str, str]:
    parts = QUESTION_TEXT_SEPARATOR_RE.split(question["question_text"], maxsplit=1)
    if len(parts) == 2:
        row_label = parts[0].strip()
        root_text = parts[1].strip()
        if row_label and root_text:
            return row_label, root_text
    return question["canonical_id"], question["question_text"]


def _row_preview(row: tuple[Any | None, ...]) -> str:
    values = [str(value) for value in row[:3] if value is not None]
    return " | ".join(values)
