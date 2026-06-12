"""
Diagnostic: find exactly where Q1's multi-select option labels (United States...)
are lost. Run from the project root:

    cd ~/workspace/artifacts/survey-insight-engine
    python3 diagnose_q1_labels.py "<path to the .xlsx survey input you uploaded>"

It prints Q1's structure at each pipeline stage so we can see which step drops
the labels. Paste the full output back.
"""
from __future__ import annotations

import sys


def main() -> None:
    if len(sys.argv) < 2:
        print("usage: python3 diagnose_q1_labels.py <survey_input.xlsx>")
        sys.exit(1)
    path = sys.argv[1]

    from src.datamap_parser import parse_datamap  # type: ignore

    # parse_datamap defaults to looking for a sheet literally named "Sheet1".
    # Real survey files name it "Datamap" (or similar), so detect it and pass
    # the hint. Allow an explicit override as the 2nd CLI arg.
    sheet_hint = sys.argv[2] if len(sys.argv) > 2 else None
    if sheet_hint is None:
        try:
            import openpyxl

            names = openpyxl.load_workbook(path, read_only=True).sheetnames
            for candidate in names:
                if str(candidate).strip().lower() in {"datamap", "data map", "map"}:
                    sheet_hint = candidate
                    break
            if sheet_hint is None:
                for candidate in names:
                    if "datamap" in str(candidate).strip().lower().replace(" ", ""):
                        sheet_hint = candidate
                        break
            print(f"[diagnostic] sheets in file: {names}")
            print(f"[diagnostic] using datamap sheet: {sheet_hint!r}\n")
        except Exception as exc:  # noqa: BLE001
            print(f"[diagnostic] could not pre-scan sheets ({exc}); trying default")

    # 1) Raw parse of the datamap
    if sheet_hint is not None:
        datamap = parse_datamap(path, datamap_sheet_hint=sheet_hint)
    else:
        datamap = parse_datamap(path)
    if isinstance(datamap, dict) and "questions" in datamap:
        questions = datamap["questions"]
    else:
        questions = getattr(datamap, "questions", datamap)

    def find(qid: str):
        for q in questions:
            cid = q["canonical_id"] if isinstance(q, dict) else getattr(q, "canonical_id", None)
            if cid == qid:
                return q
        return None

    for qid in ("Q1", "QTerms"):
        q = find(qid)
        print(f"\n================= {qid} =================")
        if q is None:
            print("  NOT FOUND in parsed datamap")
            continue
        get = (lambda k: q.get(k)) if isinstance(q, dict) else (lambda k: getattr(q, k, None))
        print("  STAGE 1 — raw parsed ParsedQuestion:")
        print("    value_range:", get("value_range"))
        print("    options    :", get("options"))
        subs = get("sub_columns") or []
        print(f"    sub_columns ({len(subs)}): first 5 = {subs[:5]}")

    # 2) After label-pattern matching (the adapter that runs before classify)
    print("\n\n========== AFTER apply_label_pattern_matching ==========")
    try:
        from src.adapters.label_pattern_subcolumn import apply_label_pattern_matching  # type: ignore
        for qid in ("Q1", "QTerms"):
            q = find(qid)
            if q is None:
                continue
            raw_order = tuple(
                (q.get("sub_columns") if isinstance(q, dict) else getattr(q, "sub_columns", []))
            )
            ids = tuple(s[0] for s in (raw_order or []))
            transformed = apply_label_pattern_matching(q, ids)
            tget = (lambda k: transformed.get(k)) if isinstance(transformed, dict) else (lambda k: getattr(transformed, k, None))
            subs = tget("sub_columns") or []
            print(f"\n  {qid}: sub_columns ({len(subs)}) first 5 = {subs[:5]}")
            print(f"       options = {tget('options')}")
    except Exception as exc:  # noqa: BLE001
        print("  (could not run adapter directly:", type(exc).__name__, exc, ")")

    # 3) Final spec after classification — the option_map the calculators see
    print("\n\n========== FINAL QuestionSpec (what _multi_select reads) ==========")
    try:
        from src.question_classifier import classify_questions  # type: ignore

        # classify_questions needs the raw-data column names. Read them from the
        # "Raw data" sheet header (first row).
        raw_columns: list[str] = []
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            raw_sheet = None
            for nm in wb.sheetnames:
                if str(nm).strip().lower() in {"raw data", "rawdata", "data"}:
                    raw_sheet = nm
                    break
            if raw_sheet is not None:
                ws = wb[raw_sheet]
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                raw_columns = [str(c) for c in first_row if c is not None]
                print(f"[diagnostic] raw-data sheet: {raw_sheet!r}, {len(raw_columns)} columns")
        except Exception as exc:  # noqa: BLE001
            print(f"[diagnostic] could not read raw-data columns ({exc})")

        schema = classify_questions(datamap, raw_columns)
        specs = getattr(schema, "questions", schema)
        for qid in ("Q1", "QTerms"):
            spec = next((s for s in specs if getattr(s, "canonical_id", None) == qid), None)
            if spec is None:
                print(f"\n  {qid}: NOT FOUND")
                continue
            om = getattr(spec, "option_map", {}) or {}
            items = list(om.items())[:5]
            print(f"\n  {qid}: type={getattr(spec,'question_type',None)}")
            print(f"       option_map ({len(om)}) first 5 = {items}")
            grl = getattr(spec, "grid_row_labels", None) or {}
            print(f"       grid_row_labels ({len(grl)}) first 5 = {list(grl.items())[:5]}")
    except Exception as exc:  # noqa: BLE001
        import traceback

        print("  (could not classify:", type(exc).__name__, exc, ")")
        traceback.print_exc()

    # 4) RUN THE REAL ENGINE on the real data and dump the live selections.
    #    This is the exact object the exporter renders. If labels are correct
    #    here but wrong in Excel, the bug is purely in the exporter/app.
    print("\n\n========== LIVE compute_single_cuts() selections ==========")
    try:
        import pandas as pd
        import openpyxl
        from src.calculation_log import CalculationLog  # type: ignore
        from src.single_cut import compute_single_cuts  # type: ignore
        from src.question_classifier import classify_questions  # type: ignore

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        raw_sheet = None
        for nm in wb.sheetnames:
            if str(nm).strip().lower() in {"raw data", "rawdata", "data"}:
                raw_sheet = nm
                break
        ws = wb[raw_sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c) if c is not None else "" for c in rows[0]]
        df = pd.DataFrame(rows[1:], columns=header)
        raw_columns = [c for c in header if c]
        print(f"[diagnostic] raw df shape: {df.shape}")

        schema = classify_questions(datamap, raw_columns)
        log = CalculationLog()
        results, skips = compute_single_cuts(schema, df, log)

        for qid in ("Q1", "QTerms"):
            res = next((r for r in results if getattr(r, "question_id", None) == qid), None)
            if res is None:
                print(f"\n  {qid}: no result (skipped?)")
                continue
            sel = getattr(res, "selections", None)
            print(f"\n  {qid}: result type = {type(res).__name__}")
            if sel is None:
                print("       (no .selections attribute)")
                continue
            print(f"       selections keys (first 5) = {list(sel.keys())[:5]}")
            for k in list(sel.keys())[:5]:
                p = sel[k]
                print(f"         {k!r} -> label={p.get('label')!r}  count={p.get('count')}")
    except Exception as exc:  # noqa: BLE001
        import traceback

        print("  (could not run live compute:", type(exc).__name__, exc, ")")
        traceback.print_exc()


if __name__ == "__main__":
    main()
