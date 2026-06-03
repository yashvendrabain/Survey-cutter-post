from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.excel_exporter import export_winners_vs_laggards_workbook
from src.models import (
    DenominatorPolicy,
    GridRatedResult,
    GridRatedRow,
    MultiSelectResult,
    QuestionSpec,
    QuestionType,
    SegmentDefinition,
    SingleSelectResult,
    SurveySchema,
)
from src.single_cut import compute_multi_select, compute_single_select


def _fixture() -> tuple[Path, SurveySchema, pd.DataFrame, list, SegmentDefinition, CalculationLog]:
    schema = SurveySchema(
        questions=(
            QuestionSpec(
                question_id="QO",
                canonical_id="QO",
                question_text="Outcome",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("QO",),
                option_map={1: "Winner", 2: "Laggard"},
            ),
            QuestionSpec(
                question_id="QO2",
                canonical_id="QO2",
                question_text="Override outcome",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("QO2",),
                option_map={1: "Winner", 2: "Laggard"},
            ),
            QuestionSpec(
                question_id="Q1",
                canonical_id="Q1",
                question_text="Choice",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q1",),
                option_map={1: "A", 2: "B"},
            ),
            QuestionSpec(
                question_id="QG",
                canonical_id="QG",
                question_text="Rated grid",
                question_type=QuestionType.GRID_RATED,
                raw_columns=("QGr1c1", "QGr2c1"),
                option_map={},
                grid_row_labels={"QGr1c1": "Row 1", "QGr2c1": "Row 2"},
            ),
        ),
        respondent_id_column="rid",
        total_respondents=4,
        source_datamap_path="map.xlsx",
        source_rawdata_path="raw.xlsx",
        parsed_at=datetime.now(timezone.utc),
    )
    df = pd.DataFrame(
        {
            "rid": [1, 2, 3, 4],
            "QO": [1, 1, 2, 2],
            "QO2": [2, 2, 1, 1],
            "Q1": [1, 2, 1, 2],
            "QGr1c1": [8, 9, 3, 4],
            "QGr2c1": [7, 8, 2, 3],
        }
    )
    results = [
        SingleSelectResult(
            question_id="Q1",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                1: {"label": "A", "count": 2, "rate": 0.5},
                2: {"label": "B", "count": 2, "rate": 0.5},
            },
        ),
        GridRatedResult(
            question_id="QG",
            question_type=QuestionType.GRID_RATED,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            question_text="Rated grid",
            column_headers=["Score"],
            rows=(
                GridRatedRow(
                    row_id="QGr1",
                    row_label="Row 1",
                    means_per_column=[6.0],
                    valid_n_per_column=[4],
                ),
                GridRatedRow(
                    row_id="QGr2",
                    row_label="Row 2",
                    means_per_column=[5.0],
                    valid_n_per_column=[4],
                ),
            ),
            total_respondents=4,
            total_responses=8,
            show_delta=False,
        ),
    ]
    output_path = Path(tempfile.gettempdir()) / "winners_vs_laggards_test.xlsx"
    log = CalculationLog()
    segment = SegmentDefinition(
        "QO",
        "categorical",
        winner_values=(1,),
        laggard_values=(2,),
        winner_label="Winners",
        loser_label="Laggards",
        laggard_label="Laggards",
    )
    return output_path, schema, df, results, segment, log


def _build_workbook(segment: SegmentDefinition | None = None, themes: dict | None = None):
    output_path, schema, df, results, default_segment, log = _fixture()
    export_winners_vs_laggards_workbook(
        output_path=output_path,
        decoded_df=df,
        schema=schema,
        single_cut_results=results,
        segment_definition=segment or default_segment,
        laggard_segment_definition=segment or default_segment,
        themes=themes
        or {"themes": [{"name": "Demographics", "question_ids": ["Q1", "QG"]}]},
        calculation_log=log,
    )
    return output_path, load_workbook(output_path, data_only=False), log


def _data_sheet(workbook):
    return workbook["Demographics"]


def _find_row(sheet, column: int, value: str) -> int:
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=column).value == value:
            return row
    raise AssertionError(f"{value!r} not found in column {column}")


def _single_select_header_row(sheet) -> int:
    return _find_row(sheet, 1, "Option ID")


def _single_select_data_row(sheet) -> int:
    return _single_select_header_row(sheet) + 1


def _raw_column_values(workbook, header: str) -> list:
    raw = workbook["_RawData"]
    headers = [raw.cell(row=1, column=col).value for col in range(1, raw.max_column + 1)]
    column = headers.index(header) + 1
    return [raw.cell(row=row, column=column).value for row in range(2, raw.max_row + 1)]


def _conditional_formatting_rules(sheet, cell_range: str) -> list:
    rules = []
    for cf_range, cf_rules in sheet.conditional_formatting._cf_rules.items():
        if cell_range in str(cf_range):
            rules.extend(cf_rules)
    return rules


def _build_mask_fixture_workbook(winner_n: int = 30, laggard_n: int = 70, other_n: int = 0):
    output_path, schema, _df, results, default_segment, log = _fixture()
    rows = winner_n + laggard_n + other_n
    df = pd.DataFrame(
        {
            "rid": list(range(1, rows + 1)),
            "QO": [1] * winner_n + [2] * laggard_n + [3] * other_n,
            "QO2": [2] * winner_n + [1] * laggard_n + [3] * other_n,
            "Q1": ([1, 2] * ((rows + 1) // 2))[:rows],
            "QGr1c1": [8] * winner_n + [3] * laggard_n + [5] * other_n,
            "QGr2c1": [7] * winner_n + [2] * laggard_n + [5] * other_n,
        }
    )
    export_winners_vs_laggards_workbook(
        output_path=output_path,
        decoded_df=df,
        schema=schema,
        single_cut_results=results,
        segment_definition=default_segment,
        laggard_segment_definition=default_segment,
        themes={"themes": [{"name": "Demographics", "question_ids": ["Q1", "QG"]}]},
        calculation_log=log,
    )
    return output_path, load_workbook(output_path, data_only=False), log


def _m01_like_question(
    qid: str,
    question_type: QuestionType,
    *,
    raw_columns: tuple[str, ...] | None = None,
    option_map: dict | None = None,
    text: str | None = None,
) -> QuestionSpec:
    return QuestionSpec(
        question_id=qid,
        canonical_id=qid,
        question_text=text or qid,
        question_type=question_type,
        raw_columns=raw_columns or (qid,),
        option_map=option_map or {},
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
    )


def _fill_binary_option(
    values: list,
    start: int,
    answered_n: int,
    selected_n: int,
) -> None:
    for offset in range(answered_n):
        values[start + offset] = 1 if offset < selected_n else 0


def _build_m01_like_wvl_workbook():
    winner_n = 100
    laggard_n = 100
    rows = winner_n + laggard_n
    q4 = [None] * rows
    q5 = [None] * rows
    q2 = [None] * rows
    _fill_binary_option(q4, 0, 70, 42)
    _fill_binary_option(q4, winner_n, 84, 39)
    _fill_binary_option(q5, 0, 68, 11)
    _fill_binary_option(q5, winner_n, 79, 7)
    q2[:67] = [1] * 29 + [2] * 38
    q2[winner_n : winner_n + 80] = [1] * 24 + [2] * 56
    schema = SurveySchema(
        questions=(
            _m01_like_question(
                "QO",
                QuestionType.SINGLE_SELECT,
                option_map={1: "Winner", 2: "Laggard"},
                text="Outcome",
            ),
            _m01_like_question(
                "Q4",
                QuestionType.MULTI_SELECT_BINARY,
                raw_columns=("Q4r1",),
                option_map={"Q4r1": "GenAI tools"},
                text="GenAI use cases",
            ),
            _m01_like_question(
                "Q2",
                QuestionType.SINGLE_SELECT,
                option_map={1: "Mostly regional", 2: "Other"},
                text="Geographic footprint",
            ),
            _m01_like_question(
                "Q5",
                QuestionType.MULTI_SELECT_BINARY,
                raw_columns=("Q5r1",),
                option_map={"Q5r1": "Talent shortage"},
                text="Barriers",
            ),
        ),
        respondent_id_column="rid",
        total_respondents=rows,
        source_datamap_path="map.xlsx",
        source_rawdata_path="raw.xlsx",
        parsed_at=datetime.now(timezone.utc),
    )
    df = pd.DataFrame(
        {
            "rid": list(range(1, rows + 1)),
            "QO": [1] * winner_n + [2] * laggard_n,
            "Q4r1": q4,
            "Q2": q2,
            "Q5r1": q5,
        }
    )
    log = CalculationLog()
    results = [
        compute_multi_select(schema.get_question("Q4"), df, log),
        compute_single_select(schema.get_question("Q2"), df, log),
        compute_multi_select(schema.get_question("Q5"), df, log),
    ]
    segment = SegmentDefinition(
        "QO",
        "categorical",
        winner_values=(1,),
        laggard_values=(2,),
        winner_label="Winners",
        loser_label="Laggards",
        laggard_label="Laggards",
    )
    output_path = Path(tempfile.gettempdir()) / "winners_vs_laggards_m01_like.xlsx"
    export_winners_vs_laggards_workbook(
        output_path=output_path,
        decoded_df=df,
        schema=schema,
        single_cut_results=results,
        segment_definition=segment,
        laggard_segment_definition=segment,
        themes={"themes": [{"name": "M01", "question_ids": ["Q4", "Q2", "Q5"]}]},
        calculation_log=CalculationLog(),
    )
    workbook = load_workbook(output_path, data_only=True)
    formula_workbook = load_workbook(output_path, data_only=False)
    return output_path, workbook, formula_workbook


class TestWinnersVsLaggardsWorkbook(unittest.TestCase):
    def test_workbook_is_produced(self) -> None:
        output_path, _workbook, _log = _build_workbook()
        self.assertTrue(output_path.exists())

    def test_run_summary_contains_labels_and_sizes(self) -> None:
        _path, workbook, _log = _build_workbook()
        values = [cell.value for row in workbook["Run_Summary"].iter_rows() for cell in row]
        self.assertIn("Winner cohort size:", values)
        self.assertIn("Laggard cohort size:", values)
        self.assertIn("Winners", values)

    def test_single_select_sheet_has_documented_columns(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        header_row = _single_select_header_row(sheet)
        headers = [sheet.cell(row=header_row, column=col).value for col in range(1, 12)]
        self.assertEqual(
            headers,
            [
                "Option ID",
                "Option",
                "Count",
                "%",
                "Count",
                "%",
                "Count",
                "%",
                "Count",
                "%",
                "Who prioritizes more?",
            ],
        )

    def test_wvl_workbook_column_headers_simplified(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        header_row = _single_select_header_row(sheet)

        self.assertEqual(
            [sheet.cell(row=header_row, column=col).value for col in range(3, 11)],
            ["Count", "%", "Count", "%", "Count", "%", "Count", "%"],
        )
        old_headers = {
            "Winners count",
            "Winners %",
            "Laggards count",
            "Laggards %",
            "Others count",
            "Others %",
            "Total count",
            "Total %",
        }
        self.assertTrue(
            old_headers.isdisjoint(
                {
                    sheet.cell(row=header_row, column=col).value
                    for col in range(1, sheet.max_column + 1)
                }
            )
        )

    def test_wvl_workbook_cohort_band_header_present(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        group_row = _single_select_header_row(sheet) - 1

        self.assertEqual(sheet.cell(row=group_row, column=3).value, "Winners")
        self.assertEqual(sheet.cell(row=group_row, column=5).value, "Laggards")
        self.assertEqual(sheet.cell(row=group_row, column=7).value, "Others")
        self.assertEqual(sheet.cell(row=group_row, column=9).value, "Total")
        self.assertEqual(sheet.cell(row=group_row, column=3).fill.fgColor.rgb, "00CC0000")
        self.assertEqual(sheet.cell(row=group_row, column=3).font.color.rgb, "00FFFFFF")

    def test_wvl_workbook_formulas_intact_after_header_simplification(self) -> None:
        output_path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)

        formula_expectations = {
            3: ("Q1_data", '"1"', "winners_mask_data"),
            5: ("Q1_data", '"1"', "laggards_mask_data"),
            7: ("Q1_data", '"1"', "others_mask_data"),
            9: ("Q1_data", '"1"', None),
        }
        for column, (data_name, criterion, mask_name) in formula_expectations.items():
            formula = str(sheet.cell(data_row, column).value)
            self.assertTrue(formula.startswith("=COUNTIFS("))
            self.assertIn(data_name, formula)
            self.assertIn(criterion, formula)
            if mask_name is not None:
                self.assertIn(mask_name, formula)
            self.assertNotIn("Winners count", formula)
            self.assertNotIn("Laggards count", formula)
            self.assertNotIn("Others count", formula)
            self.assertNotIn("Total count", formula)

        self.assertEqual(sheet.cell(data_row, 4).value, f"=IFERROR(C{data_row}/$C${data_row + 3},0)")
        self.assertEqual(sheet.cell(data_row, 6).value, f"=IFERROR(E{data_row}/$C${data_row + 4},0)")
        self.assertEqual(sheet.cell(data_row, 8).value, f"=IFERROR(G{data_row}/$C${data_row + 5},0)")
        self.assertEqual(sheet.cell(data_row, 10).value, f"=IFERROR(I{data_row}/$C${data_row + 6},0)")

        values_workbook = load_workbook(output_path, data_only=True)
        try:
            values_sheet = _data_sheet(values_workbook)
            for column in (3, 5, 7, 9):
                self.assertNotIn(values_sheet.cell(data_row, column).value, ("#NAME?", "#REF!"))
        finally:
            values_workbook.close()

    def test_extended_single_select_layout_has_11_columns(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        header_row = _single_select_header_row(sheet)
        self.assertEqual(sheet.cell(row=header_row, column=11).value, "Who prioritizes more?")
        self.assertIsNone(sheet.cell(row=header_row, column=12).value)

    def test_count_percentage_and_delta_cells_are_formulas(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)
        self.assertTrue(str(sheet.cell(data_row, 3).value).startswith("="))
        self.assertTrue(str(sheet.cell(data_row, 4).value).startswith("=IFERROR("))
        self.assertTrue(str(sheet.cell(data_row, 7).value).startswith("="))
        self.assertTrue(str(sheet.cell(data_row, 9).value).startswith("="))
        self.assertTrue(str(sheet.cell(data_row, 11).value).startswith("=IF(AND("))
        self.assertIn("winners_mask_data", sheet.cell(data_row, 3).value)
        self.assertIn("laggards_mask_data", sheet.cell(data_row, 5).value)
        self.assertIn("others_mask_data", sheet.cell(data_row, 7).value)

    def test_single_select_others_count_formula_uses_others_mask(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        formula = str(sheet.cell(_single_select_data_row(sheet), 7).value)
        self.assertIn("others_mask_data", formula)
        self.assertIn("Q1_data", formula)

    def test_who_prioritizes_more_formula_is_nested_if(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        formula = str(sheet.cell(_single_select_data_row(sheet), 11).value)
        self.assertTrue(formula.startswith("=IF(AND("))
        self.assertIn('"Winners"', formula)
        self.assertIn('"Laggards"', formula)
        self.assertIn('"Others"', formula)
        self.assertIn('"Tied"', formula)

    def test_who_prioritizes_more_threshold_is_2pp(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        formula = str(sheet.cell(_single_select_data_row(sheet), 11).value)
        self.assertIn("0.02", formula)

    def test_single_select_countifs_uses_option_code_not_label(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)
        formula = str(sheet.cell(data_row, 3).value)
        self.assertIn('Q1_data,"1"', formula)
        self.assertNotIn('Q1_data,"A"', formula)

    def test_single_select_cached_pct_uses_cohort_answered_denominator(self) -> None:
        output_path, schema, df, _results, segment, _log = _fixture()
        df = df.copy()
        df["Q1"] = [1, None, 1, 2]
        result = compute_single_select(schema.get_question("Q1"), df, CalculationLog())
        export_winners_vs_laggards_workbook(
            output_path=output_path,
            decoded_df=df,
            schema=schema,
            single_cut_results=[result],
            segment_definition=segment,
            laggard_segment_definition=segment,
            themes={"themes": [{"name": "Demographics", "question_ids": ["Q1"]}]},
            calculation_log=CalculationLog(),
        )
        workbook = load_workbook(output_path, data_only=True)
        try:
            sheet = _data_sheet(workbook)
            data_row = _single_select_data_row(sheet)
            winner_total_row = data_row + 3
            laggard_total_row = data_row + 4
            self.assertEqual(sheet.cell(data_row, 3).value, 1)
            self.assertAlmostEqual(sheet.cell(data_row, 4).value, 1.0)
            self.assertEqual(sheet.cell(winner_total_row, 3).value, 1)
            self.assertEqual(sheet.cell(data_row, 5).value, 1)
            self.assertAlmostEqual(sheet.cell(data_row, 6).value, 0.5)
            self.assertEqual(sheet.cell(laggard_total_row, 3).value, 2)
        finally:
            workbook.close()

    def test_wvl_m01_like_multi_select_uses_cohort_answered_denominator(self) -> None:
        _path, workbook, formula_workbook = _build_m01_like_wvl_workbook()
        try:
            sheet = workbook["M01"]
            formula_sheet = formula_workbook["M01"]
            q4_row = _find_row(sheet, 1, "Q4r1")
            q4_winner_total_row = q4_row + 2
            self.assertEqual(sheet.cell(q4_row, 3).value, 42)
            self.assertAlmostEqual(sheet.cell(q4_row, 4).value * 100, 60.0, places=4)
            self.assertEqual(sheet.cell(q4_row, 5).value, 39)
            self.assertAlmostEqual(sheet.cell(q4_row, 6).value * 100, 46.4286, places=4)
            self.assertEqual(sheet.cell(q4_winner_total_row, 3).value, 70)
            self.assertIn("SUMPRODUCT", formula_sheet.cell(q4_winner_total_row, 3).value)
            self.assertIn("winners_mask_data", formula_sheet.cell(q4_winner_total_row, 3).value)
        finally:
            workbook.close()
            formula_workbook.close()

    def test_wvl_m01_like_single_select_and_lift_cached_values(self) -> None:
        _path, workbook, _formula_workbook = _build_m01_like_wvl_workbook()
        try:
            sheet = workbook["M01"]
            q2_row = _find_row(sheet, 2, "Mostly regional")
            self.assertAlmostEqual(sheet.cell(q2_row, 4).value * 100, 43.2836, places=4)
            self.assertAlmostEqual(sheet.cell(q2_row, 6).value * 100, 30.0, places=4)

            q5_row = _find_row(sheet, 1, "Q5r1")
            winner_pct = sheet.cell(q5_row, 4).value
            laggard_pct = sheet.cell(q5_row, 6).value
            self.assertAlmostEqual(winner_pct * 100, 16.1765, places=4)
            self.assertAlmostEqual(laggard_pct * 100, 8.8608, places=4)
            self.assertAlmostEqual(winner_pct / laggard_pct, 1.83, delta=0.01)
        finally:
            workbook.close()
            _formula_workbook.close()

    def test_grid_rated_layout_uses_formula_means(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        row = next(
            r
            for r in range(1, sheet.max_row + 1)
            if sheet.cell(row=r, column=1).value == "Sub-question ID"
        )
        self.assertTrue(str(sheet.cell(row=row + 1, column=3).value).startswith("=IFERROR(AVERAGEIFS"))
        self.assertTrue(str(sheet.cell(row=row + 1, column=4).value).startswith("=IFERROR(AVERAGEIFS"))
        self.assertTrue(str(sheet.cell(row=row + 1, column=5).value).startswith("=IFERROR(AVERAGEIFS"))
        self.assertTrue(str(sheet.cell(row=row + 1, column=6).value).startswith("=IFERROR(AVERAGEIFS"))
        self.assertTrue(str(sheet.cell(row=row + 1, column=7).value).startswith("="))

    def test_grid_rated_extended_layout_has_means_for_3_cohorts(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        row = next(
            r
            for r in range(1, sheet.max_row + 1)
            if sheet.cell(row=r, column=1).value == "Sub-question ID"
        )
        self.assertEqual(sheet.cell(row=row, column=3).value, "Winners mean")
        self.assertEqual(sheet.cell(row=row, column=4).value, "Laggards mean")
        self.assertEqual(sheet.cell(row=row, column=5).value, "Others mean")
        self.assertEqual(sheet.cell(row=row, column=6).value, "Total mean")

    def test_named_ranges_for_masks_exist(self) -> None:
        _path, workbook, _log = _build_workbook()
        self.assertIn("winners_mask_data", workbook.defined_names)
        self.assertIn("laggards_mask_data", workbook.defined_names)
        self.assertIn("others_mask_data", workbook.defined_names)

    def test_others_mask_data_named_range_exists(self) -> None:
        _path, workbook, _log = _build_workbook()
        self.assertIn("others_mask_data", workbook.defined_names)

    def test_winners_mask_data_reflects_winner_membership(self) -> None:
        _path, workbook, _log = _build_mask_fixture_workbook()
        values = _raw_column_values(workbook, "winners_mask_data")
        self.assertEqual(sum(value is True for value in values), 30)
        self.assertEqual(sum(value is False for value in values), 70)

    def test_laggards_mask_data_reflects_laggard_membership(self) -> None:
        _path, workbook, _log = _build_mask_fixture_workbook()
        values = _raw_column_values(workbook, "laggards_mask_data")
        self.assertEqual(sum(value is True for value in values), 70)
        self.assertEqual(sum(value is False for value in values), 30)

    def test_others_mask_count_equals_complement(self) -> None:
        _path, workbook, _log = _build_mask_fixture_workbook(winner_n=30, laggard_n=40, other_n=30)
        winners = _raw_column_values(workbook, "winners_mask_data")
        laggards = _raw_column_values(workbook, "laggards_mask_data")
        others = _raw_column_values(workbook, "others_mask_data")
        self.assertEqual(sum(value is True for value in others), 30)
        self.assertEqual(
            sum(value is True for value in others),
            len(others) - sum(value is True for value in winners) - sum(value is True for value in laggards),
        )

    def test_override_outcome_records_laggard_source_column(self) -> None:
        segment = SegmentDefinition(
            "QO",
            "categorical",
            winner_values=(1,),
            laggard_values=(2,),
            laggard_outcome_question_id="QO2",
        )
        _path, _workbook, log = _build_workbook(segment)
        laggard_records = [
            record for record in log.all_records() if record.metric_name == "laggard_mask"
        ]
        self.assertEqual(laggard_records[-1].source_question_id, "QO2")

    def test_backward_compat_defaults_inverse_laggards(self) -> None:
        segment = SegmentDefinition("QO", "categorical", winner_values=(1,))
        _path, workbook, _log = _build_workbook(segment)
        raw = workbook["_RawData"]
        headers = [raw.cell(row=1, column=col).value for col in range(1, raw.max_column + 1)]
        lag_col = headers.index("laggards_mask_data") + 1
        self.assertEqual([raw.cell(row=row, column=lag_col).value for row in range(2, 6)], [False, False, True, True])

    def test_category_sheet_has_red_banner(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        row = _find_row(sheet, 1, "Demographics")
        self.assertGreater(row, 1)
        self.assertEqual(sheet.cell(row=row, column=1).fill.fgColor.rgb, "00CC0000")
        self.assertEqual(sheet.cell(row=row, column=1).font.color.rgb, "00FFFFFF")

    def test_question_block_includes_filter_ui_rows(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        title_row = _find_row(sheet, 1, "Q1 - Choice")
        self.assertEqual(
            [sheet.cell(row=title_row + 1, column=col).value for col in range(1, 6)],
            ["Per-question filter", "Filter Q", "(None)", "Value", "(All)"],
        )
        self.assertIn("MATCH(Demographics_Q1_F_Q,All_Questions,0)", sheet.cell(title_row + 1, 6).value)
        self.assertEqual(sheet.cell(title_row + 1, 7).value, '="|" & SUBSTITUTE(Demographics_Q1_F_V, ", ", "|") & "|"')
        self.assertEqual(sheet.cell(title_row + 2, 1).value, "Cross-tab by")
        self.assertEqual(sheet.cell(title_row + 2, 3).value, "(None)")
        self.assertTrue(str(sheet.cell(title_row + 3, 1).value).startswith("Note: This question was shown"))
        self.assertIsNone(sheet.cell(title_row + 4, 1).value)
        self.assertIsNone(sheet.cell(title_row + 5, 1).value)
        self.assertEqual(sheet.cell(title_row + 6, 1).value, "Option ID")

    def test_workbook_filter_rows_are_written_before_category_banner(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        self.assertEqual(sheet["A1"].value, "LOCAL FILTERS (override workbook defaults)")
        self.assertEqual(sheet["A2"].value, "Filter")
        self.assertEqual(sheet["A3"].value, "Custom 1 question")
        banner_row = _find_row(sheet, 1, "Demographics")
        self.assertGreater(banner_row, 3)
        self.assertIn("Filters", workbook.sheetnames)

    def test_theme_groups_create_category_sheets_not_one_all_questions_bucket(self) -> None:
        _path, workbook, _log = _build_workbook(
            themes={
                "themes": [
                    {"name": "Demographics", "question_ids": ["Q1"]},
                    {"name": "AI Usage", "question_ids": ["QG"]},
                ]
            }
        )
        self.assertIn("Demographics", workbook.sheetnames)
        self.assertIn("AI Usage", workbook.sheetnames)
        self.assertNotIn("All Questions", workbook.sheetnames)

    def test_data_body_cells_use_arial(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)
        self.assertEqual(sheet.cell(data_row, 3).font.name, "Arial")
        self.assertEqual(sheet.cell(data_row, 4).font.name, "Arial")
        self.assertEqual(sheet.cell(data_row, 7).font.name, "Arial")

    def test_single_select_number_formats_are_applied(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)
        self.assertEqual(sheet.cell(data_row, 3).number_format, "#,##0")
        self.assertEqual(sheet.cell(data_row, 4).number_format, "0.0%")
        self.assertEqual(sheet.cell(data_row, 5).number_format, "#,##0")
        self.assertEqual(sheet.cell(data_row, 6).number_format, "0.0%")
        self.assertEqual(sheet.cell(data_row, 7).number_format, "#,##0")
        self.assertEqual(sheet.cell(data_row, 8).number_format, "0.0%")
        self.assertEqual(sheet.cell(data_row, 9).number_format, "#,##0")
        self.assertEqual(sheet.cell(data_row, 10).number_format, "0.0%")
        self.assertEqual(sheet.cell(data_row, 11).number_format, "@")

    def test_title_bar_has_bain_red_fill(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        title_row = _find_row(sheet, 1, "Q1 - Choice")
        self.assertEqual(sheet.cell(title_row, 1).fill.fgColor.rgb, "FFCC0000")
        self.assertEqual(sheet.cell(title_row, 1).font.color.rgb, "FFFFFFFF")

    def test_percent_columns_have_heatmap_conditional_formatting(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_start = _single_select_data_row(sheet)
        data_end = data_start + 1
        ranges = [str(item) for item in sheet.conditional_formatting]
        for column in ("D", "F", "H", "J"):
            self.assertTrue(any(f"{column}{data_start}:{column}{data_end}" in item for item in ranges))

    def test_existing_heatmap_on_pct_columns_still_present(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_start = _single_select_data_row(sheet)
        data_end = data_start + 1
        ranges = [str(item) for item in sheet.conditional_formatting]
        for column in ("D", "F", "H", "J"):
            self.assertTrue(any(f"{column}{data_start}:{column}{data_end}" in item for item in ranges))

    def test_who_prioritizes_more_conditional_formatting_4_rules(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_start = _single_select_data_row(sheet)
        data_end = data_start + 1
        rules = _conditional_formatting_rules(sheet, f"K{data_start}:K{data_end}")
        self.assertGreaterEqual(sum(getattr(rule, "type", None) == "cellIs" for rule in rules), 4)

    def test_grouped_header_row_has_merged_cells(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        group_row = _single_select_header_row(sheet) - 1
        merged_ranges = {str(range_) for range_ in sheet.merged_cells.ranges}
        self.assertIn(f"C{group_row}:D{group_row}", merged_ranges)
        self.assertIn(f"E{group_row}:F{group_row}", merged_ranges)
        self.assertIn(f"G{group_row}:H{group_row}", merged_ranges)
        self.assertIn(f"I{group_row}:J{group_row}", merged_ranges)

    def test_grid_rated_number_formats_are_applied(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        header_row = next(
            row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value == "Sub-question ID"
        )
        self.assertEqual(sheet.cell(row=header_row + 1, column=3).number_format, "0.00")
        self.assertEqual(sheet.cell(row=header_row + 1, column=4).number_format, "0.00")
        self.assertEqual(sheet.cell(row=header_row + 1, column=5).number_format, "0.00")
        self.assertEqual(sheet.cell(row=header_row + 1, column=6).number_format, "0.00")
        self.assertEqual(sheet.cell(row=header_row + 1, column=7).number_format, "+0.00;-0.00")
        self.assertEqual(sheet.cell(row=header_row + 1, column=8).number_format, "#,##0")
        self.assertEqual(sheet.cell(row=header_row + 1, column=9).number_format, "#,##0")
        self.assertEqual(sheet.cell(row=header_row + 1, column=10).number_format, "#,##0")
        self.assertEqual(sheet.cell(row=header_row + 1, column=11).number_format, "#,##0")
        self.assertEqual(sheet.cell(row=header_row + 1, column=12).number_format, "@")


if __name__ == "__main__":
    unittest.main()
