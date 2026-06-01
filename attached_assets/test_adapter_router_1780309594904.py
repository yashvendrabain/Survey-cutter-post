"""Tests for the default data-map adapter router."""

from __future__ import annotations

import os
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from src.adapters.bcn_multicolumn import BcnMulticolumnAdapter
from src.adapters.compact_two_column import CompactTwoColumnAdapter
from src.adapters.registry import get_default_registry
from src.datamap_parser import parse_datamap
from tests.conftest import DATAMAP_FIXTURE_PATH


def _make_compact_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["Q1", "Employment status"])
    worksheet.append([1, "Full-time"])
    worksheet.append([2, "Part-time"])
    worksheet.append([None, None])
    worksheet.append(["Q2", "Industry"])
    worksheet.append(["a", "Technology"])
    worksheet.append(["b", "Retail"])
    return workbook


def _make_bcn_signal_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["Question ID", "Question Text", "Type"])
    worksheet.append(["Q1", "A question", "Values: 1-2"])
    return workbook


def _save_temp_workbook(workbook: Workbook) -> str:
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    handle.close()
    workbook.save(handle.name)
    workbook.close()
    return handle.name


class TestAdapterRouter(unittest.TestCase):
    def test_router_picks_bcn_for_bcn_sample(self) -> None:
        workbook = load_workbook(DATAMAP_FIXTURE_PATH, read_only=True, data_only=True)
        try:
            adapter, result = get_default_registry().pick_adapter(workbook)
        finally:
            workbook.close()

        self.assertEqual(adapter.name, "bcn_multicolumn")
        self.assertGreaterEqual(result.confidence, 0.6)

    def test_router_picks_bcn_for_raw_subcolumn_heavy_workbook(self) -> None:
        workbook = _make_bcn_signal_workbook()
        raw_columns = [f"Q14s{i}" for i in range(1, 13)]

        adapter, result = get_default_registry().pick_adapter(workbook, raw_columns)

        self.assertEqual(adapter.name, "bcn_multicolumn")
        self.assertGreaterEqual(result.confidence, 0.5)

    def test_router_picks_compact_for_compact_two_column_workbook(self) -> None:
        workbook = _make_compact_workbook()

        adapter, result = get_default_registry().pick_adapter(workbook)

        self.assertEqual(adapter.name, "compact_two_column")
        self.assertGreaterEqual(result.confidence, 0.6)

    def test_router_output_matches_bcn_adapter_on_bcn_sample(self) -> None:
        workbook = load_workbook(DATAMAP_FIXTURE_PATH, read_only=True, data_only=True)
        try:
            setattr(workbook, "_survey_source_path", str(DATAMAP_FIXTURE_PATH))
            direct = BcnMulticolumnAdapter().parse(workbook)
        finally:
            workbook.close()

        routed = parse_datamap(str(DATAMAP_FIXTURE_PATH))

        self.assertEqual(routed, direct)

    def test_router_output_matches_compact_adapter_on_compact_sample(self) -> None:
        path = _save_temp_workbook(_make_compact_workbook())
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                setattr(workbook, "_survey_source_path", path)
                direct = CompactTwoColumnAdapter().parse(workbook)
            finally:
                workbook.close()

            routed = parse_datamap(path)
        finally:
            os.remove(path)

        self.assertEqual(routed, direct)


if __name__ == "__main__":
    unittest.main()
