from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.excel_exporter import export_winners_vs_laggards_workbook
from src.io import (
    ManualCohortInput,
    detect_manual_cohort_sheet,
    parse_manual_cohort_workbook,
)
from src.models import SegmentDefinition
from src.outcome_segmentation import _build_segment_masks
from tests.test_excel_exporter_winners_vs_laggards import _fixture


def _manual_workbook_bytes(sheet_name: str = "Winners & Laggards") -> bytes:
    workbook = Workbook()
    raw = workbook.active
    raw.title = "Raw"
    raw.append(["record", "Q1"])
    for rid in ("R1", "R2", "R3", "R4"):
        raw.append([rid, 1])
    datamap = workbook.create_sheet("DataMap")
    datamap.append(["placeholder"])
    cohort = workbook.create_sheet(sheet_name)
    cohort.append(["Winners", "Laggards"])
    cohort.append(["R1", "R3"])
    cohort.append(["R2", "R4"])
    cohort.append(["0", None])
    path = Path(tempfile.gettempdir()) / "manual_cohort_fixture.xlsx"
    workbook.save(path)
    return path.read_bytes()


class TestManualCohortUpload(unittest.TestCase):
    def test_combined_xlsx_with_wl_sheet_detected(self) -> None:
        raw_df = pd.DataFrame({"record": ["R1", "R2", "R3", "R4"]})
        cohort = detect_manual_cohort_sheet(_manual_workbook_bytes(), raw_df, "record")
        self.assertIsInstance(cohort, ManualCohortInput)
        self.assertEqual(cohort.winner_uuids, ("R1", "R2"))
        self.assertEqual(cohort.laggard_uuids, ("R3", "R4"))
        self.assertIn("embedded sheet", cohort.source)

    def test_separate_wl_xlsx_parsed_correctly(self) -> None:
        path = Path(tempfile.gettempdir()) / "manual_separate.xlsx"
        path.write_bytes(_manual_workbook_bytes("winnerlaggard"))
        cohort = parse_manual_cohort_workbook(path, valid_uuids=["R1", "R2", "R3", "R4"])
        self.assertEqual(cohort.winner_uuids, ("R1", "R2"))
        self.assertEqual(cohort.laggard_uuids, ("R3", "R4"))

    def test_validation_fallback_to_uuid_column(self) -> None:
        raw_df = pd.DataFrame(
            {
                "record": [1, 2, 3, 4],
                "uuid": ["R1", "R2", "R3", "R4"],
            }
        )
        cohort = detect_manual_cohort_sheet(_manual_workbook_bytes(), raw_df, "record")
        self.assertEqual(cohort.winner_uuids, ("R1", "R2"))
        self.assertEqual(cohort.laggard_uuids, ("R3", "R4"))
        self.assertEqual(cohort.invalid_uuids, ())
        self.assertEqual(cohort.id_column, "uuid")

    def test_validation_no_fallback_when_primary_matches(self) -> None:
        path = Path(tempfile.gettempdir()) / "manual_primary_uuid.xlsx"
        path.write_bytes(_manual_workbook_bytes("winnerlaggard"))
        cohort = parse_manual_cohort_workbook(
            path,
            valid_uuids=["R1", "R2", "R3", "R4"],
            id_column="uuid",
        )
        self.assertEqual(cohort.winner_uuids, ("R1", "R2"))
        self.assertEqual(cohort.laggard_uuids, ("R3", "R4"))
        self.assertEqual(cohort.id_column, "uuid")

    def test_invalid_uuids_surfaced_as_warning(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Winners & Laggards"
        worksheet.append(["Winners", "Laggards"])
        worksheet.append(["R1", "MISSING"])
        path = Path(tempfile.gettempdir()) / "manual_invalid.xlsx"
        workbook.save(path)
        cohort = parse_manual_cohort_workbook(path, valid_uuids=["R1"])
        self.assertEqual(cohort.winner_uuids, ("R1",))
        self.assertEqual(cohort.laggard_uuids, ())
        self.assertEqual(cohort.invalid_uuids, ("MISSING",))

    def test_overlap_uuids_surfaced_as_warning(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Winners & Laggards"
        worksheet.append(["Winners", "Laggards"])
        worksheet.append(["R1", "R1"])
        path = Path(tempfile.gettempdir()) / "manual_overlap.xlsx"
        workbook.save(path)
        cohort = parse_manual_cohort_workbook(path, valid_uuids=["R1"])
        self.assertEqual(cohort.overlap_uuids, ("R1",))

    def test_manual_mask_construction(self) -> None:
        df = pd.DataFrame({"rid": ["A", "B", "C", "D"]})
        segment = SegmentDefinition(
            outcome_question_id="manual_uuid",
            segment_mode="manual_uuid",
            manual_winner_uuids=("A", "C"),
            manual_laggard_uuids=("B",),
        )
        winner, laggard, valid, _warnings = _build_segment_masks(
            df,
            None,
            segment,
            respondent_id_column="rid",
        )
        self.assertEqual(winner.tolist(), [True, False, True, False])
        self.assertEqual(laggard.tolist(), [False, True, False, False])
        self.assertEqual(valid.tolist(), [True, True, True, False])

    def test_segment_definition_accepts_manual_uuid_mode(self) -> None:
        segment = SegmentDefinition(
            outcome_question_id="manual_uuid",
            segment_mode="manual_uuid",
            manual_winner_uuids=("A",),
            manual_laggard_uuids=("B",),
        )
        self.assertEqual(segment.segment_mode, "manual_uuid")

    def test_segment_definition_accepts_manual_cohort_id_column(self) -> None:
        segment = SegmentDefinition(
            outcome_question_id="manual_uuid",
            segment_mode="manual_uuid",
            manual_winner_uuids=("A",),
            manual_laggard_uuids=("B",),
            manual_cohort_id_column="uuid",
        )
        self.assertEqual(segment.manual_cohort_id_column, "uuid")

    def test_build_segment_masks_uses_manual_cohort_id_column_when_set(self) -> None:
        df = pd.DataFrame(
            {
                "record": [1, 2, 3, 4],
                "uuid": ["A", "B", "C", "D"],
            }
        )
        segment = SegmentDefinition(
            outcome_question_id="manual_uuid",
            segment_mode="manual_uuid",
            manual_winner_uuids=("A", "C"),
            manual_laggard_uuids=("D",),
            manual_cohort_id_column="uuid",
        )
        winner, laggard, valid, _warnings = _build_segment_masks(
            df,
            None,
            segment,
            respondent_id_column="record",
        )
        self.assertEqual(winner.tolist(), [True, False, True, False])
        self.assertEqual(laggard.tolist(), [False, False, False, True])
        self.assertEqual(valid.tolist(), [True, False, True, True])

    def test_segment_definition_preserves_existing_laggard_fields(self) -> None:
        segment = SegmentDefinition(
            outcome_question_id="manual_uuid",
            segment_mode="manual_uuid",
            laggard_values=(2,),
            laggard_threshold=4.0,
            laggard_threshold_direction="lte",
            laggard_label="Trailing",
            laggard_outcome_question_id="Q2",
            laggard_outcome_sub_question_id="Q2r1",
            manual_winner_uuids=("A",),
            manual_laggard_uuids=("B",),
            manual_cohort_id_column="uuid",
        )
        self.assertEqual(segment.laggard_values, (2,))
        self.assertEqual(segment.laggard_threshold, 4.0)
        self.assertEqual(segment.laggard_threshold_direction, "lte")
        self.assertEqual(segment.laggard_label, "Trailing")
        self.assertEqual(segment.laggard_outcome_question_id, "Q2")
        self.assertEqual(segment.laggard_outcome_sub_question_id, "Q2r1")
        self.assertEqual(segment.manual_cohort_id_column, "uuid")

    def test_cohort_definition_sheet_written_in_workbook(self) -> None:
        output_path, schema, df, results, _default_segment, log = _fixture()
        segment = SegmentDefinition(
            outcome_question_id="manual_uuid",
            segment_mode="manual_uuid",
            manual_winner_uuids=("1", "2"),
            manual_laggard_uuids=("3",),
            winner_label="Winners",
            loser_label="Laggards",
            laggard_label="Laggards",
        )
        export_winners_vs_laggards_workbook(
            output_path=output_path,
            decoded_df=df,
            schema=schema,
            single_cut_results=results,
            segment_definition=segment,
            laggard_segment_definition=segment,
            themes={"themes": [{"name": "Demographics", "question_ids": ["Q1"]}]},
            calculation_log=CalculationLog() if log is None else log,
        )
        workbook = load_workbook(output_path, data_only=False)
        self.assertIn("_CohortDefinition", workbook.sheetnames)
        sheet = workbook["_CohortDefinition"]
        self.assertEqual(sheet.sheet_state, "hidden")
        self.assertEqual(sheet["B2"].value, "manual_uuid")
        self.assertEqual(sheet["B3"].value, "(N/A for manual)")
        self.assertEqual(sheet["A10"].value, "Winner uuids:")
        self.assertEqual(sheet["A11"].value, "1")


if __name__ == "__main__":
    unittest.main()
