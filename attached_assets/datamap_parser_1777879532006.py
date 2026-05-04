"""State-machine parser for Survey Insight Engine data map workbooks."""

from __future__ import annotations

from enum import Enum
import re
from typing import Any, Literal, TypedDict

from openpyxl import load_workbook

try:
    from config import (
        DATAMAP_SHEET_NAME,
        QUESTION_HEADER_PATTERN,
        VALUES_LINE_PATTERN,
        OPEN_NUMERIC_LINE,
        OPEN_TEXT_LINE,
        SUB_COLUMN_PATTERN,
    )
except ModuleNotFoundError as exc:
    if exc.name != "config":
        raise
    DATAMAP_SHEET_NAME = "Sheet1"
    QUESTION_HEADER_PATTERN = r"^\[?([A-Za-z][A-Za-z0-9_]*)\]?:\s*(.+)$"
    VALUES_LINE_PATTERN = r"^Values:\s*(-?\d+)\s*-\s*(-?\d+)$"
    OPEN_NUMERIC_LINE = "Open numeric response"
    OPEN_TEXT_LINE = "Open text response"
    SUB_COLUMN_PATTERN = r"^[A-Za-z][A-Za-z0-9_]*$"


QUESTION_HEADER_RE = re.compile(QUESTION_HEADER_PATTERN)
VALUES_LINE_RE = re.compile(VALUES_LINE_PATTERN)
SUB_COLUMN_RE = re.compile(SUB_COLUMN_PATTERN)
PARENT_ROW_OE_RE = re.compile(r"^([A-Za-z][A-Za-z0-9_]*?)r\d+oe$")
PARENT_OE_RE = re.compile(r"^([A-Za-z][A-Za-z0-9_]*?)oe$")


TypeHint = Literal["values_range", "open_numeric", "open_text"]


class ParsedQuestion(TypedDict):
    canonical_id: str
    raw_id: str
    question_text: str
    type_hint: TypeHint | None
    value_range: tuple[int, int] | None
    options: list[tuple[int, str]]
    sub_columns: list[tuple[str, str]]
    parent_canonical_id: str | None
    source_row: int
    warnings: list[str]


class DataMap(TypedDict):
    questions: list[ParsedQuestion]
    source_path: str
    sheet_name: str
    total_rows_in_sheet: int
    parser_warnings: list[str]


class _State(Enum):
    BETWEEN_BLOCKS = "BETWEEN_BLOCKS"
    IN_HEADER = "IN_HEADER"
    IN_TYPE_HINT = "IN_TYPE_HINT"
    IN_OPTIONS = "IN_OPTIONS"


class _Block(TypedDict):
    canonical_id: str
    raw_id: str
    question_text: str
    type_hint: TypeHint | None
    value_range: tuple[int, int] | None
    options: list[tuple[int, str]]
    sub_columns: list[tuple[str, str]]
    source_row: int
    warnings: list[str]


def parse_datamap(path: str) -> DataMap:
    """Parse Sheet1 from a data map workbook into observation-only records."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if DATAMAP_SHEET_NAME not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"{DATAMAP_SHEET_NAME!r} sheet not found. "
                f"Available sheets: {available}"
            )

        worksheet = workbook[DATAMAP_SHEET_NAME]
        questions: list[ParsedQuestion] = []
        parser_warnings: list[str] = []
        state = _State.BETWEEN_BLOCKS
        current_block: _Block | None = None
        total_rows = 0

        for row_number, raw_row in enumerate(
            worksheet.iter_rows(values_only=True), start=1
        ):
            total_rows = row_number
            row = _normalise_row(raw_row)
            col_a = _cell(row, 0)
            col_b = _cell(row, 1)
            col_c = _cell(row, 2)
            header_match = _header_match(col_a)
            is_blank = _is_blank_row(col_a, col_b)
            is_option_style = col_a is None and col_b is not None

            if state is _State.BETWEEN_BLOCKS:
                if is_blank:
                    continue
                if header_match:
                    current_block = _start_block(col_a, header_match, row_number)
                    state = _State.IN_HEADER
                    continue
                parser_warnings.append(
                    f"orphan row at row {row_number}: {_row_preview(row)!r}"
                )
                continue

            if current_block is None:
                raise RuntimeError("parser entered a block state without a block")

            if state is _State.IN_HEADER:
                if is_blank:
                    current_block["warnings"].append(
                        "header followed by blank row, no type hint"
                    )
                    questions.append(_finalise_block(current_block))
                    current_block = None
                    state = _State.BETWEEN_BLOCKS
                    continue
                if header_match:
                    parser_warnings.append(
                        f"header found mid-block at row {row_number} "
                        "— finalised previous block early"
                    )
                    questions.append(_finalise_block(current_block))
                    current_block = _start_block(col_a, header_match, row_number)
                    state = _State.IN_HEADER
                    continue
                if is_option_style:
                    current_block["warnings"].append(
                        f"option row encountered before type hint at row {row_number}"
                    )
                    continue

                _capture_type_hint(current_block, col_a, row_number)
                state = _State.IN_TYPE_HINT
                continue

            if state is _State.IN_TYPE_HINT:
                if is_blank:
                    questions.append(_finalise_block(current_block))
                    current_block = None
                    state = _State.BETWEEN_BLOCKS
                    continue
                if header_match:
                    parser_warnings.append(
                        f"header found mid-block at row {row_number} "
                        "— finalised previous block early"
                    )
                    questions.append(_finalise_block(current_block))
                    current_block = _start_block(col_a, header_match, row_number)
                    state = _State.IN_HEADER
                    continue
                if is_option_style:
                    _capture_option_row(current_block, col_b, col_c, row_number)
                    state = _State.IN_OPTIONS
                    continue

                current_block["warnings"].append(
                    f"unrecognised row after type hint at row {row_number}: "
                    f"{_row_preview(row)!r}"
                )
                continue

            if state is _State.IN_OPTIONS:
                if is_blank:
                    questions.append(_finalise_block(current_block))
                    current_block = None
                    state = _State.BETWEEN_BLOCKS
                    continue
                if header_match:
                    parser_warnings.append(
                        f"header found mid-block at row {row_number} "
                        "— finalised previous block early"
                    )
                    questions.append(_finalise_block(current_block))
                    current_block = _start_block(col_a, header_match, row_number)
                    state = _State.IN_HEADER
                    continue
                if is_option_style:
                    _capture_option_row(current_block, col_b, col_c, row_number)
                    continue

                current_block["warnings"].append(
                    f"unrecognised option row at row {row_number}: "
                    f"{_row_preview(row)!r}"
                )

        if current_block is not None:
            questions.append(_finalise_block(current_block))

        return {
            "questions": questions,
            "source_path": path,
            "sheet_name": DATAMAP_SHEET_NAME,
            "total_rows_in_sheet": total_rows,
            "parser_warnings": parser_warnings,
        }
    finally:
        workbook.close()


def _normalise_row(row: tuple[Any, ...]) -> tuple[Any | None, ...]:
    return tuple(_normalise_cell(value) for value in row)


def _normalise_cell(value: Any) -> Any | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _cell(row: tuple[Any | None, ...], index: int) -> Any | None:
    if index >= len(row):
        return None
    return row[index]


def _is_blank_row(col_a: Any | None, col_b: Any | None) -> bool:
    return col_a is None and col_b is None


def _header_match(value: Any | None) -> re.Match[str] | None:
    if not isinstance(value, str):
        return None
    if _looks_like_type_hint(value):
        return None
    return QUESTION_HEADER_RE.match(value)


def _looks_like_type_hint(value: str) -> bool:
    return (
        value == OPEN_NUMERIC_LINE
        or value == OPEN_TEXT_LINE
        or value.startswith("Values:")
    )


def _start_block(
    header_value: Any | None, header_match: re.Match[str], row_number: int
) -> _Block:
    if not isinstance(header_value, str):
        raise ValueError("header value must be a string")

    raw_id = header_value.split(":", 1)[0].strip()
    return {
        "canonical_id": header_match.group(1).strip(),
        "raw_id": raw_id,
        "question_text": header_match.group(2).strip(),
        "type_hint": None,
        "value_range": None,
        "options": [],
        "sub_columns": [],
        "source_row": row_number,
        "warnings": [],
    }


def _capture_type_hint(
    block: _Block, value: Any | None, row_number: int
) -> None:
    if not isinstance(value, str):
        block["warnings"].append(f"unrecognised type hint: {value!r}")
        block["type_hint"] = None
        block["value_range"] = None
        return

    if value == OPEN_NUMERIC_LINE:
        block["type_hint"] = "open_numeric"
        block["value_range"] = None
        return

    if value == OPEN_TEXT_LINE:
        block["type_hint"] = "open_text"
        block["value_range"] = None
        return

    values_match = VALUES_LINE_RE.match(value)
    if values_match:
        low = int(values_match.group(1))
        high = int(values_match.group(2))
        block["type_hint"] = "values_range"
        block["value_range"] = (low, high)
        if low > high:
            block["warnings"].append(f"value range inverted: {low} > {high}")
        return

    block["warnings"].append(f"unrecognised type hint: {value!r}")
    block["type_hint"] = None
    block["value_range"] = None


def _capture_option_row(
    block: _Block, col_b: Any | None, col_c: Any | None, row_number: int
) -> None:
    label = _parse_label(col_c, block["warnings"], row_number)
    if label is None:
        return

    if isinstance(col_b, str) and col_b.startswith("[") and col_b.endswith("]"):
        inner_id = col_b[1:-1].strip()
        if not SUB_COLUMN_RE.match(inner_id):
            block["warnings"].append(
                f"sub-column id does not match pattern: {inner_id!r} "
                f"at row {row_number}"
            )
            return
        block["sub_columns"].append((inner_id, label))
        return

    try:
        code = int(str(col_b).strip())
    except (TypeError, ValueError):
        block["warnings"].append(
            f"option code in col B is not an integer: {col_b!r} "
            f"at row {row_number}"
        )
        return

    block["options"].append((code, label))


def _parse_label(
    value: Any | None, warnings: list[str], row_number: int
) -> str | None:
    if not isinstance(value, str) or not value:
        warnings.append(f"option label is empty at row {row_number}")
        return None
    return value


def _finalise_block(block: _Block) -> ParsedQuestion:
    type_hint = block["type_hint"]
    parent_canonical_id = (
        _derive_parent_canonical_id(block["canonical_id"])
        if type_hint == "open_text"
        else None
    )
    return {
        "canonical_id": block["canonical_id"],
        "raw_id": block["raw_id"],
        "question_text": block["question_text"],
        "type_hint": type_hint,
        "value_range": block["value_range"],
        "options": list(block["options"]),
        "sub_columns": list(block["sub_columns"]),
        "parent_canonical_id": parent_canonical_id,
        "source_row": block["source_row"],
        "warnings": list(block["warnings"]),
    }


def _derive_parent_canonical_id(canonical_id: str) -> str | None:
    row_oe_match = PARENT_ROW_OE_RE.match(canonical_id)
    if row_oe_match:
        return row_oe_match.group(1)

    oe_match = PARENT_OE_RE.match(canonical_id)
    if oe_match:
        return oe_match.group(1)

    return None


def _row_preview(row: tuple[Any | None, ...]) -> str:
    values = [str(value) for value in row[:3] if value is not None]
    return " | ".join(values)
