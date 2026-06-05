from __future__ import annotations

import unittest

from openpyxl import Workbook

from src.adapters.six_column_combined import SixColumnCombinedAdapter
from src.question_classifier import classify_questions
from src.models import QuestionType


def _six_column_workbook() -> Workbook:
    workbook = Workbook()
    index = workbook.active
    index.title = "Index"
    index.append(["template page"])

    datamap = workbook.create_sheet("Datamap")
    datamap.append([
        "Brand equity",
        19,
        None,
        "Multiple choice",
        "single-select",
        "On a scale of 0-10, how likely are you to recommend <brand>?",
    ])
    datamap.append([
        "Brand equity",
        20,
        None,
        "Text input",
        None,
        "Please elaborate on why you gave <nps_score> for <brand>.",
    ])
    datamap.append([
        "Brand equity",
        22,
        None,
        "Matrix",
        "single-select",
        "On a scale from 1-5, how well does <brand> perform?",
    ])

    raw = workbook.create_sheet("Raw Data")
    raw.append([
        "Respondent",
        "Status",
        "Q19 | Covetrus",
        "Q19 | MWI Animal Health",
        "Q20 | Covetrus",
        "Q22: Quality | Covetrus",
        "Q22: Quality | MWI Animal Health",
        "Q22: Computed(Answered) | Covetrus",
    ])
    raw.append(["r1", "Qualified", 9, 6, "Great", 5, 4, 1])
    raw.append(["r2", "Qualified", 10, 8, "Fine", 4, 3, 1])
    return workbook


class TestSixColumnCombinedAdapter(unittest.TestCase):
    def test_detects_raw_data_and_six_column_datamap(self) -> None:
        result = SixColumnCombinedAdapter().detect(_six_column_workbook())

        self.assertGreaterEqual(result.confidence, 0.9)
        self.assertIn("six-column", result.reason)

    def test_parse_groups_entity_columns_and_excludes_computed_flags(self) -> None:
        parsed = SixColumnCombinedAdapter().parse(_six_column_workbook())
        questions = {question["canonical_id"]: question for question in parsed["questions"]}

        self.assertEqual(
            questions["Q19"]["sub_columns"],
            [("Q19 | Covetrus", "Covetrus"), ("Q19 | MWI Animal Health", "MWI Animal Health")],
        )
        q22_columns = [column for column, _label in questions["Q22"]["sub_columns"]]
        self.assertIn("Q22: Quality | Covetrus", q22_columns)
        self.assertNotIn("Q22: Computed(Answered) | Covetrus", q22_columns)

    def test_parsed_workbook_classifies_nps_and_grid(self) -> None:
        parsed = SixColumnCombinedAdapter().parse(_six_column_workbook())
        raw_columns = [
            "Respondent",
            "Status",
            "Q19 | Covetrus",
            "Q19 | MWI Animal Health",
            "Q20 | Covetrus",
            "Q22: Quality | Covetrus",
            "Q22: Quality | MWI Animal Health",
        ]

        schema = classify_questions(parsed, raw_columns)

        self.assertIs(schema.get_question("Q19").question_type, QuestionType.NPS)
        self.assertIs(schema.get_question("Q20").question_type, QuestionType.OPEN_TEXT)
        self.assertIs(schema.get_question("Q22").question_type, QuestionType.GRID_RATED)


if __name__ == "__main__":
    unittest.main()
