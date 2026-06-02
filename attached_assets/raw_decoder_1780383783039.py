"""Raw survey data decoder for the Survey Insight Engine."""

from __future__ import annotations

from numbers import Real
from pathlib import Path
import re
from typing import Any

import pandas as pd

from src.datamap_parser import DataMap, ParsedQuestion, derive_numeric_label_metadata
from src.models import DataQualityReport, QuestionType

try:
    from config import MISSING_VALUE_TOKENS, HIGH_MISSING_THRESHOLD
except ModuleNotFoundError as exc:
    if exc.name != "config":
        raise
    MISSING_VALUE_TOKENS = {"", "NA", "N/A", "NULL", "null", "None", "nan"}
    HIGH_MISSING_THRESHOLD = 0.5


RESPONDENT_ID_CANDIDATES = (
    "Respondent",
    "respondent",
    "RespondentID",
    "respondent_id",
    "Respondent ID",
    "Response ID",
    "response_id",
    "ResponseID",
    "ParticipantID",
    "participant_id",
    "Participant ID",
    "SubmissionID",
    "submission_id",
    "Submission ID",
    "record",
    "uuid",
    "id",
    "ID",
)
_REJECTION_PREFIXES = (
    "NO TO: ",
    "NOT: ",
    "No to: ",
    "NOT SELECTED: ",
    "NOT SELECTED - ",
    "NO - ",
    "Not selected: ",
)
_TEMPLATE_PLACEHOLDER_PATTERN = re.compile(r"^\$\s*\{.+\}$")


def decode_raw_data(
    path: str,
    data_map: DataMap,
) -> tuple[pd.DataFrame, DataQualityReport]:
    """Load raw survey data and return a decoded DataFrame plus quality report."""

    expected_columns = _expected_columns_from_datamap(data_map)
    dataframe = _load_raw_file(path, expected_columns)
    dataframe = _strip_string_values(dataframe)
    dataframe = _replace_missing_tokens(dataframe)

    original_columns = tuple(str(column) for column in dataframe.columns)
    respondent_id_column = _find_respondent_id_column(dataframe.columns.tolist())
    warnings: list[str] = []
    generated_respondent_id = False
    sheet_warning = dataframe.attrs.get("_raw_decoder_sheet_warning")
    if sheet_warning:
        warnings.append(str(sheet_warning))

    if respondent_id_column is None:
        respondent_id_column = "respondent_id"
        dataframe[respondent_id_column] = range(1, len(dataframe) + 1)
        generated_respondent_id = True
        warnings.append("no respondent ID column found; generated sequential IDs")

    columns_in_datamap = len(set(original_columns).intersection(expected_columns))
    columns_not_in_datamap = tuple(
        column for column in original_columns if column not in expected_columns
    )
    question_types = _question_types_for_decoding(
        data_map,
        dataframe.columns.tolist(),
        respondent_id_column,
        len(dataframe),
        path,
    )
    value_ranges = _value_ranges_by_column(data_map)
    no_numeric_coercion_columns = _string_preserved_columns(data_map, question_types)
    numeric_metadata_by_column = _numeric_label_metadata_by_column(data_map)

    coercion_log: list[dict] = []
    decoder_warnings: list[dict] = []
    for column, question_spec in numeric_metadata_by_column.items():
        if column not in dataframe.columns:
            continue
        dataframe[column] = _decode_numeric_series(
            dataframe[column],
            question_spec,
            decoder_warnings,
            str(column),
        )

    for column in dataframe.columns:
        if column == respondent_id_column:
            continue
        if column in no_numeric_coercion_columns:
            continue
        if column in numeric_metadata_by_column:
            continue
        dataframe[column] = _coerce_column_to_numeric(
            dataframe[column], str(column), coercion_log
        )

    missing_pct = {
        str(column): float(dataframe[column].isna().mean())
        for column in dataframe.columns
    }
    out_of_range_pct = {
        str(column): _out_of_range_pct(dataframe[column], value_ranges.get(str(column)))
        for column in dataframe.columns
    }

    for column, pct in missing_pct.items():
        if pct > HIGH_MISSING_THRESHOLD:
            warnings.append(f"column {column} has {pct:.1%} missing values")
    for column, pct in out_of_range_pct.items():
        if pct > 0.05:
            warnings.append(f"column {column} has {pct:.1%} out-of-range values")

    report = DataQualityReport(
        total_rows=int(len(dataframe)),
        total_columns=int(len(dataframe.columns)),
        columns_in_datamap=int(columns_in_datamap),
        columns_not_in_datamap=columns_not_in_datamap,
        per_column_missing_pct=missing_pct,
        per_column_out_of_range_pct=out_of_range_pct,
        coercion_log=tuple(coercion_log),
        warnings=tuple(warnings),
        decoder_warnings=tuple(decoder_warnings),
    )

    if generated_respondent_id:
        dataframe[respondent_id_column] = dataframe[respondent_id_column].astype("Int64")

    dataframe = _decode_option_columns(dataframe, data_map, question_types)
    dataframe = _normalise_dataframe_for_excel(dataframe)

    return dataframe, report


def _load_raw_file(
    path: str,
    expected_columns: set[str] | None = None,
) -> pd.DataFrame:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path, dtype=str)
    if suffix == ".xlsx":
        sheet_name: int | str = 0
        first_sheet_name = _first_sheet_name(path)
        fallback_warning = None
        if expected_columns:
            detected, fallback_warning = _find_data_sheet(path, expected_columns)
            sheet_name = detected
        dataframe = pd.read_excel(
            path,
            dtype=str,
            engine="openpyxl",
            sheet_name=sheet_name,
        )
        dataframe = _normalise_dataframe_for_excel(dataframe)
        if fallback_warning:
            dataframe.attrs["_raw_decoder_sheet_warning"] = fallback_warning
        elif isinstance(sheet_name, str) and sheet_name != first_sheet_name:
            dataframe.attrs["_raw_decoder_sheet_warning"] = (
                f"data loaded from sheet '{sheet_name}' "
                "(not the first sheet)"
            )
        return dataframe
    raise ValueError(f"unsupported raw data file extension: {suffix or '<none>'}")


def _normalise_dataframe_for_excel(dataframe: pd.DataFrame) -> pd.DataFrame:
    from src.io import _normalise_dataframe

    return _normalise_dataframe(dataframe)


def _find_data_sheet(path: str, expected_columns: set[str]) -> tuple[str, str | None]:
    """Return the sheet name whose headers best match expected_columns."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) == 1:
            return workbook.sheetnames[0], None

        first_sheet = workbook.sheetnames[0]
        best_sheet = None
        best_score = 0
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            first_row = next(
                worksheet.iter_rows(max_row=1, values_only=True),
                (),
            )
            headers = {
                str(cell).strip()
                for cell in first_row
                if cell is not None
            }
            score = len(headers.intersection(expected_columns))
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
        if best_score == 0:
            return (
                first_sheet,
                "no sheet matched expected columns; loaded first sheet "
                f"'{first_sheet}' as fallback",
            )
        return str(best_sheet), None
    finally:
        workbook.close()


def _first_sheet_name(path: str) -> str | None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames[0] if workbook.sheetnames else None
    finally:
        workbook.close()


def _strip_string_values(dataframe: pd.DataFrame) -> pd.DataFrame:
    stripped = dataframe.copy()
    for column in stripped.columns:
        stripped[column] = stripped[column].map(
            lambda value: value.strip() if isinstance(value, str) else value
        )
    return stripped


def _replace_missing_tokens(dataframe: pd.DataFrame) -> pd.DataFrame:
    replaced = dataframe.copy()
    for column in replaced.columns:
        replaced[column] = replaced[column].map(_replace_missing_value)
    return replaced


def _replace_missing_value(value: Any) -> Any:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, str) and _is_template_placeholder(value.strip()):
        return pd.NA
    if isinstance(value, str) and value in MISSING_VALUE_TOKENS:
        return pd.NA
    return value


def _is_template_placeholder(value: str) -> bool:
    return bool(_TEMPLATE_PLACEHOLDER_PATTERN.match(value))


def _find_respondent_id_column(columns: list[str]) -> str | None:
    for candidate in RESPONDENT_ID_CANDIDATES:
        candidate_key = candidate.lower()
        for column in columns:
            if str(column).lower() == candidate_key:
                return column
    return columns[0] if columns else None


def _expected_columns_from_datamap(data_map: DataMap) -> set[str]:
    expected: set[str] = set()
    for question in data_map["questions"]:
        expected.add(question["canonical_id"])
        sub_columns = question["sub_columns"]
        if sub_columns:
            expected.update(sub_column_id for sub_column_id, _ in sub_columns)
    return expected


def _value_ranges_by_column(data_map: DataMap) -> dict[str, tuple[int, int]]:
    value_ranges: dict[str, tuple[int, int]] = {}
    for question in data_map["questions"]:
        value_range = question["value_range"]
        if value_range is None:
            continue
        for column in _question_expected_columns(question):
            value_ranges[column] = value_range
    return value_ranges


def _string_preserved_columns(
    data_map: DataMap,
    question_types: dict[str, QuestionType] | None = None,
) -> set[str]:
    question_types = question_types or {}
    preserved: set[str] = set()
    for question in data_map["questions"]:
        question_type = question_types.get(question["canonical_id"])
        if question["type_hint"] == "open_text" or (
            question["type_hint"] == "open_numeric"
            and question_type
            not in {
                QuestionType.GRID_SINGLE_SELECT,
                QuestionType.GRID_RATED,
            }
        ):
            preserved.update(_question_expected_columns(question))
    return preserved


def _question_expected_columns(question: ParsedQuestion) -> tuple[str, ...]:
    if question["sub_columns"]:
        return tuple(sub_column_id for sub_column_id, _ in question["sub_columns"])
    return (question["canonical_id"],)


def _numeric_label_metadata_by_column(data_map: DataMap) -> dict[str, ParsedQuestion]:
    metadata_by_column: dict[str, ParsedQuestion] = {}
    for question in data_map["questions"]:
        if not question.get("label_to_numeric_value"):
            mapping, na_labels, allowed_range = derive_numeric_label_metadata(question)
            if mapping:
                question["label_to_numeric_value"] = mapping
                question["na_label_set"] = na_labels
                question["allowed_numeric_range"] = allowed_range
        if not question.get("label_to_numeric_value"):
            continue
        for column in _question_expected_columns(question):
            metadata_by_column[column] = question
    return metadata_by_column


def _decode_numeric_series(
    series: pd.Series,
    question_spec: ParsedQuestion,
    warnings_log: list[dict],
    column_name: str,
) -> pd.Series:
    decoded_values = [
        decode_numeric_cell(raw_value, question_spec, warnings_log, row_idx, column_name)
        for row_idx, raw_value in series.items()
    ]
    return pd.Series(decoded_values, index=series.index, dtype="Float64")


def decode_numeric_cell(
    raw_value: Any,
    question_spec: Any,
    warnings_log: list[dict],
    row_idx: Any,
    column_name: str,
) -> float | None:
    label_to_numeric_value = dict(
        _question_attr(question_spec, "label_to_numeric_value", {}) or {}
    )
    na_label_set = frozenset(_question_attr(question_spec, "na_label_set", frozenset()) or ())
    allowed_range = _question_attr(question_spec, "allowed_numeric_range", None)
    if allowed_range is None and label_to_numeric_value:
        values = tuple(label_to_numeric_value.values())
        allowed_range = (float(min(values)), float(max(values)))

    if isinstance(raw_value, bool):
        warnings_log.append(
            {
                "question_id": _question_attr(question_spec, "canonical_id", ""),
                "column": column_name,
                "row": row_idx,
                "raw_value": str(raw_value),
                "action": "unexpected_type_bool",
            }
        )
        return None

    if isinstance(raw_value, Real):
        value = float(raw_value)
        if allowed_range is None:
            return value
        lo, hi = allowed_range
        if lo <= value <= hi:
            return value
        warnings_log.append(
            {
                "question_id": _question_attr(question_spec, "canonical_id", ""),
                "column": column_name,
                "row": row_idx,
                "raw_value": raw_value,
                "action": f"out_of_range (expected {lo}-{hi})",
            }
        )
        return None

    if raw_value is None or pd.isna(raw_value):
        return None
    if raw_value == "":
        return None

    if isinstance(raw_value, str):
        if raw_value in label_to_numeric_value:
            return float(label_to_numeric_value[raw_value])

        if raw_value in na_label_set:
            return None

        try:
            value = float(raw_value)
        except ValueError:
            value = None
        if value is not None:
            if allowed_range is None:
                return value
            lo, hi = allowed_range
            if lo <= value <= hi:
                return value

        warnings_log.append(
            {
                "question_id": _question_attr(question_spec, "canonical_id", ""),
                "column": column_name,
                "row": row_idx,
                "raw_value": raw_value,
                "action": "unrecognized_value_treated_as_missing",
            }
        )
        return None

    warnings_log.append(
        {
            "question_id": _question_attr(question_spec, "canonical_id", ""),
            "column": column_name,
            "row": row_idx,
            "raw_value": str(raw_value),
            "action": f"unexpected_type_{type(raw_value).__name__}",
        }
    )
    return None


def _question_attr(question_spec: Any, name: str, default: Any = None) -> Any:
    if isinstance(question_spec, dict):
        return question_spec.get(name, default)
    return getattr(question_spec, name, default)


def _coerce_column_to_numeric(
    series: pd.Series, column_name: str, coercion_log: list[dict]
) -> pd.Series:
    original = series.copy()
    numeric_series = pd.to_numeric(original, errors="coerce")
    affected_mask = original.notna() & numeric_series.isna()
    affected_count = int(affected_mask.sum())

    if affected_count > 0:
        values_coerced = sorted(
            {str(value) for value in original[affected_mask].dropna().tolist()}
        )
        coercion_log.append(
            {
                "column": column_name,
                "from_type": "string",
                "to_type": "numeric",
                "values_coerced": values_coerced,
                "rows_affected": affected_count,
            }
        )

    if int(numeric_series.notna().sum()) == 0:
        return original
    return numeric_series


def _question_types_for_decoding(
    data_map: DataMap,
    raw_columns: list[str],
    respondent_id_column: str | None,
    total_respondents: int,
    source_rawdata_path: str,
) -> dict[str, QuestionType]:
    """Classify questions early enough to avoid decoding grid columns as options."""

    try:
        from src.question_classifier import classify_questions

        schema = classify_questions(
            data_map,
            raw_columns,
            respondent_id_column=respondent_id_column,
            total_respondents=total_respondents,
            source_rawdata_path=source_rawdata_path,
        )
    except Exception:
        return {}
    return {spec.canonical_id: spec.question_type for spec in schema.questions}


def _decode_option_columns(
    dataframe: pd.DataFrame,
    data_map: DataMap,
    question_types: dict[str, QuestionType] | None = None,
) -> pd.DataFrame:
    decoded = dataframe.copy()
    question_types = question_types or {}
    for question in data_map["questions"]:
        if question.get("label_to_numeric_value"):
            continue
        question_type = question_types.get(question["canonical_id"])
        if question_type is QuestionType.GRID_RATED:
            continue
        if (
            question_type is QuestionType.GRID_SINGLE_SELECT
            or _is_grid_single_select_question(question)
        ):
            decoded = _decode_grid_single_select_question(decoded, question)
            continue

        option_map = {code: label for code, label in question["options"]}
        if not option_map:
            continue

        columns = _question_expected_columns(question)
        for column in columns:
            if column not in decoded.columns:
                continue
            decoded[column] = decoded[column].map(
                lambda value, mapping=option_map: _decode_option_value(value, mapping)
            )
    return decoded


def _is_grid_single_select_question(question: ParsedQuestion) -> bool:
    return (
        question["type_hint"] == "values_range"
        and bool(question["sub_columns"])
        and bool(question["options"])
    )


def _decode_grid_single_select_question(
    dataframe: pd.DataFrame,
    question: ParsedQuestion,
) -> pd.DataFrame:
    decoded = dataframe.copy()
    discovered_labels: dict[str, str] = {}

    for column, _fallback_label in question["sub_columns"]:
        if column not in decoded.columns:
            continue
        if not _grid_column_uses_text_selection(decoded[column]):
            continue

        decoded_values: list[Any] = []
        for value in decoded[column].tolist():
            selected, clean_text = _decode_grid_cell_value(value)
            decoded_values.append(selected)
            if selected == 1 and clean_text and column not in discovered_labels:
                discovered_labels[column] = clean_text
        decoded[column] = decoded_values

    if discovered_labels:
        question["sub_columns"] = [
            (column, discovered_labels.get(column, label))
            for column, label in question["sub_columns"]
        ]
    return decoded


def _grid_column_uses_text_selection(series: pd.Series) -> bool:
    for value in series.dropna().tolist():
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped or _is_template_placeholder(stripped):
                continue
            try:
                float(stripped)
            except ValueError:
                return True
    return False


def _decode_grid_cell_value(value: Any) -> tuple[int | None, str]:
    if pd.isna(value):
        return None, ""
    if isinstance(value, str):
        text = value.strip()
        if not text or _is_template_placeholder(text):
            return None, ""
        was_rejected, clean_text = _strip_rejection_prefix(text)
        return (0 if was_rejected else 1), clean_text
    return (1 if _truthy_grid_value(value) else 0), str(value).strip()


def _truthy_grid_value(value: Any) -> bool:
    try:
        return float(value) != 0.0
    except (TypeError, ValueError):
        return bool(value)


def _strip_rejection_prefix(value: str) -> tuple[bool, str]:
    for prefix in _REJECTION_PREFIXES:
        if value.startswith(prefix):
            return True, value[len(prefix) :].strip()
    return False, value.strip()


def _decode_option_value(value: Any, option_map: dict[int | str, str]) -> Any:
    if pd.isna(value):
        return pd.NA

    for candidate in _option_lookup_candidates(value):
        if candidate in option_map:
            return option_map[candidate]

    return value


def _option_lookup_candidates(value: Any) -> tuple[Any, ...]:
    candidates: list[Any] = [value]

    if isinstance(value, float) and value.is_integer():
        candidates.append(int(value))

    if isinstance(value, str):
        stripped = value.strip()
        candidates.append(stripped)
        try:
            numeric = float(stripped)
        except ValueError:
            numeric = None
        if numeric is not None and numeric.is_integer():
            candidates.append(int(numeric))

    candidates.append(str(value))
    deduped: list[Any] = []
    for candidate in candidates:
        if candidate not in deduped:
            deduped.append(candidate)
    return tuple(deduped)


def _out_of_range_pct(
    series: pd.Series, value_range: tuple[int, int] | None
) -> float:
    if value_range is None or not pd.api.types.is_numeric_dtype(series):
        return 0.0

    valid_values = series.dropna()
    valid_n = int(len(valid_values))
    if valid_n == 0:
        return 0.0

    low, high = value_range
    out_of_range_count = int((~valid_values.between(low, high)).sum())
    return float(out_of_range_count / valid_n)
