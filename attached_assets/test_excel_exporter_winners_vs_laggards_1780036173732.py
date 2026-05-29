from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.excel_exporter import export_winners_vs_laggards_workbook
from src.models import (
    DenominatorPolicy,
    GridRatedResult,
    GridRatedRow,
    QuestionSpec,
    QuestionType,
    SegmentDefinition,
    SingleSelectResult,
    SurveySchema,
)


def _fixture() -> tuple[Path, SurveySchema, pd.DataFrame, list, SegmentDefinition, CalculationLog]:
    schema = SurveySchema(
        questions=(
            QuestionSpec(
                question_id="QO",
                canonical_id="QO",
                question_text="Outcome",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("QO",),
                option_map={1: "Winner", 2: "Laggard"},
            ),
            QuestionSpec(
                question_id="QO2",
                canonical_id="QO2",
                question_text="Override outcome",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("QO2",),
                option_map={1: "Winner", 2: "Laggard"},
            ),
            QuestionSpec(
                question_id="Q1",
                canonical_id="Q1",
                question_text="Choice",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q1",),
                option_map={1: "A", 2: "B"},
            ),
            QuestionSpec(
                question_id="QG",
                canonical_id="QG",
                question_text="Rated grid",
                question_type=QuestionType.GRID_RATED,
                raw_columns=("QGr1c1", "QGr2c1"),
                option_map={},
                grid_row_labels={"QGr1c1": "Row 1", "QGr2c1": "Row 2"},
            ),
        ),
        respondent_id_column="rid",
        total_respondents=4,
        source_datamap_path="map.xlsx",
        source_rawdata_path="raw.xlsx",
        parsed_at=datetime.now(timezone.utc),
    )
    df = pd.DataFrame(
        {
            "rid": [1, 2, 3, 4],
            "QO": [1, 1, 2, 2],
            "QO2": [2, 2, 1, 1],
            "Q1": [1, 2, 1, 2],
            "QGr1c1": [8, 9, 3, 4],
            "QGr2c1": [7, 8, 2, 3],
        }
    )
    results = [
        SingleSelectResult(
            question_id="Q1",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                1: {"label": "A", "count": 2, "rate": 0.5},
                2: {"label": "B", "count": 2, "rate": 0.5},
            },
        ),
        GridRatedResult(
            question_id="QG",
            question_type=QuestionType.GRID_RATED,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            question_text="Rated grid",
            column_headers=["Score"],
            rows=(
                GridRatedRow(
                    row_id="QGr1",
                    row_label="Row 1",
                    means_per_column=[6.0],
                    valid_n_per_column=[4],
                ),
                GridRatedRow(
                    row_id="QGr2",
                    row_label="Row 2",
                    means_per_column=[5.0],
                    valid_n_per_column=[4],
                ),
            ),
            total_respondents=4,
            total_responses=8,
            show_delta=False,
        ),
    ]
    output_path = Path(tempfile.gettempdir()) / "winners_vs_laggards_test.xlsx"
    log = CalculationLog()
    segment = SegmentDefinition(
        "QO",
        "categorical",
        winner_values=(1,),
        laggard_values=(2,),
        winner_label="Winners",
        loser_label="Laggards",
        laggard_label="Laggards",
    )
    return output_path, schema, df, results, segment, log


def _build_workbook(segment: SegmentDefinition | None = None, themes: dict | None = None):
    output_path, schema, df, results, default_segment, log = _fixture()
    export_winners_vs_laggards_workbook(
        output_path=output_path,
        decoded_df=df,
        schema=schema,
        single_cut_results=results,
        segment_definition=segment or default_segment,
        laggard_segment_definition=segment or default_segment,
        themes=themes
        or {"themes": [{"name": "Demographics", "question_ids": ["Q1", "QG"]}]},
        calculation_log=log,
    )
    return output_path, load_workbook(output_path, data_only=False), log


def _data_sheet(workbook):
    return workbook["Demographics"]


def _find_row(sheet, column: int, value: str) -> int:
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=column).value == value:
            return row
    raise AssertionError(f"{value!r} not found in column {column}")


def _single_select_header_row(sheet) -> int:
    return _find_row(sheet, 1, "Option ID")


def _single_select_data_row(sheet) -> int:
    return _single_select_header_row(sheet) + 1


def _raw_column_values(workbook, header: str) -> list:
    raw = workbook["_RawData"]
    headers = [raw.cell(row=1, column=col).value for col in range(1, raw.max_column + 1)]
    column = headers.index(header) + 1
    return [raw.cell(row=row, column=column).value for row in range(2, raw.max_row + 1)]


def _build_mask_fixture_workbook():
    output_path, schema, _df, results, default_segment, log = _fixture()
    rows = 100
    df = pd.DataFrame(
        {
            "rid": list(range(1, rows + 1)),
            "QO": [1] * 30 + [2] * 70,
            "QO2": [2] * 30 + [1] * 70,
            "Q1": [1, 2] * 50,
            "QGr1c1": [8] * 30 + [3] * 70,
            "QGr2c1": [7] * 30 + [2] * 70,
        }
    )
    export_winners_vs_laggards_workbook(
        output_path=output_path,
        decoded_df=df,
        schema=schema,
        single_cut_results=results,
        segment_definition=default_segment,
        laggard_segment_definition=default_segment,
        themes={"themes": [{"name": "Demographics", "question_ids": ["Q1", "QG"]}]},
        calculation_log=log,
    )
    return output_path, load_workbook(output_path, data_only=False), log


class TestWinnersVsLaggardsWorkbook(unittest.TestCase):
    def test_workbook_is_produced(self) -> None:
        output_path, _workbook, _log = _build_workbook()
        self.assertTrue(output_path.exists())

    def test_run_summary_contains_labels_and_sizes(self) -> None:
        _path, workbook, _log = _build_workbook()
        values = [cell.value for row in workbook["Run_Summary"].iter_rows() for cell in row]
        self.assertIn("Winner cohort size:", values)
        self.assertIn("Laggard cohort size:", values)
        self.assertIn("Winners", values)

    def test_single_select_sheet_has_documented_columns(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        header_row = _single_select_header_row(sheet)
        headers = [sheet.cell(row=header_row, column=col).value for col in range(1, 8)]
        self.assertEqual(headers, ["Option ID", "Option", "Winners count", "Winners %", "Laggards count", "Laggards %", "Delta %"])

    def test_count_percentage_and_delta_cells_are_formulas(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)
        self.assertTrue(str(sheet.cell(data_row, 3).value).startswith("="))
        self.assertTrue(str(sheet.cell(data_row, 4).value).startswith("=IFERROR("))
        self.assertTrue(str(sheet.cell(data_row, 7).value).startswith("="))
        self.assertIn("winners_mask_data", sheet.cell(data_row, 3).value)
        self.assertIn("laggards_mask_data", sheet.cell(data_row, 5).value)

    def test_single_select_countifs_uses_option_code_not_label(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)
        formula = str(sheet.cell(data_row, 3).value)
        self.assertIn('Q1_data,"1"', formula)
        self.assertNotIn('Q1_data,"A"', formula)

    def test_grid_rated_layout_uses_formula_means(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        row = next(
            r
            for r in range(1, sheet.max_row + 1)
            if sheet.cell(row=r, column=1).value == "Sub-question ID"
        )
        self.assertTrue(str(sheet.cell(row=row + 1, column=3).value).startswith("=IFERROR(AVERAGEIFS"))
        self.assertTrue(str(sheet.cell(row=row + 1, column=4).value).startswith("=IFERROR(AVERAGEIFS"))
        self.assertTrue(str(sheet.cell(row=row + 1, column=5).value).startswith("="))

    def test_named_ranges_for_masks_exist(self) -> None:
        _path, workbook, _log = _build_workbook()
        self.assertIn("winners_mask_data", workbook.defined_names)
        self.assertIn("laggards_mask_data", workbook.defined_names)

    def test_winners_mask_data_reflects_winner_membership(self) -> None:
        _path, workbook, _log = _build_mask_fixture_workbook()
        values = _raw_column_values(workbook, "winners_mask_data")
        self.assertEqual(sum(value is True for value in values), 30)
        self.assertEqual(sum(value is False for value in values), 70)

    def test_laggards_mask_data_reflects_laggard_membership(self) -> None:
        _path, workbook, _log = _build_mask_fixture_workbook()
        values = _raw_column_values(workbook, "laggards_mask_data")
        self.assertEqual(sum(value is True for value in values), 70)
        self.assertEqual(sum(value is False for value in values), 30)

    def test_override_outcome_records_laggard_source_column(self) -> None:
        segment = SegmentDefinition(
            "QO",
            "categorical",
            winner_values=(1,),
            laggard_values=(2,),
            laggard_outcome_question_id="QO2",
        )
        _path, _workbook, log = _build_workbook(segment)
        laggard_records = [
            record for record in log.all_records() if record.metric_name == "laggard_mask"
        ]
        self.assertEqual(laggard_records[-1].source_question_id, "QO2")

    def test_backward_compat_defaults_inverse_laggards(self) -> None:
        segment = SegmentDefinition("QO", "categorical", winner_values=(1,))
        _path, workbook, _log = _build_workbook(segment)
        raw = workbook["_RawData"]
        headers = [raw.cell(row=1, column=col).value for col in range(1, raw.max_column + 1)]
        lag_col = headers.index("laggards_mask_data") + 1
        self.assertEqual([raw.cell(row=row, column=lag_col).value for row in range(2, 6)], [False, False, True, True])

    def test_category_sheet_has_red_banner(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        row = _find_row(sheet, 1, "Demographics")
        self.assertGreater(row, 1)
        self.assertEqual(sheet.cell(row=row, column=1).fill.fgColor.rgb, "00CC0000")
        self.assertEqual(sheet.cell(row=row, column=1).font.color.rgb, "00FFFFFF")

    def test_question_block_includes_filter_ui_rows(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        title_row = _find_row(sheet, 1, "Q1 - Choice")
        self.assertEqual(
            [sheet.cell(row=title_row + 1, column=col).value for col in range(1, 6)],
            ["Per-question filter", "Filter Q", "(None)", "Value", "(All)"],
        )
        self.assertIn("MATCH(Demographics_Q1_F_Q,All_Questions,0)", sheet.cell(title_row + 1, 6).value)
        self.assertEqual(sheet.cell(title_row + 1, 7).value, '="|" & SUBSTITUTE(Demographics_Q1_F_V, ", ", "|") & "|"')
        self.assertEqual(sheet.cell(title_row + 2, 1).value, "Cross-tab by")
        self.assertEqual(sheet.cell(title_row + 2, 3).value, "(None)")
        self.assertTrue(str(sheet.cell(title_row + 3, 1).value).startswith("Note: This question was shown"))
        self.assertIsNone(sheet.cell(title_row + 4, 1).value)
        self.assertEqual(sheet.cell(title_row + 5, 1).value, "Option ID")

    def test_workbook_filter_rows_are_written_before_category_banner(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        self.assertEqual(sheet["A1"].value, "LOCAL FILTERS (override workbook defaults)")
        self.assertEqual(sheet["A2"].value, "Filter")
        self.assertEqual(sheet["A3"].value, "Custom 1 question")
        banner_row = _find_row(sheet, 1, "Demographics")
        self.assertGreater(banner_row, 3)
        self.assertIn("Filters", workbook.sheetnames)

    def test_theme_groups_create_category_sheets_not_one_all_questions_bucket(self) -> None:
        _path, workbook, _log = _build_workbook(
            themes={
                "themes": [
                    {"name": "Demographics", "question_ids": ["Q1"]},
                    {"name": "AI Usage", "question_ids": ["QG"]},
                ]
            }
        )
        self.assertIn("Demographics", workbook.sheetnames)
        self.assertIn("AI Usage", workbook.sheetnames)
        self.assertNotIn("All Questions", workbook.sheetnames)

    def test_data_body_cells_use_arial(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)
        self.assertEqual(sheet.cell(data_row, 3).font.name, "Arial")
        self.assertEqual(sheet.cell(data_row, 4).font.name, "Arial")
        self.assertEqual(sheet.cell(data_row, 7).font.name, "Arial")

    def test_single_select_number_formats_are_applied(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        data_row = _single_select_data_row(sheet)
        self.assertEqual(sheet.cell(data_row, 3).number_format, "#,##0")
        self.assertEqual(sheet.cell(data_row, 4).number_format, "0.0%")
        self.assertEqual(sheet.cell(data_row, 5).number_format, "#,##0")
        self.assertEqual(sheet.cell(data_row, 6).number_format, "0.0%")
        self.assertEqual(sheet.cell(data_row, 7).number_format, "+0.0%;-0.0%")

    def test_title_bar_has_bain_red_fill(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        title_row = _find_row(sheet, 1, "Q1 - Choice")
        self.assertEqual(sheet.cell(title_row, 1).fill.fgColor.rgb, "FFCC0000")
        self.assertEqual(sheet.cell(title_row, 1).font.color.rgb, "FFFFFFFF")

    def test_percent_columns_have_heatmap_conditional_formatting(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        ranges = [str(item) for item in sheet.conditional_formatting]
        self.assertTrue(any("D17:D18" in item for item in ranges))
        self.assertTrue(any("F17:F18" in item for item in ranges))

    def test_grid_rated_number_formats_are_applied(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        header_row = next(
            row
            for row in range(1, sheet.max_row + 1)
            if sheet.cell(row=row, column=1).value == "Sub-question ID"
        )
        self.assertEqual(sheet.cell(row=header_row + 1, column=3).number_format, "0.00")
        self.assertEqual(sheet.cell(row=header_row + 1, column=4).number_format, "0.00")
        self.assertEqual(sheet.cell(row=header_row + 1, column=5).number_format, "+0.00;-0.00")
        self.assertEqual(sheet.cell(row=header_row + 1, column=6).number_format, "#,##0")
        self.assertEqual(sheet.cell(row=header_row + 1, column=7).number_format, "#,##0")


if __name__ == "__main__":
    unittest.main()
