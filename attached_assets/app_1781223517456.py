"""Streamlit entry point for the Survey Analysis Engine."""

from __future__ import annotations

import html
import json
import os
import sys
import tempfile
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import streamlit as st
except ModuleNotFoundError:  # pragma: no cover - local import smoke test fallback.
    st = None  # type: ignore[assignment]


PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.ui_constants import (
    APP_TAGLINE,
    APP_TITLE,
    EMPTY_NO_CROSS_CUTS,
    EMPTY_NO_RESULTS,
    PIPELINE_STAGES,
    SECTION_CROSS_CUTS,
    SECTION_DOWNLOADS,
    SECTION_GLOBAL_FILTER,
    SECTION_RESULTS,
    SECTION_UPLOAD,
    STATUS_GLOBAL_FILTER_ACTIVE,
    STATUS_GLOBAL_FILTER_INACTIVE,
    TOOLTIP_BREAKDOWN,
    TOOLTIP_CROSS_CUT_SUGGESTIONS,
    TOOLTIP_GLOBAL_FILTER,
    TOOLTIP_PER_QUESTION_FILTER,
    TOOLTIP_THREE_DOWNLOADS,
)

import re as _re
from src.ai_insights import (
    generate_insight,
    generate_outlier_insight,
    generate_table_insight,
)
from src.chat_panel import render_chat_panel
from src.cross_cut_suggestions import score_suggestions_for_outcome
from src.filter_options import (
    build_filter_specs_from_selection,
    cross_cut_question_options,
    filter_question_options,
    resolve_cross_cut_question,
)
from src.models import InsightResult, OutcomeSegmentationResult, QuestionType


_INSIGHT_CACHE: dict[str, Any] = {}
RANK_CROSS_TAB_METRICS = (
    "Weighted Average",
    "Sum of ranks",
    "Rank position count",
)


SESSION_DEFAULTS = {
    "decoded_df": None,
    "active_df": None,
    "global_filter_state": None,
    "global_filter_stats": None,
    "global_filter_rows": [],
    "results": [],
    "skips": [],
    "schema": None,
    "quality_report": None,
    "log": None,
    "output_path": None,
    "raw_data_path_label": None,
    "datamap_path_label": None,
    "load_report": None,
    "cross_cut_results": [],
    "cross_cut_skips": [],
    "cross_cut_suggestions": [],
    "cross_cut_only_bytes": None,
    "rank_cross_tab_settings": {},
    "rank_cross_tab_settings_dirty": False,
    "filtered_results": {},
    "filtered_workbook_bytes": None,
    "single_cut_normal_view": False,
    "run_complete": False,
    "wiz_step": 1,
    "wiz_complete": False,
    "wiz_category_assignments": None,
    "wiz_selected_demographics": None,
    "wiz_num_custom_filters": 2,
    "wiz_num_per_question_filters": 0,
    "wizard_workbook_custom_filter_count": 2,
    "wizard_per_question_filter_count": 0,
    "wizard_embed_input_files": False,
    "skip_ai_enhancements": False,
    "input_file_embed_sources": None,
    "manual_cohort_input": None,
    "manual_cohort_source": None,
    "manual_cohort_id_column": None,
    "manual_cohort_overlap_blocked": False,
    "wiz_crosscut_consolidation": "one_sheet",
    "ss_search": "",
    "pending_global_filter": None,
    "pending_per_question_filter": {},
    "global_filter_error": None,
    "per_question_filter_errors": {},
    "data_map": None,
    "survey_type_result": None,
    "outcome_variable_id": None,
    "segment_definition": None,
    "segmentation_result": None,
    "outcome_segmented_workbook_bytes": None,
    "outcome_segmented_workbook_signature": None,
    "hypothesis_results": [],
    "pending_hypothesis_spec": None,
    "wizard_active": False,
    "wizard_step": 1,
    "wizard_config": {},
    "wizard_config_complete": False,
    "wizard_detected_scores": [],
    "wizard_parse_failure_note": "",
    "wizard_preview": {},
    "wizard_diagnostic": None,
    "wizard_saved_format": "",
}


# ---------------------------------------------------------------------------
# Infrastructure helpers
# ---------------------------------------------------------------------------


def _require_streamlit() -> Any:
    if st is None:
        raise RuntimeError(
            "Streamlit is not installed in this environment. "
            "Install project requirements before running the app."
        )
    return st


CROSS_CUT_ENGINE_VERSION = "day14.5"


# ---------------------------------------------------------------------------
# Visual theme + UI helpers (Day 16)
# ---------------------------------------------------------------------------


_THEME_CSS = """
<style>
html, body, [class*="css"], .stApp, .stMarkdown,
.stText, .stDataFrame, button, input, select,
textarea, label, p, div, span, h1, h2, h3 {
  font-family: Arial, Helvetica, sans-serif !important;
}
.stApp, .main .block-container { background-color: #FFFFFF !important; }
header[data-testid="stHeader"] {
  background: #FFFFFF !important;
  border-bottom: 3px solid #CC0000 !important;
  height: 52px !important;
}
#MainMenu, footer, .stDeployButton { display: none !important; }
[data-testid="stSidebar"] {
  background: #F8F8F8 !important;
  border-right: 1px solid #E0E0E0 !important;
}
[data-testid="stSidebar"] .stMarkdown p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] .stTextInput label {
  font-size: 12px !important; color: #333333 !important;
}
[data-testid="stMetric"] {
  background: #FFFFFF !important;
  border: 1px solid #E0E0E0 !important;
  border-top: 3px solid #CC0000 !important;
  padding: 16px 20px !important;
  border-radius: 0 !important;
}
[data-testid="stMetricValue"] {
  font-family: Arial, Helvetica, sans-serif !important;
  font-size: 32px !important; font-weight: 700 !important;
  color: #0A0A0A !important;
}
[data-testid="stMetricLabel"] {
  font-size: 10px !important; font-weight: 700 !important;
  letter-spacing: 0.12em !important; text-transform: uppercase !important;
  color: #888888 !important;
}
.section-header-box {
  display: flex; align-items: center; gap: 12px;
  padding: 10px 0; border-bottom: 2px solid #E0E0E0;
  margin-bottom: 20px;
}
.section-num {
  width: 26px; height: 26px; background: #CC0000; color: white;
  display: flex; align-items: center; justify-content: center;
  font-size: 12px; font-weight: 700; flex-shrink: 0;
  font-family: Arial, Helvetica, sans-serif;
}
.section-name {
  font-size: 28px; font-weight: 700; text-transform: uppercase;
  letter-spacing: 0.08em; color: #0A0A0A;
}
.section-meta {
  margin-left: auto; font-size: 10px; color: #888888;
  font-family: Arial, Helvetica, sans-serif;
}
.stButton > button {
  border-radius: 0 !important;
  font-family: Arial, Helvetica, sans-serif !important;
  font-weight: 700 !important; font-size: 11px !important;
  letter-spacing: 0.1em !important; text-transform: uppercase !important;
  border: none !important; transition: background 0.15s !important;
}
.stButton > button[kind="primary"], .stButton > button:first-child {
  background: #CC0000 !important; color: white !important;
  padding: 10px 28px !important;
}
.stButton > button:hover { background: #990000 !important; color: white !important; }
.stButton > button[kind="secondary"] {
  background: #F0F0F0 !important; color: #333333 !important;
  border: 1px solid #CCCCCC !important;
}
[data-testid="stFileUploader"] {
  background: #FFFFFF !important;
  border: 1px dashed #CC0000 !important;
  border-radius: 0 !important; padding: 24px !important;
}
[data-testid="stFileUploader"]:hover { background: #FFF5F5 !important; }
[data-testid="stFileUploaderDropzoneInstructions"] { color: #CC0000 !important; }
.stAlert { border-radius: 0 !important; font-size: 12px !important; }
[data-testid="stInfo"] {
  background: #FFF5F5 !important;
  border: 1px solid rgba(204,0,0,0.3) !important;
  border-left: 4px solid #CC0000 !important;
  color: #0A0A0A !important;
}
[data-testid="stExpander"] {
  border: 1px solid #E0E0E0 !important; border-radius: 0 !important;
  border-left: 3px solid #E0E0E0 !important;
  margin-bottom: 4px !important; background: #FFFFFF !important;
}
[data-testid="stExpander"]:hover { border-left-color: #CC0000 !important; }
[data-testid="stExpander"] summary {
  font-weight: 600 !important; font-size: 13px !important;
  padding: 12px 16px !important; color: #0A0A0A !important;
}
[data-testid="stExpanderDetails"] {
  border-top: 1px solid #E0E0E0 !important; padding: 16px !important;
}
[data-testid="stSelectbox"] > div, [data-testid="stMultiSelect"] > div {
  border-radius: 0 !important;
}
.stSelectbox [data-baseweb="select"] div,
.stMultiSelect [data-baseweb="select"] div {
  border-radius: 0 !important; border-color: #CCCCCC !important;
  font-size: 12px !important;
}
.stTextInput input {
  border-radius: 0 !important; border-color: #CCCCCC !important;
  font-size: 12px !important;
}
[data-testid="stDataFrame"] {
  border: 1px solid #E0E0E0 !important; border-radius: 0 !important;
}
</style>
"""


_THEME_CSS_DAY18 = """
<style>
/* Day 18 — Fixed red header bar */
header[data-testid="stHeader"] {
  background: #CC0000 !important;
  border-bottom: none !important;
  height: 64px !important;
  z-index: 999 !important;
  position: fixed !important;
  top: 0 !important; left: 0 !important; right: 0 !important;
}
header[data-testid="stHeader"]::before {
  content: ''; position: absolute; inset: 0;
  background: #CC0000; z-index: -1;
}
header[data-testid="stHeader"] [data-testid="stToolbar"] {
  display: none !important;
}
.stApp { margin-top: 0 !important; }
.main .block-container {
  padding-top: 24px !important;
  max-width: 100% !important;
}
[data-testid="stSidebar"] {
  margin-top: 64px !important;
  height: calc(100vh - 64px) !important;
}
.custom-header {
  position: fixed; top: 0; left: 0; right: 0;
  height: 64px; background: #CC0000; z-index: 1000;
  display: flex; align-items: center;
  padding: 0 32px; color: white;
  font-family: Arial, Helvetica, sans-serif;
}
.custom-header-title {
  font-size: 20px; font-weight: 700;
  color: white; letter-spacing: 0.02em; margin-right: 24px;
}
.custom-header-tagline {
  font-size: 12px; color: rgba(255,255,255,0.85);
  font-weight: 400;
  border-left: 1px solid rgba(255,255,255,0.3);
  padding-left: 24px;
}

/* Day 18 — Sidebar question navigation buttons */
[data-testid="stSidebar"] .stButton > button {
  text-align: left !important;
  justify-content: flex-start !important;
  white-space: normal !important;
  height: auto !important;
  min-height: 44px !important;
  padding: 10px 12px !important;
  font-size: 12px !important;
  font-weight: 500 !important;
  line-height: 1.4 !important;
  letter-spacing: 0 !important;
  text-transform: none !important;
  border-radius: 0 !important;
  border-left: 3px solid transparent !important;
  background: white !important;
  color: #333 !important;
  border: 1px solid #E8E8E8 !important;
  margin-bottom: 2px !important;
  word-wrap: break-word !important;
  overflow-wrap: break-word !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
  background: #FFF5F5 !important;
  border-left-color: #CC0000 !important;
  color: #0A0A0A !important;
}
[data-testid="stSidebar"] .stButton > button[kind="primary"] {
  background: #FFF0F0 !important;
  border-left: 3px solid #CC0000 !important;
  color: #CC0000 !important;
  font-weight: 700 !important;
}

/* Day 18 — Plotly chart styling */
.plotly, .plotly-chart { background: white !important; }
.plotly-chart [class*="modebar"] {
  background: rgba(255,255,255,0.8) !important;
  border: 1px solid #E0E0E0 !important;
}
.plotly-chart [class*="modebar-btn"] { color: #666 !important; }
</style>
"""


_POLISH_CSS_6D = """
<style>
/* ===== 6D polish: scoped to .main so header/sidebar are untouched ===== */
.main .block-container h1 { font-size:26px !important; font-weight:700 !important; color:#15171A !important; letter-spacing:-0.01em !important; }
.main .block-container h2 { font-size:19px !important; font-weight:600 !important; color:#15171A !important; }
.main .block-container h3 { font-size:15px !important; font-weight:600 !important; color:#15171A !important; }
.main .block-container p, .main .block-container li { font-size:14px !important; line-height:1.6 !important; color:#2B2F33 !important; }

.main .stButton > button, .main .stDownloadButton > button { border-radius:8px !important; text-transform:none !important; letter-spacing:0.01em !important; font-weight:600 !important; font-size:13px !important; padding:10px 20px !important; transition:transform 0.12s ease, box-shadow 0.12s ease, background 0.12s ease, border-color 0.12s ease, color 0.12s ease !important; }
.main .stButton > button:active, .main .stDownloadButton > button:active { transform:translateY(1px) !important; }
.main .stButton > button[kind="primary"], .main .stButton > button:first-child, .main .stDownloadButton > button { background:linear-gradient(180deg,#E00000 0%,#CC0000 100%) !important; color:#FFFFFF !important; border:1px solid #B30000 !important; box-shadow:0 1px 2px rgba(140,0,0,0.30), inset 0 1px 0 rgba(255,255,255,0.15) !important; }
.main .stButton > button[kind="primary"]:hover, .main .stButton > button:first-child:hover, .main .stDownloadButton > button:hover { background:linear-gradient(180deg,#CC0000 0%,#B30000 100%) !important; box-shadow:0 6px 16px rgba(204,0,0,0.30) !important; transform:translateY(-1px) !important; }
.main .stButton > button[kind="primary"]:focus-visible, .main .stButton > button:first-child:focus-visible, .main .stDownloadButton > button:focus-visible { outline:none !important; box-shadow:0 0 0 3px rgba(204,0,0,0.32) !important; }
.main .stButton > button[kind="secondary"] { background:#FFFFFF !important; color:#1A1A1A !important; border:1px solid #D5D7DA !important; box-shadow:0 1px 2px rgba(16,24,40,0.05) !important; }
.main .stButton > button[kind="secondary"]:hover { background:#FFF5F5 !important; border-color:#CC0000 !important; color:#CC0000 !important; box-shadow:0 4px 12px rgba(16,24,40,0.08) !important; transform:translateY(-1px) !important; }
.main .stButton > button[kind="secondary"]:focus-visible { outline:none !important; box-shadow:0 0 0 3px rgba(204,0,0,0.18) !important; }

.main .stTextInput input, .main .stNumberInput input, .main textarea { border-radius:8px !important; border:1px solid #D8D8D8 !important; font-size:13px !important; padding:8px 12px !important; }
.main .stTextInput input:focus, .main .stNumberInput input:focus, .main textarea:focus { border-color:#CC0000 !important; box-shadow:0 0 0 3px rgba(204,0,0,0.10) !important; }
.main .stSelectbox [data-baseweb="select"] > div, .main .stMultiSelect [data-baseweb="select"] > div { border-radius:8px !important; border-color:#D8D8D8 !important; font-size:13px !important; }

.main [data-testid="stExpander"] { border:1px solid #E6E6E6 !important; border-left:1px solid #E6E6E6 !important; border-radius:10px !important; box-shadow:0 1px 2px rgba(16,24,40,0.04) !important; overflow:hidden !important; margin-bottom:10px !important; }
.main [data-testid="stExpander"]:hover { border-color:#D0D0D0 !important; border-left-color:#CC0000 !important; box-shadow:0 2px 8px rgba(16,24,40,0.06) !important; }
.main [data-testid="stExpander"] summary { font-size:14px !important; font-weight:600 !important; padding:13px 16px !important; }

.main [data-testid="stMetric"] { border:1px solid #E6E6E6 !important; border-top:3px solid #CC0000 !important; border-radius:10px !important; box-shadow:0 1px 2px rgba(16,24,40,0.04) !important; padding:16px 18px !important; }

.main [data-testid="stTabs"] [data-baseweb="tab-list"] { gap:4px !important; border-bottom:1px solid #E6E6E6 !important; }
.main [data-testid="stTabs"] [data-baseweb="tab"] { font-size:13px !important; font-weight:600 !important; color:#8A8F94 !important; padding:8px 16px !important; }
.main [data-testid="stTabs"] [aria-selected="true"] { color:#CC0000 !important; }
.main [data-testid="stTabs"] [data-baseweb="tab-highlight"] { background:#CC0000 !important; height:2px !important; }

.main [data-testid="stRadio"] > div[role="radiogroup"] { display:inline-flex !important; gap:0 !important; border:1px solid #D8D8D8 !important; border-radius:8px !important; background:#F5F6F7 !important; padding:2px !important; }
.main [data-testid="stRadio"] > div[role="radiogroup"] > label { margin:0 !important; padding:5px 14px !important; border-radius:6px !important; font-size:12px !important; font-weight:600 !important; color:#5A5A5A !important; cursor:pointer !important; }
.main [data-testid="stRadio"] > div[role="radiogroup"] > label:has(input:checked) { background:#FFFFFF !important; color:#CC0000 !important; box-shadow:0 1px 2px rgba(16,24,40,0.10) !important; }
.main [data-testid="stRadio"] > div[role="radiogroup"] > label > div:first-child { display:none !important; }

.main [data-testid="stDataFrame"] { border:1px solid #E6E6E6 !important; border-radius:10px !important; overflow:hidden !important; }

.main .stAlert { border-radius:10px !important; }
.main [data-testid="stInfo"] { background:#F7F9FB !important; border:1px solid #E3E8EE !important; border-left:4px solid #CC0000 !important; color:#2B2F33 !important; }

.main .section-header-box { border-bottom:1px solid #ECECEC !important; padding:8px 0 12px !important; margin-bottom:18px !important; }
.main .section-num { border-radius:8px !important; width:40px !important; height:40px !important; font-size:20px !important; box-shadow:0 1px 2px rgba(204,0,0,0.25) !important; }
.main .section-name { font-size:28px !important; letter-spacing:0.06em !important; }

.main [data-testid="stFileUploader"] { border:1.5px dashed #D8B0B0 !important; border-radius:10px !important; background:#FAFAFA !important; }
.main [data-testid="stFileUploader"]:hover { border-color:#CC0000 !important; background:#FFF8F8 !important; }
</style>
"""


_UI_V2_CSS = """
<style>
/* ===================== UI V2 — bold, cohesive, easy to navigate ===================== */

/* ---- Hero header: charcoal -> Bain red gradient, tagline ---- */
.custom-header { background:#CC0000 !important; height:60px !important; border-bottom:1px solid #7a0000 !important; box-shadow:0 2px 10px rgba(0,0,0,0.18) !important; padding:0 28px !important; }
.custom-header-title { font-size:15px !important; letter-spacing:0.01em !important; }
.custom-header-title strong { font-weight:700 !important; letter-spacing:0.02em !important; }
.main .block-container { padding-top:24px !important; }
[id^="section-"] { scroll-margin-top:78px !important; }

/* ---- Nav: 3 tabs + a "More" dropdown ---- */
.nav-tab { height:60px !important; font-size:13px !important; font-weight:500 !important; }
.nav-more { position:relative !important; display:inline-flex !important; align-items:center !important; height:60px !important; }
.nav-more-toggle { cursor:pointer !important; }
.nav-more-menu { position:absolute !important; top:60px !important; right:0 !important; min-width:210px !important; background:#FFFFFF !important; border:1px solid #E6E6E6 !important; border-radius:10px !important; box-shadow:0 12px 30px rgba(0,0,0,0.18) !important; padding:6px !important; display:none !important; flex-direction:column !important; z-index:1000000 !important; }
.nav-more:hover .nav-more-menu { display:flex !important; }
.nav-more-item { display:flex !important; align-items:center !important; gap:8px !important; padding:9px 12px !important; font-size:13px !important; font-weight:500 !important; color:#1A1A1A !important; text-decoration:none !important; border-radius:7px !important; white-space:nowrap !important; cursor:pointer !important; }
.nav-more-item:hover { background:#FFF5F5 !important; color:#CC0000 !important; }
.nav-more-item .nav-badge { background:#CC0000 !important; color:#FFF !important; margin-left:auto !important; }

/* ---- Journey rail: 3-phase animated progress ---- */
.journey { display:flex !important; align-items:center !important; gap:0 !important; margin:2px 0 22px !important; padding:14px 18px !important; background:#FFFFFF !important; border:1px solid #ECECEC !important; border-radius:14px !important; box-shadow:0 1px 3px rgba(16,24,40,0.05) !important; font-family:Arial, sans-serif !important; }
.journey-step { display:flex !important; align-items:center !important; gap:10px !important; flex:0 0 auto !important; }
.journey-dot { width:34px !important; height:34px !important; border-radius:50% !important; display:flex !important; align-items:center !important; justify-content:center !important; font-size:15px !important; font-weight:700 !important; flex-shrink:0 !important; transition:all 0.3s ease !important; }
.journey-dot.done { background:linear-gradient(180deg,#E00000,#CC0000) !important; color:#FFF !important; box-shadow:0 3px 10px rgba(204,0,0,0.30) !important; }
.journey-dot.active { background:#FFFFFF !important; color:#CC0000 !important; border:2px solid #CC0000 !important; box-shadow:0 0 0 4px rgba(204,0,0,0.14) !important; animation:journeyPulse 1.8s ease-in-out infinite !important; }
.journey-dot.todo { background:#F2F3F5 !important; color:#A8ADB3 !important; border:2px solid #E2E4E8 !important; }
@keyframes journeyPulse { 0%,100%{box-shadow:0 0 0 4px rgba(204,0,0,0.14);} 50%{box-shadow:0 0 0 8px rgba(204,0,0,0.05);} }
.journey-text { display:flex !important; flex-direction:column !important; line-height:1.2 !important; }
.journey-title { font-size:13px !important; font-weight:700 !important; color:#1A1A1A !important; }
.journey-sub { font-size:11px !important; color:#8A8F94 !important; }
.journey-step.todo .journey-title { color:#A8ADB3 !important; }
.journey-line { flex:1 1 auto !important; height:3px !important; margin:0 14px !important; border-radius:3px !important; background:#E8EAED !important; min-width:24px !important; }
.journey-line.done { background:linear-gradient(90deg,#CC0000,#E00000) !important; }
/* Bigger, more visible variant for the 4-step guide */
.journey-big { padding:18px 22px !important; }
.journey-big .journey-dot { width:46px !important; height:46px !important; font-size:20px !important; }
.journey-big .journey-title { font-size:15px !important; }
.journey-big .journey-sub { font-size:12px !important; }
.journey-big .journey-step.active .journey-title { color:#CC0000 !important; }
.journey-next { margin:-12px 0 20px !important; padding:10px 18px !important; background:#FFF6F6 !important; border:1px solid #F6D5D5 !important; border-left:4px solid #CC0000 !important; border-radius:8px !important; font-family:Arial,sans-serif !important; font-size:13px !important; font-weight:600 !important; color:#9A1B1B !important; }

/* ---- Count-up stat tiles ---- */
.stat-row { display:flex !important; gap:14px !important; flex-wrap:wrap !important; margin:4px 0 8px !important; }
.stat-tile { flex:1 1 150px !important; background:#FFFFFF !important; border:1px solid #ECECEC !important; border-top:3px solid #CC0000 !important; border-radius:14px !important; padding:16px 18px !important; box-shadow:0 1px 3px rgba(16,24,40,0.06) !important; transition:transform 0.15s ease, box-shadow 0.15s ease !important; font-family:Arial, sans-serif !important; }
.stat-tile:hover { transform:translateY(-2px) !important; box-shadow:0 8px 20px rgba(16,24,40,0.10) !important; }
.stat-num { font-size:30px !important; font-weight:800 !important; color:#CC0000 !important; line-height:1 !important; letter-spacing:-0.02em !important; }
.stat-label { font-size:11px !important; font-weight:600 !important; text-transform:uppercase !important; letter-spacing:0.07em !important; color:#8A8F94 !important; margin-top:6px !important; }
.stat-icon { font-size:16px !important; opacity:0.9 !important; }

/* ---- Question-type color system (pills + sidebar accents) ---- */
.qtype { display:inline-flex !important; align-items:center !important; gap:5px !important; font-size:10px !important; font-weight:700 !important; letter-spacing:0.04em !important; text-transform:uppercase !important; padding:2px 8px !important; border-radius:20px !important; line-height:1.5 !important; }
.qtype-single  { background:#E8F0FE !important; color:#1A56C4 !important; }
.qtype-multi   { background:#E7F6EC !important; color:#1E7D43 !important; }
.qtype-nps     { background:#FCEBEB !important; color:#CC0000 !important; }
.qtype-grid    { background:#F3EAFB !important; color:#7A2BB8 !important; }
.qtype-numeric { background:#FFF3E0 !important; color:#B26A00 !important; }
.qtype-rank    { background:#E6F7F7 !important; color:#0E7C86 !important; }
.qtype-other   { background:#F0F1F3 !important; color:#5A5F66 !important; }

/* ---- Richer buttons: deeper gradient, glow, lift, ripple-ish press ---- */
.main .stButton > button, .main .stDownloadButton > button { border-radius:10px !important; padding:11px 22px !important; font-size:13px !important; }
.main .stButton > button[kind="primary"], .main .stButton > button:first-child, .main .stDownloadButton > button { background:linear-gradient(180deg,#EE0B0B 0%,#CC0000 55%,#B30000 100%) !important; border:1px solid #9e0000 !important; box-shadow:0 2px 4px rgba(140,0,0,0.30), inset 0 1px 0 rgba(255,255,255,0.18) !important; }
.main .stButton > button[kind="primary"]:hover, .main .stButton > button:first-child:hover, .main .stDownloadButton > button:hover { box-shadow:0 8px 22px rgba(204,0,0,0.38) !important; transform:translateY(-2px) !important; }

/* ---- Cards lift on hover ---- */
.main [data-testid="stExpander"] { transition:transform 0.15s ease, box-shadow 0.15s ease, border-color 0.15s ease !important; }
.main [data-testid="stExpander"]:hover { transform:translateY(-1px) !important; }

/* ---- Sidebar getting-started checklist ---- */
.gs-wrap { font-family:Arial, sans-serif !important; }
.gs-title { font-size:13px !important; font-weight:700 !important; color:#1A1A1A !important; margin:4px 0 10px !important; display:flex !important; align-items:center !important; gap:7px !important; }
.gs-step { display:flex !important; gap:11px !important; padding:10px 0 !important; border-bottom:1px solid #EFEFEF !important; }
.gs-step:last-child { border-bottom:none !important; }
.gs-badge { width:26px !important; height:26px !important; border-radius:50% !important; flex-shrink:0 !important; display:flex !important; align-items:center !important; justify-content:center !important; font-size:13px !important; font-weight:700 !important; }
.gs-badge.done { background:linear-gradient(180deg,#E00000,#CC0000) !important; color:#FFF !important; }
.gs-badge.todo { background:#F2F3F5 !important; color:#9AA0A6 !important; border:1.5px solid #E2E4E8 !important; }
.gs-body { line-height:1.3 !important; }
.gs-step-title { font-size:12.5px !important; font-weight:600 !important; color:#1A1A1A !important; }
.gs-step.todo .gs-step-title { color:#9AA0A6 !important; }
.gs-step-desc { font-size:11px !important; color:#8A8F94 !important; margin-top:2px !important; }
</style>
"""


_BAIN_ICONS: dict[str, str] = {
    "upload":    '<path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/>',
    "download":  '<path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/>',
    "chart":     '<line x1="12" y1="20" x2="12" y2="10"/><line x1="18" y1="20" x2="18" y2="4"/><line x1="6" y1="20" x2="6" y2="14"/>',
    "users":     '<path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 0 0-3-3.87"/><path d="M16 3.13a4 4 0 0 1 0 7.75"/>',
    "clipboard": '<path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"/><rect x="8" y="2" width="8" height="4" rx="1" ry="1"/>',
    "list":      '<line x1="8" y1="6" x2="21" y2="6"/><line x1="8" y1="12" x2="21" y2="12"/><line x1="8" y1="18" x2="21" y2="18"/><line x1="3" y1="6" x2="3.01" y2="6"/><line x1="3" y1="12" x2="3.01" y2="12"/><line x1="3" y1="18" x2="3.01" y2="18"/>',
    "spark":     '<path d="M12 3l1.6 5.4L19 10l-5.4 1.6L12 17l-1.6-5.4L5 10l5.4-1.6L12 3z"/>',
    "chev-left":  '<polyline points="15 18 9 12 15 6"/>',
    "chev-right": '<polyline points="9 18 15 12 9 6"/>',
}


def _svg_icon(name: str, size: int = 16, color: str = "currentColor", stroke: float = 1.75) -> str:
    """Inline SVG icon. color='currentColor' lets it inherit parent text color."""
    body = _BAIN_ICONS.get(name, "")
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" '
        f'stroke="{color}" stroke-width="{stroke}" stroke-linecap="round" '
        f'stroke-linejoin="round" style="vertical-align:middle;display:inline-block;flex-shrink:0;">'
        f'{body}</svg>'
    )


def _custom_header_html() -> str:
    app = _require_streamlit()
    return (
        '<div class="custom-header">'
        '<span class="custom-header-title">'
        '<strong>Survey Analysis Engine</strong>'
        '<span style="opacity:0.78;font-weight:400;margin-left:10px;">'
        'turn any survey into a consultant-ready workbook</span>'
        '</span>'
        '<div class="custom-header-nav" id="custom-header-nav"></div>'
        '</div>'
    )


def _inject_global_css() -> None:
    app = _require_streamlit()

    app.markdown(
        """
    <style>
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }

    .insight-card { display: flex; align-items: flex-start; gap: 10px; padding: 12px 14px; background: #FFF; border: 0.5px solid #E5E5E5; border-left: 3px solid #CC0000; border-radius: 6px; margin-bottom: 8px; font-family: Arial, sans-serif; }
    .insight-icon { display: flex; align-items: center; justify-content: center; width: 24px; height: 24px; background: #FCEBEB; color: #CC0000; border-radius: 4px; font-size: 14px; flex-shrink: 0; margin-top: 1px; }
    .insight-body { flex: 1; min-width: 0; }
    .insight-label { font-size: 10px; font-weight: 500; color: #888; text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 3px; }
    .insight-text { font-size: 14px; font-weight: 500; line-height: 1.45; color: #1A1A1A; }
    .insight-text .num { color: #CC0000; font-weight: 500; }
    .insight-footer { font-size: 10px; color: #999; margin-top: 4px; }
    .insight-template { background: #F5F5F5; border-left-color: #888; }
    .insight-template .insight-icon { background: #E5E5E5; color: #888; }

    .ui-panel { background: #FFF; border: 0.5px solid #E5E5E5; border-radius: 8px; margin-bottom: 12px; overflow: hidden; }
    .ui-panel-head { display: flex; align-items: center; padding: 9px 14px; border-bottom: 0.5px solid #EEE; font-size: 12px; background: #FAFAFA; }
    .ui-panel-title { font-weight: 500; color: #1A1A1A; }
    .ui-panel-meta { margin-left: auto; color: #888; font-size: 11px; }

    div[data-testid="stDataFrame"] { font-size: 12px; }

    /* Nav bar buttons */
    div[data-testid="stButton"] button {
        border-radius: 6px !important;
        font-size: 12px !important;
        padding: 6px 8px !important;
    }
    div[data-testid="stButton"] button[kind="primary"] {
        background: #CC0000 !important;
        border-color: #CC0000 !important;
        color: #FFF !important;
    }
    div[data-testid="stButton"] button[kind="primary"]:hover {
        background: #B30000 !important;
        border-color: #B30000 !important;
    }

    /* CHANGE B — merged red header with embedded nav tabs */
    .custom-header {
        position: fixed !important;
        top: 0 !important;
        left: 0 !important;
        right: 0 !important;
        z-index: 999999 !important;
        background: #CC0000 !important;
        color: #FFFFFF !important;
        padding: 0 24px !important;
        height: 56px !important;
        display: flex !important;
        align-items: center !important;
        gap: 24px !important;
        border-bottom: 1px solid #A30000 !important;
        font-family: Arial, sans-serif !important;
        box-sizing: border-box !important;
    }
    .custom-header-title {
        font-size: 14px !important;
        font-weight: 500 !important;
        color: #FFF !important;
        flex-shrink: 0 !important;
        white-space: nowrap !important;
    }
    .custom-header-title strong { font-weight: 600 !important; }
    .custom-header-nav {
        display: flex !important;
        align-items: center !important;
        gap: 0 !important;
        margin-left: auto !important;
        height: 100% !important;
    }
    .nav-bar { display: none !important; }
    .nav-tab {
        display: inline-flex !important;
        align-items: center !important;
        gap: 6px !important;
        padding: 0 14px !important;
        height: 56px !important;
        font-size: 13px !important;
        font-family: Arial, sans-serif !important;
        font-weight: 400 !important;
        color: rgba(255,255,255,0.75) !important;
        text-decoration: none !important;
        border: none !important;
        border-bottom: 3px solid transparent !important;
        border-radius: 0 !important;
        background: transparent !important;
        white-space: nowrap !important;
        cursor: pointer !important;
        transition: color 0.15s ease, border-color 0.15s ease, background 0.15s ease !important;
        box-sizing: border-box !important;
        outline: none !important;
    }
    .nav-tab:hover {
        color: #FFFFFF !important;
        background: rgba(255,255,255,0.08) !important;
        border-bottom: 3px solid rgba(255,255,255,0.4) !important;
        text-decoration: none !important;
    }
    .nav-tab.active {
        color: #FFFFFF !important;
        font-weight: 500 !important;
        border-bottom: 3px solid #FFFFFF !important;
        background: rgba(255,255,255,0.08) !important;
    }
    .nav-badge {
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        background: rgba(255,255,255,0.2) !important;
        color: #FFFFFF !important;
        font-size: 10px !important;
        font-weight: 500 !important;
        padding: 1px 6px !important;
        border-radius: 10px !important;
        min-width: 18px !important;
        line-height: 1.4 !important;
    }
    .nav-tab.active .nav-badge {
        background: #FFFFFF !important;
        color: #CC0000 !important;
    }
    .main .block-container {
        padding-top: 24px !important;
        max-width: 100% !important;
    }
    [id^="section-"] {
        scroll-margin-top: 72px !important;
        display: block !important;
    }
    header[data-testid="stHeader"] { display: none !important; }
    </style>
    """,
        unsafe_allow_html=True,
    )
    app.markdown(_THEME_CSS, unsafe_allow_html=True)
    app.markdown(_THEME_CSS_DAY18, unsafe_allow_html=True)
    app.markdown(_custom_header_html(), unsafe_allow_html=True)
    app.markdown(_POLISH_CSS_6D, unsafe_allow_html=True)
    app.markdown(_UI_V2_CSS, unsafe_allow_html=True)

def _inject_theme_css() -> None:
    """Backwards-compatible alias retained for any internal callers."""
    _inject_global_css()


def _section_header(
    num: str, title: str, anchor: str | None = None, meta: str = ""
) -> None:
    app = _require_streamlit()
    anchor_html = f"<a name='{anchor}'></a>" if anchor else ""
    meta_html = f"<div class='section-meta'>{meta}</div>" if meta else ""
    app.markdown(
        f"{anchor_html}<div class='section-header-box'>"
        f"<div class='section-num'>{num}</div>"
        f"<div class='section-name'>{title}</div>"
        f"{meta_html}</div>",
        unsafe_allow_html=True,
    )


def _red_instruction_box(title: str, body: str) -> None:
    """Short red-bordered instruction box (matches the upload step's box)."""
    app = _require_streamlit()
    app.markdown(
        "<div style='padding:12px 16px;background:#FFF6F6;border:1px solid #F6D5D5;"
        "border-left:4px solid #CC0000;border-radius:10px;margin:2px 0 16px;"
        "font-family:Arial,sans-serif;'>"
        f"<div style='font-weight:700;font-size:13px;color:#1A1A1A;margin-bottom:3px;'>{title}</div>"
        f"<div style='font-size:12.5px;color:#444;line-height:1.5;'>{body}</div></div>",
        unsafe_allow_html=True,
    )


def _section_header_no_num(title: str, meta: str = "") -> None:
    """Header in the same style as _section_header but with no number box.

    Used for sub-steps like the output-structure config so it reads as the same
    visual tier as '1. UPLOAD YOUR SURVEY' without repeating a step number.
    """
    app = _require_streamlit()
    meta_html = f"<div class='section-meta'>{meta}</div>" if meta else ""
    app.markdown(
        f"<div class='section-header-box'>"
        f"<div class='section-name'>{title}</div>"
        f"{meta_html}</div>",
        unsafe_allow_html=True,
    )


def _compute_outlier_flags(values: list) -> list:
    """Return 'high' / 'low' / '' per value using a guarded z-score rule.

    Never raises — wrapped in a defensive try/except so a malformed value
    list cannot break a render path.
    """
    try:
        import statistics

        flags = ["" for _ in values]
        numeric = [
            v for v in values
            if v is not None and isinstance(v, (int, float)) and v == v
        ]
        if len(numeric) < 4:
            return flags
        if max(numeric) < 10:
            return flags
        if len(set(numeric)) <= 1:
            return flags
        mean = statistics.mean(numeric)
        std = statistics.stdev(numeric)
        if std == 0:
            return flags
        val_range = max(numeric) - min(numeric)
        min_val = min(numeric)
        range_significant = (
            val_range > mean * 0.5
            and (
                min_val == 0
                or max(numeric) / max(min_val, 0.001) > 3
            )
        )
        for i, v in enumerate(values):
            if v is None or not isinstance(v, (int, float)):
                continue
            z = (v - mean) / std
            if z > 2.0 and v >= mean + val_range * 0.5:
                flags[i] = "high"
            elif z < -1.5 and range_significant and v < mean * 0.3:
                flags[i] = "low"
        return flags
    except Exception:
        return ["" for _ in values]


_ICONS = {
    "analysis": '<svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 3v18h18"/><path d="M7 12l3-3 3 3 5-5"/></svg>',
    "insight": '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 18h6"/><path d="M10 22h4"/><path d="M15.09 14c.18-.98.65-1.74 1.41-2.5A4.65 4.65 0 0 0 18 8 6 6 0 0 0 6 8c0 1 .23 2.23 1.5 3.5A4.61 4.61 0 0 1 8.91 14"/></svg>',
    "outlier": '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10.29 3.86 1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"/><line x1="12" y1="9" x2="12" y2="13"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg>',
    "winner": '<svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="8" r="6"/><path d="M15.477 12.89 17 22l-5-3-5 3 1.523-9.11"/></svg>',
}


def _icon(name: str, color: str = "currentColor") -> str:
    """Render an inline SVG icon as an HTML string."""
    svg = _ICONS.get(name, "")
    if not svg:
        return ""
    return (
        f'<span style="display:inline-flex;align-items:center;'
        f'color:{color};margin-right:8px;vertical-align:middle;">{svg}</span>'
    )


def _style_outliers(df: Any) -> Any:
    """Per-column outlier styling for st.dataframe Styler.apply(axis=None)."""

    try:
        styled = pd.DataFrame("", index=df.index, columns=df.columns)
        for col in df.columns:
            try:
                vals = pd.to_numeric(df[col], errors="coerce").tolist()
                col_flags = _compute_outlier_flags(vals)
                for i, flag in enumerate(col_flags):
                    if flag == "high":
                        styled.iloc[i, styled.columns.get_loc(col)] = (
                            "background-color:#FFF0F0;color:#CC0000;"
                            "font-weight:bold;border-left:3px solid #CC0000"
                        )
                    elif flag == "low":
                        styled.iloc[i, styled.columns.get_loc(col)] = (
                            "background-color:#FFFBE6;color:#E65100;"
                            "font-weight:600"
                        )
            except Exception:
                continue
        return styled
    except Exception:
        return pd.DataFrame("", index=df.index, columns=df.columns)


def _style_outliers_by_row(df: Any) -> Any:
    """Per-row outlier styling — applies _style_outliers logic across each row.

    For row-percent cross-tabs, outliers should be detected within each row
    (since row values sum to 100%). Implemented by transposing, reusing the
    per-column logic, and transposing the result back. Values are scaled to
    0-100 first because _compute_outlier_flags ignores series whose max < 10.
    """
    try:
        scaled = df * 100
        styled_t = _style_outliers(scaled.T)
        return styled_t.T
    except Exception:
        return pd.DataFrame("", index=df.index, columns=df.columns)


def _style_outliers_pct(df: Any) -> Any:
    """Per-column outlier styling for fractional (0-1) percent dataframes.

    Scales values to 0-100 before delegating to _style_outliers so the
    `max < 10` guard in _compute_outlier_flags doesn't silently disable
    highlighting on percent tables.
    """
    try:
        scaled = df * 100
        return _style_outliers(scaled)
    except Exception:
        return pd.DataFrame("", index=df.index, columns=df.columns)


def _styled_dataframe(df: Any, **kwargs: Any) -> None:
    """Render a dataframe with outlier styling, falling back gracefully."""
    app = _require_streamlit()
    try:
        app.dataframe(
            df.style.apply(_style_outliers, axis=None), **kwargs
        )
    except Exception:
        app.dataframe(df, **kwargs)


def _clear_all_cached_insights() -> None:
    """Drop every cached AI insight from session state.

    Called whenever the underlying data context changes (global filter apply,
    pipeline rerun, new upload) so users never see insights stale relative to
    the current data.
    """
    app = _require_streamlit()
    stale = [
        k for k in list(app.session_state.keys())
        if isinstance(k, str) and (
            k.startswith("insight_sc_")
            or k.startswith("insight_cc_")
            or k.startswith("insight_breakdown_")
        )
    ]
    for k in stale:
        app.session_state.pop(k, None)


def _build_insight_payload_single_cut(
    result: Any, spec: Any, filters_applied: list | None = None
) -> dict:
    """Build payload dict for a SingleCutResult (any subtype)."""
    rows: list[dict[str, Any]] = []
    if hasattr(result, "distribution") and getattr(result, "distribution", None):
        for code, payload in result.distribution.items():
            label = (
                spec.option_map.get(code, str(code))
                if getattr(spec, "option_map", None)
                else str(code)
            )
            rows.append(
                {
                    "label": label,
                    "count": payload.get("count", 0) if isinstance(payload, dict) else 0,
                    "rate": payload.get("rate", 0.0) if isinstance(payload, dict) else 0.0,
                }
            )
    elif hasattr(result, "selections") and getattr(result, "selections", None):
        for sub_id, payload in result.selections.items():
            if not isinstance(payload, dict):
                continue
            label = payload.get("label", sub_id) or sub_id
            label_lower = str(label).lower()
            if "unchecked" in label_lower or "not selected" in label_lower:
                continue
            rate = payload.get("selection_rate")
            if rate is None:
                rate = payload.get("rate", 0.0)
            rows.append(
                {
                    "label": label,
                    "count": payload.get("count", 0),
                    "rate": float(rate) if rate is not None else 0.0,
                }
            )
    elif hasattr(result, "rows") and getattr(result, "rows", None):
        for sub_id, row_result in result.rows.items():
            row_label = (
                spec.grid_row_labels.get(sub_id, sub_id)
                if getattr(spec, "grid_row_labels", None)
                else sub_id
            )
            dist = (
                row_result.distribution
                if hasattr(row_result, "distribution")
                else {}
            )
            for code, payload in dist.items():
                if not isinstance(payload, dict):
                    continue
                rows.append(
                    {
                        "label": f"{row_label} \u2014 {payload.get('label', str(code))}",
                        "count": payload.get("count", 0),
                        "rate": payload.get("rate", 0.0),
                    }
                )

    summary: dict[str, Any] = {}
    if hasattr(result, "mean") and result.mean is not None:
        try:
            summary["mean"] = round(float(result.mean), 4)
        except (TypeError, ValueError):
            pass
    if hasattr(result, "median") and result.median is not None:
        try:
            summary["median"] = round(float(result.median), 4)
        except (TypeError, ValueError):
            pass
    if hasattr(result, "std") and result.std is not None:
        try:
            summary["std"] = round(float(result.std), 4)
        except (TypeError, ValueError):
            pass

    # Numeric subtypes have no row-wise distribution; synthesize summary rows
    # so the AI insight pipeline does not fall back to template-only mode.
    if not rows and summary:
        for stat_name, stat_value in summary.items():
            rows.append(
                {"label": stat_name, "count": int(getattr(result, "valid_n", 0) or 0), "rate": stat_value}
            )

    return {
        "table_kind": "single_cut",
        "question_id": spec.canonical_id,
        "question_text": spec.question_text,
        "valid_n": int(getattr(result, "valid_n", 0) or 0),
        "missing_n": int(getattr(result, "missing_n", 0) or 0),
        "filters_applied": filters_applied or [],
        "rows": rows,
        "summary": summary,
    }


def _build_insight_payload_cross_cut(result: Any) -> dict:
    """Build payload dict for a CrossCutResult."""
    rt = result.result_table or {}
    rows: list[dict[str, Any]] = []

    if "counts" in rt:
        row_labels = rt.get("row_label_map", {}) or {}
        col_labels = rt.get("column_label_map", {}) or {}
        row_pcts = rt.get("row_pct", {}) or {}
        for row_code, col_dict in (rt.get("counts", {}) or {}).items():
            if not isinstance(col_dict, dict):
                continue
            for col_code, count in col_dict.items():
                rows.append(
                    {
                        "row_label": row_labels.get(row_code, str(row_code)),
                        "col_label": col_labels.get(col_code, str(col_code)),
                        "count": count,
                        "row_pct": (row_pcts.get(row_code, {}) or {}).get(
                            col_code, 0.0
                        ),
                    }
                )
    elif "per_segment" in rt:
        for seg_val, seg_data in (rt.get("per_segment", {}) or {}).items():
            if not isinstance(seg_data, dict):
                continue
            try:
                mean_v = round(float(seg_data.get("mean", 0) or 0), 4)
            except (TypeError, ValueError):
                mean_v = 0
            try:
                median_v = round(float(seg_data.get("median", 0) or 0), 4)
            except (TypeError, ValueError):
                median_v = 0
            rows.append(
                {
                    "segment_label": seg_data.get("label", str(seg_val)),
                    "n": seg_data.get("n", 0),
                    "mean": mean_v,
                    "median": median_v,
                }
            )
    elif "target_result" in rt:
        tr = rt.get("target_result", {}) or {}
        if "distribution" in tr:
            for code, payload in tr["distribution"].items():
                if not isinstance(payload, dict):
                    continue
                rows.append(
                    {
                        "label": payload.get("label", str(code)),
                        "count": payload.get("count", 0),
                        "rate": payload.get("rate", 0.0),
                    }
                )
        elif "selections" in tr:
            for sub_id, payload in tr["selections"].items():
                if not isinstance(payload, dict):
                    continue
                label = payload.get("label", sub_id) or sub_id
                if "unchecked" in str(label).lower():
                    continue
                rows.append(
                    {
                        "label": label,
                        "count": payload.get("count", 0),
                    }
                )

    title = (
        getattr(result, "synthetic_question_title", None)
        or getattr(result, "title", None)
        or result.cross_cut_id
    )
    filter_expr = getattr(result, "filter_expr", None) or rt.get("filter_expr")
    overall = rt.get("overall") or {}
    valid_n = (
        rt.get("grand_total")
        or rt.get("filter_n")
        or rt.get("paired_n")
        or (overall.get("n") if isinstance(overall, dict) else 0)
        or 0
    )
    return {
        "table_kind": result.analysis_type.value.lower(),
        "question_id": result.cross_cut_id,
        "question_text": title,
        "valid_n": int(valid_n),
        "missing_n": 0,
        "filters_applied": [filter_expr] if filter_expr else [],
        "rows": rows,
        "summary": {},
    }


def _render_insight_section(
    insight_key: str, payload_factory: Any, table_kind: str, title_hint: str
) -> None:
    """Render the AI insight section: button + cached result display."""
    app = _require_streamlit()
    app.divider()
    cols = app.columns([3, 1.4])
    with cols[0]:
        app.markdown("**AI insight**")
    with cols[1]:
        app.markdown(
            "<style>.ai-insight-btn + div .stButton button p,"
            ".ai-insight-btn + div .stButton button{white-space:nowrap !important;}</style>"
            "<div class='ai-insight-btn'></div>",
            unsafe_allow_html=True,
        )
        btn_label = (
            "Generate insight"
            if insight_key not in app.session_state
            else "Regenerate"
        )
        if app.button(
            btn_label, key=f"btn_{insight_key}", use_container_width=True
        ):
            with app.spinner("Generating insight..."):
                from src.ai_insights import generate_insight
                payload = payload_factory()
                ir = generate_insight(
                    payload, table_kind=table_kind, title_hint=title_hint
                )
                app.session_state[insight_key] = ir
            app.rerun()

    if insight_key in app.session_state:
        ir = app.session_state[insight_key]
        import html as _html_mod
        title_html = _html_mod.escape(str(ir.title))
        insight_html = _html_mod.escape(str(ir.insight))
        app.markdown(
            f"<div style='font-size: 14px; font-weight: 500; "
            f"color: #1a1a1a; margin-bottom: 4px;'>{title_html}</div>",
            unsafe_allow_html=True,
        )
        if ir.was_template:
            app.caption(
                f"{ir.insight}  *(template \u2014 AI unavailable)*"
            )
            if ir.error_message:
                with app.expander("Why was AI unavailable?"):
                    app.code(ir.error_message)
        else:
            app.markdown(
                f"<div style='font-size: 13px; line-height: 1.7; "
                f"color: #333; background: #f8f9fa; "
                f"border-left: 3px solid #185FA5; padding: 10px 14px; "
                f"border-radius: 0;'>{insight_html}</div>",
                unsafe_allow_html=True,
            )
            app.caption(
                f"Generated by {ir.model_used} \u00b7 "
                f"{ir.tokens_used} tokens"
            )


def _render_insight_card(
    insight: InsightResult,
    label: str = "Key insight",
    icon: str = "ti-bulb",
) -> None:
    app = _require_streamlit()
    if not insight or not insight.insight:
        return

    headline_raw = insight.insight
    number_pattern = _re.compile(r"(\d+(?:\.\d+)?\s*(?:%|x|X|×)?)")
    escaped = html.escape(headline_raw)
    headline_highlighted = number_pattern.sub(
        r'<span class="num">\1</span>', escaped
    )
    label_html = html.escape(label)

    if insight.was_template:
        app.markdown(
            f"""
        <div class="insight-card insight-template">
            <div class="insight-icon"><i class="ti ti-info-circle"></i></div>
            <div class="insight-body">
                <div class="insight-label">{label_html}</div>
                <div class="insight-text">{html.escape(headline_raw)}</div>
                <div class="insight-footer">Template fallback · AI unavailable</div>
            </div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        return

    model = html.escape(insight.model_used or "")
    footer = f"AI insight · {model}" if model else "AI insight"
    app.markdown(
        f"""
    <div class="insight-card">
        <div class="insight-icon"><i class="ti {icon}"></i></div>
        <div class="insight-body">
            <div class="insight-label">{label_html}</div>
            <div class="insight-text">{headline_highlighted}</div>
            <div class="insight-footer">{footer}</div>
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )


def _copy_button(df: Any, key: str) -> None:
    """Render a Copy Table button that copies the DataFrame as TSV.

    Tab-separated values paste cleanly into Excel preserving columns
    and rows. Uses navigator.clipboard for one-click copy.
    """
    app = _require_streamlit()
    try:
        if df is None or len(df) == 0:
            return
        lines = []
        headers = list(df.columns)
        if df.index.name:
            headers = [df.index.name] + headers
        lines.append("\t".join(str(h) for h in headers))
        for idx, row in df.iterrows():
            if df.index.name:
                row_vals = [str(idx)] + [str(v) for v in row.values]
            else:
                row_vals = [str(v) for v in row.values]
            lines.append("\t".join(row_vals))
        tsv = "\n".join(lines)
        tsv_escaped = (
            tsv.replace("\\", "\\\\")
            .replace("`", "\\`")
            .replace("$", "\\$")
            .replace("</", "<\\/")
        )
        safe_key = "".join(
            c if (c.isalnum() or c in "_-") else "_" for c in str(key)
        )
        button_id = f"copy_btn_{safe_key}"
        app.markdown(
            f"""
<div style="display:flex;justify-content:flex-end;margin-bottom:4px;">
  <button id="{button_id}"
    onclick="navigator.clipboard.writeText(`{tsv_escaped}`).then(()=>{{
      var b=document.getElementById('{button_id}');
      b.textContent='\u2713 Copied!';b.style.background='#2E7D32';
      setTimeout(()=>{{b.textContent='\u2398 Copy Table';b.style.background='#0A0A0A';}},2000);
    }}).catch(()=>{{document.getElementById('{button_id}').textContent='Copy failed';}});"
    style="background:#0A0A0A;color:white;border:none;padding:6px 14px;
      font-family:Arial,Helvetica,sans-serif;font-size:11px;font-weight:600;
      letter-spacing:0.05em;cursor:pointer;display:flex;align-items:center;
      gap:6px;transition:background 0.15s;">\u2398 Copy Table</button>
</div>
""",
            unsafe_allow_html=True,
        )
    except Exception:
        return


def _render_sc_table_html(
    distribution: dict,
    display_mode: str,
    valid_n: int,
    flags: list | None = None,
    key_suffix: str = "",
) -> None:
    """Branded HTML table for SingleSelect / MultiSelect distributions.

    Labels are HTML-escaped to prevent any XSS via maliciously crafted
    data-map files (defense in depth — labels are analyst-supplied).

    ``flags`` may be precomputed by the caller so that highlighting is
    identical across display-mode toggles. If omitted, flags are derived
    from the distribution counts.
    """
    import html as _html

    app = _require_streamlit()
    if flags is None:
        counts = [p.get("count", 0) for _, p in distribution.items()]
        flags = _compute_outlier_flags(counts)
    rows_html = ""
    for i, (code, payload) in enumerate(distribution.items()):
        label = _html.escape(str(payload.get("label", code)))
        count = payload.get("count", 0)
        rate = payload.get("rate")
        if rate is None:
            rate = (count / valid_n) if valid_n else 0
        pct = f"{rate * 100:.1f}%"
        flag = flags[i] if i < len(flags) else ""
        bar_width = max(0, min(160, int(rate * 160)))
        bar_html = (
            f'<div style="display:flex;align-items:center;gap:8px;">'
            f'<div style="width:160px;height:4px;background:#F0F0F0;'
            f'position:relative;flex-shrink:0;">'
            f'<div style="position:absolute;top:0;left:0;bottom:0;'
            f'width:{bar_width}px;background:#CC0000;"></div>'
            f"</div></div>"
        )
        if flag == "high":
            row_style = "background:#FFF0F0;border-left:3px solid #CC0000;"
            count_style = "font-weight:700;color:#CC0000;"
            flag_cell = "\u2B06"
        elif flag == "low":
            row_style = "background:#FFFBE6;"
            count_style = "font-weight:600;color:#E65100;"
            flag_cell = "\u2193"
        else:
            row_style = ""
            count_style = "color:#333;"
            flag_cell = ""
        td = "padding:7px 10px;border-bottom:1px solid #F5F5F5;"
        if display_mode == "Counts":
            row = (
                f'<tr style="{row_style}">'
                f'<td style="{td}font-size:12px;">{label}</td>'
                f'<td style="{td}{count_style}font-size:12px;text-align:right;">'
                f"{count:,}</td>"
                f'<td style="{td}">{bar_html}</td>'
                f'<td style="{td}font-size:11px;color:#CC0000;">{flag_cell}</td>'
                f"</tr>"
            )
        elif display_mode == "Counts + %":
            row = (
                f'<tr style="{row_style}">'
                f'<td style="{td}font-size:12px;">{label}</td>'
                f'<td style="{td}{count_style}font-size:12px;text-align:right;'
                f'width:60px;">{count:,}</td>'
                f'<td style="{td}font-size:12px;font-weight:700;color:#CC0000;'
                f'text-align:right;width:54px;">{pct}</td>'
                f'<td style="{td}">{bar_html}</td>'
                f'<td style="{td}font-size:11px;color:#CC0000;">{flag_cell}</td>'
                f"</tr>"
            )
        else:  # "% only"
            row = (
                f'<tr style="{row_style}">'
                f'<td style="{td}font-size:12px;">{label}</td>'
                f'<td style="{td}font-size:13px;font-weight:700;color:#CC0000;'
                f'text-align:right;">{pct}</td>'
                f'<td style="{td}">{bar_html}</td>'
                f'<td style="{td}font-size:11px;color:#CC0000;">{flag_cell}</td>'
                f"</tr>"
            )
        rows_html += row

    th = (
        "padding:8px 10px;font-size:10px;text-transform:uppercase;"
        "letter-spacing:0.1em;color:#888;"
    )
    if display_mode == "Counts":
        headers = (
            f"<th style='{th}text-align:left;'>Label</th>"
            f"<th style='{th}text-align:right;'>Count</th>"
            f"<th style='{th}'>Bar</th><th></th>"
        )
    elif display_mode == "Counts + %":
        headers = (
            f"<th style='{th}text-align:left;'>Label</th>"
            f"<th style='{th}text-align:right;'>Count</th>"
            f"<th style='{th}text-align:right;'>%</th>"
            f"<th style='{th}'>Bar</th><th></th>"
        )
    else:
        headers = (
            f"<th style='{th}text-align:left;'>Label</th>"
            f"<th style='{th}text-align:right;'>%</th>"
            f"<th style='{th}'>Bar</th><th></th>"
        )

    app.markdown(
        f'<table style="width:100%;border-collapse:collapse;'
        f'font-family:Arial,Helvetica,sans-serif;">'
        f'<thead><tr style="background:#F8F8F8;'
        f'border-bottom:2px solid #E0E0E0;">{headers}</tr></thead>'
        f"<tbody>{rows_html}</tbody></table>"
        f'<div style="font-size:10px;color:#888;margin-top:6px;'
        f'font-family:Arial,Helvetica,sans-serif;">'
        f"Valid N: {valid_n:,} &nbsp;\u00b7&nbsp; "
        f"\u2B06 high outlier &nbsp;\u00b7&nbsp; \u2193 low outlier"
        f"</div>",
        unsafe_allow_html=True,
    )

    # Secondary, fully-interactive table: sortable / copyable / CSV-downloadable.
    # The HTML table above remains the primary visual (bars cannot live inside
    # st.dataframe). We use a checkbox toggle (NOT st.expander) because this
    # helper is itself rendered inside an st.expander in the single-cut card,
    # and Streamlit forbids nested expanders.
    show_table = app.checkbox(
        "Show table view (sortable \u00b7 copyable \u00b7 downloadable)",
        key=f"sc_tableview_{key_suffix}" if key_suffix else None,
        value=False,
    )
    if show_table:
        rows_for_df = []
        for i, (code, payload) in enumerate(distribution.items()):
            label = payload.get("label", str(code))
            count = payload.get("count", 0)
            rate = payload.get("rate")
            if rate is None:
                rate = (count / valid_n) if valid_n else 0
            pct = round(rate * 100, 1)
            flag = flags[i] if i < len(flags) else ""
            flag_label = (
                "High outlier" if flag == "high"
                else "Low outlier" if flag == "low"
                else ""
            )
            if display_mode == "Counts":
                rows_for_df.append(
                    {"Label": label, "Count": count, "Flag": flag_label}
                )
            elif display_mode == "Counts + %":
                rows_for_df.append(
                    {"Label": label, "Count": count, "%": pct, "Flag": flag_label}
                )
            else:
                rows_for_df.append(
                    {"Label": label, "%": pct, "Flag": flag_label}
                )
        df_plain = pd.DataFrame(rows_for_df)
        _copy_button(df_plain, f"sc_{key_suffix or id(distribution)}_{display_mode}")
        try:
            app.dataframe(
                df_plain.style.apply(_style_outliers, axis=None),
                use_container_width=True,
                hide_index=True,
            )
        except Exception:
            app.dataframe(
                df_plain, use_container_width=True, hide_index=True
            )


# ---------------------------------------------------------------------------
# Day 18 — Plotly chart helpers
# ---------------------------------------------------------------------------


def _chart_text_label(text: str, limit: int) -> str:
    text = (text or "").strip()
    return text[: limit - 3] + "\u2026" if len(text) > limit else text


def _chart_type_options() -> list[Any]:
    from src.chart_recommender import ChartType

    return [
        ChartType.COLUMN_STACKED,
        ChartType.COLUMN_CLUSTERED,
        ChartType.BAR_CLUSTERED,
        ChartType.BAR_STACKED,
        ChartType.LINE,
        ChartType.HEATMAP_TABLE,
        ChartType.COMBO,
    ]


def _format_chart_type_label(chart_type: Any) -> str:
    value = getattr(chart_type, "value", str(chart_type))
    labels = {
        "COLUMN_STACKED": "Stacked column",
        "COLUMN_CLUSTERED": "Clustered column",
        "BAR_CLUSTERED": "Clustered bar",
        "BAR_STACKED": "Stacked bar",
        "LINE": "Line chart",
        "HEATMAP_TABLE": "Heatmap table",
        "COMBO": "Combo chart",
    }
    return labels.get(value, value.replace("_", " ").title())


def _coerce_chart_type(value: Any, default_type: Any) -> Any:
    from src.chart_recommender import ChartType

    if isinstance(value, ChartType):
        return value
    try:
        return ChartType(str(value))
    except Exception:  # noqa: BLE001
        return default_type


def _chart_type_override_key(canonical_id: str) -> str:
    return f"chart_type_override_{canonical_id}"


def _clear_chart_type_overrides() -> None:
    app = _require_streamlit()
    for key in list(app.session_state.keys()):
        if str(key).startswith("chart_type_override_"):
            del app.session_state[key]


def _render_chart_type_override_control(
    canonical_id: str,
    default_type: Any,
    location: str = "",
) -> Any:
    app = _require_streamlit()
    options = _chart_type_options()
    default_type = _coerce_chart_type(default_type, options[0])
    override_key = _chart_type_override_key(canonical_id)
    selected_type = _coerce_chart_type(
        app.session_state.get(override_key, default_type),
        default_type,
    )
    widget_key = (
        override_key
        if not location
        else f"{override_key}_{location}"
    )
    if widget_key not in app.session_state:
        app.session_state[widget_key] = selected_type
    selected_type = app.selectbox(
        "Chart type",
        options=options,
        format_func=_format_chart_type_label,
        key=widget_key,
    )
    selected_type = _coerce_chart_type(selected_type, default_type)
    app.session_state[override_key] = selected_type
    return selected_type


def _chart_recommendation_for_screen(recommendation: Any, selected_chart_type: Any) -> Any:
    from dataclasses import replace

    selected_chart_type = _coerce_chart_type(selected_chart_type, recommendation.chart_type)
    if selected_chart_type == recommendation.chart_type:
        return recommendation
    return replace(recommendation, chart_type=selected_chart_type)


def _chart_override_caption(selected_type: Any, default_type: Any) -> None:
    from src.chart_recommender import ChartType

    if selected_type in (ChartType.HEATMAP_TABLE, ChartType.LINE) and selected_type != default_type:
        app = _require_streamlit()
        app.caption(
            "This chart type isn't ideal for this question. Consider using "
            f"{_format_chart_type_label(default_type)}."
        )


def _render_chart_for_distribution(
    distribution: dict,
    spec: Any,
    valid_n: int,
    display_mode: str,
    key_suffix: str = "",
) -> None:
    """Interactive Plotly bar chart for an SS or MS distribution."""
    try:
        import plotly.graph_objects as go
    except Exception:  # noqa: BLE001
        return

    app = _require_streamlit()
    labels: list[str] = []
    counts: list[float] = []
    pcts: list[float] = []
    for code, payload in distribution.items():
        labels.append(str(payload.get("label", code)))
        c = payload.get("count", 0) or 0
        counts.append(c)
        rate = payload.get("rate")
        if rate is None:
            rate = (c / valid_n) if valid_n else 0
        pcts.append(rate * 100)

    if not labels:
        return

    from src.chart_recommender import ChartType

    selected_chart_type = _render_chart_type_override_control(
        str(getattr(spec, "canonical_id", key_suffix or "distribution")),
        ChartType.BAR_CLUSTERED,
        f"distribution_{key_suffix or 'main'}",
    )
    _chart_override_caption(selected_chart_type, ChartType.BAR_CLUSTERED)

    flags = _compute_outlier_flags(counts)
    bar_colors = [
        "#CC0000" if f == "high" else "#E65100" if f == "low" else "#888888"
        for f in flags
    ]
    if display_mode == "% only":
        y_values, y_label = pcts, "%"
    else:
        y_values, y_label = counts, "Count"

    text_vals: list[str] = []
    for c, p in zip(counts, pcts):
        if display_mode == "Counts":
            text_vals.append(f"{int(c):,}")
        elif display_mode == "% only":
            text_vals.append(f"{p:.1f}%")
        else:
            text_vals.append(f"{int(c):,} ({p:.1f}%)")

    if selected_chart_type == ChartType.LINE:
        fig = go.Figure(
            data=[
                go.Scatter(
                    x=labels,
                    y=y_values,
                    mode="lines+markers+text",
                    text=text_vals,
                    textposition="top center",
                    line={"color": "#CC0000", "width": 2},
                    marker={"size": 7},
                    customdata=list(zip(counts, pcts)),
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "Count: %{customdata[0]:,}<br>"
                        "%: %{customdata[1]:.1f}%<extra></extra>"
                    ),
                )
            ]
        )
    elif selected_chart_type == ChartType.HEATMAP_TABLE:
        fig = go.Figure(
            data=[
                go.Heatmap(
                    z=[y_values],
                    x=labels,
                    y=[y_label],
                    text=[text_vals],
                    texttemplate="%{text}",
                    colorscale=[[0, "#FFFFFF"], [1, "#CC0000"]],
                    showscale=False,
                )
            ]
        )
    elif selected_chart_type in (ChartType.BAR_CLUSTERED, ChartType.BAR_STACKED):
        fig = go.Figure(
            data=[
                go.Bar(
                    y=labels,
                    x=y_values,
                    orientation="h",
                    marker_color=bar_colors,
                    text=text_vals,
                    textposition="outside",
                    customdata=list(zip(counts, pcts)),
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "Count: %{customdata[0]:,}<br>"
                        "%: %{customdata[1]:.1f}%<extra></extra>"
                    ),
                )
            ]
        )
    else:
        fig = go.Figure(
            data=[
                go.Bar(
                    x=labels,
                    y=y_values,
                    marker_color=bar_colors,
                    text=text_vals,
                    textposition="outside",
                    customdata=list(zip(counts, pcts)),
                    hovertemplate=(
                        "<b>%{x}</b><br>"
                        "Count: %{customdata[0]:,}<br>"
                        "%: %{customdata[1]:.1f}%<extra></extra>"
                    ),
                )
            ]
        )
    fig.update_layout(
        title={
            "text": _chart_text_label(
                spec.question_text or spec.canonical_id, 80
            ),
            "font": {"family": "Arial", "size": 14, "color": "#0A0A0A"},
        },
        xaxis={"title": "", "tickfont": {"family": "Arial", "size": 11}},
        yaxis={
            "title": y_label,
            "tickfont": {"family": "Arial", "size": 10},
            "gridcolor": "#F0F0F0",
        },
        plot_bgcolor="white",
        paper_bgcolor="white",
        showlegend=False,
        barmode="stack" if selected_chart_type in (ChartType.COLUMN_STACKED, ChartType.BAR_STACKED) else "group",
        height=380,
        margin={"t": 50, "b": 80, "l": 60, "r": 40},
        font={"family": "Arial"},
    )
    try:
        app.plotly_chart(
            fig,
            use_container_width=True,
            key=f"chart_sc_{spec.canonical_id}_{key_suffix}",
            config={
                "displayModeBar": True,
                "displaylogo": False,
                "modeBarButtonsToRemove": ["lasso2d", "select2d"],
                "toImageButtonOptions": {
                    "format": "png",
                    "filename": f"{spec.canonical_id}_chart",
                    "height": 600,
                    "width": 1000,
                    "scale": 2,
                },
            },
        )
    except Exception as exc:  # noqa: BLE001
        app.caption(f"Chart unavailable: {type(exc).__name__}: {exc}")


def _thinkcell_chart_payload(result: Any, spec: Any):
    """Compute (recommendation, payload) for the recommended chart, or None on error."""
    app = _require_streamlit()
    try:
        from src.chart_recommender import recommend_chart
        from src.thinkcell_table_formatter import format_for_thinkcell

        recommendation = recommend_chart(result, spec)
        schema = app.session_state.get("schema")
        survey_name = (
            getattr(schema, "source_datamap_path", None)
            or app.session_state.get("datamap_path_label")
            or "Survey"
        )
        payload = format_for_thinkcell(
            result,
            recommendation,
            question_text=getattr(spec, "question_text", "") or "",
            survey_name=str(survey_name),
        )
        return recommendation, payload
    except Exception as exc:  # noqa: BLE001
        app.caption(f"Recommended chart unavailable: {type(exc).__name__}: {exc}")
        return None


def _render_recommended_chart(result: Any, spec: Any) -> None:
    """Render the rule-recommended Plotly chart with a chart-type override."""
    app = _require_streamlit()
    computed = _thinkcell_chart_payload(result, spec)
    if computed is None:
        return
    recommendation, payload = computed
    try:
        from src.chart_renderer import render_chart
    except Exception as exc:  # noqa: BLE001
        app.caption(f"Recommended chart unavailable: {type(exc).__name__}: {exc}")
        return

    canonical_id = str(getattr(spec, "canonical_id", result.question_id))
    selected_chart_type = _render_chart_type_override_control(
        canonical_id,
        recommendation.chart_type,
        "recommended",
    )
    _chart_override_caption(selected_chart_type, recommendation.chart_type)
    screen_recommendation = _chart_recommendation_for_screen(
        recommendation,
        selected_chart_type,
    )

    try:
        fig = render_chart(screen_recommendation, payload)
        app.plotly_chart(
            fig,
            use_container_width=True,
            key=f"recommended_chart_{result.question_id}",
            config={
                "displayModeBar": True,
                "displaylogo": False,
                "modeBarButtonsToRemove": ["lasso2d", "select2d"],
                "toImageButtonOptions": {
                    "format": "png",
                    "filename": f"{result.question_id}_recommended_chart",
                    "height": 700,
                    "width": 1100,
                    "scale": 2,
                },
            },
        )
    except Exception as exc:  # noqa: BLE001
        if selected_chart_type != recommendation.chart_type:
            app.caption(
                "That chart type could not render for this question; showing "
                f"{_format_chart_type_label(recommendation.chart_type)} instead. "
                f"Reason: {type(exc).__name__}: {exc}"
            )
            try:
                fig = render_chart(recommendation, payload)
                app.plotly_chart(
                    fig,
                    use_container_width=True,
                    key=f"recommended_chart_{result.question_id}_fallback",
                    config={
                        "displayModeBar": True,
                        "displaylogo": False,
                        "modeBarButtonsToRemove": ["lasso2d", "select2d"],
                        "toImageButtonOptions": {
                            "format": "png",
                            "filename": f"{result.question_id}_recommended_chart",
                            "height": 700,
                            "width": 1100,
                            "scale": 2,
                        },
                    },
                )
            except Exception as fallback_exc:  # noqa: BLE001
                app.caption(
                    "Inline recommended chart unavailable: "
                    f"{type(fallback_exc).__name__}: {fallback_exc}"
                )
        else:
            app.caption(f"Inline recommended chart unavailable: {type(exc).__name__}: {exc}")


def _render_thinkcell_layout(result: Any, spec: Any) -> None:
    """Render the copy-paste think-cell datasheet table (no template, no download)."""
    app = _require_streamlit()
    computed = _thinkcell_chart_payload(result, spec)
    if computed is None:
        return
    recommendation, payload = computed
    app.write(f"**Chart type:** {recommendation.chart_type.value}")
    app.write(
        "**Layout for think-cell:** copy the table below and paste directly "
        "into the think-cell datasheet for the recommended chart type."
    )
    app.code(payload.to_tsv(), language="tsv")
    app.caption(
        "Click in the box above, Ctrl+A, Ctrl+C to copy. Then in PowerPoint: "
        "insert a think-cell "
        + recommendation.chart_type.value.lower().replace("_", " ")
        + " chart and paste into its datasheet."
    )


def _render_recommended_chart_and_thinkcell(result: Any, spec: Any) -> None:
    """Chart up front; think-cell datasheet behind a toggle.

    Uses a toggle (not an expander) because this renders INSIDE the single-cut
    card's expander, and Streamlit forbids nesting expanders.
    """
    app = _require_streamlit()
    _render_recommended_chart(result, spec)
    tc_key = f"show_tc_{getattr(spec, 'canonical_id', id(spec))}"
    if app.toggle("Show Think-cell layout", key=tc_key, value=False):
        with app.container(border=True):
            _render_thinkcell_layout(result, spec)


def _render_chart_for_cross_tab(
    result: Any, schema: Any, display_mode: str = "Counts"
) -> None:
    """Clustered bar chart for a cross-tab result."""
    try:
        import plotly.graph_objects as go
    except Exception:  # noqa: BLE001
        return

    app = _require_streamlit()
    rt = result.result_table
    counts = rt.get("counts", {}) or {}
    row_labels = rt.get("row_label_map", {}) or {}
    col_labels = rt.get("column_label_map", {}) or {}
    row_pct = rt.get("row_pct", {}) or {}
    if not counts:
        return

    row_codes = sorted(counts.keys(), key=lambda v: str(v))
    col_codes = sorted(
        {c for row in counts.values() if isinstance(row, dict) for c in row.keys()},
        key=lambda v: str(v),
    )
    x_categories = [str(row_labels.get(rc, rc)) for rc in row_codes]
    palette = [
        "#CC0000", "#0A0A0A", "#666666", "#990000",
        "#999999", "#330000", "#444444", "#FF6666",
    ]

    a, b = result.source_question_ids
    from src.chart_recommender import ChartType

    selected_chart_type = _render_chart_type_override_control(
        str(a),
        ChartType.BAR_CLUSTERED,
        f"cross_tab_{result.cross_cut_id}_{display_mode}",
    )
    _chart_override_caption(selected_chart_type, ChartType.BAR_CLUSTERED)
    a_spec = schema.get_question(a) if schema is not None else None
    b_spec = schema.get_question(b) if schema is not None else None
    a_text = a_spec.question_text if (a_spec and a_spec.question_text) else a
    b_text = b_spec.question_text if (b_spec and b_spec.question_text) else b
    title = (
        f"{_chart_text_label(a_text, 50)} \u00d7 {_chart_text_label(b_text, 50)}"
    )

    fig = go.Figure()
    if selected_chart_type == ChartType.HEATMAP_TABLE:
        z_values = []
        text_values = []
        for rc in row_codes:
            if display_mode == "Row %":
                row_vals = [(row_pct.get(rc, {}) or {}).get(cc, 0) * 100 for cc in col_codes]
                row_text = [f"{v:.1f}%" for v in row_vals]
            else:
                row_vals = [(counts.get(rc, {}) or {}).get(cc, 0) for cc in col_codes]
                row_text = [f"{v:,}" for v in row_vals]
            z_values.append(row_vals)
            text_values.append(row_text)
        fig.add_trace(
            go.Heatmap(
                z=z_values,
                x=[str(col_labels.get(cc, cc)) for cc in col_codes],
                y=x_categories,
                text=text_values,
                texttemplate="%{text}",
                colorscale=[[0, "#FFFFFF"], [1, "#CC0000"]],
                showscale=False,
            )
        )
    else:
        for i, cc in enumerate(col_codes):
            col_label = str(col_labels.get(cc, cc))
            if display_mode == "Row %":
                y_vals = [
                    (row_pct.get(rc, {}) or {}).get(cc, 0) * 100 for rc in row_codes
                ]
                text_vals = [f"{v:.1f}%" for v in y_vals]
            else:
                y_vals = [(counts.get(rc, {}) or {}).get(cc, 0) for rc in row_codes]
                text_vals = [f"{v:,}" for v in y_vals]
            if selected_chart_type == ChartType.LINE:
                fig.add_trace(
                    go.Scatter(
                        name=col_label,
                        x=x_categories,
                        y=y_vals,
                        mode="lines+markers+text",
                        text=text_vals,
                        textposition="top center",
                        line={"color": palette[i % len(palette)], "width": 2},
                        hovertemplate=f"<b>{col_label}</b><br>%{{x}}: %{{y}}<extra></extra>",
                    )
                )
            elif selected_chart_type in (ChartType.BAR_CLUSTERED, ChartType.BAR_STACKED):
                fig.add_trace(
                    go.Bar(
                        name=col_label,
                        y=x_categories,
                        x=y_vals,
                        orientation="h",
                        text=text_vals,
                        textposition="outside",
                        marker_color=palette[i % len(palette)],
                        hovertemplate=f"<b>{col_label}</b><br>%{{y}}: %{{x}}<extra></extra>",
                    )
                )
            else:
                fig.add_trace(
                    go.Bar(
                        name=col_label,
                        x=x_categories,
                        y=y_vals,
                        text=text_vals,
                        textposition="outside",
                        marker_color=palette[i % len(palette)],
                        hovertemplate=f"<b>{col_label}</b><br>%{{x}}: %{{y}}<extra></extra>",
                    )
                )

    fig.update_layout(
        title={"text": title, "font": {"family": "Arial", "size": 13}},
        barmode="stack" if selected_chart_type in (ChartType.COLUMN_STACKED, ChartType.BAR_STACKED) else "group",
        xaxis={"title": a, "tickfont": {"family": "Arial", "size": 10}},
        yaxis={
            "title": "Count" if display_mode == "Counts" else "%",
            "gridcolor": "#F0F0F0",
            "tickfont": {"family": "Arial", "size": 10},
        },
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend={"title": b, "font": {"family": "Arial", "size": 10}},
        height=440,
        margin={"t": 60, "b": 80, "l": 60, "r": 40},
        font={"family": "Arial"},
    )
    try:
        app.plotly_chart(
            fig,
            use_container_width=True,
            key=f"chart_ct_{result.cross_cut_id}_{display_mode}",
            config={
                "displayModeBar": True,
                "displaylogo": False,
                "toImageButtonOptions": {
                    "format": "png",
                    "filename": f"{result.cross_cut_id}_chart",
                    "height": 600,
                    "width": 1200,
                    "scale": 2,
                },
            },
        )
    except Exception as exc:  # noqa: BLE001
        app.caption(f"Chart unavailable: {type(exc).__name__}: {exc}")


def _render_chart_for_expected_vs_realized(result: Any) -> None:
    """Side-by-side bars comparing Expected, Realized, and Gap means."""
    try:
        import plotly.graph_objects as go
    except Exception:  # noqa: BLE001
        return

    app = _require_streamlit()
    rt = result.result_table
    expected = rt.get("expected", {}) or {}
    realized = rt.get("realized", {}) or {}
    gap = rt.get("gap", {}) or {}
    series = [
        ("Expected", expected.get("mean")),
        ("Realized", realized.get("mean")),
        ("Gap", gap.get("mean")),
    ]
    labels = [name for name, val in series if val is not None]
    values = [float(val) for name, val in series if val is not None]
    if not values:
        return

    from src.chart_recommender import ChartType

    selected_chart_type = _render_chart_type_override_control(
        str(getattr(result, "source_question_ids", ("expected_vs_realized",))[0]),
        ChartType.COLUMN_CLUSTERED,
        f"expected_vs_realized_{result.cross_cut_id}",
    )
    _chart_override_caption(selected_chart_type, ChartType.COLUMN_CLUSTERED)

    palette = ["#666666", "#0A0A0A", "#CC0000"]
    text_values = [f"{v:.2f}" for v in values]
    if selected_chart_type == ChartType.LINE:
        fig = go.Figure(
            data=[
                go.Scatter(
                    x=labels,
                    y=values,
                    mode="lines+markers+text",
                    text=text_values,
                    textposition="top center",
                    line={"color": "#CC0000", "width": 2},
                )
            ]
        )
    elif selected_chart_type == ChartType.HEATMAP_TABLE:
        fig = go.Figure(
            data=[
                go.Heatmap(
                    z=[values],
                    x=labels,
                    y=["Mean"],
                    text=[text_values],
                    texttemplate="%{text}",
                    colorscale=[[0, "#FFFFFF"], [1, "#CC0000"]],
                    showscale=False,
                )
            ]
        )
    elif selected_chart_type in (ChartType.BAR_CLUSTERED, ChartType.BAR_STACKED):
        fig = go.Figure(
            data=[
                go.Bar(
                    y=labels,
                    x=values,
                    orientation="h",
                    marker_color=palette[: len(values)],
                    text=text_values,
                    textposition="outside",
                )
            ]
        )
    else:
        fig = go.Figure(
            data=[
                go.Bar(
                    x=labels,
                    y=values,
                    marker_color=palette[: len(values)],
                    text=text_values,
                    textposition="outside",
                )
            ]
        )
    fig.update_layout(
        title="Mean: Expected vs Realized (Gap = Realized \u2212 Expected)",
        barmode="stack" if selected_chart_type in (ChartType.COLUMN_STACKED, ChartType.BAR_STACKED) else "group",
        xaxis={"tickfont": {"family": "Arial", "size": 11}},
        yaxis={
            "title": "Mean",
            "gridcolor": "#F0F0F0",
            "tickfont": {"family": "Arial", "size": 10},
        },
        plot_bgcolor="white",
        paper_bgcolor="white",
        showlegend=False,
        height=360,
        margin={"t": 50, "b": 60, "l": 60, "r": 40},
        font={"family": "Arial"},
    )
    try:
        app.plotly_chart(
            fig,
            use_container_width=True,
            key=f"chart_evr_{result.cross_cut_id}",
            config={"displaylogo": False},
        )
    except Exception as exc:  # noqa: BLE001
        app.caption(f"Chart unavailable: {type(exc).__name__}: {exc}")


def _render_chart_for_segment_metric(result: Any, schema: Any) -> None:
    """Simple bar chart of mean-by-segment for GROUP_COMPARISON-style results."""
    try:
        import plotly.graph_objects as go
    except Exception:  # noqa: BLE001
        return

    app = _require_streamlit()
    rt = result.result_table
    per_seg = rt.get("per_segment", {}) or {}
    labels: list[str] = []
    values: list[float] = []
    for seg_val, seg_data in per_seg.items():
        if not isinstance(seg_data, dict):
            continue
        mean = seg_data.get("mean")
        if mean is None:
            continue
        labels.append(str(seg_data.get("label", seg_val)))
        values.append(float(mean))
    if not values:
        return

    from src.chart_recommender import ChartType

    rt = result.result_table
    canonical_id = str(rt.get("metric_question_id") or getattr(result, "cross_cut_id", "segment_metric"))
    selected_chart_type = _render_chart_type_override_control(
        canonical_id,
        ChartType.BAR_CLUSTERED,
        f"segment_metric_{result.cross_cut_id}",
    )
    _chart_override_caption(selected_chart_type, ChartType.BAR_CLUSTERED)

    flags = _compute_outlier_flags(values)
    colors = [
        "#CC0000" if f == "high" else "#E65100" if f == "low" else "#666666"
        for f in flags
    ]
    text_values = [f"{v:.2f}" for v in values]
    if selected_chart_type == ChartType.LINE:
        fig = go.Figure(
            data=[
                go.Scatter(
                    x=labels,
                    y=values,
                    mode="lines+markers+text",
                    text=text_values,
                    textposition="top center",
                    line={"color": "#CC0000", "width": 2},
                )
            ]
        )
    elif selected_chart_type == ChartType.HEATMAP_TABLE:
        fig = go.Figure(
            data=[
                go.Heatmap(
                    z=[values],
                    x=labels,
                    y=["Mean"],
                    text=[text_values],
                    texttemplate="%{text}",
                    colorscale=[[0, "#FFFFFF"], [1, "#CC0000"]],
                    showscale=False,
                )
            ]
        )
    elif selected_chart_type in (ChartType.BAR_CLUSTERED, ChartType.BAR_STACKED):
        fig = go.Figure(
            data=[
                go.Bar(
                    y=labels,
                    x=values,
                    orientation="h",
                    marker_color=colors,
                    text=text_values,
                    textposition="outside",
                )
            ]
        )
    else:
        fig = go.Figure(
            data=[
                go.Bar(
                    x=labels,
                    y=values,
                    marker_color=colors,
                    text=text_values,
                    textposition="outside",
                )
            ]
        )
    fig.update_layout(
        title="Mean by segment",
        barmode="stack" if selected_chart_type in (ChartType.COLUMN_STACKED, ChartType.BAR_STACKED) else "group",
        xaxis={"tickfont": {"family": "Arial", "size": 10}},
        yaxis={
            "gridcolor": "#F0F0F0",
            "tickfont": {"family": "Arial", "size": 10},
        },
        plot_bgcolor="white",
        paper_bgcolor="white",
        showlegend=False,
        height=380,
        margin={"t": 50, "b": 80, "l": 60, "r": 40},
        font={"family": "Arial"},
    )
    try:
        app.plotly_chart(
            fig,
            use_container_width=True,
            key=f"chart_seg_{result.cross_cut_id}",
            config={"displaylogo": False},
        )
    except Exception as exc:  # noqa: BLE001
        app.caption(f"Chart unavailable: {type(exc).__name__}: {exc}")


def _drain_pending_actions() -> None:
    """Process queued filter actions before any section renders.

    Implements the "set flag in click handler, compute at top of next render"
    pattern that eliminates the double-click bug: both the global filter and
    every per-question filter become single-click because the new state is
    fully populated before sections 2-5 render.
    """
    app = _require_streamlit()

    pending_gf = app.session_state.get("pending_global_filter")
    if pending_gf is not None:
        app.session_state["pending_global_filter"] = None
        app.session_state["global_filter_error"] = None
        try:
            from src.global_filter import apply_global_filter
            from src.models import GlobalFilterState

            state = GlobalFilterState(
                filters=tuple(pending_gf["filter_specs"])
            )
            filtered_df, stats = apply_global_filter(
                app.session_state["decoded_df"],
                state,
                app.session_state.get("schema"),
            )
            app.session_state["global_filter_state"] = state
            app.session_state["global_filter_stats"] = stats
            app.session_state["active_df"] = filtered_df
            _invalidate_stage_c_state()
            _rerun_single_cuts_on_active_df()
            _clear_all_cached_insights()
            _INSIGHT_CACHE.clear()
        except Exception as exc:  # noqa: BLE001
            app.session_state["global_filter_error"] = (
                f"{type(exc).__name__}: {exc}"
            )

    pending_pq = app.session_state.get("pending_per_question_filter") or {}
    if pending_pq:
        from src.filtered_single_cut import compute_filtered_single_cut

        schema = app.session_state.get("schema")
        active_df = app.session_state.get("active_df")
        log = app.session_state.get("log")
        errors: dict[str, str] = {}
        for cid, payload in list(pending_pq.items()):
            try:
                filtered_result = compute_filtered_single_cut(
                    cid, payload["filter_specs"], schema, active_df, log
                )
                app.session_state.setdefault("filtered_results", {})[cid] = (
                    filtered_result
                )
            except Exception as exc:  # noqa: BLE001
                errors[cid] = f"{type(exc).__name__}: {exc}"
        app.session_state["pending_per_question_filter"] = {}
        app.session_state["per_question_filter_errors"] = errors
        app.session_state["filtered_workbook_bytes"] = None


def _initialise_session_state() -> None:
    app = _require_streamlit()
    for key, value in SESSION_DEFAULTS.items():
        app.session_state.setdefault(key, value)
    stamped = app.session_state.get("cross_cut_engine_version")
    if stamped != CROSS_CUT_ENGINE_VERSION:
        app.session_state["cross_cut_results"] = []
        app.session_state["cross_cut_skips"] = []
        app.session_state["cross_cut_only_bytes"] = None
        app.session_state["filtered_results"] = {}
        app.session_state["filtered_workbook_bytes"] = None
        app.session_state["cross_cut_engine_version"] = CROSS_CUT_ENGINE_VERSION


_NAV_VIEWS = {"upload", "single", "cross", "outcome", "download", "filters"}


def _current_nav_view() -> str:
    app = _require_streamlit()
    try:
        raw = app.query_params.get("view", "upload")
    except Exception:
        raw = "upload"
    if isinstance(raw, list):
        raw = raw[0] if raw else "upload"
    view = str(raw or "upload").strip().lower()
    return view if view in _NAV_VIEWS else "upload"


def _set_current_nav_view(view: str) -> None:
    if view not in _NAV_VIEWS:
        return
    app = _require_streamlit()
    try:
        app.query_params["view"] = view
    except Exception:
        try:
            app.experimental_set_query_params(view=view)
        except Exception:
            pass


def _nav_href(view: str) -> str:
    return "#"


def _nav_onclick(view: str) -> str:
    v = view.lower()
    return (
        "var b=document.querySelectorAll('button');"
        "for(var i=0;i<b.length;i++){"
        "if((b[i].textContent||'').trim().toLowerCase()"
        ".indexOf('navjump_" + v + "')===0){b[i].click();break;}"
        "}return false;"
    )



def _upload_status(uploaded_file: Any | None) -> str:
    if uploaded_file is None:
        return "not uploaded"
    size_kb = uploaded_file.size / 1024
    return f"{uploaded_file.name} ({size_kb:.1f} KB)"


def _temp_dir() -> str | None:
    tmp_path = Path("/tmp")
    return str(tmp_path) if tmp_path.exists() else None


def _write_upload_to_temp(uploaded_file: Any) -> str:
    suffix = Path(uploaded_file.name).suffix
    with tempfile.NamedTemporaryFile(
        suffix=suffix,
        delete=False,
        dir=_temp_dir(),
    ) as temp_file:
        temp_file.write(uploaded_file.getbuffer())
        return temp_file.name


def _cleanup_temp_files(*paths: str | None) -> None:
    for path in paths:
        if path and os.path.exists(path):
            os.unlink(path)


def _persist_uploaded_input_file_sources(uploaded_files: list[Any], load_report: Any) -> dict[str, Any]:
    """Persist uploaded files so exporters can embed verbatim audit sheets."""

    by_name = {getattr(file, "name", ""): file for file in uploaded_files}

    def find_upload(name: str | None, *, prefer_docx: bool | None = None) -> Any | None:
        if name and name in by_name:
            return by_name[name]
        lowered = (name or "").lower()
        for uploaded in uploaded_files:
            upload_name = getattr(uploaded, "name", "")
            if lowered and upload_name.lower() == lowered:
                return uploaded
        for uploaded in uploaded_files:
            is_docx = getattr(uploaded, "name", "").lower().endswith(".docx")
            if prefer_docx is None or is_docx == prefer_docx:
                return uploaded
        return None

    def persist(uploaded: Any | None) -> str | None:
        return _write_upload_to_temp(uploaded) if uploaded is not None else None

    scenario = getattr(load_report, "scenario", "")
    raw_source = getattr(load_report, "raw_data_source", "")
    datamap_source = getattr(load_report, "datamap_source", "")
    sources: dict[str, Any] = {"scenario": scenario}

    if scenario == "B_combined_xlsx":
        combined = next(
            (
                uploaded
                for uploaded in uploaded_files
                if getattr(uploaded, "name", "").lower().endswith(".xlsx")
            ),
            None,
        )
        combined_path = persist(combined)
        sources.update(
            {
                "raw_path": combined_path,
                "datamap_path": combined_path,
                "raw_sheet": raw_source.removeprefix("sheet:") if isinstance(raw_source, str) and raw_source.startswith("sheet:") else None,
                "datamap_sheet": datamap_source.removeprefix("sheet:") if isinstance(datamap_source, str) and datamap_source.startswith("sheet:") else None,
            }
        )
        return sources

    raw_upload = find_upload(raw_source, prefer_docx=False)
    datamap_upload = find_upload(datamap_source, prefer_docx=(scenario == "C_word_datamap"))
    sources.update(
        {
            "raw_path": persist(raw_upload),
            "datamap_path": persist(datamap_upload),
        }
    )
    return sources


# ---------------------------------------------------------------------------
# Pipeline orchestration
# ---------------------------------------------------------------------------


def _run_with_status_heartbeat(
    status: Any,
    *,
    start_label: str,
    heartbeat_labels: tuple[str, ...],
    work: Any,
    complete_label: str | None = None,
    interval_seconds: float = 5.0,
) -> Any:
    """Run a blocking phase while emitting Streamlit status frames."""
    from concurrent.futures import ThreadPoolExecutor, TimeoutError

    status.update(label=start_label, state="running")
    labels = heartbeat_labels or (start_label,)
    heartbeat_index = 0
    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(work)
        while True:
            try:
                result = future.result(timeout=interval_seconds)
            except TimeoutError:
                status.update(
                    label=labels[heartbeat_index % len(labels)],
                    state="running",
                )
                heartbeat_index += 1
                continue
            if complete_label:
                status.update(label=complete_label, state="running")
            return result


def _fallback_stage4_themes(questions: list[dict[str, Any]], error: str = "") -> dict:
    demo_qids = [
        str(question["question_id"])
        for question in questions
        if question.get("question_id") and question.get("is_demographic")
    ]
    other_qids = [
        str(question["question_id"])
        for question in questions
        if question.get("question_id") and not question.get("is_demographic")
    ]
    themes: list[dict[str, Any]] = []
    if demo_qids:
        themes.append({"name": "Demographics", "question_ids": demo_qids})
    if other_qids:
        themes.append({"name": "All Questions", "question_ids": other_qids})
    return {
        "themes": themes,
        "was_template": True,
        "error_message": error,
    }


def _fallback_stage4_demo_priority(questions: list[dict[str, Any]], error: str = "") -> dict:
    return {
        "priority_ordered": [
            str(question["question_id"])
            for question in questions
            if question.get("question_id")
        ],
        "categories": {},
        "was_template": True,
        "error_message": error,
    }


def _fallback_stage4_short_labels(questions: list[dict[str, Any]]) -> dict[str, str]:
    labels: dict[str, str] = {}
    for question in questions:
        qid = question.get("question_id")
        if not qid:
            continue
        text = str(question.get("question_text", ""))
        labels[str(qid)] = text[:40] + ("..." if len(text) > 40 else "")
    return labels


def _run_pipeline(
    data_map: Any,
    dataframe: Any,
    load_report: Any,
    status: Any,
) -> None:
    from src.calculation_log import CalculationLog
    from src.cross_cut_suggestions import suggest_cross_cuts
    from src.excel_exporter import export_single_cuts
    from src.models import DataQualityReport, GlobalFilterState
    from src.question_classifier import classify_questions, reconcile_multiselect_value_subtypes
    from src.single_cut import compute_single_cuts
    from src.ui.wizard import normalise_custom_filter_count, normalise_per_question_filter_count

    app = _require_streamlit()
    if app.session_state.get("wizard_active") and app.session_state.get("wizard_config_complete"):
        app.session_state["wizard_pipeline_adapter"] = "WizardConfiguredAdapter"
    _clear_chart_type_overrides()
    status.update(label="Stage 1/5: Data files loaded.", state="complete")
    # The unified io layer parses + decodes upstream and discards the
    # decoder's quality report. Reconstruct a minimal one so downstream
    # exporters don't crash while still surfacing parser-side warnings.
    parser_warnings = tuple(
        f"parser: {w}" for w in (load_report.parser_warnings or [])
    )
    quality_report = DataQualityReport(
        total_rows=int(len(dataframe)),
        total_columns=int(len(dataframe.columns)),
        columns_in_datamap=int(len(dataframe.columns)),
        columns_not_in_datamap=tuple(),
        per_column_missing_pct={col: 0.0 for col in dataframe.columns},
        per_column_out_of_range_pct={col: 0.0 for col in dataframe.columns},
        coercion_log=tuple(),
        warnings=parser_warnings,
    )

    classified_schema = _run_with_status_heartbeat(
        status,
        start_label="Stage 2/5: Classifying questions...",
        heartbeat_labels=(
            "Stage 2/5: Reading question metadata...",
            "Stage 2/5: Classifying survey questions...",
            "Stage 2/5: Checking demographic fields...",
        ),
        work=lambda: classify_questions(
            data_map,
            dataframe.columns.tolist(),
            respondent_id_column="record",
            total_respondents=len(dataframe),
            source_rawdata_path=load_report.raw_data_source,
        ),
        complete_label="Stage 2/5: Questions classified.",
    )
    schema = reconcile_multiselect_value_subtypes(classified_schema, dataframe)

    log = CalculationLog()
    results, skips = _run_with_status_heartbeat(
        status,
        start_label="Stage 3/5: Computing single cuts...",
        heartbeat_labels=(
            "Stage 3/5: Computing response distributions...",
            "Stage 3/5: Calculating question summaries...",
            "Stage 3/5: Recording skipped cuts...",
        ),
        work=lambda: compute_single_cuts(schema, dataframe, log),
        complete_label="Stage 3/5: Single cuts computed.",
    )

    output_path = "/tmp/survey_analysis.xlsx"
    from src.ai_insights import (
        categorize_demographic_questions,
        categorize_questions_into_themes,
        generate_short_labels,
    )

    schema, _ = _wizard_pipeline_overrides(schema, {})
    questions_for_themes = [
        {
            "question_id": q.canonical_id,
            "question_text": q.question_text,
            "question_type": q.question_type.value,
            "is_demographic": getattr(q, "is_demographic", False),
        }
        for q in schema.questions
    ]
    demographic_questions_for_priority = [
        {
            "question_id": q.canonical_id,
            "question_text": q.question_text,
            "question_type": q.question_type.value,
        }
        for q in schema.demographic_questions()
    ]
    questions_for_labels = [
        {"question_id": q.canonical_id, "question_text": q.question_text}
        for q in schema.questions
    ]
    if app.session_state.get("skip_ai_enhancements", False):
        status.update(label="Stage 4/5: Skipping AI themes and labels.", state="running")
        themes = _fallback_stage4_themes(
            questions_for_themes,
            error="AI enhancements skipped by user.",
        )
        demo_priority = _fallback_stage4_demo_priority(
            demographic_questions_for_priority,
            error="AI enhancements skipped by user.",
        )
        short_labels = _fallback_stage4_short_labels(questions_for_labels)
        status.update(label="Stage 4/5: AI enhancements skipped.", state="running")
    else:
        def _run_stage4_ai_enhancements() -> tuple[dict, dict, dict[str, str]]:
            from concurrent.futures import ThreadPoolExecutor

            theme_cache: dict[str, Any] = {}
            demo_cache: dict[str, Any] = {}
            label_cache: dict[str, Any] = {}
            with ThreadPoolExecutor(max_workers=3) as executor:
                themes_future = executor.submit(
                    categorize_questions_into_themes,
                    questions_for_themes,
                    cache=theme_cache,
                )
                demo_future = executor.submit(
                    categorize_demographic_questions,
                    demographic_questions_for_priority,
                    cache=demo_cache,
                )
                labels_future = executor.submit(
                    generate_short_labels,
                    questions_for_labels,
                    cache=label_cache,
                )
                return (
                    themes_future.result(),
                    demo_future.result(),
                    labels_future.result(),
                )

        themes, demo_priority, short_labels = _run_with_status_heartbeat(
            status,
            start_label="Stage 4/5: Generating AI themes, labels, and priorities...",
            heartbeat_labels=(
                "Stage 4/5: Grouping related questions...",
                "Stage 4/5: Ranking demographic fields...",
                "Stage 4/5: Shortening workbook labels...",
                "Stage 4/5: Waiting for AI fallbacks if needed...",
            ),
            work=_run_stage4_ai_enhancements,
            complete_label="Stage 4/5: AI enhancements ready.",
        )
    schema, themes = _wizard_pipeline_overrides(schema, themes)
    workbook_custom_filter_count = normalise_custom_filter_count(
        app.session_state.get(
            "wizard_workbook_custom_filter_count",
            app.session_state.get("wiz_num_custom_filters"),
        )
    )
    per_question_filter_count = normalise_per_question_filter_count(
        app.session_state.get(
            "wizard_per_question_filter_count",
            app.session_state.get("wiz_num_per_question_filters"),
        )
    )
    hypothesis_results = app.session_state.get("hypothesis_results", [])
    embed_input_files = app.session_state.get("wizard_embed_input_files", False)
    input_file_sources = app.session_state.get("input_file_embed_sources")
    _run_with_status_heartbeat(
        status,
        start_label="Stage 5/5: Building Excel workbook...",
        heartbeat_labels=(
            "Stage 5/5: Writing single-cut sheets...",
            "Stage 5/5: Adding workbook metadata...",
            "Stage 5/5: Finalizing Excel export...",
        ),
        work=lambda: export_single_cuts(
            results=results,
            skips=skips,
            schema=schema,
            quality_report=quality_report,
            log=log,
            output_path=output_path,
            themes=themes,
            decoded_df=dataframe,
            demo_priority=demo_priority,
            short_labels=short_labels,
            workbook_custom_filter_count=workbook_custom_filter_count,
            per_question_filter_count=per_question_filter_count,
            rank_cross_tab_settings=app.session_state.get("rank_cross_tab_settings"),
            hypothesis_results=hypothesis_results,
            embed_input_files=embed_input_files,
            input_file_sources=input_file_sources,
        ),
        complete_label="Stage 5/5: Excel workbook built.",
    )
    app.session_state["manual_cohort_input"] = getattr(
        load_report,
        "manual_cohort_input",
        None,
    )
    app.session_state["manual_cohort_source"] = (
        "embedded sheet"
        if app.session_state["manual_cohort_input"] is not None
        else None
    )
    app.session_state["manual_cohort_id_column"] = (
        getattr(app.session_state["manual_cohort_input"], "id_column", None)
        if app.session_state["manual_cohort_input"] is not None
        else None
    )

    _INSIGHT_CACHE.clear()
    app.session_state["decoded_df"] = dataframe
    app.session_state["active_df"] = dataframe
    app.session_state["data_map"] = data_map
    app.session_state["survey_type_result"] = None
    app.session_state["outcome_variable_id"] = None
    app.session_state.pop("outcome_variable_selector", None)
    _invalidate_stage_c_state()
    app.session_state["global_filter_state"] = GlobalFilterState()
    app.session_state["global_filter_stats"] = None
    app.session_state["global_filter_rows"] = []
    app.session_state["results"] = results
    app.session_state["skips"] = skips
    app.session_state["schema"] = schema
    app.session_state["quality_report"] = quality_report
    app.session_state["log"] = log
    app.session_state["output_path"] = output_path
    app.session_state["cross_cut_results"] = []
    app.session_state["cross_cut_skips"] = []
    app.session_state["cross_cut_suggestions"] = suggest_cross_cuts(schema)
    app.session_state["cross_cut_only_bytes"] = None
    app.session_state["rank_cross_tab_settings"] = {}
    app.session_state["rank_cross_tab_settings_dirty"] = False
    app.session_state["hypothesis_results"] = []
    app.session_state["pending_hypothesis_spec"] = None
    app.session_state["filtered_results"] = {}
    app.session_state["filtered_workbook_bytes"] = None
    _clear_all_cached_insights()
    app.session_state["run_complete"] = True
    status.update(label="Analysis complete.", state="complete")


def _refresh_full_workbook() -> None:
    from src.ai_insights import (
        categorize_demographic_questions,
        categorize_questions_into_themes,
        generate_short_labels,
    )
    from src.excel_exporter import export_single_cuts
    from src.ui.wizard import normalise_custom_filter_count, normalise_per_question_filter_count

    app = _require_streamlit()
    schema = app.session_state["schema"]
    questions_for_themes = [
        {
            "question_id": q.canonical_id,
            "question_text": q.question_text,
            "question_type": q.question_type.value,
            "is_demographic": getattr(q, "is_demographic", False),
        }
        for q in schema.questions
    ]
    themes = categorize_questions_into_themes(
        questions_for_themes,
        cache=_INSIGHT_CACHE,
    )
    demographic_questions_for_priority = [
        {
            "question_id": q.canonical_id,
            "question_text": q.question_text,
            "question_type": q.question_type.value,
        }
        for q in schema.demographic_questions()
    ]
    demo_priority = categorize_demographic_questions(
        demographic_questions_for_priority,
        cache=_INSIGHT_CACHE,
    )
    short_labels = generate_short_labels(
        [
            {"question_id": q.canonical_id, "question_text": q.question_text}
            for q in schema.questions
        ],
        cache=_INSIGHT_CACHE,
    )
    # One-sheet mode: drop AI themes so everything lands on a single
    # "All Questions" sheet (+ Demographics) rather than themed sheets.
    if app.session_state.get("wiz_output_structure") == "one_sheet":
        themes = None
    export_single_cuts(
        results=app.session_state["results"],
        skips=app.session_state["skips"],
        schema=schema,
        quality_report=app.session_state["quality_report"],
        log=app.session_state["log"],
        output_path=app.session_state["output_path"],
        cross_cut_results=app.session_state["cross_cut_results"],
        cross_cut_skips=app.session_state["cross_cut_skips"],
        themes=themes,
        decoded_df=app.session_state.get("decoded_df"),
        demo_priority=demo_priority,
        short_labels=short_labels,
        workbook_custom_filter_count=normalise_custom_filter_count(
            app.session_state.get(
                "wizard_workbook_custom_filter_count",
                app.session_state.get("wiz_num_custom_filters"),
            )
        ),
        per_question_filter_count=normalise_per_question_filter_count(
            app.session_state.get(
                "wizard_per_question_filter_count",
                app.session_state.get("wiz_num_per_question_filters"),
            )
        ),
        rank_cross_tab_settings=app.session_state.get("rank_cross_tab_settings"),
        hypothesis_results=app.session_state.get("hypothesis_results", []),
        embed_input_files=app.session_state.get("wizard_embed_input_files", False),
        input_file_sources=app.session_state.get("input_file_embed_sources"),
    )
    app.session_state["rank_cross_tab_settings_dirty"] = False


def _run_cross_cut_specs(specs: list[Any]) -> None:
    from src.cross_cut_engine import compute_cross_cuts

    if not specs:
        return

    app = _require_streamlit()
    results, skips = compute_cross_cuts(
        specs,
        app.session_state["schema"],
        app.session_state["active_df"],
        app.session_state["log"],
    )
    existing = {
        result.cross_cut_id: result
        for result in app.session_state["cross_cut_results"]
    }
    for result in results:
        existing[result.cross_cut_id] = result
        app.session_state.setdefault(f"cc_select_{result.cross_cut_id}", True)
    app.session_state["cross_cut_results"] = list(existing.values())
    app.session_state["cross_cut_skips"].extend(skips)
    app.session_state["cross_cut_only_bytes"] = None
    _refresh_full_workbook()


_STAGE_C_WIDGET_KEYS = (
    "seg_mode_radio",
    "winner_values_multiselect",
    "winner_label_input",
    "loser_label_input",
    "numeric_threshold_input",
    "threshold_direction_radio",
    "winner_label_num_input",
    "loser_label_num_input",
    "laggard_values_multiselect",
    "laggard_threshold_input",
    "laggard_threshold_direction_radio",
    "laggard_label_quartile_input",
    "laggard_override_checkbox",
    "laggard_outcome_variable_selector",
    "laggard_outcome_sub_question_selector",
)


def _invalidate_stage_c_state() -> None:
    """Clear Stage C segment definition + result + widget keys.

    Called whenever the underlying data context changes (new upload, global
    filter applied/cleared, outcome variable swapped) so the displayed
    segmentation result always matches the current context.
    """
    app = _require_streamlit()
    app.session_state["segment_definition"] = None
    app.session_state["segmentation_result"] = None
    for key in _STAGE_C_WIDGET_KEYS:
        app.session_state.pop(key, None)


def _rerun_single_cuts_on_active_df() -> None:
    """Recompute single cuts after the active DataFrame changed."""
    from src.calculation_log import CalculationLog
    from src.single_cut import compute_single_cuts

    app = _require_streamlit()
    schema = app.session_state["schema"]
    active_df = app.session_state["active_df"]

    log = CalculationLog()
    results, skips = compute_single_cuts(schema, active_df, log)

    app.session_state["log"] = log
    app.session_state["results"] = results
    app.session_state["skips"] = skips
    app.session_state["cross_cut_results"] = []
    app.session_state["cross_cut_skips"] = []
    app.session_state["cross_cut_only_bytes"] = None
    app.session_state["hypothesis_results"] = []
    app.session_state["pending_hypothesis_spec"] = None
    app.session_state["filtered_results"] = {}
    app.session_state["filtered_workbook_bytes"] = None
    _clear_all_cached_insights()
    _refresh_full_workbook()


# ---------------------------------------------------------------------------
# Schema helpers
# ---------------------------------------------------------------------------


def _eligible_question_options() -> list[str]:
    schema = _require_streamlit().session_state["schema"]
    if schema is None:
        return []
    return [option.question_id for option in cross_cut_question_options(schema)]


def _question_label_map() -> dict[str, str]:
    """Full ID + question text, truncated to keep dropdowns readable."""
    schema = _require_streamlit().session_state["schema"]
    out: dict[str, str] = {}
    for spec in schema.questions:
        text = (spec.question_text or "").strip()
        if len(text) > 80:
            text = text[:77] + "\u2026"
        out[spec.canonical_id] = (
            f"{spec.canonical_id} \u2014 {text}" if text else spec.canonical_id
        )
    for option in cross_cut_question_options(schema):
        out[option.question_id] = option.label
    return out


def _eligible_filter_questions() -> list[Any]:
    """Questions with addressable filter values."""
    schema = _require_streamlit().session_state["schema"]
    if schema is None:
        return []
    return filter_question_options(schema)


# ---------------------------------------------------------------------------
# Cross-cut preview helpers (preserved from Day 10.6)
# ---------------------------------------------------------------------------


def _preview_cross_tab(result: Any) -> None:
    app = _require_streamlit()
    ct = result.result_table
    a, b = result.source_question_ids
    counts = ct.get("counts", {})
    row_pct = ct.get("row_pct", {})
    column_pct = ct.get("column_pct", {})
    row_label_map = ct.get("row_label_map", {})
    col_label_map = ct.get("column_label_map", {})
    row_codes = sorted(counts.keys(), key=lambda v: str(v))
    col_codes = sorted(
        {c for row in counts.values() if isinstance(row, dict) for c in row.keys()},
        key=lambda v: str(v),
    )

    display_mode = app.radio(
        "Display",
        options=["Counts", "Row %", "Column %"],
        horizontal=True,
        key=f"preview_mode_{result.cross_cut_id}",
    )
    if display_mode == "Row %":
        source = row_pct
    elif display_mode == "Column %":
        source = column_pct
    else:
        source = counts

    df = pd.DataFrame(
        index=[row_label_map.get(rc, str(rc)) for rc in row_codes],
        columns=[col_label_map.get(cc, str(cc)) for cc in col_codes],
        data=[
            [source.get(rc, {}).get(cc, 0) for cc in col_codes]
            for rc in row_codes
        ],
    )
    df.index.name = f"\u2193 {a}"
    df.columns.name = f"\u2192 {b}"
    app.caption(f"Rows: {a}   Columns: {b}")
    schema = app.session_state.get("schema")
    chart_mode = "Row %" if display_mode == "Row %" else "Counts"
    _render_chart_for_cross_tab(result, schema, chart_mode)
    _copy_button(df, f"ct_{result.cross_cut_id}_{display_mode}")
    if display_mode == "Counts":
        _styled_dataframe(df, use_container_width=True)
    else:
        # Row %: outliers per row (rows sum to 100%); Column %: per column.
        # Both use scaled (0-100) helpers so the outlier detector's
        # `max < 10` guard doesn't suppress highlights on fractional values.
        styler_fn = (
            _style_outliers_by_row if display_mode == "Row %" else _style_outliers_pct
        )
        try:
            app.dataframe(
                df.style.apply(styler_fn, axis=None).format("{:.1%}"),
                use_container_width=True,
            )
        except Exception:
            app.dataframe(df.style.format("{:.1%}"), use_container_width=True)
    app.caption(f"Grand total: {ct.get('grand_total', 0):,} responses")
    app.caption(
        "Tip: hover the table to use the built-in toolbar "
        "(search, fullscreen, download as CSV)."
    )


def _preview_segment_profile(result: Any) -> None:
    app = _require_streamlit()
    rt = result.result_table
    app.caption(
        f"Filter: {rt.get('filter_expr', '<no filter>')}  \u00b7  "
        f"Filter N: {rt.get('filter_n', 0):,}"
    )
    tr = rt.get("target_result", {}) or {}
    if "distribution" in tr:
        rows = [
            {
                "Code": code,
                "Label": payload.get("label", ""),
                "Count": payload.get("count", 0),
            }
            for code, payload in sorted(
                tr["distribution"].items(), key=lambda x: str(x[0])
            )
        ]
        _df = pd.DataFrame(rows)
        _copy_button(_df, f"sp_dist_{result.cross_cut_id}")
        _styled_dataframe(_df, use_container_width=True, hide_index=True)
    elif "selections" in tr:
        rows = []
        for sub_id, payload in tr["selections"].items():
            label = payload.get("label", "") or ""
            label_lower = label.lower()
            if "unchecked" in label_lower or "not selected" in label_lower:
                continue
            rows.append(
                {
                    "Sub-column": sub_id,
                    "Label": label,
                    "Selected count": payload.get("count", 0),
                }
            )
        rows.sort(key=lambda r: r["Selected count"], reverse=True)
        _df = pd.DataFrame(rows)
        _copy_button(_df, f"sp_sel_{result.cross_cut_id}")
        _styled_dataframe(_df, use_container_width=True, hide_index=True)
        app.caption(
            "Selected counts only. Unchecked counts remain in the audit trail "
            "of the downloaded workbook."
        )
    elif "mean" in tr:
        df = pd.DataFrame(
            [
                {"Statistic": "Valid N", "Count": tr.get("valid_n", 0)},
                {"Statistic": "Missing N", "Count": tr.get("missing_n", 0)},
            ]
        )
        _copy_button(df, f"sp_num_{result.cross_cut_id}")
        app.dataframe(df, use_container_width=True, hide_index=True)
        app.caption(
            "Numeric statistics (mean, median, std) in the downloaded workbook."
        )
    elif "rows" in tr:
        app.caption(f"Grid with {len(tr['rows'])} rows. Per-row counts:")
        grid_rows = []
        for sub_id, row_result in tr["rows"].items():
            dist = row_result.get("distribution", {}) if isinstance(row_result, dict) else {}
            row_dict: dict[str, Any] = {"Row": sub_id}
            for code, payload in dist.items():
                row_dict[f"{code}: {payload.get('label', '')}"] = payload.get("count", 0)
            grid_rows.append(row_dict)
        _df = pd.DataFrame(grid_rows)
        _copy_button(_df, f"sp_grid_{result.cross_cut_id}")
        _styled_dataframe(_df, use_container_width=True, hide_index=True)
    else:
        app.info("Preview not available for this target type.")


def _rank_rows_k(rank_rows: dict[Any, Any]) -> int:
    for row_payload in rank_rows.values():
        if not isinstance(row_payload, dict):
            continue
        payloads = [row_payload.get("overall")]
        per_segment = row_payload.get("per_segment", {}) or {}
        if isinstance(per_segment, dict):
            payloads.extend(per_segment.values())
        for payload in payloads:
            if not isinstance(payload, dict):
                continue
            rank_k = int(payload.get("rank_k", 0) or 0)
            if rank_k > 0:
                return rank_k
            counts = payload.get("counts_per_rank")
            if isinstance(counts, list) and counts:
                return len(counts)
    return 1


def _default_rank_cross_tab_settings(rank_k: int) -> dict[str, Any]:
    return {
        "metric": "Weighted Average",
        "points": [float(rank_k - offset) for offset in range(rank_k)],
        "rank_position": 1,
    }


def _normalise_rank_cross_tab_settings(raw: Any, rank_k: int) -> dict[str, Any]:
    raw = raw if isinstance(raw, dict) else {}
    defaults = _default_rank_cross_tab_settings(rank_k)
    metric = str(raw.get("metric") or defaults["metric"])
    if metric == "Weighted average":
        metric = "Weighted Average"
    if metric not in RANK_CROSS_TAB_METRICS:
        metric = defaults["metric"]
    points_raw = raw.get("points", raw.get("weights"))
    points: list[float] = []
    if isinstance(points_raw, (list, tuple)):
        for value in points_raw[:rank_k]:
            try:
                points.append(float(value))
            except (TypeError, ValueError):
                points.append(0.0)
    if len(points) != rank_k:
        points = list(defaults["points"])
    try:
        rank_position = int(raw.get("rank_position") or defaults["rank_position"])
    except (TypeError, ValueError):
        rank_position = int(defaults["rank_position"])
    rank_position = max(1, min(rank_k, rank_position))
    return {
        "metric": metric,
        "points": points,
        "rank_position": rank_position,
    }


def _render_rank_points_inputs(
    question_id: str,
    rank_k: int,
    key_prefix: str,
    points: list[float],
) -> list[float]:
    app = _require_streamlit()
    rank_k = max(1, int(rank_k or 1))
    points = list(points[:rank_k])
    if len(points) != rank_k:
        points = [float(rank_k - offset) for offset in range(rank_k)]
    cols = app.columns(min(rank_k, 4))
    for offset in range(rank_k):
        with cols[offset % len(cols)]:
            points[offset] = float(
                app.number_input(
                    f"Rank {offset + 1} points",
                    value=float(points[offset]),
                    step=1.0,
                    format="%.2f",
                    key=f"{key_prefix}_point_{offset + 1}",
                )
            )
    return points


def _render_rank_cross_tab_metric_controls(
    question_id: str,
    rank_k: int,
    key_prefix: str,
) -> dict[str, Any]:
    app = _require_streamlit()
    rank_k = max(1, int(rank_k or 1))
    state = app.session_state.setdefault("rank_cross_tab_settings", {})
    prior = _normalise_rank_cross_tab_settings(state.get(question_id), rank_k)
    metric = app.selectbox(
        "Rank metric",
        options=list(RANK_CROSS_TAB_METRICS),
        index=list(RANK_CROSS_TAB_METRICS).index(prior["metric"]),
        key=f"{key_prefix}_metric",
    )
    points = list(prior["points"])
    if metric == "Weighted Average":
        points = _render_rank_points_inputs(
            question_id,
            rank_k,
            key_prefix,
            points,
        )
    rank_position = int(prior["rank_position"])
    if metric == "Rank position count":
        rank_position = int(
            app.selectbox(
                "Rank position",
                options=list(range(1, rank_k + 1)),
                index=rank_position - 1,
                key=f"{key_prefix}_position",
            )
        )
    settings = {
        "metric": metric,
        "points": points,
        "rank_position": rank_position,
    }
    default_settings = _default_rank_cross_tab_settings(rank_k)
    existing = state.get(question_id)
    if existing != settings:
        state[question_id] = settings
        if existing is not None or settings != default_settings:
            app.session_state["rank_cross_tab_settings_dirty"] = True
    return settings


def _rank_cross_tab_metric_value(
    payload: dict[str, Any],
    metric: str,
    points: list[float],
    rank_position: int,
) -> Any:
    counts_raw = payload.get("counts_per_rank")
    counts = [int(value or 0) for value in counts_raw] if isinstance(counts_raw, list) else []
    ranked_n = int(payload.get("n", 0) or sum(counts))
    answered_n = int(payload.get("answered_n", 0) or ranked_n)
    rank_sum = payload.get("rank_sum")
    if rank_sum is None:
        mean = payload.get("mean")
        rank_sum = float(mean) * ranked_n if isinstance(mean, (int, float)) else None
    if metric == "Weighted Average":
        if not counts or answered_n <= 0:
            return None
        numerator = sum(float(point) * count for point, count in zip(points, counts))
        return round(numerator / answered_n, 3)
    if metric == "Sum of ranks":
        return int(round(float(rank_sum or 0.0)))
    if metric == "Rank position count":
        index = max(0, min(len(counts) - 1, int(rank_position) - 1)) if counts else 0
        return counts[index] if counts else 0
    value = payload.get("weighted_average")
    if value is None and counts:
        rank_k = len(counts)
        points_sum = sum(
            count * (rank_k - rank + 1)
            for rank, count in enumerate(counts, start=1)
        )
        value = (points_sum / answered_n) if answered_n else None
    return round(float(value), 3) if isinstance(value, (int, float)) else None


def _rank_cross_tab_metric_caption(metric: str) -> str:
    if metric == "Weighted Average":
        return "Weighted Average per option by segment. Higher is preferred."
    if metric == "Sum of ranks":
        return "Sum of raw ranks per option by segment. Lower sums indicate stronger preference when bases are comparable."
    if metric == "Rank position count":
        return "Count of respondents placing each option at the selected rank position."
    return "Weighted Average per option by segment. Higher is preferred. Audit trail in the downloaded workbook."


def _preview_group_comparison(result: Any) -> None:
    app = _require_streamlit()
    rt = result.result_table
    seg_q = rt.get("segment_question_id", "")
    met_q = rt.get("metric_question_id", "")
    app.caption(f"Metric: {met_q}   Segments: {seg_q}")
    if rt.get("grid_rows"):
        segment_columns: list[tuple[Any, str]] = []
        for row_payload in (rt.get("grid_rows", {}) or {}).values():
            if not isinstance(row_payload, dict):
                continue
            for seg_val, seg_data in (row_payload.get("per_segment", {}) or {}).items():
                label = seg_data.get("label", str(seg_val)) if isinstance(seg_data, dict) else str(seg_val)
                key = (seg_val, label)
                if key not in segment_columns:
                    segment_columns.append(key)

        rows = []
        for row_id, row_payload in (rt.get("grid_rows", {}) or {}).items():
            row_label = row_payload.get("label", row_id) if isinstance(row_payload, dict) else row_id
            record = {"Row": row_label}
            per_segment = row_payload.get("per_segment", {}) if isinstance(row_payload, dict) else {}
            for seg_val, label in segment_columns:
                seg_data = per_segment.get(seg_val, {}) if isinstance(per_segment, dict) else {}
                record[f"{label} Mean"] = seg_data.get("mean") if isinstance(seg_data, dict) else None
                record[f"{label} Median"] = seg_data.get("median") if isinstance(seg_data, dict) else None
                record[f"{label} N"] = seg_data.get("n", 0) if isinstance(seg_data, dict) else 0
            overall = row_payload.get("overall", {}) if isinstance(row_payload, dict) else {}
            record["Overall Mean"] = overall.get("mean") if isinstance(overall, dict) else None
            record["Overall Median"] = overall.get("median") if isinstance(overall, dict) else None
            rows.append(record)
        _df = pd.DataFrame(rows)
        _copy_button(_df, f"gc_grid_{result.cross_cut_id}")
        _styled_dataframe(_df, use_container_width=True, hide_index=True)
        app.caption("Grid-rated row means by segment in the downloaded workbook.")
        return
    if rt.get("selection_rate_rows"):
        segment_columns: list[tuple[Any, str]] = []
        for row_payload in (rt.get("selection_rate_rows", {}) or {}).values():
            if not isinstance(row_payload, dict):
                continue
            for seg_val, seg_data in (row_payload.get("per_segment", {}) or {}).items():
                label = seg_data.get("label", str(seg_val)) if isinstance(seg_data, dict) else str(seg_val)
                key = (seg_val, label)
                if key not in segment_columns:
                    segment_columns.append(key)
        rows = []
        for row_id, row_payload in (rt.get("selection_rate_rows", {}) or {}).items():
            row_label = row_payload.get("label", row_id) if isinstance(row_payload, dict) else row_id
            record = {"Option": row_label}
            per_segment = row_payload.get("per_segment", {}) if isinstance(row_payload, dict) else {}
            for seg_val, label in segment_columns:
                seg_data = per_segment.get(seg_val, {}) if isinstance(per_segment, dict) else {}
                rate = seg_data.get("selection_rate") if isinstance(seg_data, dict) else None
                record[f"{label} %"] = round(rate * 100, 1) if isinstance(rate, (int, float)) else None
                record[f"{label} N"] = seg_data.get("n", 0) if isinstance(seg_data, dict) else 0
            overall = row_payload.get("overall", {}) if isinstance(row_payload, dict) else {}
            overall_rate = overall.get("selection_rate") if isinstance(overall, dict) else None
            record["Overall %"] = round(overall_rate * 100, 1) if isinstance(overall_rate, (int, float)) else None
            record["Overall N"] = overall.get("n", 0) if isinstance(overall, dict) else 0
            rows.append(record)
        _df = pd.DataFrame(rows)
        _copy_button(_df, f"gc_sel_{result.cross_cut_id}")
        _styled_dataframe(_df, use_container_width=True, hide_index=True)
        app.caption("Selection rate per option by segment. Per-cell counts and audit trail in the downloaded workbook.")
        return
    if rt.get("allocation_rows"):
        segment_columns: list[tuple[Any, str]] = []
        for row_payload in (rt.get("allocation_rows", {}) or {}).values():
            if not isinstance(row_payload, dict):
                continue
            for seg_val, seg_data in (row_payload.get("per_segment", {}) or {}).items():
                label = seg_data.get("label", str(seg_val)) if isinstance(seg_data, dict) else str(seg_val)
                key = (seg_val, label)
                if key not in segment_columns:
                    segment_columns.append(key)
        rows = []
        for row_id, row_payload in (rt.get("allocation_rows", {}) or {}).items():
            row_label = row_payload.get("label", row_id) if isinstance(row_payload, dict) else row_id
            record = {"Option": row_label}
            per_segment = row_payload.get("per_segment", {}) if isinstance(row_payload, dict) else {}
            for seg_val, label in segment_columns:
                seg_data = per_segment.get(seg_val, {}) if isinstance(per_segment, dict) else {}
                record[f"{label} Mean"] = seg_data.get("mean") if isinstance(seg_data, dict) else None
                record[f"{label} Median"] = seg_data.get("median") if isinstance(seg_data, dict) else None
                record[f"{label} N"] = seg_data.get("n", 0) if isinstance(seg_data, dict) else 0
            overall = row_payload.get("overall", {}) if isinstance(row_payload, dict) else {}
            record["Overall Mean"] = overall.get("mean") if isinstance(overall, dict) else None
            record["Overall Median"] = overall.get("median") if isinstance(overall, dict) else None
            rows.append(record)
        _df = pd.DataFrame(rows)
        _copy_button(_df, f"gc_alloc_{result.cross_cut_id}")
        _styled_dataframe(_df, use_container_width=True, hide_index=True)
        app.caption("Mean allocation per option by segment. Audit trail in the downloaded workbook.")
        return
    if rt.get("rank_rows"):
        rank_k = _rank_rows_k(rt.get("rank_rows", {}) or {})
        rank_settings = _render_rank_cross_tab_metric_controls(
            str(met_q),
            rank_k,
            f"rank_metric_{result.cross_cut_id}",
        )
        metric = rank_settings["metric"]
        points = rank_settings["points"]
        rank_position = rank_settings["rank_position"]
        segment_columns: list[tuple[Any, str]] = []
        for row_payload in (rt.get("rank_rows", {}) or {}).values():
            if not isinstance(row_payload, dict):
                continue
            for seg_val, seg_data in (row_payload.get("per_segment", {}) or {}).items():
                label = seg_data.get("label", str(seg_val)) if isinstance(seg_data, dict) else str(seg_val)
                key = (seg_val, label)
                if key not in segment_columns:
                    segment_columns.append(key)
        rows = []
        for row_id, row_payload in (rt.get("rank_rows", {}) or {}).items():
            row_label = row_payload.get("label", row_id) if isinstance(row_payload, dict) else row_id
            record = {"Option": row_label}
            per_segment = row_payload.get("per_segment", {}) if isinstance(row_payload, dict) else {}
            for seg_val, label in segment_columns:
                seg_data = per_segment.get(seg_val, {}) if isinstance(per_segment, dict) else {}
                record[label] = (
                    _rank_cross_tab_metric_value(seg_data, metric, points, rank_position)
                    if isinstance(seg_data, dict)
                    else None
                )
            overall = row_payload.get("overall", {}) if isinstance(row_payload, dict) else {}
            record["Overall"] = (
                _rank_cross_tab_metric_value(overall, metric, points, rank_position)
                if isinstance(overall, dict)
                else None
            )
            rows.append(record)
        _df = pd.DataFrame(rows)
        _copy_button(_df, f"gc_rank_{result.cross_cut_id}")
        _styled_dataframe(_df, use_container_width=True, hide_index=True)
        app.caption(_rank_cross_tab_metric_caption(metric))
        return
    if rt.get("nps_entities"):
        rows = []
        for entity_id, entity_payload in (rt.get("nps_entities", {}) or {}).items():
            row = {
                "Entity": entity_payload.get("label", entity_id)
                if isinstance(entity_payload, dict)
                else entity_id
            }
            if not isinstance(entity_payload, dict):
                rows.append(row)
                continue
            for seg_val, seg_data in (entity_payload.get("per_segment", {}) or {}).items():
                if not isinstance(seg_data, dict):
                    continue
                label = str(seg_data.get("label", seg_val))
                row[f"{label} NPS"] = seg_data.get("nps_score")
                row[f"{label} Promoters %"] = seg_data.get("pct_promoters")
                row[f"{label} Detractors %"] = seg_data.get("pct_detractors")
                row[f"{label} Valid N"] = seg_data.get("valid_n", seg_data.get("n", 0))
            overall = entity_payload.get("overall", {}) or {}
            row["Overall NPS"] = overall.get("nps_score")
            row["Overall Valid N"] = overall.get("valid_n", overall.get("n", 0))
            rows.append(row)
        _df = pd.DataFrame(rows)
        _copy_button(_df, f"gc_nps_{result.cross_cut_id}")
        _styled_dataframe(_df, use_container_width=True, hide_index=True)
        app.caption("NPS by entity and segment in the downloaded workbook.")
        return
    _render_chart_for_segment_metric(result, app.session_state.get("schema"))
    rows = []
    for seg_val, seg_data in (rt.get("per_segment", {}) or {}).items():
        rows.append(
            {
                "Segment": seg_data.get("label", str(seg_val)) if isinstance(seg_data, dict) else str(seg_val),
                "N": seg_data.get("n", 0) if isinstance(seg_data, dict) else 0,
            }
        )
    overall = rt.get("overall", {}) or {}
    rows.append(
        {
            "Segment": "Overall",
            "N": overall.get("valid_n", overall.get("n", 0)),
        }
    )
    _df = pd.DataFrame(rows)
    _copy_button(_df, f"gc_{result.cross_cut_id}")
    _styled_dataframe(_df, use_container_width=True, hide_index=True)
    app.caption("Group means in the downloaded workbook.")


def _preview_expected_vs_realized(result: Any) -> None:
    app = _require_streamlit()
    rt = result.result_table
    exp_q = rt.get("expected_question_id", "")
    real_q = rt.get("realized_question_id", "")
    app.caption(f"Expected: {exp_q}   Realized: {real_q}")
    _render_chart_for_expected_vs_realized(result)
    df = pd.DataFrame(
        [
            {"Metric": "Paired N", "Count": rt.get("paired_n", 0)},
            {"Metric": "Expected valid N", "Count": (rt.get("expected", {}) or {}).get("valid_n", 0)},
            {"Metric": "Realized valid N", "Count": (rt.get("realized", {}) or {}).get("valid_n", 0)},
        ]
    )
    _copy_button(df, f"evr_{result.cross_cut_id}")
    _styled_dataframe(df, use_container_width=True, hide_index=True)
    app.caption(
        "Mean expected, mean realized, gap statistics in the downloaded workbook."
    )


def _render_cross_cut_preview(result: Any) -> None:
    app = _require_streamlit()
    from src.models import AnalysisType
    if app.checkbox(
        "Show preview (counts only)",
        value=True,
        key=f"cc_preview_{result.cross_cut_id}",
    ):
        try:
            at = result.analysis_type
            if at == AnalysisType.CROSS_TAB:
                _preview_cross_tab(result)
            elif at == AnalysisType.SEGMENT_PROFILE:
                _preview_segment_profile(result)
            elif at == AnalysisType.GROUP_COMPARISON:
                _preview_group_comparison(result)
            elif at == AnalysisType.EXPECTED_VS_REALIZED:
                _preview_expected_vs_realized(result)
            else:
                app.info(f"Preview not implemented for {at.value}.")
        except Exception as exc:  # noqa: BLE001
            app.error(f"Could not render preview: {type(exc).__name__}: {exc}")


# ---------------------------------------------------------------------------
# Single-cut display helpers (preserved from Day 12)
# ---------------------------------------------------------------------------


def _format_filter(filter_spec: Any) -> str:
    values = filter_spec.get_effective_values()
    if values is None:
        return f"{filter_spec.filter_question_id} (breakdown)"
    if len(values) == 1:
        return f"{filter_spec.filter_question_id} = {values[0]}"
    joined = ", ".join(str(v) for v in values)
    return f"{filter_spec.filter_question_id} \u2208 {{{joined}}}"


def _resolve_filter_value(schema: Any, df: Any, q_id: str, raw_value: Any) -> Any:
    """Reconcile UI option-map codes against actual raw-data column values.

    Some Word-derived data maps keep option labels ("India") in the raw data
    while the option_map exposes ``code -> label`` pairs. The UI hands back
    the code; if that code does not appear in the column we try str(code) and
    finally the option label, returning whichever value actually filters rows.
    """
    if df is None or q_id not in df.columns:
        return raw_value
    col_values = set(df[q_id].dropna().unique())
    if raw_value in col_values:
        return raw_value
    if str(raw_value) in col_values:
        return str(raw_value)
    q_spec = schema.get_question(q_id) if schema is not None else None
    if q_spec is not None and getattr(q_spec, "option_map", None):
        for key in (raw_value, str(raw_value)):
            label = q_spec.option_map.get(key)
            if label and label in col_values:
                return label
    return raw_value


def _normalize_value_list(val: Any) -> list:
    """Coerce a row's value-state (legacy scalar/None or new list) into a list."""
    if val is None:
        return []
    if isinstance(val, (list, tuple)):
        return [v for v in val if v is not None]
    return [val]


def _build_filter_spec(
    schema: Any, df: Any, q_id: str, vals: list
) -> Any:
    """Build a FilterSpec, resolving label/code mismatches per value."""
    from src.models import FilterSpec

    if not vals:
        return FilterSpec(filter_question_id=q_id)
    specs = build_filter_specs_from_selection(schema, q_id, vals) if schema is not None else []
    if len(specs) == 1:
        return specs[0]
    resolved = [_resolve_filter_value(schema, df, q_id, v) for v in vals]
    if len(resolved) == 1:
        return FilterSpec(filter_question_id=q_id, filter_value=resolved[0])
    return FilterSpec(filter_question_id=q_id, filter_values=tuple(resolved))


def _build_filter_specs(
    schema: Any, df: Any, q_id: str, vals: list
) -> list[Any]:
    """Build one or more FilterSpecs from a parent question/value selection."""
    if schema is not None:
        specs = build_filter_specs_from_selection(schema, q_id, vals)
        if specs:
            return specs
    from src.models import FilterSpec

    if not vals:
        return [FilterSpec(filter_question_id=q_id)]
    resolved = [_resolve_filter_value(schema, df, q_id, v) for v in vals]
    if len(resolved) == 1:
        return [FilterSpec(filter_question_id=q_id, filter_value=resolved[0])]
    return [FilterSpec(filter_question_id=q_id, filter_values=tuple(resolved))]


def _build_grid_display_rows(
    result: Any, schema: Any
) -> list[dict[str, Any]]:
    """Build UI display rows for a binary grid result.

    Returns one ``{"Option": label, "Count": count}`` dict per sub-question,
    using the schema's ``grid_row_labels`` for display labels and the
    highest-coded distribution entry (typically value=1, "Checked") for
    counts. Sub-questions with a zero count are omitted.
    """
    spec = next(
        (q for q in schema.questions if q.canonical_id == result.question_id),
        None,
    )
    grid_row_labels = getattr(spec, "grid_row_labels", None) or {}

    rows: list[dict[str, Any]] = []
    for sub_id, row_result in result.rows.items():
        label = grid_row_labels.get(sub_id, sub_id)
        dist = getattr(row_result, "distribution", {}) or {}
        if 1 in dist:
            count = dist[1].get("count", 0)
        elif dist:
            max_code = max(dist.keys(), key=lambda k: str(k))
            count = dist[max_code].get("count", 0)
        else:
            count = 0
        if not count:
            continue
        rows.append({"Option": label, "Count": count})
    return rows


def _render_single_cut_result(result: Any, spec: Any) -> None:
    """Display a SingleCutResult with branded HTML for SS/MS distributions."""
    from src.models import (
        GridBinaryPivotResult,
        GridRatedResult,
        GridSingleSelectResult,
        MultiSelectResult,
        NPSResult,
        NumericResult,
        RankOrderResult,
        SingleSelectResult,
    )

    app = _require_streamlit()
    app.caption(
        f"Valid N: {result.valid_n:,}  \u00b7  Missing N: {result.missing_n:,}"
    )
    if isinstance(result, GridSingleSelectResult):
        grid_rows: list[dict[str, Any]] = []
        for sub_id, row_result in result.rows.items():
            row_label = (
                spec.grid_row_labels.get(sub_id, sub_id)
                if spec.grid_row_labels
                else sub_id
            )
            row_dict: dict[str, Any] = {"Row": row_label}
            for code, payload in row_result.distribution.items():
                row_dict[f"{code}: {payload.get('label', '')}"] = payload.get(
                    "count", 0
                )
            grid_rows.append(row_dict)
        _grid_df = pd.DataFrame(grid_rows)
        _copy_button(_grid_df, f"grid_{result.question_id}")
        _styled_dataframe(_grid_df, use_container_width=True, hide_index=True)
        _render_recommended_chart_and_thinkcell(result, spec)
    elif isinstance(result, SingleSelectResult):
        sorted_dist = dict(
            sorted(result.distribution.items(), key=lambda kv: str(kv[0]))
        )
        # Compute flags ONCE from counts so highlighting stays identical
        # across Counts / Counts+% / % only toggles.
        ss_counts = [p.get("count", 0) for _, p in sorted_dist.items()]
        ss_flags = _compute_outlier_flags(ss_counts)
        tab_chart, tab_tc = app.tabs(["Chart", "Think-cell"])
        with tab_chart:
            _render_recommended_chart(result, spec)
            display_mode = app.radio(
                "Display",
                options=["Counts", "Counts + %", "% only"],
                format_func=lambda x: {"Counts": "Count only", "Counts + %": "Count", "% only": "%"}[x],
                index=1,
                horizontal=True,
                key=f"sc_display_{result.question_id}",
                label_visibility="collapsed",
            )
            _render_sc_table_html(
                sorted_dist,
                display_mode,
                result.valid_n,
                flags=ss_flags,
                key_suffix=str(result.question_id),
            )
        with tab_tc:
            _render_thinkcell_layout(result, spec)
    elif isinstance(result, MultiSelectResult):
        ms_dist: dict[Any, dict[str, Any]] = {}
        for sub_id, payload in result.selections.items():
            label = payload.get("label", "") or ""
            label_lower = label.lower()
            if "unchecked" in label_lower or "not selected" in label_lower:
                continue
            # MultiSelectResult payloads use "selection_rate"; tolerate "rate"
            # for forward-compat. Either way, _render_sc_table_html will fall
            # back to count/valid_n if the rate is missing.
            ms_rate = payload.get("selection_rate")
            if ms_rate is None:
                ms_rate = payload.get("rate")
            ms_dist[sub_id] = {
                "label": label or sub_id,
                "count": payload.get("count", 0),
                "rate": ms_rate,
            }
        ms_dist = dict(
            sorted(ms_dist.items(), key=lambda kv: kv[1]["count"], reverse=True)
        )
        # Compute flags ONCE from selected counts.
        ms_counts = [p.get("count", 0) for _, p in ms_dist.items()]
        ms_flags = _compute_outlier_flags(ms_counts)
        tab_chart, tab_tc = app.tabs(["Chart", "Think-cell"])
        with tab_chart:
            _render_recommended_chart(result, spec)
            display_mode = app.radio(
                "Display",
                options=["Counts", "Counts + %", "% only"],
                format_func=lambda x: {"Counts": "Count only", "Counts + %": "Count", "% only": "%"}[x],
                index=1,
                horizontal=True,
                key=f"sc_display_{result.question_id}",
                label_visibility="collapsed",
            )
            _render_sc_table_html(
                ms_dist,
                display_mode,
                result.valid_n,
                flags=ms_flags,
                key_suffix=f"ms_{result.question_id}",
            )
            app.caption(
                "Selected counts only. Unchecked counts remain in the audit trail "
                "of the downloaded workbook."
            )
        with tab_tc:
            _render_thinkcell_layout(result, spec)
    elif isinstance(result, NumericResult):
        rows = [
            {"Statistic": "Valid N", "Value": result.valid_n},
            {"Statistic": "Missing N", "Value": result.missing_n},
        ]
        _num_df = pd.DataFrame(rows)
        _copy_button(_num_df, f"num_{result.question_id}")
        app.dataframe(_num_df, use_container_width=True, hide_index=True)
        _render_recommended_chart_and_thinkcell(result, spec)
        app.caption("Mean, median, std and percentiles in the downloaded workbook.")
    elif isinstance(result, NPSResult):
        nps_rows = []
        for e in result.entities:
            nps_rows.append(
                {
                    "Entity": e.entity_label,
                    "Promoters %": round(e.pct_promoters * 100, 1),
                    "Passives %": round(e.pct_passives * 100, 1),
                    "Detractors %": round(e.pct_detractors * 100, 1),
                    "NPS": round(e.nps_score),
                    "Valid N": e.valid_n,
                }
            )
        nps_df = pd.DataFrame(nps_rows)
        _copy_button(nps_df, f"nps_{result.question_id}")
        _styled_dataframe(nps_df, use_container_width=True, hide_index=True)
        if nps_rows:
            try:
                app.bar_chart(nps_df.set_index("Entity")[["NPS"]])
            except Exception:
                pass
        app.caption(
            "NPS = % promoters (9\u201310) \u2212 % detractors (0\u20136). "
            "Per-entity counts and full audit trail are in the downloaded workbook."
        )
    elif isinstance(result, RankOrderResult):
        rank_k = max(1, int(getattr(result, "K", 1) or 1))
        state = app.session_state.setdefault("rank_cross_tab_settings", {})
        prior = _normalise_rank_cross_tab_settings(
            state.get(result.question_id),
            rank_k,
        )
        app.caption("Rank points seed the workbook's editable Weighted Average cells.")
        points = _render_rank_points_inputs(
            result.question_id,
            rank_k,
            f"rank_points_{result.question_id}",
            list(prior["points"]),
        )
        settings = {
            "metric": prior["metric"],
            "points": points,
            "rank_position": prior["rank_position"],
        }
        existing = state.get(result.question_id)
        if existing != settings:
            state[result.question_id] = settings
            if existing is not None or settings != _default_rank_cross_tab_settings(rank_k):
                app.session_state["rank_cross_tab_settings_dirty"] = True
        rank_rows: list[dict[str, Any]] = []
        for row in result.rows:
            numerator = sum(
                int(count) * float(point)
                for count, point in zip(row.counts_per_rank, points)
            )
            weighted_average = (
                numerator / result.total_respondents
                if result.total_respondents
                else 0.0
            )
            record: dict[str, Any] = {
                "Option": row.option_label,
                "Weighted Average": round(float(weighted_average), 3),
            }
            for rank_number, count in enumerate(row.counts_per_rank, start=1):
                record[f"Rank {rank_number} Count"] = count
            rank_rows.append(record)
        rank_df = pd.DataFrame(rank_rows)
        _copy_button(rank_df, f"rank_{result.question_id}")
        _styled_dataframe(rank_df, use_container_width=True, hide_index=True)
        _render_recommended_chart_and_thinkcell(result, spec)
    elif isinstance(result, (GridRatedResult, GridBinaryPivotResult)):
        _render_recommended_chart_and_thinkcell(result, spec)
    else:
        app.info("Preview not available for this result type.")


# ---------------------------------------------------------------------------
# Per-question multi-row filter panel (Day 14)
# ---------------------------------------------------------------------------


def _find_q_index(q_id: str | None, options: list[tuple[str, Any]]) -> int:
    for i, (_label, value) in enumerate(options):
        if value == q_id:
            return i
    return 0


def _purge_widget_keys(*prefixes: str) -> None:
    """Delete widget-state keys with the given prefixes.

    Streamlit's selectbox honors the ``index=`` arg only on the first render
    of a given key; once that key exists in ``session_state`` the stored value
    overrides ``index=``. When the list of rows changes (add/delete), the
    index-based keys point at the wrong rows, so we purge them and let the
    row-state seed the next render.
    """
    app = _require_streamlit()
    for key in list(app.session_state.keys()):
        if any(key.startswith(prefix) for prefix in prefixes):
            del app.session_state[key]


def _qtype_pill(question_type: Any) -> str:
    """Color-coded HTML pill for a question type (consistent color + icon)."""
    val = question_type.value if hasattr(question_type, "value") else str(question_type)
    v = str(val).upper()
    if "NPS" in v:
        cls, icon, label = "qtype-nps", "\u2605", "NPS"
    elif "MULTI" in v:
        cls, icon, label = "qtype-multi", "\u2611", "Multi-select"
    elif "SINGLE" in v:
        cls, icon, label = "qtype-single", "\u25C9", "Single-select"
    elif "GRID" in v or "MATRIX" in v:
        cls, icon, label = "qtype-grid", "\u25A6", "Grid"
    elif "RANK" in v:
        cls, icon, label = "qtype-rank", "\u2261", "Rank"
    elif "NUMERIC" in v or "ALLOCATION" in v:
        cls, icon, label = "qtype-numeric", "\u2116", "Numeric"
    else:
        cls, icon, label = "qtype-other", "\u2022", str(val).replace("_", " ").title()
    return f'<span class="qtype {cls}">{icon} {html.escape(label)}</span>'


def _question_type_is_grid(question_type: Any) -> bool:
    value = question_type.value if hasattr(question_type, "value") else str(question_type)
    return str(value).upper() in {
        "GRID_RATED",
        "GRID_SINGLE_SELECT",
        "GRID_BINARY_SELECT",
    }


def _render_single_cut_card(
    result: Any, spec: Any, expanded: bool = False
) -> None:
    from src.models import FilterSpec  # noqa: F401  (used by transitive helpers)

    app = _require_streamlit()
    schema = app.session_state["schema"]
    short_text = (spec.question_text or "")[:80]

    expander_label = f"{spec.canonical_id} \u2014 {short_text}"
    pending_insight: dict[str, Any] | None = None
    breakdown_allowed = not _question_type_is_grid(spec.question_type)
    with app.expander(expander_label, expanded=expanded):
        app.markdown(
            _qtype_pill(spec.question_type)
            + f'&nbsp;<span style="color:#8A8F94;font-size:12px;">'
            f'Valid N: {result.valid_n:,} &nbsp;\u00b7&nbsp; Missing: {result.missing_n:,}</span>',
            unsafe_allow_html=True,
        )

        filter_key = f"filters_{spec.canonical_id}"
        if filter_key not in app.session_state:
            app.session_state[filter_key] = []
        filter_rows_current = app.session_state.get(filter_key, [])
        has_active_filter = bool(app.session_state.get("filtered_results", {}).get(spec.canonical_id))
        show_filter_ui = bool(filter_rows_current) or has_active_filter

        if show_filter_ui:
            app.markdown("**Filters for this question**")
            app.caption(TOOLTIP_PER_QUESTION_FILTER)

        filter_rows: list[tuple[str | None, Any]] = (
            app.session_state[filter_key] if show_filter_ui else []
        )

        eligible = _eligible_filter_questions()
        filter_options_by_id = {q.question_id: q for q in eligible}
        question_options: list[tuple[str, Any]] = [("None", None)] + [
            (
                q.label,
                q.question_id,
            )
            for q in eligible
            if q.question_id != spec.canonical_id
        ]

        new_rows: list[tuple[str | None, Any]] = []
        delete_index: int | None = None
        for i, (q_id, val) in enumerate(filter_rows):
            cols = app.columns([4, 4, 1])
            with cols[0]:
                q_pick = app.selectbox(
                    "Filter question",
                    options=question_options,
                    format_func=lambda x: x[0],
                    index=_find_q_index(q_id, question_options),
                    key=f"{filter_key}_q_{i}",
                    label_visibility="visible" if i == 0 else "collapsed",
                )
            picked_q_id = q_pick[1]
            with cols[1]:
                if picked_q_id is not None:
                    q_spec = filter_options_by_id[picked_q_id]
                    value_codes = [choice.value for choice in q_spec.values]
                    value_labels = {choice.value: choice.label for choice in q_spec.values}
                    prior = _normalize_value_list(val)
                    default_codes = [v for v in prior if v in value_codes]
                    widget_key = f"{filter_key}_v_{i}_{picked_q_id}"
                    if widget_key not in app.session_state:
                        app.session_state[widget_key] = default_codes
                    else:
                        app.session_state[widget_key] = [
                            v for v in app.session_state[widget_key]
                            if v in value_codes
                        ]
                    v_pick = app.multiselect(
                        "Values (leave empty for breakdown)" if breakdown_allowed else "Values",
                        options=value_codes,
                        format_func=lambda v: value_labels.get(v, str(v)),
                        key=widget_key,
                        label_visibility="visible" if i == 0 else "collapsed",
                        help=TOOLTIP_BREAKDOWN if i == 0 and breakdown_allowed else None,
                        placeholder=(
                            "All values (breakdown)"
                            if breakdown_allowed
                            else "Select one or more values"
                        ),
                    )
                    new_val = list(v_pick)
                else:
                    app.multiselect(
                        "Values",
                        options=[],
                        disabled=True,
                        key=f"{filter_key}_v_disabled_{i}",
                        label_visibility="visible" if i == 0 else "collapsed",
                        placeholder="Pick a question first",
                    )
                    new_val = []
            with cols[2]:
                if i == 0:
                    app.markdown("&nbsp;", unsafe_allow_html=True)
                if app.button(
                    "\u2715",
                    key=f"{filter_key}_del_{i}",
                    help="Remove this filter",
                ):
                    delete_index = i
            new_rows.append((picked_q_id, new_val))

        app.session_state[filter_key] = new_rows

        if delete_index is not None:
            new_rows.pop(delete_index)
            app.session_state[filter_key] = new_rows
            _purge_widget_keys(f"{filter_key}_q_", f"{filter_key}_v_")
            app.rerun()

        apply_clicked = False
        if show_filter_ui:
            cols_btn = app.columns([2, 2, 4])
            with cols_btn[0]:
                if app.button("+ Add filter", key=f"{filter_key}_add"):
                    app.session_state[filter_key] = new_rows + [(None, None)]
                    _purge_widget_keys(f"{filter_key}_q_", f"{filter_key}_v_")
                    app.rerun()
            with cols_btn[1]:
                has_filter_question = any(q is not None for q, _v in new_rows)
                grid_filter_missing_value = (
                    not breakdown_allowed
                    and any(
                        q is not None and not _normalize_value_list(v)
                        for q, v in new_rows
                    )
                )
                apply_clicked = app.button(
                    "Apply filters",
                    key=f"{filter_key}_apply",
                    type="primary",
                    disabled=not has_filter_question or grid_filter_missing_value,
                )
        elif app.button("+ Add filter just for this question", key=f"{filter_key}_reveal"):
            app.session_state[filter_key] = [(None, None)]
            app.rerun()

        pq_errors = app.session_state.get("per_question_filter_errors") or {}
        if show_filter_ui and spec.canonical_id in pq_errors:
            app.error(f"Filter failed: {pq_errors[spec.canonical_id]}")

        if apply_clicked:
            active_df = app.session_state["active_df"]
            specs = []
            for q, v in new_rows:
                if q is not None:
                    specs.extend(_build_filter_specs(schema, active_df, q, _normalize_value_list(v)))
            seen: set[str] = set()
            duplicate = next(
                (f.filter_question_id for f in specs if f.filter_question_id in seen
                 or seen.add(f.filter_question_id)),
                None,
            )
            breakdowns = [f for f in specs if f.is_breakdown()]
            if duplicate is not None:
                app.error(
                    f"Duplicate filter on {duplicate}. "
                    "Each filter question can only be used once per card."
                )
            elif len(breakdowns) > 1:
                app.error(
                    "Only one breakdown filter at a time. "
                    "Pick a value for at least all but one of the breakdowns."
                )
            elif breakdowns and not breakdown_allowed:
                app.error("Grid questions support value filters, not breakdowns.")
            else:
                # Queue for processing in _drain_pending_actions on next rerun.
                # Inline compute would require a second click for state to
                # propagate; deferring fixes the double-click bug.
                app.session_state.setdefault(
                    "pending_per_question_filter", {}
                )[spec.canonical_id] = {"filter_specs": specs}
                app.session_state.setdefault(
                    "per_question_filter_errors", {}
                ).pop(spec.canonical_id, None)
                # Filter context changed — drop stale cached insights for this card.
                for k in (
                    f"insight_sc_{spec.canonical_id}",
                    f"insight_breakdown_{spec.canonical_id}",
                ):
                    app.session_state.pop(k, None)
                app.rerun()

        app.divider()

        filtered = app.session_state.get("filtered_results", {}).get(
            spec.canonical_id
        )
        if filtered is not None:
            _fcol1, _fcol2 = app.columns([5, 1])
            with _fcol1:
                app.info(
                    "Filtered: "
                    + ", ".join(_format_filter(f) for f in filtered.filters_applied)
                    + f"  \u00b7  N = {filtered.filtered_n:,}"
                )
            with _fcol2:
                if app.button("Remove filter", key=f"{filter_key}_remove"):
                    # Drop the stored filtered result + reset the filter rows so the
                    # card reverts to unfiltered values.
                    app.session_state.get("filtered_results", {}).pop(
                        spec.canonical_id, None
                    )
                    app.session_state[filter_key] = []
                    app.session_state.get("per_question_filter_errors", {}).pop(
                        spec.canonical_id, None
                    )
                    for k in (
                        f"insight_sc_{spec.canonical_id}",
                        f"insight_breakdown_{spec.canonical_id}",
                    ):
                        app.session_state.pop(k, None)
                    app.rerun()
            for warning in filtered.warnings:
                app.warning(warning)
            active_filter_strs = [
                _format_filter(f) for f in filtered.filters_applied
            ]
            if filtered.dispatch_mode == "single_cut_filtered":
                _render_single_cut_result(filtered.single_cut_result, spec)
                pending_insight = {
                    "insight_key": f"insight_sc_{spec.canonical_id}",
                    "payload_factory": lambda: _build_insight_payload_single_cut(
                        filtered.single_cut_result, spec, active_filter_strs
                    ),
                    "table_kind": "filtered_single_cut",
                    "title_hint": spec.question_text,
                }
            elif filtered.dispatch_mode == "cross_cut_breakdown":
                _render_cross_cut_preview(filtered.cross_cut_result)
                pending_insight = {
                    "insight_key": f"insight_breakdown_{spec.canonical_id}",
                    "payload_factory": lambda: _build_insight_payload_cross_cut(
                        filtered.cross_cut_result
                    ),
                    "table_kind": filtered.cross_cut_result.analysis_type.value.lower(),
                    "title_hint": spec.question_text,
                }
            app.checkbox(
                "Include in filtered workbook download",
                value=True,
                key=f"fsc_select_{spec.canonical_id}",
            )
            if app.button(
                "Clear filters", key=f"clear_filter_{spec.canonical_id}"
            ):
                del app.session_state["filtered_results"][spec.canonical_id]
                app.session_state[filter_key] = []
                app.session_state["filtered_workbook_bytes"] = None
                for k in (
                    f"insight_sc_{spec.canonical_id}",
                    f"insight_breakdown_{spec.canonical_id}",
                ):
                    app.session_state.pop(k, None)
                app.rerun()
        else:
            _render_single_cut_result(result, spec)
            pending_insight = {
                "insight_key": f"insight_sc_{spec.canonical_id}",
                "payload_factory": lambda: _build_insight_payload_single_cut(
                    result, spec, []
                ),
                "table_kind": "single_cut",
                "title_hint": spec.question_text,
            }

    # Render insight section OUTSIDE the expander above. Streamlit forbids
    # nesting expanders, and _render_insight_section uses an expander to show
    # the API error detail when the AI call fails.
    if pending_insight is not None:
        # Auto-create an insight when a per-question filter is active;
        # otherwise keep the manual "Generate AI Insight" button.
        if pending_insight["table_kind"] != "single_cut":
            with app.spinner("Generating insight..."):
                auto_payload = pending_insight["payload_factory"]()
                auto_insight = generate_insight(
                    table_payload=auto_payload,
                    table_kind=pending_insight["table_kind"],
                    title_hint=str(pending_insight["title_hint"]),
                    cache=_INSIGHT_CACHE,
                )
            _render_insight_card(auto_insight)
        else:
            _render_insight_section(**pending_insight)


# ---------------------------------------------------------------------------
# Cross-cut helpers (preserved)
# ---------------------------------------------------------------------------


def _render_suggested_cross_cuts() -> None:
    app = _require_streamlit()
    suggestions = app.session_state["cross_cut_suggestions"]
    if not suggestions:
        app.write("No rule-based suggestions available for this schema.")
        return

    import html as _html

    for index, (spec, reason) in enumerate(suggestions[:15], start=1):
        col_text, col_button = app.columns([4, 1])
        with col_text:
            app.markdown(
                f'<div style="font-size:13px;font-weight:600;color:#0A0A0A;'
                f'font-family:Arial;">{index}. {_html.escape(spec.title)}</div>'
                f'<div style="font-size:10px;color:#888;margin-top:2px;'
                f'font-family:Arial;">'
                f'{_html.escape(" · ".join(spec.source_question_ids))}'
                f'</div>',
                unsafe_allow_html=True,
            )
            app.caption(reason)
        if col_button.button("Run this", key=f"run_suggestion_{spec.cross_cut_id}"):
            _run_cross_cut_specs([spec])
            app.success(f"Ran {spec.cross_cut_id}")


def _render_manual_cross_cut() -> None:
    from src.models import AnalysisType, CrossCutSpec, QuestionType

    app = _require_streamlit()
    options = _eligible_question_options()
    labels = _question_label_map()
    if len(options) < 2:
        app.write("At least two analysis-eligible questions are required.")
        return

    with app.form("manual_cross_cut_form"):
        analysis_type_name = app.selectbox(
            "Analysis type",
            [
                AnalysisType.CROSS_TAB.value,
                AnalysisType.SEGMENT_PROFILE.value,
                AnalysisType.GROUP_COMPARISON.value,
                AnalysisType.EXPECTED_VS_REALIZED.value,
            ],
        )
        first = app.selectbox(
            "First source question",
            options,
            format_func=lambda value: labels.get(value, value),
        )
        second = app.selectbox(
            "Second source question",
            options,
            index=1,
            format_func=lambda value: labels.get(value, value),
        )
        filter_expr = app.text_input(
            "Filter expression for segment profile",
            value=f"{first} == 1",
            help="Required only for SEGMENT_PROFILE. Supports equality, e.g. Q3 == 1.",
        )
        submitted = app.form_submit_button("Run manual cross cut", type="primary")

    if not submitted:
        return

    analysis_type = AnalysisType(analysis_type_name)
    schema = app.session_state["schema"]
    first_spec = resolve_cross_cut_question(schema, first)
    second_spec = resolve_cross_cut_question(schema, second)
    if first_spec is None or second_spec is None:
        app.error("Selected questions are no longer available in the current schema.")
        return

    numeric_types = {
        QuestionType.DIRECT_NUMERIC,
        QuestionType.NUMERIC_ALLOCATION,
        QuestionType.GRID_RATED,
        QuestionType.NPS,
        QuestionType.MULTI_SELECT_BINARY,
        QuestionType.RANK_ORDER,
        QuestionType.GRID_BINARY_SELECT,
    }
    categorical_types = {
        QuestionType.SINGLE_SELECT,
        QuestionType.DEMOGRAPHIC_OR_SEGMENT,
        QuestionType.GRID_SINGLE_SELECT,
        QuestionType.NPS,
    }
    first_is_numeric = first_spec.question_type in numeric_types
    second_is_numeric = second_spec.question_type in numeric_types
    first_is_categorical = first_spec.question_type in categorical_types
    second_is_categorical = second_spec.question_type in categorical_types
    source_ids = (first, second)

    if first_is_numeric ^ second_is_numeric:
        if first_is_numeric and second_is_categorical:
            segment_id, metric_id = second, first
            metric_spec = first_spec
        elif second_is_numeric and first_is_categorical:
            segment_id, metric_id = first, second
            metric_spec = second_spec
        else:
            app.error(
                "Numeric cross-cuts require one categorical segment and one numeric metric."
            )
            return
        analysis_type = AnalysisType.GROUP_COMPARISON
        source_ids = (segment_id, metric_id)
    elif analysis_type is AnalysisType.CROSS_TAB:
        if not (first_is_categorical and second_is_categorical):
            app.error(
                "CROSS_TAB requires two categorical questions. Use GROUP_COMPARISON "
                "for one categorical segment and one numeric metric."
            )
            return
    elif analysis_type is AnalysisType.GROUP_COMPARISON:
        app.error(
            "GROUP_COMPARISON requires one categorical segment and one supported metric."
        )
        return
    elif analysis_type is AnalysisType.EXPECTED_VS_REALIZED:
        if not (
            first_spec.question_type is QuestionType.DIRECT_NUMERIC
            and second_spec.question_type is QuestionType.DIRECT_NUMERIC
        ):
            app.error("EXPECTED_VS_REALIZED requires two direct numeric questions.")
            return

    try:
        spec = CrossCutSpec(
            cross_cut_id=f"MANUAL_{analysis_type.value}_{source_ids[0]}_{source_ids[1]}",
            title=f"{analysis_type.value}: {source_ids[0]} x {source_ids[1]}",
            analysis_type=analysis_type,
            source_question_ids=source_ids,
            filter_expr=filter_expr if analysis_type is AnalysisType.SEGMENT_PROFILE else None,
            filter_mask_description=filter_expr if analysis_type is AnalysisType.SEGMENT_PROFILE else None,
        )
        _run_cross_cut_specs([spec])
        app.success(f"Ran {spec.cross_cut_id}")
    except Exception as exc:  # noqa: BLE001
        app.error(f"{type(exc).__name__}: {exc}")


def _render_selected_question_cross_cut_builder(target_id: str) -> None:
    from src.models import AnalysisType, CrossCutSpec, QuestionType

    app = _require_streamlit()
    schema = app.session_state.get("schema")
    if schema is None:
        app.info("Load a survey before creating cross cuts.")
        return

    target_spec = resolve_cross_cut_question(schema, target_id)
    if target_spec is None:
        app.info("Selected question is not available for cross cuts.")
        return
    if _question_type_is_grid(target_spec.question_type):
        app.info("Grid questions show single-cut results only.")
        return

    labels = _question_label_map()
    dimension_types = {
        QuestionType.SINGLE_SELECT,
        QuestionType.DEMOGRAPHIC_OR_SEGMENT,
    }
    metric_types = {
        QuestionType.DIRECT_NUMERIC,
        QuestionType.NUMERIC_ALLOCATION,
        QuestionType.NPS,
        QuestionType.MULTI_SELECT_BINARY,
        QuestionType.RANK_ORDER,
    }
    categorical_types = set(dimension_types)
    dimension_options: list[str] = []
    for qid in _eligible_question_options():
        if qid == target_id:
            continue
        spec = resolve_cross_cut_question(schema, qid)
        if spec is not None and spec.question_type in dimension_types:
            dimension_options.append(qid)

    app.markdown("**Cross cut this question**")
    app.caption(f"Target: {labels.get(target_id, target_id)}")
    if not dimension_options:
        app.info("No eligible dimension questions are available.")
        return

    with app.form(f"selected_question_cross_cut_{target_id}"):
        dimension_id = app.selectbox(
            "Dimension",
            dimension_options,
            format_func=lambda value: labels.get(value, value),
            key=f"selected_question_dimension_{target_id}",
        )
        submitted = app.form_submit_button("Run cross cut", type="primary")

    result_key = f"selected_question_cross_cut_result_{target_id}"
    if submitted:
        dimension_spec = resolve_cross_cut_question(schema, dimension_id)
        if dimension_spec is None:
            app.error("Selected dimension is no longer available.")
            return
        if target_spec.question_type in metric_types:
            analysis_type = AnalysisType.GROUP_COMPARISON
            source_ids = (dimension_id, target_id)
        elif (
            target_spec.question_type in categorical_types
            and dimension_spec.question_type in categorical_types
        ):
            analysis_type = AnalysisType.CROSS_TAB
            source_ids = (target_id, dimension_id)
        else:
            app.error("Pick a categorical dimension for this target.")
            return

        try:
            spec = CrossCutSpec(
                cross_cut_id=f"INLINE_{analysis_type.value}_{source_ids[0]}_{source_ids[1]}",
                title=f"{target_id} by {dimension_id}",
                analysis_type=analysis_type,
                source_question_ids=source_ids,
            )
            _run_cross_cut_specs([spec])
            app.session_state[result_key] = spec.cross_cut_id
            app.success("Cross cut created.")
        except Exception as exc:  # noqa: BLE001
            app.error(f"{type(exc).__name__}: {exc}")
            return

    result_id = app.session_state.get(result_key)
    result = next(
        (
            item
            for item in app.session_state.get("cross_cut_results", [])
            if item.cross_cut_id == result_id
        ),
        None,
    )
    if result is not None:
        app.markdown("**Result**")
        _render_cross_cut_preview(result)


# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------


DRAFT_ADAPTER_DIR = "/tmp/draft_adapters"
FORMAT_WIZARD_STEPS = (
    "Identify sheets",
    "Respondent ID column",
    "Question ID format",
    "Sub-column separator",
    "Option code location",
    "Section prefixes",
    "Helper / metadata columns",
)


def _load_saved_wizard_configs() -> dict[str, dict[str, Any]]:
    """Load session-local draft adapter configs from /tmp/draft_adapters."""

    saved: dict[str, dict[str, Any]] = {}
    draft_dir = Path(DRAFT_ADAPTER_DIR)
    if not draft_dir.exists():
        return saved
    for path in sorted(draft_dir.glob("*.json")):
        try:
            saved[path.stem] = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
    return saved


def _save_wizard_config_draft(config: dict[str, Any]) -> str:
    draft_dir = Path(DRAFT_ADAPTER_DIR)
    draft_dir.mkdir(parents=True, exist_ok=True)
    name = str(config.get("config_name") or "wizard_format")
    safe_name = _re.sub(r"[^A-Za-z0-9_.-]+", "_", name).strip("_") or "wizard_format"
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    path = draft_dir / f"{timestamp}_{safe_name}.json"
    path.write_text(json.dumps(config, indent=2, sort_keys=True), encoding="utf-8")
    return str(path)


def _render_saved_wizard_config_picker() -> None:
    app = _require_streamlit()
    saved_configs = _load_saved_wizard_configs()
    if not saved_configs:
        return
    options = [""] + list(saved_configs)
    selected = app.selectbox(
        "Saved survey formats",
        options,
        index=0,
        key="wizard_saved_format",
        format_func=lambda value: "Auto-detect format" if value == "" else value,
    )
    if selected:
        app.session_state["wizard_config"] = saved_configs[selected]
        app.session_state["wizard_config_complete"] = True
        app.session_state["wizard_active"] = False
        app.caption("Saved format selected; the format wizard will be skipped.")


def _uploaded_files_signature(uploaded_files: list[Any]) -> tuple[tuple[str, int], ...]:
    signature: list[tuple[str, int]] = []
    for file in uploaded_files:
        try:
            content = _read_uploaded_bytes(file)
            signature.append((str(file.name), len(content)))
        except Exception:
            signature.append((str(getattr(file, "name", "<upload>")), 0))
    return tuple(signature)


def _read_uploaded_bytes(file: Any) -> bytes:
    file.seek(0)
    content = file.read()
    file.seek(0)
    return content


def _probe_format_wizard(uploaded_files: list[Any]) -> dict[str, Any]:
    """Inspect uploads and return wizard defaults plus adapter scores."""

    import io as stdlib_io
    from openpyxl import load_workbook
    from src.adapters.registry import NoAdapterError, get_default_registry
    from src.adapters.wizard_configured import (
        default_helper_columns,
        default_respondent_id_column,
        detect_question_id_pattern,
        detect_sub_column_separator,
    )

    probe: dict[str, Any] = {
        "needs_wizard": False,
        "scores": [],
        "sheet_names": [],
        "raw_columns": [],
        "raw_data_sheet_name": "",
        "data_map_sheet_name": "",
        "respondent_id_column": "respondent_id",
        "question_id_pattern": r"^Q\d+",
        "sub_column_separator": "none",
        "helper_columns": [],
        "parse_failure_note": "",
    }
    xlsx_files = [file for file in uploaded_files if str(file.name).lower().endswith(".xlsx")]
    if not xlsx_files:
        return probe

    workbook_file = _select_datamap_workbook_upload(xlsx_files)
    content = _read_uploaded_bytes(workbook_file)
    workbook = load_workbook(stdlib_io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        raw_sheet = _default_raw_sheet_name(sheet_names)
        map_sheet = _default_datamap_sheet_name(sheet_names)
        raw_df = _read_probe_raw_dataframe(content, raw_sheet)
        scores_raw_df = raw_df if raw_df is not None else tuple()
        router = get_default_registry()
        needs_wizard, scores = router.needs_wizard(workbook, scores_raw_df)
        parse_failure_note = ""
        if not needs_wizard:
            try:
                router.parse(workbook, scores_raw_df)
            except NoAdapterError as exc:
                needs_wizard = True
                parse_failure_note = str(exc)
        raw_columns = [str(column) for column in getattr(raw_df, "columns", [])]
        data_map_values = _sample_sheet_first_column(workbook, map_sheet)
        probe.update(
            {
                "needs_wizard": needs_wizard,
                "scores": scores,
                "sheet_names": sheet_names,
                "raw_columns": raw_columns,
                "raw_data_sheet_name": raw_sheet,
                "data_map_sheet_name": map_sheet,
                "respondent_id_column": default_respondent_id_column(raw_columns),
                "question_id_pattern": detect_question_id_pattern(data_map_values),
                "sub_column_separator": detect_sub_column_separator(raw_columns[:100]),
                "helper_columns": list(default_helper_columns(raw_columns)),
                "parse_failure_note": parse_failure_note,
            }
        )
    finally:
        workbook.close()
    return probe


def _read_probe_raw_dataframe(content: bytes, sheet_name: str) -> pd.DataFrame | None:
    import io as stdlib_io

    try:
        return pd.read_excel(stdlib_io.BytesIO(content), sheet_name=sheet_name, dtype=str)
    except Exception:
        return None


def _select_datamap_workbook_upload(xlsx_files: list[Any]) -> Any:
    for file in xlsx_files:
        lowered = str(file.name).lower()
        if any(token in lowered for token in ("map", "datamap", "question", "schema")):
            return file
    return xlsx_files[0]


def _default_raw_sheet_name(sheet_names: list[str]) -> str:
    ignored = ("datamap", "data map", "map", "index", "instructions", "questions", "schema")
    for sheet_name in sheet_names:
        lowered = _re.sub(r"[\s_]+", " ", sheet_name.strip().lower())
        if not any(token in lowered for token in ignored):
            return sheet_name
    return sheet_names[0] if sheet_names else ""


def _default_datamap_sheet_name(sheet_names: list[str]) -> str:
    preferred = ("data map", "datamap", "questions", "schema")
    for sheet_name in sheet_names:
        lowered = _re.sub(r"[\s_]+", " ", sheet_name.strip().lower())
        if any(token in lowered for token in preferred):
            return sheet_name
    return sheet_names[1] if len(sheet_names) > 1 else (sheet_names[0] if sheet_names else "")


def _sample_sheet_first_column(workbook: Any, sheet_name: str) -> list[str]:
    if sheet_name not in getattr(workbook, "sheetnames", []):
        return []
    worksheet = workbook[sheet_name]
    values: list[str] = []
    for row in worksheet.iter_rows(min_row=1, max_row=30, values_only=True):
        if row and row[0] is not None:
            values.append(str(row[0]))
    return values


def _maybe_start_format_wizard(uploaded_files: list[Any]) -> bool:
    """Render the survey format wizard when auto-detection confidence is low."""

    app = _require_streamlit()
    if not uploaded_files or app.session_state.get("wizard_config_complete"):
        return False

    signature = _uploaded_files_signature(uploaded_files)
    if app.session_state.get("wizard_upload_signature") != signature:
        app.session_state["wizard_upload_signature"] = signature
        app.session_state["wizard_active"] = False
        app.session_state["wizard_step"] = 1
        app.session_state["wizard_config"] = {}
        app.session_state["wizard_config_complete"] = False
        app.session_state["wizard_preview"] = {}
        app.session_state["wizard_diagnostic"] = None
        try:
            probe = _probe_format_wizard(uploaded_files)
            app.session_state["wizard_detected_scores"] = probe.get("scores", [])
            app.session_state["wizard_parse_failure_note"] = probe.get("parse_failure_note", "")
            app.session_state["wizard_sheet_names"] = probe.get("sheet_names", [])
            app.session_state["wizard_raw_columns"] = probe.get("raw_columns", [])
            if probe.get("needs_wizard"):
                app.session_state["wizard_active"] = True
                app.session_state["wizard_config"] = _wizard_config_from_probe(probe)
        except Exception as exc:
            app.session_state["wizard_active"] = True
            app.session_state["wizard_detected_scores"] = []
            app.session_state["wizard_parse_failure_note"] = str(exc)
            app.session_state["wizard_config"] = _default_wizard_config()
            app.session_state["wizard_diagnostic"] = _build_wizard_diagnostic_panel_text(
                app.session_state["wizard_config"],
                str(exc),
                raw_columns=[],
                first_column_values=[],
            )

    if app.session_state.get("wizard_active"):
        _render_format_wizard(uploaded_files)
        return not app.session_state.get("wizard_config_complete")
    return False


def _wizard_config_from_probe(probe: dict[str, Any]) -> dict[str, Any]:
    return {
        "raw_data_sheet_name": probe.get("raw_data_sheet_name") or "",
        "data_map_sheet_name": probe.get("data_map_sheet_name") or "",
        "respondent_id_column": probe.get("respondent_id_column") or "respondent_id",
        "question_id_pattern": probe.get("question_id_pattern") or r"^Q\d+",
        "sub_column_separator": probe.get("sub_column_separator") or "none",
        "option_code_position": "column_b",
        "section_prefixes": ("Q",),
        "config_name": None,
        "helper_columns": tuple(probe.get("helper_columns") or []),
    }


def _default_wizard_config() -> dict[str, Any]:
    return {
        "raw_data_sheet_name": "",
        "data_map_sheet_name": "",
        "respondent_id_column": "respondent_id",
        "question_id_pattern": r"^Q\d+",
        "sub_column_separator": "none",
        "option_code_position": "column_b",
        "section_prefixes": ("Q",),
        "config_name": None,
        "helper_columns": tuple(),
    }


def _render_format_wizard(uploaded_files: list[Any]) -> None:
    app = _require_streamlit()
    config = dict(app.session_state.get("wizard_config") or _default_wizard_config())
    step = int(app.session_state.get("wizard_step", 1))
    step = max(1, min(7, step))

    with app.container():
        app.subheader("Survey format wizard")
        app.caption(
            "We couldn't automatically detect this survey's format. "
            "Help us read it correctly."
        )
        _render_format_wizard_scores()
        app.progress(step / 7)
        app.caption(f"Step {step}/7: {FORMAT_WIZARD_STEPS[step - 1]}")

        with app.expander("Previous answers", expanded=False):
            app.json(config)

        if step == 1:
            _render_format_wizard_step_1(config)
        elif step == 2:
            _render_format_wizard_step_2(config)
        elif step == 3:
            _render_format_wizard_step_3(config)
        elif step == 4:
            _render_format_wizard_step_4(config)
        elif step == 5:
            _render_format_wizard_step_5(config)
        elif step == 6:
            _render_format_wizard_step_6(config)
        elif step == 7:
            _render_format_wizard_step_7(config)
            _render_format_wizard_preview(uploaded_files, config)

        app.session_state["wizard_config"] = config
        back, next_col = app.columns(2)
        if back.button("Back", disabled=step <= 1, key="format_wizard_back"):
            app.session_state["wizard_step"] = max(1, step - 1)
            _wizard_rerun(app)
        if next_col.button("Next", type="primary", disabled=step >= 7, key="format_wizard_next"):
            app.session_state["wizard_step"] = min(7, step + 1)
            _wizard_rerun(app)


def _render_format_wizard_scores() -> None:
    app = _require_streamlit()
    scores = app.session_state.get("wizard_detected_scores") or []
    app.markdown("Auto-detection results:")
    if not scores:
        app.caption("No registered adapter produced a confident match. [Wizard handling required]")
        return
    for adapter_name, confidence, reason in scores:
        app.caption(f"{adapter_name}: {confidence:.2f} ({reason})")
    parse_failure_note = app.session_state.get("wizard_parse_failure_note")
    if parse_failure_note:
        app.caption(
            "BUT parse produced too few questions — adapter could not actually "
            f"read this format. {parse_failure_note}"
        )
    app.caption("[Wizard handling required]")


def _render_format_wizard_step_1(config: dict[str, Any]) -> None:
    app = _require_streamlit()
    sheet_names = _wizard_sheet_names()
    if sheet_names:
        raw_index = _safe_index(sheet_names, config.get("raw_data_sheet_name"))
        map_index = _safe_index(sheet_names, config.get("data_map_sheet_name"))
        config["raw_data_sheet_name"] = app.selectbox(
            "Which sheet contains the raw survey responses?",
            sheet_names,
            index=raw_index,
            key="wizard_raw_data_sheet_name",
        )
        config["data_map_sheet_name"] = app.selectbox(
            "Which sheet contains the data map?",
            sheet_names,
            index=map_index,
            key="wizard_data_map_sheet_name",
        )
    else:
        config["raw_data_sheet_name"] = app.text_input(
            "Which sheet contains the raw survey responses?",
            value=str(config.get("raw_data_sheet_name") or ""),
            key="wizard_raw_data_sheet_name",
        )
        config["data_map_sheet_name"] = app.text_input(
            "Which sheet contains the data map?",
            value=str(config.get("data_map_sheet_name") or ""),
            key="wizard_data_map_sheet_name",
        )


def _render_format_wizard_step_2(config: dict[str, Any]) -> None:
    app = _require_streamlit()
    raw_columns = _wizard_raw_columns()
    if raw_columns:
        config["respondent_id_column"] = app.selectbox(
            "Which column in the raw data identifies each respondent?",
            raw_columns,
            index=_safe_index(raw_columns, config.get("respondent_id_column")),
            key="wizard_respondent_id_column",
        )
    else:
        config["respondent_id_column"] = app.text_input(
            "Which column in the raw data identifies each respondent?",
            value=str(config.get("respondent_id_column") or "respondent_id"),
            key="wizard_respondent_id_column",
        )


def _render_format_wizard_step_3(config: dict[str, Any]) -> None:
    app = _require_streamlit()
    labels = [
        "Q followed by number (Q1, Q2, Q3)",
        "Lowercase q followed by number (q1, q2, q3)",
        "Q with underscore (Q_1, Q_2)",
        "Question_N (Question_1, Question_2)",
        "Other (specify regex)",
    ]
    mapping = {
        labels[0]: r"^Q\d+",
        labels[1]: r"^q\d+",
        labels[2]: r"^Q_\d+",
        labels[3]: r"^Question_\d+",
    }
    current = str(config.get("question_id_pattern") or r"^Q\d+")
    reverse = {value: label for label, value in mapping.items()}
    choice = app.radio(
        "How are question IDs written in this survey?",
        labels,
        index=labels.index(reverse.get(current, labels[-1])),
        key="wizard_question_id_choice",
    )
    if choice == labels[-1]:
        config["question_id_pattern"] = app.text_input(
            "Question ID regex",
            value=current,
            placeholder=r"e.g. ^SQ\d+ or ^Item_\d+",
            key="wizard_question_id_pattern_custom",
        )
    else:
        config["question_id_pattern"] = mapping[choice]


def _render_format_wizard_step_4(config: dict[str, Any]) -> None:
    app = _require_streamlit()
    labels = [
        "Q1r1, Q1r2, ... (lowercase r separator)",
        "Q1s1, Q1s2, ... (lowercase s separator)",
        "Q1_1, Q1_2, ... (underscore separator)",
        "Q1.1, Q1.2, ... (dot separator)",
        "Q1: option label, Q1: another label (colon followed by label)",
        "Other (specify)",
        "No multi-part questions in this survey",
    ]
    mapping = {
        labels[0]: "r",
        labels[1]: "s",
        labels[2]: "_",
        labels[3]: r"\.",
        labels[4]: ":",
        labels[6]: "none",
    }
    current = str(config.get("sub_column_separator") or "none")
    reverse = {value: label for label, value in mapping.items()}
    choice = app.radio(
        "How are multi-part questions (grids, multi-select) named in the raw data?",
        labels,
        index=labels.index(reverse.get(current, labels[5])),
        key="wizard_sub_column_separator_choice",
    )
    if choice == labels[5]:
        example = app.text_input(
            "Example column name or separator regex",
            value=current if current not in {"r", "s", "_", r"\.", ":", "none"} else "",
            key="wizard_sub_column_separator_custom",
        )
        config["sub_column_separator"] = _infer_separator_from_example(example)
    else:
        config["sub_column_separator"] = mapping[choice]


def _render_format_wizard_step_5(config: dict[str, Any]) -> None:
    app = _require_streamlit()
    choices = {
        "Numbered in column A: 1, 2, 3 in column A; labels in column B": "column_b",
        "Same row as question: codes alongside the question": "same_row",
        "Below question header, indented: rows after question header are options": "indented_below",
        "Custom (describe)": "custom",
    }
    labels = list(choices)
    current = str(config.get("option_code_position") or "column_b")
    current_label = next((label for label, value in choices.items() if value == current), labels[0])
    selected = app.radio(
        "In the data map, where are option codes written?",
        labels,
        index=labels.index(current_label),
        key="wizard_option_code_position_choice",
    )
    config["option_code_position"] = choices[selected]
    if config["option_code_position"] == "custom":
        app.text_input("Describe the custom option-code layout", key="wizard_option_custom_description")


def _render_format_wizard_step_6(config: dict[str, Any]) -> None:
    app = _require_streamlit()
    enabled = app.checkbox(
        "This survey has demographic (D), screener (S), or other section-prefixed questions",
        value=bool(config.get("section_prefixes")),
        key="wizard_section_prefix_enabled",
    )
    if enabled:
        options = ["Q", "D", "S", "F", "I", "Custom"]
        selected = app.multiselect(
            "Which prefixes should the parser look for?",
            options,
            default=[value for value in config.get("section_prefixes", ("Q",)) if value in options],
            key="wizard_section_prefixes",
        )
        if "Custom" in selected:
            custom = app.text_input("Custom prefix", key="wizard_custom_section_prefix")
            selected = [value for value in selected if value != "Custom"] + ([custom] if custom else [])
        config["section_prefixes"] = tuple(value for value in selected if value)
    else:
        config["section_prefixes"] = ("Q",)


def _render_format_wizard_step_7(config: dict[str, Any]) -> None:
    app = _require_streamlit()
    raw_columns = _wizard_raw_columns()
    defaults = [column for column in config.get("helper_columns", []) if column in raw_columns]
    config["helper_columns"] = tuple(
        app.multiselect(
            "Columns to skip",
            raw_columns,
            default=defaults,
            key="wizard_helper_columns",
        )
        if raw_columns
        else []
    )
    config["config_name"] = app.text_input(
        "Format name",
        value=str(config.get("config_name") or ""),
        key="wizard_config_name",
    ) or None


def _render_format_wizard_preview(uploaded_files: list[Any], config: dict[str, Any]) -> None:
    app = _require_streamlit()
    app.markdown("Wizard configured. Detected:")
    if app.button("Refresh preview", key="wizard_refresh_preview"):
        app.session_state["wizard_preview"] = _build_wizard_preview(uploaded_files, config)
    preview = app.session_state.get("wizard_preview") or _build_wizard_preview(uploaded_files, config)
    app.session_state["wizard_preview"] = preview
    for line in preview.get("lines", []):
        app.caption(line)
    app.markdown("Question type breakdown:")
    for label, count in preview.get("breakdown", {}).items():
        app.caption(f"- {count} {label}")

    proceed, save, back = app.columns(3)
    if proceed.button("Proceed to analysis", type="primary", key="wizard_proceed_to_analysis"):
        app.session_state["wizard_config"] = config
        app.session_state["wizard_config_complete"] = True
        app.session_state["wizard_active"] = False
        app.session_state["wiz_category_assignments"] = None
        _wizard_rerun(app)
    if save.button("Save as draft adapter", key="wizard_save_draft_adapter"):
        path = _save_wizard_config_draft(config)
        app.success(f"Saved draft adapter to {path}")
    if back.button("Go back to wizard", key="wizard_go_back_to_wizard"):
        app.session_state["wizard_step"] = 1
        _wizard_rerun(app)

    if preview.get("diagnostic"):
        _render_wizard_diagnostic_panel(preview["diagnostic"])


def _build_wizard_preview(uploaded_files: list[Any], config: dict[str, Any]) -> dict[str, Any]:
    try:
        data_map, raw_df, _load_report = _load_survey_inputs_with_format_wizard(uploaded_files, config)
        questions = data_map.get("questions", [])
        matched_columns = {
            column
            for question in questions
            for column, _label in question.get("sub_columns", [])
        }
        matched_columns.update(
            question["canonical_id"]
            for question in questions
            if question["canonical_id"] in set(str(column) for column in raw_df.columns)
        )
        breakdown = _wizard_question_type_breakdown(questions)
        return {
            "lines": [
                f"- {len(questions)} questions from data map",
                f"- {len(matched_columns)} raw data columns matched to questions",
                f"- {len(config.get('helper_columns') or [])} columns skipped as helpers",
                f"- {len(raw_df)} respondents identified by {config.get('respondent_id_column')!r} column",
            ],
            "breakdown": breakdown,
            "diagnostic": None if questions else _build_wizard_diagnostic_panel_text(
                config,
                "0 questions parsed",
                raw_columns=[str(column) for column in raw_df.columns],
                first_column_values=[],
            ),
        }
    except Exception as exc:
        diagnostic = _build_wizard_diagnostic_panel_text(
            config,
            str(exc),
            raw_columns=_wizard_raw_columns(),
            first_column_values=[],
        )
        return {"lines": ["- Preview failed"], "breakdown": {}, "diagnostic": diagnostic}


def _wizard_question_type_breakdown(questions: list[dict[str, Any]]) -> dict[str, int]:
    breakdown = {
        "single-select": 0,
        "multi-select": 0,
        "grid-categorical-row": 0,
        "grid-rated": 0,
        "direct-numeric": 0,
    }
    for question in questions:
        if question.get("sub_columns") and question.get("options"):
            breakdown["grid-categorical-row"] += 1
        elif question.get("sub_columns"):
            breakdown["multi-select"] += 1
        elif question.get("options"):
            breakdown["single-select"] += 1
        else:
            breakdown["direct-numeric"] += 1
    return breakdown


def _build_wizard_diagnostic_panel_text(
    config: dict[str, Any],
    error: str,
    *,
    raw_columns: list[str],
    first_column_values: list[str],
) -> str:
    pattern = config.get("question_id_pattern", r"^Q\d+")
    separator = config.get("sub_column_separator", "none")
    samples = first_column_values or ["Item_1: What is your age?", "Item_2: What is your gender?"]
    sample_columns = raw_columns[:4] or ["Item_1", "Item_2", "Item_2_sub_1", "Item_2_sub_2"]
    return "\n".join(
        [
            "Could not parse this survey with the wizard configuration. Here's what happened:",
            f"Error: {error}",
            f"Data map sheet: {config.get('data_map_sheet_name')!r}",
            f"- 0 rows matching question ID pattern: {pattern}",
            "- Rows with content in column A were present.",
            "Tip: try a different question ID pattern.",
            "The first few rows of column A contained:",
            *[f"- {value!r}" for value in samples[:5]],
            "Suggested pattern: ^Item_\\d+",
            f"Raw data sheet: {config.get('raw_data_sheet_name')!r}",
            f"Sub-column pattern: {separator} - no matches found",
            "Tip: try a different sub-column separator.",
            "Sample columns:",
            *[f"- {value!r}" for value in sample_columns],
            "Suggested separator: '_sub_'",
            f"Helper columns skipped: {', '.join(config.get('helper_columns') or []) or 'none'}",
        ]
    )


def _render_wizard_diagnostic_panel(diagnostic: str) -> None:
    app = _require_streamlit()
    app.error("Could not parse this survey with the wizard configuration.")
    with app.expander("Wizard diagnostic", expanded=True):
        app.write(diagnostic)
        adjust, reset = app.columns(2)
        if adjust.button("Adjust wizard and retry", key="wizard_adjust_retry"):
            app.session_state["wizard_active"] = True
            app.session_state["wizard_config_complete"] = False
            _wizard_rerun(app)
        if reset.button("Start over", key="wizard_start_over"):
            app.session_state["wizard_active"] = False
            app.session_state["wizard_config"] = {}
            app.session_state["wizard_config_complete"] = False
            _wizard_rerun(app)


def _wizard_sheet_names() -> list[str]:
    app = _require_streamlit()
    sheet_names = app.session_state.get("wizard_sheet_names")
    if sheet_names:
        return list(sheet_names)
    config = app.session_state.get("wizard_config") or {}
    names = [config.get("raw_data_sheet_name"), config.get("data_map_sheet_name")]
    return [str(name) for name in names if name]


def _wizard_raw_columns() -> list[str]:
    app = _require_streamlit()
    raw_columns = app.session_state.get("wizard_raw_columns")
    if raw_columns:
        return list(raw_columns)
    preview = app.session_state.get("wizard_preview") or {}
    raw_columns = preview.get("raw_columns")
    if raw_columns:
        return list(raw_columns)
    config = app.session_state.get("wizard_config") or {}
    helper_columns = list(config.get("helper_columns") or [])
    return helper_columns


def _safe_index(options: list[str], value: Any) -> int:
    try:
        return options.index(str(value))
    except ValueError:
        return 0


def _infer_separator_from_example(example: str) -> str:
    value = str(example or "").strip()
    if not value:
        return "none"
    if ":" in value:
        return ":"
    if "_sub_" in value:
        return "_sub_"
    if "_" in value:
        return "_"
    if "." in value:
        return r"\."
    return value


def _load_survey_inputs_for_current_format(uploaded_files: list[Any]):
    app = _require_streamlit()
    if app.session_state.get("wizard_config_complete"):
        return _load_survey_inputs_with_format_wizard(
            uploaded_files,
            dict(app.session_state.get("wizard_config") or {}),
        )
    from src.io import load_survey_inputs

    return load_survey_inputs(uploaded_files)


def _load_survey_inputs_with_format_wizard(
    uploaded_files: list[Any],
    config_dict: dict[str, Any],
):
    import io as stdlib_io
    from openpyxl import load_workbook
    from src.adapters.wizard_configured import WizardConfig, WizardConfiguredAdapter
    from src.io import (
        _detect_scenario,
        _normalise_dataframe,
        _read_upload,
        _safe_unlink,
        _write_bytes_to_temp,
        _write_upload_to_temp,
    )
    from src.models import LoadReport
    from src.raw_decoder import decode_raw_data

    cfg = WizardConfig(**config_dict)
    scenario = _detect_scenario(uploaded_files)
    temp_paths: list[str] = []
    try:
        if scenario == "B_combined_xlsx":
            xlsx_file = next(file for file in uploaded_files if str(file.name).lower().endswith(".xlsx"))
            content = _read_upload(xlsx_file)
            raw_path = _write_bytes_to_temp(content, ".xlsx")
            temp_paths.append(raw_path)
            workbook = load_workbook(stdlib_io.BytesIO(content), read_only=True, data_only=True)
            try:
                raw_probe_df = pd.read_excel(
                    stdlib_io.BytesIO(content),
                    sheet_name=cfg.raw_data_sheet_name,
                    dtype=str,
                )
                data_map = WizardConfiguredAdapter(cfg).parse(workbook, raw_probe_df)
            finally:
                workbook.close()
            raw_df, _quality_report = decode_raw_data(raw_path, data_map)
        else:
            datamap_file, raw_file = _wizard_identify_separate_uploads(uploaded_files, cfg)
            datamap_path = _write_upload_to_temp(datamap_file, Path(datamap_file.name).suffix)
            raw_path = _write_upload_to_temp(raw_file, Path(raw_file.name).suffix)
            temp_paths.extend([datamap_path, raw_path])
            workbook = load_workbook(datamap_path, read_only=True, data_only=True)
            try:
                raw_probe_df = (
                    pd.read_csv(raw_path, dtype=str)
                    if raw_path.lower().endswith(".csv")
                    else pd.read_excel(raw_path, sheet_name=cfg.raw_data_sheet_name, dtype=str)
                )
                data_map = WizardConfiguredAdapter(cfg).parse(workbook, raw_probe_df)
            finally:
                workbook.close()
            raw_df, _quality_report = decode_raw_data(raw_path, data_map)
        raw_df = _normalise_dataframe(raw_df)
        return data_map, raw_df, LoadReport(
            scenario=f"{scenario}_wizard_configured",
            raw_data_source=f"sheet:{cfg.raw_data_sheet_name}",
            datamap_source=f"sheet:{cfg.data_map_sheet_name}",
            raw_rows=int(len(raw_df)),
            raw_columns=int(len(raw_df.columns)),
            questions_parsed=int(len(data_map["questions"])),
            parser_warnings=data_map["parser_warnings"],
            detection_notes=["Wizard configured adapter used for survey parsing."],
        )
    finally:
        for path in temp_paths:
            _safe_unlink(path)


def _wizard_identify_separate_uploads(uploaded_files: list[Any], cfg: Any) -> tuple[Any, Any]:
    xlsx_files = [file for file in uploaded_files if str(file.name).lower().endswith(".xlsx")]
    non_map_files = [
        file
        for file in uploaded_files
        if str(file.name).lower().endswith((".csv", ".xlsx"))
    ]
    datamap_file = _select_datamap_workbook_upload(xlsx_files)
    for file in non_map_files:
        if file is not datamap_file:
            return datamap_file, file
    raise ValueError(f"raw data file for sheet {cfg.raw_data_sheet_name!r} not found")


def _section_upload() -> None:
    app = _require_streamlit()
    app.markdown('<div id="section-upload"></div>', unsafe_allow_html=True)
    _section_header("1", "UPLOAD YOUR SURVEY", anchor="section-1", meta="CSV \u00b7 XLSX \u00b7 DOCX")
    app.markdown(
        "<div style='padding:14px 18px;background:#FFF6F6;border:1px solid #F6D5D5;"
        "border-left:4px solid #CC0000;border-radius:10px;margin-bottom:14px;"
        "font-family:Arial,sans-serif;'>"
        "<div style='font-weight:700;font-size:14px;color:#1A1A1A;margin-bottom:4px;'>"
        "Step 1 \u2014 Upload raw data and data map</div>"
        "<div style='font-size:13px;color:#444;line-height:1.55;'>"
        "Upload two things: your <b>raw data</b> (the responses, .csv or .xlsx) and your "
        "<b>data map</b> (the question/answer codebook, .xlsx). You can also drop a single "
        "combined Excel, and the tool will detect the format automatically.</div></div>",
        unsafe_allow_html=True,
    )

    uploaded_files = app.file_uploader(
        "Upload your survey files",
        type=["csv", "xlsx", "docx"],
        accept_multiple_files=True,
        key="unified_upload",
        help=(
            "**Three scenarios are supported:**\n\n"
            "1. Raw data (.csv or .xlsx) + data map (.xlsx) \u2014 two "
            "separate files\n\n"
            "2. Combined .xlsx with raw data and data map on separate "
            "sheets\n\n"
            "3. Raw data (.csv or .xlsx) + Word survey document (.docx) "
            "\u2014 tool auto-builds the data map from the Word doc"
        ),
    )
    _render_saved_wizard_config_picker()
    embed_inputs = app.checkbox(
        "Include original raw data and data map as sheets in exported workbooks",
        value=False,
        key="wizard_embed_input_files",
        help=(
            "Adds two new sheets to every exported workbook: Raw Data (Input) "
            "and Data Map (Input). Useful for audit and reproducibility. "
            "Increases workbook size."
        ),
    )

    detected_scenario: str | None = None
    docx_only = False
    if uploaded_files:
        names = [f.name for f in uploaded_files]
        app.caption(f"Uploaded: {', '.join(names)}")
        try:
            from src.io import _detect_scenario
            detected_scenario = _detect_scenario(uploaded_files)
        except Exception:
            detected_scenario = None
        scenario_labels = {
            "A_separate_files":
                "\u2713 Two-file input detected (raw data + data map)",
            "B_combined_xlsx":
                "\u2713 Combined Excel detected (data + map on separate sheets)",
            "C_word_datamap":
                "\u2713 Word survey document detected (data map will be parsed "
                "from Word file)",
        }
        if detected_scenario in scenario_labels:
            app.info(scenario_labels[detected_scenario])

        # Scenario C without raw data: graceful handling, no exception.
        docx_files = [f for f in uploaded_files if f.name.lower().endswith(".docx")]
        non_docx = [f for f in uploaded_files if not f.name.lower().endswith(".docx")]
        docx_only = bool(docx_files) and not non_docx

    # Update legacy status labels for the sidebar "files uploaded" indicator.
    if uploaded_files:
        app.session_state["raw_data_path_label"] = ", ".join(
            f.name for f in uploaded_files
        )
        app.session_state["datamap_path_label"] = ", ".join(
            f.name for f in uploaded_files
        )

    if docx_only:
        app.warning(
            "Data map will be parsed from the Word document, but no raw "
            "data file was uploaded. Please also upload your raw data "
            "(.csv or .xlsx) to run the analysis."
        )

    format_wizard_blocking = _maybe_start_format_wizard(uploaded_files) if uploaded_files else False

    # -- Setup wizard (appears after upload, before Run Analysis) --
    if (
        uploaded_files
        and not format_wizard_blocking
        and not app.session_state.get("wiz_complete")
        and not app.session_state.get("run_complete")
    ):
        _render_setup_wizard(uploaded_files)

    ready = bool(uploaded_files) and not docx_only and not format_wizard_blocking
    centre_left, centre_mid, centre_right = app.columns([2, 3, 2])
    with centre_mid:
        generate_ai_enhancements = app.checkbox(
            "Generate AI themes & labels (adds ~30s)",
            value=not bool(app.session_state.get("skip_ai_enhancements", False)),
            key="generate_ai_enhancements",
            disabled=not ready,
            help="Turn off to use deterministic themes and labels without AI calls.",
        )
        app.session_state["skip_ai_enhancements"] = not generate_ai_enhancements
        run_clicked = app.button(
            "Run analysis",
            type="primary",
            disabled=not ready,
            use_container_width=True,
        )

    if app.session_state.pop("_wizard_run_requested", False):
        run_clicked = True

    if run_clicked:
        if uploaded_files and not app.session_state.get("wiz_complete") and app.session_state.get("wiz_category_assignments") is not None:
            _wizard_apply_overrides()
            app.session_state["wiz_complete"] = True
        import logging
        logging.warning("RUN_CLICKED uploaded=%s docx_only=%s", bool(uploaded_files), docx_only)
        app.session_state["run_complete"] = False
        try:
            with app.status("Starting analysis...", expanded=True) as status:
                data_map, raw_df, load_report = _run_with_status_heartbeat(
                    status,
                    start_label="Stage 1/5: Loading uploaded files...",
                    heartbeat_labels=(
                        "Stage 1/5: Reading uploaded files...",
                        "Stage 1/5: Parsing data map...",
                        "Stage 1/5: Decoding raw data...",
                        "Stage 1/5: Normalizing decoded data...",
                    ),
                    work=lambda: _load_survey_inputs_for_current_format(uploaded_files),
                    complete_label="Stage 1/5: Uploaded files loaded.",
                )
                app.session_state["data_map"] = data_map
                app.session_state["load_report"] = load_report
                app.session_state["input_file_embed_sources"] = (
                    _persist_uploaded_input_file_sources(uploaded_files, load_report)
                    if embed_inputs
                    else None
                )
                for note in load_report.detection_notes:
                    app.caption(f"\u2139\ufe0f {note}")
                _run_pipeline(data_map, raw_df, load_report, status)
        except Exception as exc:  # noqa: BLE001
            app.session_state["run_complete"] = False
            app.error(f"{type(exc).__name__}: {exc}")
            with app.expander("Show full traceback"):
                app.code(traceback.format_exc())
        if app.session_state.get("run_complete"):
            # Force one rerun so the sidebar navigator (rendered earlier in
            # main()) repaints with the freshly produced single cuts instead
            # of waiting for the next user interaction.
            app.rerun()

    if app.session_state["run_complete"]:
        schema = app.session_state["schema"]
        results = app.session_state["results"]
        skips = app.session_state["skips"]
        log = app.session_state["log"]
        _render_stat_tiles([
            ("users",     int(getattr(schema, "total_respondents", 0) or 0), "Respondents"),
            ("clipboard", len(schema.questions), "Total questions"),
            ("chart",     len(results), "Single cuts"),
        ])
        # #3: directly below the counts, show the global filter as a SECTION
        # (not a button), then the first question's single cut + cross-cut.
        app.divider()
        _section_global_filter()
        app.divider()
        _section_single_cuts()


# ---------------------------------------------------------------------------
# Section 2 — Global filter
# ---------------------------------------------------------------------------


def _apply_global_filter_action(rows: list[tuple[str | None, Any]]) -> None:
    """Queue the global filter for processing on the next rerun.

    The actual ``apply_global_filter`` + single-cut recompute runs in
    ``_drain_pending_actions()`` at the top of ``main()``. Storing only the
    spec list here (rather than computing inline) eliminates the double-click
    bug: the click handler returns immediately, ``st.rerun()`` fires, and the
    recompute happens before any section renders so sections 2-5 all see the
    filtered state in a single round trip.
    """
    app = _require_streamlit()
    schema = app.session_state.get("schema")
    decoded_df = app.session_state.get("decoded_df")
    spec_list = []
    for q, v in rows:
        vals = _normalize_value_list(v)
        if q is None or not vals:
            continue
        spec_list.extend(_build_filter_specs(schema, decoded_df, q, vals))
    app.session_state["pending_global_filter"] = {"filter_specs": spec_list}
    app.session_state["global_filter_error"] = None
    app.rerun()


def _clear_global_filter_action() -> None:
    from src.models import GlobalFilterState

    app = _require_streamlit()
    app.session_state["global_filter_state"] = GlobalFilterState()
    app.session_state["global_filter_stats"] = None
    app.session_state["global_filter_rows"] = []
    app.session_state["active_df"] = app.session_state["decoded_df"]
    _invalidate_stage_c_state()
    _rerun_single_cuts_on_active_df()
    _INSIGHT_CACHE.clear()
    app.rerun()


def _section_global_filter() -> None:
    app = _require_streamlit()
    app.markdown('<div id="section-filter"></div>', unsafe_allow_html=True)
    _section_header("3", "Apply filters", anchor="section-filter", meta="Across all cuts")

    if not app.session_state["run_complete"]:
        app.caption("Available after you run the analysis.")
        return

    app.markdown(TOOLTIP_GLOBAL_FILTER)

    schema = app.session_state["schema"]
    gf_state = app.session_state.get("global_filter_state")
    stats = app.session_state.get("global_filter_stats") or {}

    if gf_state is not None and gf_state.is_active():
        rows_before = stats.get("rows_before", 0)
        rows_after = stats.get("rows_after", 0)
        app.info(
            f"\U0001F535 Global filter active: {gf_state.description()}. "
            f"All analyses below restricted to {rows_after:,} of "
            f"{rows_before:,} respondents."
        )
    else:
        app.caption(
            f"No global filter. All analyses run on the full "
            f"{schema.total_respondents:,} respondents."
        )

    eligible = _eligible_filter_questions()
    if not eligible:
        app.warning("No categorical questions available to use as global filters.")
        return

    filter_options_by_id = {q.question_id: q for q in eligible}
    question_options: list[tuple[str, Any]] = [("None", None)] + [
        (
            q.label,
            q.question_id,
        )
        for q in eligible
    ]

    rows: list[tuple[str | None, Any]] = list(
        app.session_state.get("global_filter_rows", [])
    )
    if not rows:
        rows = [(None, None)]

    new_rows: list[tuple[str | None, Any]] = []
    delete_index: int | None = None
    for i, (q_id, val) in enumerate(rows):
        cols = app.columns([4, 4, 1])
        with cols[0]:
            q_pick = app.selectbox(
                "Filter question",
                options=question_options,
                format_func=lambda x: x[0],
                index=_find_q_index(q_id, question_options),
                key=f"gf_q_{i}",
                label_visibility="visible" if i == 0 else "collapsed",
            )
        picked_q_id = q_pick[1]
        with cols[1]:
            if picked_q_id is not None:
                q_spec = filter_options_by_id[picked_q_id]
                value_codes = [choice.value for choice in q_spec.values]
                value_labels = {choice.value: choice.label for choice in q_spec.values}
                prior = _normalize_value_list(val)
                default_codes = [v for v in prior if v in value_codes]
                widget_key = f"gf_v_{i}_{picked_q_id}"
                if widget_key not in app.session_state:
                    app.session_state[widget_key] = default_codes
                else:
                    app.session_state[widget_key] = [
                        v for v in app.session_state[widget_key]
                        if v in value_codes
                    ]
                v_pick = app.multiselect(
                    "Values (select one or more)",
                    options=value_codes,
                    format_func=lambda v: value_labels.get(v, str(v)),
                    key=widget_key,
                    label_visibility="visible" if i == 0 else "collapsed",
                    placeholder="Pick value(s)",
                )
                new_val = list(v_pick)
            else:
                app.multiselect(
                    "Values",
                    options=[],
                    disabled=True,
                    key=f"gf_v_disabled_{i}",
                    label_visibility="visible" if i == 0 else "collapsed",
                    placeholder="Pick a question first",
                )
                new_val = []
        with cols[2]:
            if i == 0:
                app.markdown("&nbsp;", unsafe_allow_html=True)
            if app.button("\u2715", key=f"gf_del_{i}", help="Remove this filter"):
                delete_index = i
        new_rows.append((picked_q_id, new_val))

    app.session_state["global_filter_rows"] = new_rows

    if delete_index is not None:
        new_rows.pop(delete_index)
        app.session_state["global_filter_rows"] = new_rows
        _purge_widget_keys("gf_q_", "gf_v_")
        app.rerun()

    btn_cols = app.columns([2, 2, 2, 4])
    with btn_cols[0]:
        if app.button("+ Add another filter", key="gf_add"):
            app.session_state["global_filter_rows"] = new_rows + [(None, [])]
            _purge_widget_keys("gf_q_", "gf_v_")
            app.rerun()
    with btn_cols[1]:
        complete_rows = [
            (q, v) for q, v in new_rows
            if q is not None and _normalize_value_list(v)
        ]
        if app.button(
            "Apply global filter",
            key="gf_apply",
            type="primary",
            disabled=not complete_rows,
        ):
            _apply_global_filter_action(new_rows)
    with btn_cols[2]:
        if gf_state is not None and gf_state.is_active():
            if app.button("Clear global filter", key="gf_clear"):
                _clear_global_filter_action()


# ---------------------------------------------------------------------------
# Section 2.5 — Survey type classification + outcome variable selector
# ---------------------------------------------------------------------------


def _section_survey_classification(show_toggle: bool = True) -> None:
    """Detect survey type after first run and let the user pick an outcome variable.

    Detection is deterministic and cached in ``session_state.survey_type_result``;
    it only runs once per uploaded survey. The outcome variable defaults to the
    detector's top suggestion but can be overridden via the dropdown below.
    """
    app = _require_streamlit()
    if not app.session_state.get("run_complete"):
        return

    data_map = app.session_state.get("data_map")
    if data_map is None:
        return

    toggle_key = "show_winners_vs_laggards"
    if toggle_key not in app.session_state:
        app.session_state[toggle_key] = False
    if show_toggle:
        arrow = "v" if app.session_state[toggle_key] else ">"
        if app.button(
            f"{arrow}  Outcome Segmentation",
            key="toggle_winners_vs_laggards",
            use_container_width=True,
        ):
            app.session_state[toggle_key] = not app.session_state[toggle_key]
            app.rerun()
        if not app.session_state["show_winners_vs_laggards"]:
            return
    else:
        app.session_state[toggle_key] = True

    if app.toggle("Open Advanced", key="outcome_seg_advanced",
                  help="Multi-metric group builder with balance and composition."):
        from src.advanced_segmentation_ui import render_advanced_outcome_segmentation
        render_advanced_outcome_segmentation()
        return

    app.markdown(
        "<div style='padding:14px 18px;background:#F7F8FA;border:1px solid #E8EAED;"
        "border-radius:10px;margin:6px 0 14px;font-family:Arial,sans-serif;'>"
        "<div style='font-weight:700;font-size:14px;color:#1A1A1A;margin-bottom:4px;'>"
        "What this does</div>"
        "<div style='font-size:13px;color:#444;line-height:1.6;'>"
        "Split your respondents into <b>two groups</b> based on any outcome you pick "
        "(for example, a high-scoring group vs. a lower-scoring group), then "
        "automatically find the questions where the two groups answer most "
        "differently. It's a fast way to see <i>what sets one group apart</i> \u2014 "
        "the groups can be anything your data defines.</div></div>",
        unsafe_allow_html=True,
    )

    if app.session_state.get("survey_type_result") is None:
        with app.spinner("Detecting survey type..."):
            from src.survey_type_detector import detect_survey_type

            try:
                survey_type_result = detect_survey_type(
                    schema=data_map,
                    decoded_df=app.session_state["decoded_df"],
                )
            except Exception as exc:  # noqa: BLE001 - never block the UI on detection.
                app.warning(f"Survey type detection failed: {type(exc).__name__}: {exc}")
                return

            # Penalise imbalanced outcome candidates so they sort lower in
            # the dropdown and don't get auto-selected as the primary
            # outcome. The detector itself is data-distribution-blind.
            from dataclasses import replace

            decoded_df = app.session_state["decoded_df"]

            def _imbalance_penalty(opt) -> float:
                if opt.question_id not in decoded_df.columns:
                    return 0.0
                col = decoded_df[opt.question_id].dropna()
                if len(col) == 0:
                    return 0.0
                top_pct = col.value_counts(normalize=True).iloc[0]
                if top_pct >= 0.95:
                    return 0.4
                if top_pct >= 0.90:
                    return 0.2
                return 0.0

            # Filter out demographic/metadata questions when picking the
            # AUTO-SELECTED default. The full all_eligible_questions list
            # (including demographics) is preserved in the dropdown so the
            # user can still manually choose one if they want.
            classified_schema = app.session_state.get("schema")

            def _is_demographic_or_metadata(opt) -> bool:
                if classified_schema is None:
                    return False
                spec = classified_schema.get_question(opt.question_id)
                if spec is None:
                    return False
                return spec.is_demographic or spec.question_type.value in (
                    "METADATA_OR_ID",
                    "OPEN_TEXT",
                    "UNKNOWN",
                )

            eligible_for_outcome = [
                opt
                for opt in survey_type_result.all_eligible_questions
                if not _is_demographic_or_metadata(opt)
            ]

            rebalanced = sorted(
                eligible_for_outcome,
                key=lambda opt: (
                    -(opt.relevance_score - _imbalance_penalty(opt)),
                    _imbalance_penalty(opt),
                    opt.question_id,
                ),
            )
            best_id = rebalanced[0].question_id if rebalanced else None
            survey_type_result = replace(
                survey_type_result,
                outcome_question_id=best_id,
            )

            app.session_state["survey_type_result"] = survey_type_result
            if app.session_state.get("outcome_variable_id") is None:
                app.session_state["outcome_variable_id"] = (
                    survey_type_result.outcome_question_id
                )

    result = app.session_state["survey_type_result"]
    app.markdown("### Step 1 \u2014 Pick your outcome")
    app.caption(
        "The outcome is the single measure that defines 'success' for this analysis "
        "\u2014 for example revenue growth %, overall satisfaction, or an NPS score. "
        "The tool suggests one automatically; change it if a different measure better "
        "captures what you care about."
    )
    manual_cohort = app.session_state.get("manual_cohort_input")
    if manual_cohort is not None and "seg_mode_radio" not in app.session_state:
        app.session_state["seg_mode_radio"] = "manual_uuid"
    if app.session_state.get("seg_mode_radio") == "manual_uuid" and manual_cohort is not None:
        app.success(
            "Groups detected from your uploaded file: "
            f"{len(manual_cohort.winner_uuids)} in the top group, "
            f"{len(manual_cohort.laggard_uuids)} in the comparison group"
        )
        app.caption("Manual UID mode hides the outcome variable picker.")
        _render_segment_definition_ui()
        _render_segmentation_results()
        app.divider()
        return
    app.info(
        "Select the primary outcome variable for segmentation analysis. "
        "Auto-selected based on survey type; you can override."
    )

    all_options = result.all_eligible_questions
    if not all_options:
        app.warning(
            "No measurable questions detected. Segmentation analysis will be unavailable."
        )
        _render_segment_definition_ui()
        _render_segmentation_results()
        app.divider()
        return

    decoded_df_for_labels = app.session_state.get("decoded_df")

    def _outcome_label(opt) -> str:
        base = f"{opt.question_id}: {opt.question_text}"
        score = f"score: {opt.relevance_score:.2f}"
        if (
            decoded_df_for_labels is not None
            and opt.question_id in decoded_df_for_labels.columns
        ):
            col = decoded_df_for_labels[opt.question_id].dropna()
            if len(col) > 0:
                top_val_pct = col.value_counts(normalize=True).iloc[0]
                if top_val_pct >= 0.90:
                    return (
                        f"{base} ({score} \u2014 "
                        f"{top_val_pct:.0%} single value, likely too imbalanced)"
                    )
        return f"{base} ({score})"

    option_labels: list[str] = []
    option_ids: list[str | None] = []
    for opt in all_options:
        option_labels.append(_outcome_label(opt))
        option_ids.append(opt.question_id)
    option_labels.append("None \u2014 no outcome variable")
    option_ids.append(None)

    current_outcome = app.session_state.get("outcome_variable_id")
    default_index = (
        option_ids.index(current_outcome) if current_outcome in option_ids else 0
    )

    selected_label = app.selectbox(
        "Outcome Variable",
        options=option_labels,
        index=default_index,
        help="Auto-selected based on survey type. Override if needed.",
        key="outcome_variable_selector",
    )
    selected_id = option_ids[option_labels.index(selected_label)]
    if selected_id != app.session_state.get("outcome_variable_id"):
        app.session_state["outcome_variable_id"] = selected_id
        # Outcome variable changed — invalidate stale Stage C state so the
        # results displayed always match the currently selected outcome.
        _invalidate_stage_c_state()
        app.rerun()

    if selected_id is not None:
        selected_opt = next(
            opt for opt in all_options if opt.question_id == selected_id
        )
        app.success(
            f"**Selected:** {selected_opt.question_id} \u2014 {selected_opt.reason}"
        )
    else:
        app.warning(
            "No outcome variable selected. Segmentation analysis will be "
            "unavailable in later stages."
        )

    if len(result.candidate_outcome_questions) > 1:
        with app.expander(
            f"Top {len(result.candidate_outcome_questions)} "
            "Outcome Candidates (auto-ranked)"
        ):
            for opt in result.candidate_outcome_questions:
                app.markdown(f"**{opt.question_id}** \u2014 {opt.question_text}")
                app.caption(
                    f"Relevance: {opt.relevance_score:.2f} | {opt.reason}"
                )
                app.markdown("")

    _render_segment_definition_ui()
    _render_segmentation_results()

    app.divider()


def _render_segment_definition_ui() -> None:
    """Stage C: define winner/loser segments and run outcome segmentation."""
    app = _require_streamlit()
    outcome_id = app.session_state.get("outcome_variable_id")

    schema = app.session_state.get("schema")
    if schema is None:
        return
    manual_cohort = app.session_state.get("manual_cohort_input")
    manual_available = manual_cohort is not None
    outcome_spec = schema.get_question(outcome_id) if outcome_id else None
    if outcome_id and outcome_spec is None and not manual_available:
        app.error(f"Outcome question {outcome_id} not found in schema.")
        return

    from src.models import SegmentDefinition
    from src.io import parse_manual_cohort_workbook

    app.markdown("---")
    app.markdown("### Step 2 \u2014 Define the top group")
    app.caption(
        "Decide who counts as the **top group** on that outcome. Hover the "
        "\u24d8 on the mode selector for what each option means."
    )
    if manual_available:
        app.success(
            "Groups detected from your uploaded file: "
            f"{len(manual_cohort.winner_uuids)} in the top group, "
            f"{len(manual_cohort.laggard_uuids)} in the comparison group"
        )
        if "seg_mode_radio" not in app.session_state:
            app.session_state["seg_mode_radio"] = "manual_uuid"

    _SEG_MODE_LABELS = {
        "manual_uuid": "Manual UID upload",
        "categorical": "Categorical",
        "numeric_threshold": "Numeric threshold",
        "quartile": "Quartile",
    }
    _SEG_MODE_HELP = (
        "How to define your two groups to compare:\n\n"
        "\u2022 Categorical \u2014 pick the answer(s) that define your top group "
        "(e.g. those who chose '20%+' on an outcome).\n\n"
        "\u2022 Numeric threshold \u2014 top group scored at/above a value "
        "(e.g. score \u2265 15).\n\n"
        "\u2022 Quartile \u2014 top 25% vs bottom 25% on a numeric outcome.\n\n"
        "\u2022 Manual UID upload \u2014 you provide the exact list of respondent "
        "IDs in each group."
    )
    seg_mode = app.radio(
        "Segmentation mode",
        options=["manual_uuid", "categorical", "numeric_threshold", "quartile"],
        index=0 if manual_available or outcome_spec is None else 1,
        format_func=lambda x: _SEG_MODE_LABELS[x],
        key="seg_mode_radio",
        horizontal=True,
        help=_SEG_MODE_HELP,
    )

    app.markdown("#### Define the comparison group")
    laggard_outcome_question_id = None
    laggard_outcome_sub_question_id = None
    override_laggard_outcome = False
    if seg_mode != "manual_uuid":
        override_laggard_outcome = app.checkbox(
            "Use a different question for the comparison group",
            value=False,
            key="laggard_override_checkbox",
        )
    if seg_mode != "manual_uuid" and override_laggard_outcome:
        laggard_questions = [
            question
            for question in schema.analysis_eligible_questions()
            if not question.is_demographic
        ]
        laggard_labels = [
            f"{question.canonical_id}: {question.question_text}"
            for question in laggard_questions
        ]
        if laggard_labels:
            selected_laggard_label = app.selectbox(
                "Comparison group question",
                options=laggard_labels,
                key="laggard_outcome_variable_selector",
            )
            laggard_question = laggard_questions[
                laggard_labels.index(selected_laggard_label)
            ]
            laggard_outcome_question_id = laggard_question.canonical_id
            if laggard_question.question_type.value == "GRID_RATED":
                sub_options = list(laggard_question.raw_columns)
                if sub_options:
                    laggard_outcome_sub_question_id = app.selectbox(
                        "Comparison group grid sub-question",
                        options=sub_options,
                        key="laggard_outcome_sub_question_selector",
                    )

    segment_definition = None

    if seg_mode == "manual_uuid":
        if manual_cohort is None:
            uploaded_manual = app.file_uploader(
                "Upload group definitions (.xlsx)",
                type=["xlsx"],
                key="manual_winners_laggards_upload",
            )
            if uploaded_manual is not None:
                try:
                    active_df = app.session_state["active_df"]
                    primary_id_column = schema.respondent_id_column
                    raw_ids_primary = active_df[primary_id_column].astype(str).str.strip()
                    uploaded_bytes = uploaded_manual.read()
                    manual_cohort = parse_manual_cohort_workbook(
                        uploaded_bytes,
                        valid_uuids=set(raw_ids_primary),
                        source="separate upload",
                        id_column=primary_id_column,
                    )
                    resolved_id_column = primary_id_column
                    if (
                        len(manual_cohort.winner_uuids) == 0
                        and len(manual_cohort.laggard_uuids) == 0
                        and len(manual_cohort.invalid_uuids) > 0
                    ):
                        candidate_columns = [
                            column
                            for column in active_df.columns
                            if str(column).lower() in ("uuid", "respondent_id", "id")
                            and str(column) != str(primary_id_column)
                        ]
                        for fallback_column in candidate_columns:
                            raw_ids_fallback = active_df[fallback_column].astype(str).str.strip()
                            fallback_cohort = parse_manual_cohort_workbook(
                                uploaded_bytes,
                                valid_uuids=set(raw_ids_fallback),
                                source="separate upload",
                                id_column=str(fallback_column),
                            )
                            if (
                                len(fallback_cohort.winner_uuids) > 0
                                or len(fallback_cohort.laggard_uuids) > 0
                            ):
                                manual_cohort = fallback_cohort
                                resolved_id_column = str(fallback_column)
                                break
                    app.session_state["manual_cohort_input"] = manual_cohort
                    app.session_state["manual_cohort_source"] = "separate upload"
                    app.session_state["manual_cohort_id_column"] = resolved_id_column
                    manual_available = True
                except Exception as exc:  # noqa: BLE001
                    app.error(f"Manual cohort upload failed: {type(exc).__name__}: {exc}")
        if manual_cohort is not None:
            manual_cohort_id_column = (
                app.session_state.get("manual_cohort_id_column")
                or getattr(manual_cohort, "id_column", None)
                or schema.respondent_id_column
            )
            primary_id_column = schema.respondent_id_column
            if (
                manual_cohort_id_column != primary_id_column
                and manual_cohort_id_column in app.session_state["active_df"].columns
            ):
                app.info(
                    f"Manual cohort matched against '{manual_cohort_id_column}' column "
                    f"(the survey's primary respondent_id is '{primary_id_column}' "
                    f"but uploaded values are {manual_cohort_id_column}-shaped)."
                )
            raw_ids = {
                str(value).strip()
                for value in app.session_state["active_df"][manual_cohort_id_column].tolist()
            }
            winner_set = set(manual_cohort.winner_uuids)
            laggard_set = set(manual_cohort.laggard_uuids)
            overlap = tuple(sorted(winner_set & laggard_set))
            others_n = max(0, len(raw_ids - winner_set - laggard_set))
            invalid_n = len(getattr(manual_cohort, "invalid_uuids", ()) or ())
            valid_winners = len([uuid for uuid in manual_cohort.winner_uuids if uuid in raw_ids])
            valid_laggards = len([uuid for uuid in manual_cohort.laggard_uuids if uuid in raw_ids])
            source_label = getattr(manual_cohort, "source", "") or app.session_state.get("manual_cohort_source") or "uploaded file"
            app.markdown("#### Manual cohort definition")
            app.write(f"Source: {source_label}")
            app.write(
                f"Top group: {len(manual_cohort.winner_uuids)} respondents "
                f"({valid_winners} valid, {len(manual_cohort.winner_uuids) - valid_winners} not found in raw data)"
            )
            app.write(
                f"Comparison group: {len(manual_cohort.laggard_uuids)} respondents "
                f"({valid_laggards} valid, {len(manual_cohort.laggard_uuids) - valid_laggards} not found in raw data)"
            )
            app.write(f"Others: {others_n} respondents (all remaining in raw data)")
            app.write(f"Overlap (in both groups): {len(overlap)} respondents")
            if invalid_n:
                app.warning(f"{invalid_n} uploaded uuid values were not found in raw data.")
            manual_overlap_blocked = len(overlap) > 0
            app.session_state["manual_cohort_overlap_blocked"] = manual_overlap_blocked
            if manual_overlap_blocked:
                app.error(
                    "Manual cohort overlap detected. Remove respondents from one list "
                    "before running or exporting."
                )
            segment_definition = SegmentDefinition(
                outcome_question_id="manual_uuid",
                segment_mode="manual_uuid",
                winner_label="Top group",
                loser_label="Comparison group",
                laggard_label="Comparison group",
                manual_winner_uuids=tuple(manual_cohort.winner_uuids),
                manual_laggard_uuids=tuple(manual_cohort.laggard_uuids),
                manual_cohort_id_column=manual_cohort_id_column,
            )
        else:
            app.info("Upload a workbook defining your two groups to continue.")
            app.session_state["manual_cohort_overlap_blocked"] = True
    elif seg_mode == "categorical":
        app.session_state["manual_cohort_overlap_blocked"] = False
        if outcome_spec is None:
            app.error("Select an outcome variable before using categorical mode.")
            return
        options = list(outcome_spec.option_map.items())
        if not options:
            app.warning(
                "No option codes found for this question. Try numeric threshold mode."
            )
        else:
            option_labels = [f"{code}: {label}" for code, label in options]
            selected_labels = app.multiselect(
                "Select top group values",
                options=option_labels,
                help=(
                    "Respondents matching these codes are 'winners'. "
                    "All others are 'losers'."
                ),
                key="winner_values_multiselect",
            )
            winner_codes = tuple(
                code
                for code, label in options
                if f"{code}: {label}" in selected_labels
            )
            laggard_labels_selected = app.multiselect(
                "Comparison group option values",
                options=option_labels,
                help=(
                    "Respondents matching these codes are the comparison group. "
                    "Leave blank to use everyone not in the top group."
                ),
                key="laggard_values_multiselect",
            )
            laggard_codes = tuple(
                code
                for code, label in options
                if f"{code}: {label}" in laggard_labels_selected
            )
            if winner_codes:
                winner_label = app.text_input(
                    "Top group label", value="Top group", key="winner_label_input"
                )
                loser_label = app.text_input(
                    "Comparison group label", value="Comparison group", key="loser_label_input"
                )
                segment_definition = SegmentDefinition(
                    outcome_question_id=outcome_id,
                    segment_mode="categorical",
                    winner_values=winner_codes,
                    winner_label=winner_label or "Top group",
                    loser_label=loser_label or "Comparison group",
                    laggard_values=laggard_codes,
                    laggard_label=loser_label or "Comparison group",
                    laggard_outcome_question_id=laggard_outcome_question_id,
                    laggard_outcome_sub_question_id=laggard_outcome_sub_question_id,
                )
            else:
                app.warning("Select at least one value for the top group to continue.")
    elif seg_mode == "numeric_threshold":
        app.session_state["manual_cohort_overlap_blocked"] = False
        if outcome_spec is None:
            app.error("Select an outcome variable before using numeric threshold mode.")
            return
        threshold = app.number_input(
            "Threshold value",
            value=50.0,
            key="numeric_threshold_input",
        )
        direction = app.radio(
            "Top group = respondents who scored",
            options=["gte", "lte"],
            format_func=lambda x: (
                f"\u2265 {threshold} (at or above threshold)"
                if x == "gte"
                else f"\u2264 {threshold} (at or below threshold)"
            ),
            key="threshold_direction_radio",
            horizontal=True,
        )
        winner_label = app.text_input(
            "Top group label", value="High", key="winner_label_num_input"
        )
        loser_label = app.text_input(
            "Comparison group label", value="Low", key="loser_label_num_input"
        )
        laggard_threshold = app.number_input(
            "Comparison group threshold",
            value=float(threshold),
            key="laggard_threshold_input",
        )
        laggard_direction = app.selectbox(
            "Comparison group direction",
            options=["lte", "gte"],
            key="laggard_threshold_direction_radio",
        )
        segment_definition = SegmentDefinition(
            outcome_question_id=outcome_id,
            segment_mode="numeric_threshold",
            winner_threshold=float(threshold),
            threshold_direction=direction,
            winner_label=winner_label or "High",
            loser_label=loser_label or "Low",
            laggard_threshold=float(laggard_threshold),
            laggard_threshold_direction=laggard_direction,
            laggard_label=loser_label or "Comparison group",
            laggard_outcome_question_id=laggard_outcome_question_id,
            laggard_outcome_sub_question_id=laggard_outcome_sub_question_id,
        )
    else:
        app.session_state["manual_cohort_overlap_blocked"] = False
        if outcome_spec is None:
            app.error("Select an outcome variable before using quartile mode.")
            return
        quartile_winner = app.radio(
            "Top quartile",
            options=["top", "bottom"],
            format_func=lambda value: (
                "Top quartile" if value == "top" else "Bottom quartile"
            ),
            key=f"quartile_winner_{outcome_id}",
            horizontal=True,
        )
        laggard_label = app.text_input(
            "Comparison group label",
            value="Comparison group",
            key="laggard_label_quartile_input",
        )
        segment_definition = SegmentDefinition(
            outcome_question_id=outcome_id,
            segment_mode="quartile",
            quartile_winner=quartile_winner,
            laggard_label=laggard_label or "Comparison group",
            loser_label=laggard_label or "Comparison group",
            laggard_outcome_question_id=laggard_outcome_question_id,
            laggard_outcome_sub_question_id=laggard_outcome_sub_question_id,
        )

    if segment_definition is None:
        return

    app.session_state["segment_definition"] = segment_definition

    if app.button(
        "\u25B6 Run Outcome Segmentation",
        key="run_segmentation_btn",
        type="primary",
        disabled=app.session_state.get("manual_cohort_overlap_blocked", False),
    ):
        from src.outcome_segmentation import compute_outcome_segmentation

        # compute_outcome_segmentation expects a list[AuditRecord] it can
        # `.append()` to. session_state["log"] is a CalculationLog (record-only),
        # so we collect into a local list and forward into the log afterwards.
        local_audit: list = []
        with app.spinner("Running outcome segmentation..."):
            try:
                seg_result = compute_outcome_segmentation(
                    decoded_df=app.session_state["active_df"],
                    schema=schema,
                    outcome_question_id=segment_definition.outcome_question_id,
                    segment_definition=segment_definition,
                    audit_log=local_audit,
                    min_sample_size=30,
                )
            except Exception as exc:  # noqa: BLE001 - surface error to UI
                app.error(f"Segmentation failed: {type(exc).__name__}: {exc}")
                return
            calc_log = app.session_state.get("log")
            if calc_log is not None:
                for audit in local_audit:
                    calc_log.record(audit)
            app.session_state["segmentation_result"] = seg_result
            app.session_state["outcome_segmented_workbook_bytes"] = None
            app.session_state["outcome_segmented_workbook_signature"] = None
            app.rerun()


def _differentiator_gap(diff: Any) -> float:
    return float(diff.top_option_winner_rate - diff.top_option_loser_rate)


def _differentiator_lift_for_sort(diff: Any) -> float:
    lift = float(diff.top_option_lift)
    if lift >= 999.0:
        return float("inf")
    return lift


def _sorted_differentiators_for_display(seg: OutcomeSegmentationResult) -> list[Any]:
    return sorted(
        seg.differentiators,
        key=lambda diff: (
            -_differentiator_lift_for_sort(diff),
            -_differentiator_gap(diff),
            str(diff.question_id),
        ),
    )


def _differentiator_table_rows(
    differentiators: list[Any],
    winner_label: str,
    laggard_label: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for rank, diff in enumerate(differentiators, start=1):
        rows.append(
            {
                "#": rank,
                "Question": f"{diff.question_id}: {_truncate(diff.question_text, 80)}",
                "Top Option": diff.top_option_label,
                f"{winner_label} Rate": f"{diff.top_option_winner_rate:.1%}",
                f"{laggard_label} Rate": f"{diff.top_option_loser_rate:.1%}",
                "Gap": _format_rate(_differentiator_gap(diff), signed=True),
                "Lift": "\u221e" if diff.top_option_lift >= 999.0 else f"{diff.top_option_lift:.2f}x",
                "Cramér's V": f"{diff.cramers_v:.3f}",
                "p-value": f"{diff.p_value:.3f}" if diff.p_value is not None else "N/A",
                "Notes": " | ".join(diff.warnings) if diff.warnings else "",
            }
        )
    return rows


def _question_cell_matches(question_cell: Any, question_id: str | None) -> bool:
    if not question_id:
        return False
    question_text = str(question_cell)
    question_id = str(question_id)
    return question_text == question_id or question_text.startswith(f"{question_id}:")


def _style_differentiator_source(
    df: pd.DataFrame,
    highlight_question_id: str | None,
    winner_rate_col: str,
) -> pd.DataFrame:
    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    if not highlight_question_id or "Question" not in df.columns:
        return styles

    matches = df["Question"].apply(
        lambda value: _question_cell_matches(value, highlight_question_id)
    )
    styles.loc[matches, :] = "background-color: #FFF3CD;"
    for column in ("Lift", winner_rate_col):
        if column in styles.columns:
            styles.loc[
                matches, column
            ] = "background-color: #F8D7DA; font-weight: 700;"
    return styles


def _render_differentiator_dataframe(
    rows: list[dict[str, Any]],
    *,
    highlight_question_id: str | None = None,
    winner_rate_col: str = "",
) -> None:
    app = _require_streamlit()
    df = pd.DataFrame(rows)
    if df.empty:
        return

    if highlight_question_id:
        app.dataframe(
            df.style.apply(
                lambda frame: _style_differentiator_source(
                    frame,
                    highlight_question_id,
                    winner_rate_col,
                ),
                axis=None,
            ),
            use_container_width=True,
            hide_index=True,
        )
        return

    app.dataframe(df, use_container_width=True, hide_index=True)


def _render_differentiator_table_with_show_more(
    differentiators: list[Any],
    winner_label: str,
    laggard_label: str,
    *,
    default_count: int = 5,
) -> None:
    app = _require_streamlit()
    rows = _differentiator_table_rows(
        differentiators,
        winner_label,
        laggard_label,
    )
    _render_differentiator_dataframe(
        rows[:default_count],
        winner_rate_col=f"{winner_label} Rate",
    )
    remaining = rows[default_count:]
    if remaining:
        with app.expander(
            f"Show more insights ({len(remaining)} more)",
            expanded=False,
        ):
            _render_differentiator_dataframe(
                remaining,
                winner_rate_col=f"{winner_label} Rate",
            )


def _render_differentiator_source_table_expander(
    differentiators: list[Any],
    winner_label: str,
    laggard_label: str,
    highlight_question_id: str | None,
) -> None:
    app = _require_streamlit()
    rows = _differentiator_table_rows(
        differentiators,
        winner_label,
        laggard_label,
    )
    if not rows:
        return

    with app.expander("Show source table", expanded=False):
        _render_differentiator_dataframe(
            rows,
            highlight_question_id=highlight_question_id,
            winner_rate_col=f"{winner_label} Rate",
        )


_DIFF_INSIGHTS_PER_PAGE = 10


def _outcome_diff_payload(
    diff: Any,
    winner_label: str,
    laggard_label: str,
) -> dict[str, Any]:
    winner_rate = float(diff.top_option_winner_rate)
    loser_rate = float(diff.top_option_loser_rate)
    return {
        "question_text": diff.question_text,
        "option_label": diff.top_option_label,
        "winner_rate": winner_rate,
        "loser_rate": loser_rate,
        "lift": float(diff.top_option_lift),
        "rate_gap": winner_rate - loser_rate,
        "winner_label": winner_label,
        "laggard_label": laggard_label,
    }


def _outcome_diff_insight_cache_key(diff: Any) -> str:
    return (
        "outcome_diff_insight:"
        f"{diff.question_id}:"
        f"{diff.top_option_label}"
    )


def _cohort_label_plural(label: str) -> str:
    label = str(label or "Top group")
    return label if label.lower().endswith("s") else f"{label}s"


def _fallback_outcome_diff_insight(
    payload: dict[str, Any],
    *,
    reason: str = "",
) -> InsightResult:
    winner_label = _cohort_label_plural(str(payload.get("winner_label", "Winner")))
    laggard_label = _cohort_label_plural(str(payload.get("laggard_label", "Laggard")))
    option_label = str(payload.get("option_label", "this option"))
    winner_rate = float(payload.get("winner_rate", 0.0) or 0.0)
    loser_rate = float(payload.get("loser_rate", 0.0) or 0.0)
    lift = float(payload.get("lift", 0.0) or 0.0)
    return InsightResult(
        title="",
        insight=(
            f"{winner_label} are {lift:.1f}\u00d7 more likely than {laggard_label} "
            f"to select \u201c{option_label}\u201d "
            f"({winner_rate:.0%} vs {loser_rate:.0%})"
        ),
        was_template=True,
        error_message=reason,
    )


def _extract_percent_tokens(text: str) -> list[float]:
    tokens: list[float] = []
    pattern = r"(?<![A-Za-z0-9])([+-]?\d+(?:\.\d+)?)\s*(?:%|pp|percentage points?)"
    for match in _re.finditer(pattern, text, flags=_re.IGNORECASE):
        try:
            tokens.append(float(match.group(1)))
        except ValueError:
            continue
    return tokens


def _extract_lift_tokens(text: str) -> list[float]:
    tokens: list[float] = []
    pattern = r"(?<![A-Za-z0-9])([+-]?\d+(?:\.\d+)?)\s*(?:x|X|×)"
    for match in _re.finditer(pattern, text):
        try:
            tokens.append(float(match.group(1)))
        except ValueError:
            continue
    return tokens


def _outcome_diff_insight_numbers_are_valid(
    insight_text: str,
    payload: dict[str, Any],
) -> bool:
    winner_pct = float(payload.get("winner_rate", 0.0) or 0.0) * 100.0
    loser_pct = float(payload.get("loser_rate", 0.0) or 0.0) * 100.0
    gap_pct = (
        float(payload.get("rate_gap", 0.0) or 0.0) * 100.0
    )
    lift = float(payload.get("lift", 0.0) or 0.0)
    percent_allowed = (winner_pct, loser_pct, gap_pct, abs(gap_pct))
    lift_allowed = (lift,)

    percent_tokens = _extract_percent_tokens(insight_text)
    lift_tokens = _extract_lift_tokens(insight_text)
    if not percent_tokens and not lift_tokens:
        return False

    for token in percent_tokens:
        if not any(abs(token - allowed) <= 0.1 for allowed in percent_allowed):
            return False
    for token in lift_tokens:
        if not any(abs(token - allowed) <= 0.05 for allowed in lift_allowed):
            return False
    return True


def _outcome_diff_insight_framing_is_safe(text: str) -> bool:
    unsafe_patterns = (
        r"(allocat\w*|spend|spent|invest\w*|dedicat(?:e|es|ing)|devot\w*)\b[^.]{0,40}\d+(?:\.\d+)?\s*%",
        r"\d+(?:\.\d+)?\s*%\s*(of\s+(their|the)\s+)?(budget|revenue|spend|margin|sales|portfolio|growth|profit)",
        r"(expect|expects|expecting|forecast|grow|grew|growth of)\b[^.]{0,30}\d+(?:\.\d+)?\s*%",
        r"\d+(?:\.\d+)?\s*%\s+(growth|increase|gain|expansion)",
    )
    return not any(
        _re.search(pattern, text, flags=_re.IGNORECASE)
        for pattern in unsafe_patterns
    )


def _validated_outcome_diff_insight(
    insight: InsightResult,
    payload: dict[str, Any],
) -> InsightResult:
    numbers_are_valid = _outcome_diff_insight_numbers_are_valid(
        insight.insight,
        payload,
    )
    framing_is_safe = _outcome_diff_insight_framing_is_safe(insight.insight)
    if (
        insight.was_template
        or not insight.insight
        or not numbers_are_valid
        or not framing_is_safe
    ):
        reason = insight.error_message
        if not reason:
            reason = (
                "AI insight reframed selection share as a magnitude."
                if not framing_is_safe
                else "AI insight numbers did not match the source table."
            )
        return _fallback_outcome_diff_insight(
            payload,
            reason=reason,
        )
    return insight


def _get_outcome_diff_insight(
    diff: Any,
    winner_label: str,
    laggard_label: str,
) -> InsightResult:
    cache_key = _outcome_diff_insight_cache_key(diff)
    cached = _INSIGHT_CACHE.get(cache_key)
    if isinstance(cached, InsightResult):
        return cached

    payload = _outcome_diff_payload(diff, winner_label, laggard_label)
    try:
        generated = generate_insight(
            table_payload=payload,
            table_kind="winner_profile_trait",
            title_hint=str(diff.question_id),
            cache=_INSIGHT_CACHE,
        )
    except Exception as exc:  # noqa: BLE001
        generated = _fallback_outcome_diff_insight(
            payload,
            reason=f"AI insight failed: {type(exc).__name__}: {exc}",
        )
    validated = _validated_outcome_diff_insight(generated, payload)
    _INSIGHT_CACHE[cache_key] = validated
    return validated


def _render_outcome_diff_insight_card(
    seg: OutcomeSegmentationResult,
    diff: Any,
    winner_label: str,
    laggard_label: str,
) -> None:
    app = _require_streamlit()
    cols = app.columns([3, 1])
    winner_rate = float(diff.top_option_winner_rate)
    loser_rate = float(diff.top_option_loser_rate)
    gap = winner_rate - loser_rate
    lift = float(diff.top_option_lift)
    with cols[0]:
        app.markdown(f"#### {diff.question_id}")
        app.caption(diff.question_text)
        insight = _get_outcome_diff_insight(diff, winner_label, laggard_label)
        _render_insight_card(insight, label=diff.top_option_label)
    with cols[1]:
        app.metric(winner_label, f"{winner_rate:.1%}")
        app.metric(laggard_label, f"{loser_rate:.1%}")
        app.metric("Gap", f"{gap:+.1%}")
        lift_label = "∞" if lift >= 999.0 else f"{lift:.2f}x"
        app.metric("Lift", lift_label)
    _render_outcome_diff_source_table_expander(seg, diff, winner_label)


def _style_outcome_diff_source_table(
    df: pd.DataFrame,
    highlight_option_label: str,
    winner_pct_col: str,
) -> pd.DataFrame:
    styles = pd.DataFrame("", index=df.index, columns=df.columns)
    if "Option" not in df.columns:
        return styles

    target = str(highlight_option_label).strip()
    matches = df["Option"].apply(lambda value: str(value).strip() == target)
    styles.loc[matches, :] = "background-color: #FFF3CD;"
    if winner_pct_col in styles.columns:
        styles.loc[matches, winner_pct_col] = (
            "background-color: #F8D7DA; font-weight: 700;"
        )
    return styles


def _render_outcome_diff_source_table_expander(
    seg: OutcomeSegmentationResult,
    diff: Any,
    winner_label: str,
) -> None:
    app = _require_streamlit()
    rows = _winner_profile_trait_backing_rows(seg, diff.question_id)
    if not rows:
        return

    df = pd.DataFrame(rows)
    if df.empty:
        return

    with app.expander("Show source table", expanded=False):
        winner_pct_col = f"{winner_label} %"
        app.dataframe(
            df.style.apply(
                lambda frame: _style_outcome_diff_source_table(
                    frame,
                    str(diff.top_option_label),
                    winner_pct_col,
                ),
                axis=None,
            ),
            use_container_width=True,
            hide_index=True,
        )


def _render_paginated_outcome_diff_insights(
    seg: OutcomeSegmentationResult,
    sorted_diffs: list[Any],
    winner_label: str,
    laggard_label: str,
    outcome_question_id: str,
) -> None:
    app = _require_streamlit()
    if not sorted_diffs:
        return

    page_key = f"diff_insight_page_{outcome_question_id}"
    max_page = max(0, (len(sorted_diffs) - 1) // _DIFF_INSIGHTS_PER_PAGE)
    page_index = int(app.session_state.get(page_key, 0) or 0)
    page_index = min(max(page_index, 0), max_page)
    app.session_state[page_key] = page_index
    start = page_index * _DIFF_INSIGHTS_PER_PAGE
    end = min(start + _DIFF_INSIGHTS_PER_PAGE, len(sorted_diffs))

    app.markdown("### Differentiator insights")
    app.caption(f"Showing {start + 1}-{end} of {len(sorted_diffs)}")
    for diff in sorted_diffs[start:end]:
        _render_outcome_diff_insight_card(
            seg,
            diff,
            winner_label,
            laggard_label,
        )
        app.divider()

    prev_col, next_col = app.columns(2)
    with prev_col:
        if app.button(
            "◀ Previous",
            key=f"{page_key}_prev",
            disabled=page_index <= 0,
            use_container_width=True,
        ):
            app.session_state[page_key] = max(page_index - 1, 0)
            app.rerun()
    with next_col:
        if app.button(
            "Next 10 ▶",
            key=f"{page_key}_next",
            disabled=page_index >= max_page,
            use_container_width=True,
        ):
            app.session_state[page_key] = min(page_index + 1, max_page)
            app.rerun()


def _render_segmentation_results() -> None:
    """Stage C: display winner/loser metrics and differentiators."""
    app = _require_streamlit()
    seg = app.session_state.get("segmentation_result")
    if seg is None:
        return

    app.markdown("---")
    app.markdown("### Segmentation Results")

    _winner_lbl = seg.segment_definition.winner_label
    _loser_lbl = seg.segment_definition.loser_label
    col1, col2, col3 = app.columns(3)
    with col1:
        app.metric(f"{_winner_lbl}s", seg.winner_n)
    with col2:
        app.metric(f"{_loser_lbl}s", seg.loser_n)
    with col3:
        app.metric("Differentiators Found", len(seg.differentiators))

    for warning in seg.warnings:
        app.warning(warning)

    if seg.differentiators:
        app.markdown("#### Top Differentiators (ranked by lift, then gap)")
        sorted_diffs = _sorted_differentiators_for_display(seg)
        _render_differentiator_table_with_show_more(
            sorted_diffs,
            _winner_lbl,
            _loser_lbl,
        )

        infinite_lift_diffs = [
            diff
            for diff in sorted_diffs
            if diff.top_option_lift >= 999.0
        ]
        if infinite_lift_diffs:
            app.caption(
                f"{len(infinite_lift_diffs)} question(s) show "
                "infinite lift (\u221E) \u2014 the loser segment had 0 "
                "respondents select this option. Interpret with caution; "
                "check sample sizes."
            )

    if seg.skipped_questions:
        with app.expander(
            f"{len(seg.skipped_questions)} questions skipped"
        ):
            for qid, reason in seg.skipped_questions:
                app.caption(f"\u2022 {qid}: {reason}")


# ---------------------------------------------------------------------------
# Section 3 — Single cuts
# ---------------------------------------------------------------------------


def _section_single_cuts() -> None:
    """Day 18 master-detail: show only the selected question's analysis."""
    import html as _html

    app = _require_streamlit()
    app.markdown('<div id="section-singlecuts"></div>', unsafe_allow_html=True)
    _section_header("3", "Single cut", anchor="section-singlecuts", meta="Create Cuts")
    _red_instruction_box(
        "Step 3 \u2014 Run &amp; review",
        "Pick a question from the sidebar to see how people answered. "
        "Standout values are highlighted automatically.",
    )

    if not app.session_state["run_complete"]:
        app.info(EMPTY_NO_RESULTS)
        return

    results = app.session_state["results"]
    schema = app.session_state["schema"]
    if not results:
        app.info(EMPTY_NO_RESULTS)
        return

    selected = app.session_state.get("selected_question_id")
    valid_ids = {r.question_id for r in results}
    if selected not in valid_ids:
        selected = results[0].question_id
        app.session_state["selected_question_id"] = selected

    result = next((r for r in results if r.question_id == selected), results[0])
    spec = schema.get_question(result.question_id)
    if spec is None:
        app.info("Selected question is not available.")
        return

    title = _html.escape(spec.question_text or spec.canonical_id)
    qtype = spec.question_type.value if hasattr(spec.question_type, "value") else str(spec.question_type)
    app.markdown(
        f"""
        <div style="background:white;border:1px solid #E0E0E0;
          border-left:4px solid #CC0000;padding:18px 22px;margin-bottom:16px;
          font-family:Arial;">
          <div style="font-size:11px;font-weight:700;letter-spacing:0.1em;
            color:#CC0000;text-transform:uppercase;">
            {_html.escape(spec.canonical_id)} &middot; {_html.escape(qtype)}
            &middot; Valid N {result.valid_n:,}
          </div>
          <div style="font-size:18px;font-weight:600;color:#0A0A0A;
            margin-top:6px;line-height:1.4;">{title}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    normal_view = app.toggle(
        "Single pane view",
        value=bool(app.session_state.get("single_cut_normal_view", False)),
        key="single_cut_normal_view",
        help="On: single-cut and its cross-cut builder stack in one column. Off: side-by-side split workspace.",
    )
    if normal_view:
        _render_single_cut_card(result, spec, expanded=True)
        app.divider()
        _render_selected_question_cross_cut_builder(result.question_id)
    else:
        left, right = app.columns([1.1, 0.9], gap="large")
        with left:
            _render_single_cut_card(result, spec, expanded=True)
        with right:
            _render_selected_question_cross_cut_builder(result.question_id)
    app.caption(
        f"Question {results.index(result) + 1} of {len(results)}. "
        "Use the sidebar to switch questions."
    )


# ---------------------------------------------------------------------------
# Section 4 — Cross cuts
# ---------------------------------------------------------------------------


def _source_question_labels(
    question_ids: tuple[str, ...],
    schema: Any,
    *,
    include_ids: bool = True,
    max_chars: int = 60,
) -> str:
    parts: list[str] = []
    for question_id in question_ids:
        spec = schema.get_question(question_id) if schema is not None else None
        if spec is None:
            parts.append(question_id)
            continue
        text = spec.question_text[:max_chars]
        if len(spec.question_text) > max_chars:
            text += "..."
        parts.append(f"{question_id}: {text}" if include_ids else text)
    return " \u00d7 ".join(parts)


def _suggestion_label(suggestion: Any, schema: Any) -> str:
    if schema is None:
        return suggestion.synthetic_question_title
    return _source_question_labels(suggestion.source_question_ids, schema, max_chars=60)


def _cross_cut_display_title(result: Any, schema: Any) -> str:
    source_question_ids = getattr(result, "source_question_ids", ())
    if schema is None or not source_question_ids:
        return getattr(
            result,
            "synthetic_question_title",
            getattr(result, "cross_cut_id", str(result)),
        )
    return _source_question_labels(
        source_question_ids, schema, include_ids=False, max_chars=50
    )


def _section_cross_cuts() -> None:
    app = _require_streamlit()
    app.markdown('<div id="section-crosscuts"></div>', unsafe_allow_html=True)
    _section_header("2", "Cross cuts", anchor="section-crosscuts", meta="Create Cuts")

    if not app.session_state["run_complete"]:
        app.info(EMPTY_NO_CROSS_CUTS)
        return

    with app.expander("Build a cross cut", expanded=True):
        _render_manual_cross_cut()

    with app.expander("Suggestions", expanded=False):
        _render_suggested_cross_cuts()

    results = app.session_state["cross_cut_results"]
    if not results:
        app.caption("No cross cuts run yet. Build one above to populate this list.")
        return

    import html as _html

    schema = app.session_state.get("schema")
    app.markdown("**Cross-cut results**")
    for result in results:
        cc_pending_insight: dict[str, Any] | None = None
        with app.expander(
            f"{result.cross_cut_id} \u2014 {_cross_cut_display_title(result, schema)}",
            expanded=False,
        ):
            # Show full source-question text as a header so analysts know
            # exactly which questions feed this cross cut.
            source_lines = []
            for qid in result.source_question_ids:
                q = schema.get_question(qid) if schema is not None else None
                if q and q.question_text:
                    source_lines.append(
                        f"{_html.escape(qid)}: {_html.escape(q.question_text.strip())}"
                    )
                else:
                    source_lines.append(_html.escape(qid))
            app.markdown(
                f'<div style="font-size:13px;font-weight:700;color:#0A0A0A;'
                f'font-family:Arial;margin-bottom:4px;">'
                f'{_html.escape(result.synthetic_question_title)}</div>'
                f'<div style="font-size:10px;color:#888;font-family:Arial;'
                f'line-height:1.8;">{"<br>".join(source_lines)}</div>',
                unsafe_allow_html=True,
            )
            app.caption(
                f"Type: {result.analysis_type.value}  \u00b7  "
                f"Display: {result.display_mode}"
            )
            _render_cross_cut_preview(result)
            cc_pending_insight = {
                "insight_key": f"insight_cc_{result.cross_cut_id}",
                "payload_factory": (
                    lambda r=result: _build_insight_payload_cross_cut(r)
                ),
                "table_kind": result.analysis_type.value.lower(),
                "title_hint": getattr(result, "synthetic_question_title", "")
                or result.cross_cut_id,
            }
            if result.warnings:
                if app.checkbox(
                    f"Show {len(result.warnings)} warning(s)",
                    key=f"cc_warn_{result.cross_cut_id}",
                ):
                    for warning in result.warnings:
                        app.write(f"\u2022 {warning}")
            cb_col, rm_col = app.columns([3, 1])
            with cb_col:
                app.checkbox(
                    "Include in cross-cut workbook",
                    value=app.session_state.get(
                        f"cc_select_{result.cross_cut_id}", True
                    ),
                    key=f"cc_select_{result.cross_cut_id}",
                )
            with rm_col:
                if app.button(
                    "Remove",
                    key=f"cc_remove_{result.cross_cut_id}",
                    help="Remove this cross cut from the session",
                ):
                    app.session_state["cross_cut_results"] = [
                        r
                        for r in app.session_state["cross_cut_results"]
                        if r.cross_cut_id != result.cross_cut_id
                    ]
                    app.session_state.pop(
                        f"insight_cc_{result.cross_cut_id}", None
                    )
                    app.session_state["cross_cut_only_bytes"] = None
                    _refresh_full_workbook()
                    app.rerun()

        # Render insight section OUTSIDE the per-card expander above.
        # Streamlit forbids nesting expanders, and _render_insight_section
        # uses an expander to show the API error detail when the AI call fails.
        if cc_pending_insight is not None:
            _render_insight_section(**cc_pending_insight)


# ---------------------------------------------------------------------------
# Section 5 — Downloads
# ---------------------------------------------------------------------------


def _section_downloads() -> None:
    app = _require_streamlit()
    app.markdown('<div id="section-downloads"></div>', unsafe_allow_html=True)
    _section_header("4", "Download workbook", anchor="section-downloads", meta="Workbooks")

    if not app.session_state["run_complete"]:
        app.info("Run an analysis to generate downloadable workbooks.")
        return

    if app.session_state.get("rank_cross_tab_settings_dirty"):
        _refresh_full_workbook()

    output_path = app.session_state.get("output_path")
    cross_cut_results = app.session_state.get("cross_cut_results", [])
    selected_cross_cuts = [
        r
        for r in cross_cut_results
        if app.session_state.get(f"cc_select_{r.cross_cut_id}", True)
    ]
    filtered_results = app.session_state.get("filtered_results", {})
    selected_filtered = [
        r
        for cid, r in filtered_results.items()
        if app.session_state.get(f"fsc_select_{cid}", True)
    ]

    # Invalidate cached workbook bytes if the selection signature changed
    # (otherwise the user could click Download and get an outdated workbook
    # that was generated against a previous selection).
    embed_inputs_enabled = app.session_state.get("wizard_embed_input_files", False)
    input_sources = app.session_state.get("input_file_embed_sources")
    cc_signature = (
        tuple(sorted(r.cross_cut_id for r in selected_cross_cuts)),
        bool(embed_inputs_enabled),
    )
    if app.session_state.get("cross_cut_only_signature") != cc_signature:
        app.session_state["cross_cut_only_bytes"] = None
    fsc_signature = (
        tuple(
            sorted(
                cid
                for cid in filtered_results
                if app.session_state.get(f"fsc_select_{cid}", True)
            )
        ),
        bool(embed_inputs_enabled),
    )
    if app.session_state.get("filtered_workbook_signature") != fsc_signature:
        app.session_state["filtered_workbook_bytes"] = None

    col_full, col_cc, col_fsc, col_outcome = app.columns(4)

    with col_full:
        app.markdown("**Single-cut workbook**")
        app.caption("All single cuts + cross cuts inline.")
        if output_path and os.path.exists(output_path):
            with open(output_path, "rb") as workbook_file:
                workbook_bytes = workbook_file.read()
            app.download_button(
                label="Download single-cut workbook",
                data=workbook_bytes,
                file_name="survey_analysis.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
            )
        else:
            app.button(
                "Download single-cut workbook",
                disabled=True,
                use_container_width=True,
            )

    with col_cc:
        app.markdown("**Cross-cut-only workbook**")
        app.caption("Just the cross cuts you've ticked above.")
        if app.button(
            "Generate cross-cut workbook",
            disabled=(len(selected_cross_cuts) == 0),
            use_container_width=True,
            key="gen_cc_workbook",
            help=(
                f"{len(selected_cross_cuts)} cross cuts selected"
                if selected_cross_cuts
                else "Tick at least one cross cut to enable"
            ),
        ):
            from src.excel_exporter import export_cross_cuts_only

            cc_path = "/tmp/cross_cuts.xlsx"
            try:
                export_cross_cuts_only(
                    cross_cut_results=selected_cross_cuts,
                    schema=app.session_state["schema"],
                    log=app.session_state["log"],
                    output_path=cc_path,
                    embed_input_files=embed_inputs_enabled,
                    input_file_sources=input_sources,
                )
                with open(cc_path, "rb") as f:
                    app.session_state["cross_cut_only_bytes"] = f.read()
                app.session_state["cross_cut_only_signature"] = cc_signature
            except Exception as exc:  # noqa: BLE001
                app.error(f"Cross-cut export failed: {type(exc).__name__}: {exc}")
                with app.expander("Show traceback"):
                    app.code(traceback.format_exc())
        if app.session_state.get("cross_cut_only_bytes"):
            app.download_button(
                label="Download cross-cut workbook",
                data=app.session_state["cross_cut_only_bytes"],
                file_name="cross_cut_analysis.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
            )

    with col_fsc:
        app.markdown("**Filtered workbook**")
        app.caption("Per-question filtered analyses you've ticked.")
        if app.button(
            "Generate filtered workbook",
            disabled=(len(selected_filtered) == 0),
            use_container_width=True,
            key="gen_fsc_workbook",
            help=(
                f"{len(selected_filtered)} filtered analyses selected"
                if selected_filtered
                else "Apply at least one per-question filter to enable"
            ),
        ):
            from src.excel_exporter import export_filtered_single_cuts

            fsc_path = "/tmp/filtered_single_cuts.xlsx"
            try:
                export_filtered_single_cuts(
                    filtered_results=selected_filtered,
                    schema=app.session_state["schema"],
                    log=app.session_state["log"],
                    output_path=fsc_path,
                    embed_input_files=embed_inputs_enabled,
                    input_file_sources=input_sources,
                )
                with open(fsc_path, "rb") as f:
                    app.session_state["filtered_workbook_bytes"] = f.read()
                app.session_state["filtered_workbook_signature"] = fsc_signature
            except Exception as exc:  # noqa: BLE001
                app.error(f"Filtered export failed: {type(exc).__name__}: {exc}")
                with app.expander("Show traceback"):
                    app.code(traceback.format_exc())
        if app.session_state.get("filtered_workbook_bytes"):
            app.download_button(
                label="Download filtered workbook",
                data=app.session_state["filtered_workbook_bytes"],
                file_name="filtered_single_cuts.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
            )

    with col_outcome:
        app.markdown("**Outcome segmented workbook**")
        app.caption("Every single cut side-by-side for winners vs laggards.")
        seg_result = app.session_state.get("segmentation_result")
        outcome_signature = (
            getattr(seg_result, "outcome_question_id", None),
            str(getattr(seg_result, "segment_definition", "")),
            len(app.session_state.get("results", []) or []),
            bool(embed_inputs_enabled),
        )
        if app.session_state.get("outcome_segmented_workbook_signature") != outcome_signature:
            app.session_state["outcome_segmented_workbook_bytes"] = None
        if app.button(
            "Generate outcome segmented workbook",
            disabled=(
                seg_result is None
                or app.session_state.get("manual_cohort_overlap_blocked", False)
            ),
            use_container_width=True,
            key="gen_outcome_segmented_workbook",
            help=(
                "Resolve manual cohort overlap before export"
                if app.session_state.get("manual_cohort_overlap_blocked", False)
                else
                "Run outcome segmentation first"
                if seg_result is None
                else "Build winners vs laggards workbook"
            ),
        ):
            from src.excel_exporter import export_winners_vs_laggards_workbook

            outcome_path = "/tmp/outcome_winners_vs_laggards.xlsx"
            try:
                export_winners_vs_laggards_workbook(
                    output_path=outcome_path,
                    decoded_df=app.session_state["active_df"],
                    schema=app.session_state["schema"],
                    single_cut_results=app.session_state.get("results", []),
                    segment_definition=seg_result.segment_definition,
                    laggard_outcome_question_id=(
                        seg_result.segment_definition.laggard_outcome_question_id
                    ),
                    laggard_segment_definition=seg_result.segment_definition,
                    themes=app.session_state.get("auto_themes"),
                    workbook_custom_filter_count=app.session_state.get(
                        "wizard_workbook_custom_filter_count", 2
                    ),
                    per_question_filter_count=app.session_state.get(
                        "wizard_per_question_filter_count", 1
                    ),
                    calculation_log=app.session_state.get("log"),
                    embed_input_files=embed_inputs_enabled,
                    input_file_sources=input_sources,
                )
                with open(outcome_path, "rb") as f:
                    app.session_state["outcome_segmented_workbook_bytes"] = f.read()
                app.session_state["outcome_segmented_workbook_signature"] = outcome_signature
            except Exception as exc:  # noqa: BLE001
                app.error(
                    f"Outcome segmented export failed: {type(exc).__name__}: {exc}"
                )
                with app.expander("Show traceback"):
                    app.code(traceback.format_exc())
        if app.session_state.get("outcome_segmented_workbook_bytes"):
            app.download_button(
                label="Outcome Segmentation Workbook",
                data=app.session_state["outcome_segmented_workbook_bytes"],
                file_name="outcome_winners_vs_laggards.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
            )

    app.caption(TOOLTIP_THREE_DOWNLOADS)


# ---------------------------------------------------------------------------
# Stage D: AI Analysis section
# ---------------------------------------------------------------------------


def _truncate(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 3] + "..."


def _bar_colors(values: list[float]) -> list[str]:
    """Return color list with max-value bar in red, others black."""
    if not values:
        return []
    max_val = max(values)
    return ["#CC0000" if v == max_val else "#1A1A1A" for v in values]


def _insight_rows_from_payload(
    payload: dict[str, Any],
    table_kind: str,
) -> list[dict[str, Any]]:
    if table_kind == "differentiator":
        return [
            {
                "label": payload.get("top_option", ""),
                "winner_rate": payload.get("winner_rate", 0),
                "loser_rate": payload.get("loser_rate", 0),
                "lift": payload.get("lift", 0),
                "cramers_v": payload.get("cramers_v", 0),
            }
        ]
    if table_kind == "winner_profile":
        return list(payload.get("traits", []))
    return []


def _normalise_insight_payload(
    table_payload: dict[str, Any],
    table_kind: str,
    title_hint: str,
) -> dict[str, Any]:
    payload = dict(table_payload)
    payload.setdefault("table_kind", table_kind)
    payload.setdefault("question_id", payload.get("question_id", ""))
    payload.setdefault("question_text", payload.get("question_text", title_hint))
    payload.setdefault("valid_n", payload.get("winner_n", 0))
    payload.setdefault("missing_n", 0)
    payload.setdefault("filters_applied", [])
    payload.setdefault("summary", {})
    payload.setdefault("rows", _insight_rows_from_payload(payload, table_kind))
    return payload


def _section_ai_analysis() -> None:
    app = _require_streamlit()
    app.markdown('<div id="section-outcome"></div>', unsafe_allow_html=True)
    _section_header(
        "3",
        "Outcome Segmentation",
        anchor="section-outcome",
        meta="Group comparison",
    )
    _section_survey_classification(show_toggle=False)
    if not app.session_state.get("show_winners_vs_laggards", False):
        return
    app.divider()
    if not app.session_state.get("run_complete"):
        return

    app.markdown(
        f'<h2 style="display:flex;align-items:center;margin:1rem 0;">'
        f'{_icon("analysis", "#CC0000")}Outcome Segmentation Insights</h2>',
        unsafe_allow_html=True,
    )
    app.info(
        "**What this does:** it splits your respondents into a top group and the "
        "rest based on an outcome you choose (e.g. revenue growth), then finds the "
        "questions where the two groups answered most differently \u2014 i.e. what "
        "separates the best from the rest.\n\n"
        "**How to read it:** *Lift* = how many times more likely the top group is "
        "to pick an option than the rest (2\u00d7 = twice as often). *Gap* = the raw "
        "percentage-point difference. Every number here is computed from your data, "
        "not estimated."
    )

    seg: OutcomeSegmentationResult | None = app.session_state.get(
        "segmentation_result"
    )
    if seg is None:
        app.info("Available after you run the analysis.")
        return

    _render_outcome_summary_panel(seg)
    _render_smart_cross_cut_suggestions_panel(seg)


def _render_outcome_summary_panel(seg: OutcomeSegmentationResult) -> None:
    app = _require_streamlit()

    if not seg.differentiators:
        app.info("No differentiators found. Try a different segment definition.")
        return

    laggard_label = seg.segment_definition.loser_label
    winner_label = seg.segment_definition.winner_label
    sorted_diffs = _sorted_differentiators_for_display(seg)

    avg_lift_values = [
        d.top_option_lift for d in sorted_diffs[:5] if d.top_option_lift < 900
    ]
    avg_lift = (
        sum(avg_lift_values) / len(avg_lift_values) if avg_lift_values else 0.0
    )

    segment_mode = str(getattr(seg.segment_definition, "segment_mode", "")).replace(
        "_", " "
    ).title()
    app.markdown(
        f"""
    <div class="ui-panel">
        <div class="ui-panel-head">
            <span class="ui-panel-title">Segment summary</span>
            <span class="ui-panel-meta">Outcome: {seg.outcome_question_id} · {segment_mode}</span>
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    col1, col2, col3, col4 = app.columns(4)
    col1.metric("Top group", f"{seg.winner_n:,}")
    col2.metric("Comparison group", f"{seg.loser_n:,}")
    col3.metric("Differentiators", len(seg.differentiators))
    col4.metric("Avg lift (top 5)", f"{avg_lift:.1f}x")

    app.markdown(
        """
    <div class="ui-panel">
        <div class="ui-panel-head">
            <span class="ui-panel-title">Top differentiators</span>
            <span class="ui-panel-meta">Ranked by lift, then gap</span>
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    _render_differentiator_table_with_show_more(
        sorted_diffs,
        winner_label,
        laggard_label,
    )
    _render_paginated_outcome_diff_insights(
        seg,
        sorted_diffs,
        winner_label,
        laggard_label,
        seg.outcome_question_id,
    )


def _render_smart_cross_cut_suggestions_panel(
    seg: OutcomeSegmentationResult,
) -> None:
    app = _require_streamlit()
    app.markdown("### Smart Cross-cut Suggestions")
    app.caption("Ranked by relevance to your outcome variable")

    if not app.session_state.get("cross_cut_results"):
        app.info(
            "Run cross cuts in Section 4 first to see outcome-ranked "
            "suggestions here."
        )
        return

    suggestions = app.session_state.get("cross_cut_suggestions", [])
    if not suggestions:
        app.info("No rule-based cross-cut suggestions are available.")
        return

    scored_suggestions = score_suggestions_for_outcome(suggestions, seg)
    schema = app.session_state.get("schema")
    for suggestion in scored_suggestions[:10]:
        analysis_type = suggestion.analysis_type
        analysis_type_label = getattr(analysis_type, "value", str(analysis_type))
        with app.expander(
            f"[{suggestion.outcome_relevance_score:.2f}] "
            f"{_suggestion_label(suggestion, schema)}"
        ):
            app.markdown(
                f"**Business question:** {suggestion.business_question}"
            )
            app.markdown(
                "**Source questions:** "
                f"{', '.join(suggestion.source_question_ids)}"
            )
            app.markdown(
                "**Outcome relevance:** "
                f"{suggestion.outcome_relevance_score:.2f}"
            )
            app.markdown(f"**Analysis type:** {analysis_type_label}")


def _render_winner_profile_panel(seg: OutcomeSegmentationResult) -> None:
    app = _require_streamlit()
    app.markdown(
        f'<h3 style="display:flex;align-items:center;margin:1rem 0;">'
        f'{_icon("winner", "#CC0000")}Top Group Profile</h3>',
        unsafe_allow_html=True,
    )

    profile = seg.winner_profile
    if not profile.defining_traits:
        app.warning(
            "Not enough strong differentiators to build a top-group profile. "
            "Try selecting a different outcome variable or segment definition."
        )
        return

    app.markdown(
        f"**{profile.winner_label} Profile** "
        f"(n={profile.winner_n} vs {profile.loser_label} n={profile.loser_n})"
    )

    for trait in profile.defining_traits:
        app.markdown(f"#### {trait.question_id}")
        app.caption(trait.question_text)

        trait_payload = {
            "question_text": trait.question_text,
            "option_label": trait.option_label,
            "winner_rate": trait.winner_rate,
            "loser_rate": trait.loser_rate,
            "lift": trait.lift,
            "rate_gap": trait.rate_gap,
            "winner_label": profile.winner_label,
            "laggard_label": profile.loser_label,
            "laggard_top_option_label": trait.laggard_top_option_label,
            "laggard_top_option_winner_rate": trait.laggard_top_option_winner_rate,
            "laggard_top_option_loser_rate": trait.laggard_top_option_loser_rate,
        }

        try:
            trait_insight = generate_insight(
                table_payload=trait_payload,
                table_kind="winner_profile_trait",
                title_hint=trait.question_id,
                cache=_INSIGHT_CACHE,
            )
            _render_insight_card(trait_insight, label="")
            _render_trait_backing_table_expander(seg, trait)
        except Exception:
            pass

        app.markdown(
            f'<div style="font-size:11px; color:#888; text-transform:uppercase; '
            f'letter-spacing:0.05em; margin-top:8px; margin-bottom:4px;">'
            f'{html.escape(profile.winner_label)}s chose: '
            f'<strong style="color:#1A1A1A;">'
            f'{html.escape(trait.option_label)}</strong></div>',
            unsafe_allow_html=True,
        )
        col_a1, col_a2, col_a3 = app.columns(3)
        col_a1.metric(profile.winner_label, f"{trait.winner_rate:.1%}")
        col_a2.metric(profile.loser_label, f"{trait.loser_rate:.1%}")
        col_a3.metric("Gap", f"{trait.rate_gap:+.1%}")

        if trait.laggard_top_option_label:
            laggard_gap = (
                trait.laggard_top_option_loser_rate
                - trait.laggard_top_option_winner_rate
            )
            app.markdown(
                f'<div style="font-size:11px; color:#888; text-transform:uppercase; '
                f'letter-spacing:0.05em; margin-top:8px; margin-bottom:4px;">'
                f'{html.escape(profile.loser_label)}s chose: '
                f'<strong style="color:#1A1A1A;">'
                f'{html.escape(trait.laggard_top_option_label)}</strong></div>',
                unsafe_allow_html=True,
            )
            col_b1, col_b2, col_b3 = app.columns(3)
            col_b1.metric(
                profile.winner_label,
                f"{trait.laggard_top_option_winner_rate:.1%}",
            )
            col_b2.metric(
                profile.loser_label,
                f"{trait.laggard_top_option_loser_rate:.1%}",
            )
            col_b3.metric("Gap", f"{laggard_gap:+.1%}")

        app.divider()


def _render_trait_backing_table_expander(
    seg: OutcomeSegmentationResult,
    trait: Any,
) -> None:
    app = _require_streamlit()
    rows = _winner_profile_trait_backing_rows(seg, trait)
    if not rows:
        return

    with app.expander("Show table", expanded=False):
        app.dataframe(rows, use_container_width=True, hide_index=True)


def _winner_profile_trait_backing_rows(
    seg: OutcomeSegmentationResult,
    question_id: str | Any,
) -> list[dict[str, Any]]:
    app = _require_streamlit()
    schema = app.session_state.get("schema")
    dataframe = app.session_state.get("active_df")
    if dataframe is None:
        dataframe = app.session_state.get("decoded_df")
    if schema is None or dataframe is None:
        return []

    question_id = str(getattr(question_id, "question_id", question_id))
    question = schema.get_question(question_id)
    if question is None:
        return []

    from src.outcome_segmentation import _build_segment_masks

    outcome_spec = schema.get_question(seg.outcome_question_id)
    try:
        winner_mask, laggard_mask, valid_mask, _warnings = _build_segment_masks(
            dataframe,
            outcome_spec,
            seg.segment_definition,
            respondent_id_column=getattr(schema, "respondent_id_column", None),
        )
    except Exception:
        return []

    option_rows = _winner_profile_trait_option_masks(dataframe, question)
    if not option_rows:
        return []

    others_mask = valid_mask & ~winner_mask & ~laggard_mask
    rows: list[dict[str, Any]] = []
    winner_label = seg.segment_definition.winner_label
    laggard_label = seg.segment_definition.loser_label
    for option_label, option_mask, base_valid_mask in option_rows:
        winner_count, winner_pct = _cohort_option_stats(
            option_mask,
            winner_mask & base_valid_mask,
        )
        laggard_count, laggard_pct = _cohort_option_stats(
            option_mask,
            laggard_mask & base_valid_mask,
        )
        others_count, others_pct = _cohort_option_stats(
            option_mask,
            others_mask & base_valid_mask,
        )
        total_count, total_pct = _cohort_option_stats(
            option_mask,
            valid_mask & base_valid_mask,
        )
        rows.append(
            {
                "Option": option_label,
                f"{winner_label} count": winner_count,
                f"{winner_label} %": _format_rate(winner_pct),
                f"{laggard_label} count": laggard_count,
                f"{laggard_label} %": _format_rate(laggard_pct),
                "Others count": others_count,
                "Others %": _format_rate(others_pct),
                "Total count": total_count,
                "Total %": _format_rate(total_pct),
                "Gap": _format_rate(winner_pct - laggard_pct, signed=True),
                "Lift": _format_lift(winner_pct, laggard_pct),
            }
        )
    return rows


def _winner_profile_trait_option_masks(
    dataframe: pd.DataFrame,
    question: Any,
) -> list[tuple[str, pd.Series, pd.Series]]:
    if question.question_type is QuestionType.SINGLE_SELECT:
        column = _question_primary_column(question)
        if column not in dataframe.columns:
            return []
        series = dataframe[column]
        base_valid_mask = series.notna()
        return [
            (
                str(label),
                _option_value_mask(series, code, label),
                base_valid_mask,
            )
            for code, label in (question.option_map or {}).items()
        ]

    if question.question_type is QuestionType.MULTI_SELECT_BINARY:
        base_valid_mask = pd.Series(True, index=dataframe.index)
        rows: list[tuple[str, pd.Series, pd.Series]] = []
        for column in question.raw_columns:
            if column not in dataframe.columns:
                continue
            label = str((question.option_map or {}).get(column, column))
            option_mask = (dataframe[column] != 0) & dataframe[column].notna()
            rows.append((label, option_mask, base_valid_mask))
        return rows

    if question.question_type is QuestionType.DIRECT_NUMERIC:
        column = _question_primary_column(question)
        if column not in dataframe.columns:
            return []
        numeric = pd.to_numeric(dataframe[column], errors="coerce")
        base_valid_mask = numeric.notna()
        try:
            binned = pd.qcut(numeric, q=4, labels=False, duplicates="drop")
        except ValueError:
            return []
        valid_bins = sorted(binned.dropna().unique())
        if len(valid_bins) < 2:
            return []
        max_bin = int(max(valid_bins))
        return [
            (
                _quartile_label(int(bin_id), max_bin),
                binned == bin_id,
                base_valid_mask,
            )
            for bin_id in valid_bins
        ]

    return []


def _question_primary_column(question: Any) -> str:
    return question.raw_columns[0] if question.raw_columns else question.canonical_id


def _option_value_mask(series: pd.Series, code: object, label: object) -> pd.Series:
    mask = series == code
    mask = mask | (series == label)
    code_text = str(code)
    label_text = str(label)
    return mask | (series.astype("string") == code_text) | (
        series.astype("string") == label_text
    )


def _cohort_option_stats(
    option_mask: pd.Series,
    cohort_valid_mask: pd.Series,
) -> tuple[int, float]:
    denominator = int(cohort_valid_mask.sum())
    count = int((option_mask & cohort_valid_mask).sum())
    rate = count / denominator if denominator else 0.0
    return count, rate


def _format_rate(value: float, *, signed: bool = False) -> str:
    prefix = "+" if signed and value > 0 else ""
    return f"{prefix}{value:.1%}"


def _format_lift(winner_pct: float, laggard_pct: float) -> str:
    if laggard_pct == 0:
        return "\u221e" if winner_pct > 0 else "0.00x"
    return f"{winner_pct / laggard_pct:.2f}x"


def _quartile_label(bin_id: int, max_bin: int) -> str:
    if bin_id == 0:
        return "Q1 (bottom quartile)"
    if bin_id == max_bin:
        return "Q4 (top quartile)"
    return f"Q{bin_id + 1}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def _render_sidebar() -> None:
    """Branded, structured sidebar showing only what is useful mid-session."""
    import html as _html

    app = _require_streamlit()
    with app.sidebar:
        # ---- LOGO / TITLE ---------------------------------------------------
        app.markdown(
            """
            <div style="padding:16px 0 8px 0;border-bottom:3px solid #CC0000;
              margin-bottom:16px;">
              <div style="font-size:16px;font-weight:700;color:#0A0A0A;
                font-family:Arial;letter-spacing:0.03em;">
                Survey Analysis Engine
              </div>
              <div style="font-size:10px;color:#888;font-family:Arial;
                margin-top:2px;">Bain &amp; Company</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        # Survey assistant lives in the sidebar, just below the tool name.
        render_chat_panel()

        if not app.session_state.get("run_complete"):
            with app.expander("Getting started", expanded=True):
                app.markdown(
                    f"""
                    <div class="gs-wrap">
                      <div class="gs-step">
                        <div class="gs-badge done">1</div>
                        <div class="gs-body">
                          <div class="gs-step-title">Upload your survey</div>
                          <div class="gs-step-desc">Add your raw data + data map, or a single combined Excel.</div>
                        </div>
                      </div>
                      <div class="gs-step todo">
                        <div class="gs-badge todo">2</div>
                        <div class="gs-body">
                          <div class="gs-step-title">Run the analysis</div>
                          <div class="gs-step-desc">Hit Run analysis \u2014 every cut is computed in seconds.</div>
                        </div>
                      </div>
                      <div class="gs-step todo">
                        <div class="gs-badge todo">3</div>
                        <div class="gs-body">
                          <div class="gs-step-title">Review &amp; download</div>
                          <div class="gs-step-desc">Explore results here, then download the workbook.</div>
                        </div>
                      </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        if app.session_state.get("run_complete"):
            lr = app.session_state.get("load_report")
            gf_stats = app.session_state.get("global_filter_stats")
            results = app.session_state.get("results", [])

            # ---- QUESTIONS (master-detail nav) ----------------------------
            app.markdown("**Questions**")
            selected_qid = app.session_state.get("selected_question_id")
            if selected_qid is None and results:
                selected_qid = results[0].question_id
                app.session_state["selected_question_id"] = selected_qid

            app.text_input(
                "Search",
                placeholder="Filter by ID or text",
                key="sidebar_q_search",
                label_visibility="collapsed",
            )
            needle = (
                app.session_state.get("sidebar_q_search") or ""
            ).strip().lower()

            schema_obj = app.session_state.get("schema")
            shown = 0
            for result in results:
                if schema_obj is None:
                    continue
                spec = schema_obj.get_question(result.question_id)
                if spec is None:
                    continue
                if needle:
                    hay = f"{spec.canonical_id} {spec.question_text or ''}".lower()
                    if needle not in hay:
                        continue
                shown += 1
                label = (
                    f"{spec.canonical_id} \u2014 "
                    f"{spec.question_text or ''}".strip()
                )
                btn_type = (
                    "primary"
                    if result.question_id == selected_qid
                    else "secondary"
                )
                if app.button(
                    label,
                    key=f"sidebar_qbtn_{result.question_id}",
                    type=btn_type,
                    use_container_width=True,
                ):
                    app.session_state["selected_question_id"] = (
                        result.question_id
                    )
                    # Stay on the main post-run view so the global filter stays
                    # visible above the question's single cut.
                    _set_current_nav_view("upload")
                    app.rerun()
            if needle and shown == 0:
                app.caption(f"No questions match '{needle}'.")
            else:
                app.caption(f"{shown} of {len(results)} shown")

            app.markdown("**Navigate**")
            app.markdown(
                """
                <div style="font-size:11px;font-family:Arial;line-height:2.2;
                  color:#333;">
                  <a href="#section-1" style="color:#CC0000;
                    text-decoration:none;">↑ 1. Upload</a><br>
                  <a href="#section-2" style="color:#CC0000;
                    text-decoration:none;">⊕ 2. Global Filter</a><br>
                  <a href="#section-3" style="color:#CC0000;
                    text-decoration:none;">— 3. Single Cuts</a><br>
                  <a href="#section-4" style="color:#CC0000;
                    text-decoration:none;">╫ 4. Cross Cuts</a><br>
                  <a href="#section-5" style="color:#CC0000;
                    text-decoration:none;">↓ 5. Downloads</a>
                </div>
                """,
                unsafe_allow_html=True,
            )

        # Session summary — pinned to the BOTTOM of the sidebar.
        if app.session_state.get("run_complete"):
            lr = app.session_state.get("load_report")
            gf_stats = app.session_state.get("global_filter_stats")
            results = app.session_state.get("results", [])
            app.markdown("**Session**")

            if lr is not None:
                app.markdown(
                    f"""
                    <div style="font-size:10px;color:#666;font-family:Arial;
                      line-height:1.9;background:#F8F8F8;padding:10px;
                      border-left:3px solid #E0E0E0;margin-bottom:12px;">
                      <b>File</b><br>{_html.escape(str(lr.raw_data_source))}<br>
                      <b>Input type</b><br>
                      {_html.escape(lr.scenario.replace('_', ' '))}<br>
                      <b>Respondents</b><br>{lr.raw_rows:,}<br>
                      <b>Questions</b><br>{lr.questions_parsed}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            gf_state = app.session_state.get("global_filter_state")
            if (
                gf_state is not None
                and gf_state.is_active()
                and gf_stats
            ):
                app.markdown(
                    f"""
                    <div style="background:#FFF5F5;border-left:3px solid #CC0000;
                      padding:10px;font-size:10px;font-family:Arial;
                      line-height:1.9;margin-bottom:12px;">
                      <b style="color:#CC0000;">● Global Filter Active</b><br>
                      {_html.escape(gf_state.description())}<br>
                      <b>{gf_stats.get('rows_after', 0):,}</b>&nbsp;of&nbsp;
                      {gf_stats.get('rows_before', 0):,}&nbsp;respondents
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            else:
                app.markdown(
                    """
                    <div style="font-size:10px;color:#888;font-family:Arial;
                      padding:8px 0;margin-bottom:8px;">
                      ○ No global filter · full dataset
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            skips = app.session_state.get("skips", [])
            errors = [s for s in skips if s.skip_reason == "calculation_error"]
            log = app.session_state.get("log")
            log_len = len(log) if log else 0
            cc_results = app.session_state.get("cross_cut_results", [])
            err_color = "#CC0000" if errors else "#2E7D32"
            app.markdown(
                f"""
                <div style="font-size:10px;color:#666;font-family:Arial;
                  line-height:2.0;border-top:1px solid #E0E0E0;
                  padding-top:10px;margin-bottom:12px;">
                  <b>Single cuts</b>&nbsp;{len(results)}<br>
                  <b>Skipped</b>&nbsp;{len(skips)}<br>
                  <b>Errors</b>&nbsp;
                  <span style="color:{err_color};">{len(errors)}</span><br>
                  <b>Cross cuts</b>&nbsp;{len(cc_results)}
                </div>
                """,
                unsafe_allow_html=True,
            )


        app.markdown("---")
        with app.expander("Help", expanded=False):
            app.markdown(
                """
                <div style="font-size:10px;font-family:Arial;color:#444;
                  line-height:1.9;">
                  <b>Supported file types</b><br>CSV, XLSX, DOCX<br><br>
                  <b>Three input scenarios</b><br>
                  A: Raw data + data map<br>
                  B: Combined XLSX (2 sheets)<br>
                  C: Raw data + Word survey doc<br><br>
                  <b>Filters</b><br>
                  Global filter restricts everything.<br>
                  Per-question filter applies only<br>
                  to that single question.<br><br>
                  <b>Outlier flags</b><br>
                  ⬆ Red = high outlier (&gt;2σ above mean)<br>
                  ↓ Amber = low outlier (&gt;1.5σ below)
                </div>
                """,
                unsafe_allow_html=True,
            )


def _render_nav_bar() -> None:
    """Inject route-style nav tabs into the merged red header."""
    app = _require_streamlit()

    n_singlecuts = len(app.session_state.get("results", []))
    n_crosscuts = len(app.session_state.get("cross_cut_results", []))
    seg = app.session_state.get("segmentation_result")
    n_diffs = len(seg.differentiators) if seg else 0
    gf_state = app.session_state.get("global_filter_state")
    n_filters = len(gf_state.filters) if gf_state is not None and hasattr(gf_state, "filters") else 0
    has_data = bool(app.session_state.get("run_complete")) or app.session_state.get("schema") is not None
    active_view = _current_nav_view()

    def _badge(badge, badge_type):
        if not badge:
            return ""
        if badge_type == "check":
            return (
                '<span class="nav-badge" '
                'style="background:rgba(76,175,80,0.95);color:#FFF;">\u2713</span>'
            )
        return f'<span class="nav-badge">{html.escape(badge)}</span>'

    def _tab(view: str, label: str, badge: Any = None, badge_type: str | None = None) -> str:
        active = " active" if active_view == view else ""
        return (
            f'<a href="{html.escape(_nav_href(view))}" data-view="{html.escape(view)}" onclick="{html.escape(_nav_onclick(view))}" class="nav-tab{active}">'
            f'{html.escape(label)}{_badge(badge, badge_type)}</a>'
        )

    def _menu_item(view: str, label: str, badge: Any = None, badge_type: str | None = None) -> str:
        active = " active" if active_view == view else ""
        return (
            f'<a href="{html.escape(_nav_href(view))}" data-view="{html.escape(view)}" onclick="{html.escape(_nav_onclick(view))}" class="nav-more-item{active}">'
            f'{html.escape(label)}{_badge(badge, badge_type)}</a>'
        )

    create_active = " active" if active_view in {"single", "cross"} else ""
    create_badge = str(n_singlecuts) if n_singlecuts else None
    create_menu = (
        _menu_item("single", "Single cut", str(n_singlecuts) if n_singlecuts else None, None)
        + _menu_item("cross", "Cross cuts", str(n_crosscuts) if n_crosscuts else None, None)
    )
    more_items = _menu_item("filters", "Apply filters", str(n_filters) if n_filters else None, None)
    nav_items = (
        _tab("upload", "Upload survey", "\u2713" if has_data else None, "check")
        + (
            '<div class="nav-more">'
            f'<span class="nav-tab nav-more-toggle{create_active}">'
            f'Create Cuts \u25BE{_badge(create_badge, None)}</span>'
            f'<div class="nav-more-menu">{create_menu}</div>'
            '</div>'
        )
        + _tab("outcome", "Outcome Segmentation", str(n_diffs) if n_diffs else None, None)
        + (_tab("download", "Download workbook", None, None) if app.session_state.get("run_complete") else "")
        + _tab("filters", "Apply global filter", str(n_filters) if n_filters else None, None)
    )

    try:
        from streamlit.components.v1 import html as components_html
    except Exception:
        components_html = None

    for _v in _NAV_VIEWS:
        if app.button(f"navjump_{_v}", key=f"_navjump_{_v}"):
            _set_current_nav_view(_v)
            app.rerun()

    nav_json = json.dumps(nav_items)
    script = f"""
<script>
(function() {{
    var navHTML = {nav_json};
    var doc = window.parent.document;

    function injectNavTabs() {{
        var navContainer = doc.getElementById('custom-header-nav');
        if (!navContainer) return false;
        if (navContainer.dataset.signature !== navHTML) {{
            navContainer.innerHTML = navHTML;
            navContainer.dataset.signature = navHTML;
        }}
        if (!doc.defaultView.__surveyNavBridgeInstalled) {{
            doc.defaultView.__surveyNavBridgeInstalled = true;
            doc.addEventListener('click', function(e) {{
                var target = e.target;
                var link = target && target.closest ? target.closest('.nav-tab[data-view], .nav-more-item[data-view]') : null;
                if (!link) return;
                var view = link.getAttribute('data-view');
                if (!view) return;
                e.preventDefault();
                var buttonLabel = ('navjump_' + view).toLowerCase();
                var buttons = doc.querySelectorAll('button');
                for (var i = 0; i < buttons.length; i++) {{
                    if ((buttons[i].textContent || '').trim().toLowerCase().indexOf(buttonLabel) === 0) {{
                        e.preventDefault();
                        buttons[i].click();
                        return;
                    }}
                }}
            }}, true);
        }}
        Array.prototype.forEach.call(doc.querySelectorAll('button'), function(button) {{
            var _t = (button.textContent || '').trim().toLowerCase();
            if (_t.indexOf('navjump_') === 0 || _t.indexOf('chat_toggle_signal') === 0) {{
                // Hide visually but keep CLICKABLE — display:none elements cannot be
                // clicked, which was breaking the nav bridge. Move off-screen instead.
                var wrap = button.closest('[data-testid="stButton"]') || button;
                wrap.style.position = 'absolute';
                wrap.style.width = '1px'; wrap.style.height = '1px';
                wrap.style.overflow = 'hidden'; wrap.style.opacity = '0';
                wrap.style.left = '-9999px'; wrap.style.pointerEvents = 'auto';
                // Also collapse the OUTER element container so it reserves no
                // vertical space (this was the large empty gap at the top).
                var outer = button.closest('[data-testid="stElementContainer"]')
                    || button.closest('[data-testid="element-container"]');
                if (outer) {{
                    outer.style.height = '0px'; outer.style.minHeight = '0px';
                    outer.style.margin = '0px'; outer.style.padding = '0px';
                    outer.style.overflow = 'hidden';
                }}
            }}
        }});
        return true;
    }}

    injectNavTabs();
    setTimeout(injectNavTabs, 100);
    setTimeout(injectNavTabs, 500);
    setTimeout(injectNavTabs, 1500);
    setInterval(injectNavTabs, 1000);
}})();
</script>
"""
    if components_html is not None:
        components_html(script, height=0)
    else:
        app.markdown(script, unsafe_allow_html=True)
    return

def _render_stat_tiles(stats: list[tuple[str, int, str]]) -> None:
    """Render count-up stat tiles. stats = [(icon, value:int, label), ...].

    The final value is in the HTML (so it shows even if JS is blocked); the
    script animates 0 -> value once per distinct value-set.
    """
    app = _require_streamlit()
    tiles = ['<div class="stat-row">']
    for icon, value, label in stats:
        v = int(value)
        # icon can be an SVG icon name (resolved via _svg_icon) or raw HTML string for back-compat.
        icon_html = _svg_icon(icon, size=20, color="#CC0000") if icon in _BAIN_ICONS else str(icon)
        tiles.append(
            '<div class="stat-tile">'
            f'<div class="stat-icon">{icon_html}</div>'
            f'<div class="stat-num" data-target="{v}">{v:,}</div>'
            f'<div class="stat-label">{html.escape(label)}</div></div>'
        )
    tiles.append("</div>")
    app.markdown("".join(tiles), unsafe_allow_html=True)

    script = """
<script>
(function(){
  var doc = window.parent.document;
  function run(){
    var els = doc.querySelectorAll('.stat-num[data-target]');
    if(!els.length) return false;
    var sig = Array.prototype.map.call(els, function(e){return e.getAttribute('data-target');}).join(',');
    if(doc.defaultView.__statSig === sig) return true;
    doc.defaultView.__statSig = sig;
    els.forEach(function(el){
      var target = parseInt(el.getAttribute('data-target'),10) || 0;
      var dur = 700, start = null;
      function step(ts){
        if(!start) start = ts;
        var p = Math.min((ts-start)/dur, 1);
        el.textContent = Math.round(p*target).toLocaleString();
        if(p<1){ requestAnimationFrame(step); } else { el.textContent = target.toLocaleString(); }
      }
      requestAnimationFrame(step);
    });
    return true;
  }
  if(!run()){ var n=0; var iv=setInterval(function(){ if(run()||++n>25) clearInterval(iv); },150); }
})();
</script>
"""
    try:
        from streamlit.components.v1 import html as components_html
        components_html(script, height=0)
    except Exception:
        pass


def _render_journey_rail() -> None:
    """A big, highlighted 4-step guide: Upload -> Configure -> Run & review -> Download.

    Highlights whichever step matches the current view so a first-time user always
    knows where they are and what's next.
    """
    app = _require_streamlit()
    # "Detected" = files uploaded and parsed (wiz_schema set during _wizard_prepare),
    # OR a full run completed. Either advances the rail to the Configure step.
    detected = (
        app.session_state.get("wiz_schema") is not None
        or app.session_state.get("schema") is not None
        or bool(app.session_state.get("wiz_output_structure"))
    )
    has_data = detected
    done = bool(app.session_state.get("run_complete"))
    view = _current_nav_view()

    # Map the 4 guide steps to views; figure out which is "active".
    steps = [
        ("upload",   "1", "Upload survey",        "Add your raw data + data map"),
        ("upload",   "2", "Configure output",     "Categories, filters, cross-cuts"),
        ("single",   "3", "Run &amp; review",     "Single cuts, cross cuts, segmentation"),
        ("download", "4", "Download",             "Consultant-ready workbook"),
    ]
    # Determine active index from the current view + progress.
    if view == "download":
        active = 3
    elif view in ("single", "cross", "outcome", "filters") or done:
        active = 2
    elif has_data:
        active = 1
    else:
        active = 0

    parts = ['<div class="journey journey-big">']
    for i, (_v, num, title, sub) in enumerate(steps):
        if i < active:
            stt = "done"; dot = "\u2713"
        elif i == active:
            stt = "active"; dot = num
        else:
            stt = "todo"; dot = num
        parts.append(
            f'<div class="journey-step {stt}">'
            f'<div class="journey-dot {stt}">{dot}</div>'
            f'<div class="journey-text"><span class="journey-title">{title}</span>'
            f'<span class="journey-sub">{sub}</span></div></div>'
        )
        if i < len(steps) - 1:
            line_state = "done" if i < active else "todo"
            parts.append(f'<div class="journey-line {line_state}"></div>')
    parts.append("</div>")
    # "What's next" banner under the rail.
    next_hint = {
        0: "\u2192 Next: upload your raw data file and data map below.",
        1: "\u2192 Next: choose how your Excel output is organised, then set filters &amp; cross-cuts.",
        2: "",
        3: "\u2192 You're set \u2014 generate and download your workbook below.",
    }.get(active, "")
    if next_hint:
        parts_banner = (
            f'<div class="journey-next">{next_hint}</div>'
        )
    else:
        parts_banner = ""
    app.markdown("".join(parts) + parts_banner, unsafe_allow_html=True)


def main() -> None:
    app = _require_streamlit()
    app.set_page_config(
        page_title=APP_TITLE,
        page_icon="\u25A0",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    _initialise_session_state()
    _inject_global_css()
    _drain_pending_actions()
    _render_sidebar()

    _render_nav_bar()

    _render_journey_rail()

    gf_error = app.session_state.get("global_filter_error")
    if gf_error:
        app.error(f"Global filter failed: {gf_error}")

    view = _current_nav_view()
    if view == "upload":
        _section_upload()
    elif view == "filters":
        _section_global_filter()
    elif view == "single":
        _section_single_cuts()
    elif view == "cross":
        _section_cross_cuts()
    elif view == "outcome":
        _section_ai_analysis()
    elif view == "download":
        _section_downloads()
    else:
        _section_upload()

    # Persistent FULL download panel — pinned as a footer at the bottom of EVERY
    # page, but ONLY after an analysis has run (no download step before that).
    if view != "download" and app.session_state.get("run_complete"):
        _section_downloads()


# ---------------------------------------------------------------------------
# Setup wizard
# ---------------------------------------------------------------------------


def _render_output_structure_gate() -> bool:
    """WS1: ask how the Excel output should be organised BEFORE building categories.

    Returns True once the user has chosen (so the wizard may proceed). The choice
    is stored in session_state["wiz_output_structure"] as "grouped" | "one_sheet".
    Truthful preview: "one_sheet" means questions are NOT grouped into topic
    categories \u2014 they go into a single "All Questions" sheet (a separate
    "Demographics" sheet is still added when demographic questions exist).
    """
    app = _require_streamlit()
    if app.session_state.get("wiz_output_structure"):
        # Keep the heading visible (with the current choice) so the user always
        # sees where they are, even after picking.
        choice = app.session_state.get("wiz_output_structure")
        choice_label = "Grouped into categories" if choice == "grouped" else "One sheet (no categories)"
        _section_header("2", "CONFIGURE YOUR EXCEL OUTPUT STRUCTURE")
        _red_instruction_box(
            "Step 2 \u2014 Configure your output",
            "Choose how the workbook is organised, then set filters and cross-cuts. "
            "Defaults are fine if you're unsure.",
        )
        app.caption(f"Current choice: **{choice_label}**. You can revisit this anytime.")
        if app.button("Change output structure", key="wiz_change_structure"):
            app.session_state["wiz_output_structure"] = None
            app.session_state["wiz_structure_user_touched"] = True
            app.session_state.pop("wiz_skip_categories", None)
            _wizard_rerun(app)
        return True

    _section_header("2", "CONFIGURE YOUR EXCEL OUTPUT STRUCTURE")
    _red_instruction_box(
        "Step 2 \u2014 Configure your output",
        "Choose how the workbook is organised, then set filters and cross-cuts. "
        "Defaults are fine if you're unsure.",
    )

    col_a, col_b = app.columns(2)
    with col_a:
        app.markdown(
            "<div style='border:1px solid #E0E0E0;border-top:4px solid #CC0000;"
            "border-radius:10px;padding:14px 16px;min-height:230px;'>"
            "<div style='display:flex;align-items:center;gap:8px;margin-bottom:6px;'>"
            "<span style='font-weight:700;font-size:14px;'>Keep all questions on one sheet</span>"
            "<span style='background:#CC0000;color:#fff;font-size:10px;font-weight:700;"
            "padding:2px 7px;border-radius:10px;letter-spacing:0.03em;'>DEFAULT</span></div>"
            "<div style='font-size:12px;color:#444;line-height:1.5;'>"
            "All questions go onto a single <b>All Questions</b> sheet in the order they "
            "appear \u2014 simplest layout. (A separate <b>Demographics</b> sheet is still "
            "added if your survey has demographic questions.)"
            "</div>"
            "<div style='margin-top:12px;font-size:11px;color:#888;'>Excel tabs look like:</div>"
            "<div style='margin-top:4px;'>"
            "<span style='background:#CC0000;color:#fff;font-size:11px;padding:3px 9px;"
            "border-radius:4px 4px 0 0;margin-right:2px;'>All Questions</span>"
            "<span style='background:#F0F0F0;color:#333;font-size:11px;padding:3px 9px;"
            "border-radius:4px 4px 0 0;'>Demographics</span>"
            "</div></div>",
            unsafe_allow_html=True,
        )
        if app.button("Keep All Questions on One Sheet", key="wiz_struct_one",
                      type="primary", use_container_width=True):
            app.session_state["wiz_output_structure"] = "one_sheet"
            # One-sheet path: skip topic categorisation entirely.
            app.session_state["wiz_category_assignments"] = None
            app.session_state["wiz_skip_categories"] = True
            _wizard_rerun(app)
    with col_b:
        app.markdown(
            "<div style='border:1px solid #E0E0E0;border-top:4px solid #888;"
            "border-radius:10px;padding:14px 16px;min-height:230px;'>"
            "<div style='font-weight:700;font-size:14px;margin-bottom:6px;'>"
            "Group into categories</div>"
            "<div style='font-size:12px;color:#444;line-height:1.5;'>"
            "Questions are sorted into themed sheets \u2014 e.g. <b>Pricing</b>, "
            "<b>Brand</b>, <b>Loyalty</b> \u2014 so a reader jumps straight to a topic. "
            "You can rename categories and move questions between them in the next step."
            "</div>"
            "<div style='margin-top:12px;font-size:11px;color:#888;'>Excel tabs look like:</div>"
            "<div style='margin-top:4px;'>"
            "<span style='background:#CC0000;color:#fff;font-size:11px;padding:3px 9px;"
            "border-radius:4px 4px 0 0;margin-right:2px;'>Pricing</span>"
            "<span style='background:#F0F0F0;color:#333;font-size:11px;padding:3px 9px;"
            "border-radius:4px 4px 0 0;margin-right:2px;'>Brand</span>"
            "<span style='background:#F0F0F0;color:#333;font-size:11px;padding:3px 9px;"
            "border-radius:4px 4px 0 0;'>Loyalty</span>"
            "</div></div>",
            unsafe_allow_html=True,
        )
        if app.button("Use categories", key="wiz_struct_grouped", use_container_width=True):
            app.session_state["wiz_output_structure"] = "grouped"
            _wizard_rerun(app)
    return False


def _render_setup_wizard(uploaded_files: list[Any]) -> None:
    app = _require_streamlit()
    _render_wizard_css()
    if not _render_output_structure_gate():
        return
    # One-sheet mode skips the CATEGORIES step but still needs the schema built
    # (the filters step reads wiz_schema). So prepare, then null the category
    # assignments so categorisation is bypassed at export time.
    if app.session_state.get("wiz_skip_categories"):
        if app.session_state.get("wiz_schema") is None:
            _wizard_prepare(uploaded_files)
            app.session_state["wiz_category_assignments"] = None  # keep one-sheet
        if app.session_state.get("wiz_schema") is None:
            return
        app.markdown("<div class='wiz-shell'>", unsafe_allow_html=True)
        _render_wizard_top_nav()
        app.info(
            "Output set to a single **All Questions** sheet (plus Demographics if "
            "present). No category step needed \u2014 continue to filters."
        )
        step = int(app.session_state.get("wiz_step", 2) or 2)
        if step < 2:
            step = 2
            app.session_state["wiz_step"] = 2
        if step == 2:
            _render_wizard_step_filters()
        else:
            _render_wizard_step_crosscut()
        _render_wizard_nav("bottom")
        app.markdown("</div>", unsafe_allow_html=True)
        return
    if app.session_state.get("wiz_category_assignments") is None:
        _wizard_prepare(uploaded_files)
    if app.session_state.get("wiz_category_assignments") is None:
        return

    app.markdown("<div class='wiz-shell'>", unsafe_allow_html=True)
    _render_wizard_top_nav()
    step = int(app.session_state.get("wiz_step", 1))
    if step == 1:
        _render_wizard_step_categories()
    elif step == 2:
        _render_wizard_step_filters()
    else:
        _render_wizard_step_crosscut()
    _render_wizard_nav("bottom")
    app.markdown("</div>", unsafe_allow_html=True)


def _wizard_prepare(uploaded_files: list[Any]) -> None:
    app = _require_streamlit()
    try:
        from src.ai_insights import categorize_questions_into_themes
        from src.question_classifier import classify_questions
        from src.ui.wizard import (
            category_assignments_from_themes,
            selected_demographics_from_schema,
        )

        data_map, decoded_df, load_report = _load_survey_inputs_for_current_format(uploaded_files)
        schema = classify_questions(
            data_map,
            decoded_df.columns.tolist(),
            respondent_id_column="record",
            total_respondents=len(decoded_df),
            source_rawdata_path=load_report.raw_data_source,
        )
        questions_for_themes = [
            {
                "question_id": question.canonical_id,
                "question_text": question.question_text,
                "question_type": question.question_type.value,
                "is_demographic": getattr(question, "is_demographic", False),
            }
            for question in schema.questions
        ]
        try:
            auto_themes = categorize_questions_into_themes(
                questions_for_themes,
                cache=_INSIGHT_CACHE,
            )
        except Exception:
            auto_themes = {
                "themes": [
                    {
                        "name": "All Questions",
                        "question_ids": [q.canonical_id for q in schema.questions if q.analysis_eligible],
                    }
                ],
                "was_template": True,
                "error_message": "",
            }
        app.session_state["wiz_schema"] = schema
        app.session_state["wiz_decoded_df"] = decoded_df
        app.session_state["wiz_load_report"] = load_report
        app.session_state["wiz_category_assignments"] = category_assignments_from_themes(schema, auto_themes)
        app.session_state["wiz_selected_demographics"] = selected_demographics_from_schema(schema)
    except Exception as exc:  # noqa: BLE001
        app.session_state["wiz_category_assignments"] = None
        app.error(f"Setup wizard could not read the uploaded files: {type(exc).__name__}: {exc}")


def _render_wizard_css() -> None:
    _require_streamlit().markdown(
        """
<style>
.wiz-shell { background: #FFFFFF; padding: 24px 0; font-family: Arial, sans-serif; }
.wiz-stepper { display: flex; justify-content: space-between; margin: 24px 0; }
.wiz-step-circle { width: 48px; height: 48px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-weight: 700; font-size: 20px; margin: 0 auto 10px; }
.wiz-step-active { border: 3px solid #CC0000; color: #CC0000; }
.wiz-step-complete { background: #CC0000; color: white; }
.wiz-step-upcoming { border: 2px solid #999; color: #999; }
.wiz-step-label { text-align: center; font-size: 14px; font-weight: 600; font-family: Arial, sans-serif; }
.wiz-category-cell { background: #CC0000; color: white; padding: 12px 16px; font-weight: 700; font-size: 14px; font-family: Arial, sans-serif; border-radius: 4px; }
.wiz-card { border: 1px solid #E5E5E5; border-radius: 6px; padding: 12px; margin: 4px 0; min-height: 92px; }
.wiz-card-id { color: #666; font-size: 11px; font-family: Arial, sans-serif; }
.wiz-card-text { font-size: 13px; font-family: Arial, sans-serif; margin: 4px 0; }
.wiz-card-type { color: #666; font-size: 11px; font-style: italic; font-family: Arial, sans-serif; }
.wiz-divider { border-top: 1px solid #E5E5E5; margin: 16px 0; }
.wiz-stepper { position: relative; }
.wiz-step-circle { box-shadow: 0 1px 2px rgba(16,24,40,0.08); transition: all 0.15s ease; }
.wiz-step-complete { box-shadow: 0 1px 3px rgba(204,0,0,0.25); }
.wiz-step-label { color: #5A5A5A; }
</style>
""",
        unsafe_allow_html=True,
    )


def _render_wizard_stepper() -> None:
    app = _require_streamlit()
    labels = ["Categories", "Filters", "Cross-cuts"]
    current = int(app.session_state.get("wiz_step", 1))
    # Scoped CSS to render these specific buttons as round step circles.
    app.markdown(
        "<style>"
        ".wizstep-row div[data-testid='column'] .stButton button{"
        "width:46px;height:46px;border-radius:50%;font-size:18px;font-weight:700;"
        "padding:0;margin:0 auto;display:flex;align-items:center;justify-content:center;}"
        "</style>",
        unsafe_allow_html=True,
    )
    app.markdown("<div class='wizstep-row'>", unsafe_allow_html=True)
    cols = app.columns(3)
    for index, label in enumerate(labels, start=1):
        with cols[index - 1]:
            is_current = index == index and index == current
            # The circle IS the button (clickable). Primary = current step.
            if app.button(
                str(index),
                key=f"wizstep_circle_{index}",
                type=("primary" if index == current else "secondary"),
            ):
                app.session_state["wiz_step"] = index
                _wizard_rerun(app)
            weight = "700" if index == current else "400"
            color = "#CC0000" if index == current else "#888"
            app.markdown(
                f"<div style='text-align:center;font-size:13px;font-weight:{weight};"
                f"color:{color};font-family:Arial,sans-serif;margin-top:2px;'>{html.escape(label)}</div>",
                unsafe_allow_html=True,
            )
    app.markdown("</div>", unsafe_allow_html=True)


def _render_wizard_top_nav() -> None:
    """Pretty, NON-clickable 3-step indicator (Categories / Filter counts /
    Cross-cut), rendered as one aligned HTML flexbox. Navigation is via the
    Back/Next buttons at the bottom of each step."""
    app = _require_streamlit()
    step = int(app.session_state.get("wiz_step", 1))
    labels = ["Categories", "Filter counts", "Cross-cut"]

    def _circle(n: int) -> str:
        if n < step:
            bg = "linear-gradient(180deg,#E00000,#CC0000)"; col = "#fff"; bd = "none"; mark = "\u2713"
        elif n == step:
            bg = "linear-gradient(180deg,#E00000,#CC0000)"; col = "#fff"; bd = "none"; mark = str(n)
        else:
            bg = "#F4F5F7"; col = "#A8ADB3"; bd = "1.5px solid #E2E4E8"; mark = str(n)
        lab_color = "#CC0000" if n == step else "#9AA0A6"
        lab_weight = "700" if n == step else "500"
        ring = "box-shadow:0 0 0 4px rgba(204,0,0,0.10);" if n == step else "box-shadow:0 1px 3px rgba(16,24,40,0.06);"
        return (
            "<div style='display:flex;flex-direction:column;align-items:center;flex:0 0 auto;width:96px;'>"
            f"<div style='width:38px;height:38px;border-radius:50%;background:{bg};color:{col};"
            f"border:{bd};{ring}display:flex;align-items:center;justify-content:center;"
            f"font-weight:700;font-size:15px;font-family:Arial,sans-serif;'>{mark}</div>"
            f"<div style='margin-top:8px;font-size:11.5px;font-weight:{lab_weight};color:{lab_color};"
            f"font-family:Arial,sans-serif;white-space:nowrap;letter-spacing:0.02em;'>{html.escape(labels[n-1])}</div>"
            "</div>"
        )

    def _conn(after: int) -> str:
        done = after < step
        return (
            "<div style='flex:1 1 40px;height:2px;margin:0 6px;margin-top:18px;border-radius:2px;"
            f"background:{'linear-gradient(90deg,#CC0000,#E00000)' if done else '#E6E8EB'};'></div>"
        )

    # Clickable arrows (Streamlit buttons) flank the HTML stepper.
    app.markdown(
        "<style>"
        ".wizarrow + div [data-testid='stHorizontalBlock']{align-items:center;}"
        ".wizarrow .stButton button{background:transparent !important;border:none !important;"
        "color:#CC0000 !important;font-size:26px !important;font-weight:700 !important;"
        "box-shadow:none !important;padding:0 !important;line-height:1 !important;min-height:40px !important;}"
        ".wizarrow .stButton button:disabled{color:#D8DADE !important;}"
        "</style><div class='wizarrow'></div>",
        unsafe_allow_html=True,
    )
    a_back, a_mid, a_next = app.columns([1, 10, 1])
    with a_back:
        if app.button("\u2039", disabled=(step <= 1), key="wiz_arrow_back", help="Back"):
            app.session_state["wiz_step"] = max(1, step - 1)
            _wizard_rerun(app)
    with a_mid:
        app.markdown(
            "<div style='display:flex;align-items:flex-start;justify-content:center;"
            "gap:6px;max-width:620px;margin:6px auto 10px;'>"
            + _circle(1) + _conn(1) + _circle(2) + _conn(2) + _circle(3)
            + "</div>",
            unsafe_allow_html=True,
        )
    with a_next:
        if app.button("\u203a", disabled=(step >= 3), key="wiz_arrow_next", help="Next"):
            app.session_state["wiz_step"] = min(3, step + 1)
            _wizard_rerun(app)


def _render_wizard_nav(location: str) -> None:
    app = _require_streamlit()
    step = int(app.session_state.get("wiz_step", 1))
    # Right-align the Next button within its column so Back (left) and Next
    # (right) are symmetric across the row.
    app.markdown(
        f"<style>.wiznav-{location} + div [data-testid='stHorizontalBlock'] "
        f"[data-testid='column']:last-child .stButton{{display:flex;justify-content:flex-end;}}"
        f"</style><div class='wiznav-{location}'></div>",
        unsafe_allow_html=True,
    )
    back_col, step_col, next_col = app.columns([1, 2, 1])
    if back_col.button("Back", disabled=(step <= 1), key=f"wiz_back_{location}"):
        app.session_state["wiz_step"] = max(1, step - 1)
        _wizard_rerun(app)
    step_col.markdown(f"<div style='text-align:center;color:#666;font-size:13px;'>Step {step} of 3</div>", unsafe_allow_html=True)
    if step < 3:
        if next_col.button("Next", type="primary", key=f"wiz_next_{location}"):
            app.session_state["wiz_step"] = min(3, step + 1)
            _wizard_rerun(app)


def _wizard_schema():
    return _require_streamlit().session_state.get("wiz_schema")


def _wizard_dataframe():
    return _require_streamlit().session_state.get("wiz_decoded_df")


def _wizard_questions_by_category() -> dict[str, list[Any]]:
    schema = _wizard_schema()
    assignments = dict(_require_streamlit().session_state.get("wiz_category_assignments") or {})
    grouped: dict[str, list[Any]] = {}
    if schema is None:
        return grouped
    questions = {question.canonical_id: question for question in schema.questions}
    for question_id, category in assignments.items():
        question = questions.get(question_id)
        if question is not None:
            grouped.setdefault(category, []).append(question)
    for category in _require_streamlit().session_state.get("wiz_empty_categories", []):
        grouped.setdefault(category, [])
    return grouped


def _render_wizard_step_categories() -> None:
    app = _require_streamlit()
    from src.ui.wizard import question_display_text, question_type_label

    app.markdown(
        """
<style>
.wiz-cat-sidebar-item { display: block; padding: 10px 14px; margin-bottom: 6px; border-radius: 4px; border: 1px solid #E5E5E5; background: #FFFFFF; color: #1A1A1A; font-family: Arial, sans-serif; font-size: 13px; cursor: pointer; }
.wiz-cat-sidebar-item-active { background: #CC0000; color: #FFFFFF; font-weight: 700; border-color: #CC0000; }
.wiz-cat-sidebar-badge { float: right; font-size: 11px; opacity: 0.7; }
.wiz-cat-questions-heading { font-size: 16px; font-weight: 700; color: #1A1A1A; margin-bottom: 16px; }
</style>
""",
        unsafe_allow_html=True,
    )
    app.markdown("### Review & edit categories")
    pending = app.session_state.get("wiz_pending_remove")
    if pending:
        question = (_wizard_schema() or None).get_question(pending) if _wizard_schema() else None
        label = question.question_text if question else pending
        app.warning(f"Remove {label} from the analysis? It will not appear in the output workbook.")
        yes, no, _ = app.columns([1, 1, 4])
        if yes.button("Confirm remove", key="wiz_confirm_remove", type="primary"):
            assignments = dict(app.session_state.get("wiz_category_assignments") or {})
            removed_category = assignments.get(pending)
            assignments.pop(pending, None)
            app.session_state["wiz_category_assignments"] = assignments
            app.session_state["wiz_pending_remove"] = None
            grouped_after_remove = _wizard_questions_by_category()
            selected_after_remove = app.session_state.get("wiz_selected_category")
            if selected_after_remove == removed_category and selected_after_remove not in grouped_after_remove:
                app.session_state["wiz_selected_category"] = next(iter(grouped_after_remove), None)
            _wizard_rerun(app)
        if no.button("Cancel", key="wiz_cancel_remove"):
            app.session_state["wiz_pending_remove"] = None
            _wizard_rerun(app)

    if "wiz_selected_category" not in app.session_state:
        app.session_state["wiz_selected_category"] = None

    grouped = _wizard_questions_by_category()
    categories = list(grouped)
    selected_category = app.session_state.get("wiz_selected_category")
    if selected_category not in grouped:
        selected_category = categories[0] if categories else None
        app.session_state["wiz_selected_category"] = selected_category

    sidebar_col, questions_col = app.columns([1, 3])
    with sidebar_col:
        app.markdown("#### Categories")
        if categories:
            selected_index = categories.index(selected_category) if selected_category in categories else 0
            if app.session_state.get("wiz_category_sidebar_choice") not in categories:
                app.session_state["wiz_category_sidebar_choice"] = selected_category
            chosen_category = app.radio(
                "Categories",
                categories,
                index=selected_index,
                format_func=lambda category: f"{category}    {len(grouped.get(category, []))}",
                key="wiz_category_sidebar_choice",
                label_visibility="collapsed",
            )
            if chosen_category != selected_category:
                app.session_state["wiz_selected_category"] = chosen_category
                _wizard_rerun(app)

        app.markdown("<div class='wiz-divider'></div>", unsafe_allow_html=True)
        if app.button("+ Add category", key="wiz_show_add_category"):
            app.session_state["wiz_show_add_category_input"] = True
        if app.session_state.get("wiz_show_add_category_input"):
            new_category = app.text_input("Category name", key="wiz_new_category_name")
            if app.button("Create", key="wiz_create_category", type="primary"):
                name = new_category.strip()
                if name:
                    app.session_state.setdefault("wiz_empty_categories", [])
                    if name not in app.session_state["wiz_empty_categories"]:
                        app.session_state["wiz_empty_categories"].append(name)
                    app.session_state["wiz_selected_category"] = name
                app.session_state["wiz_show_add_category_input"] = False
                _wizard_rerun(app)

    with questions_col:
        if selected_category is None:
            app.caption("No categories yet. Add a category to begin.")
            return

        heading_col, rename_col = app.columns([2, 3])
        heading_col.markdown(
            f"<div class='wiz-cat-questions-heading'>{html.escape(selected_category)}</div>",
            unsafe_allow_html=True,
        )
        renamed = rename_col.text_input(
            "Rename category",
            value=selected_category,
            key=f"wiz_rename_selected_{selected_category}",
            label_visibility="collapsed",
        ).strip()
        if renamed and renamed != selected_category:
            assignments = dict(app.session_state.get("wiz_category_assignments") or {})
            app.session_state["wiz_category_assignments"] = {
                question_id: (renamed if assigned == selected_category else assigned)
                for question_id, assigned in assignments.items()
            }
            app.session_state["wiz_empty_categories"] = [
                renamed if item == selected_category else item
                for item in app.session_state.get("wiz_empty_categories", [])
            ]
            if renamed not in app.session_state.get("wiz_empty_categories", []) and not any(
                assigned == renamed for assigned in app.session_state["wiz_category_assignments"].values()
            ):
                app.session_state.setdefault("wiz_empty_categories", []).append(renamed)
            app.session_state["wiz_selected_category"] = renamed
            _wizard_rerun(app)

        questions = grouped.get(selected_category, [])
        if not questions:
            app.caption("No questions in this category yet.")
            return

        for row_start in range(0, len(questions), 3):
            cols = app.columns(3)
            for card_col, question in zip(cols, questions[row_start : row_start + 3]):
                with card_col:
                    app.markdown(
                        "<div class='wiz-card'>"
                        f"<div class='wiz-card-id'>{html.escape(question.canonical_id)}</div>"
                        f"<div class='wiz-card-text' title='{html.escape(question.question_text)}'>{html.escape(question_display_text(question, 40))}</div>"
                        f"<div class='wiz-card-type'>{html.escape(question_type_label(question))}</div>"
                        "</div>",
                        unsafe_allow_html=True,
                    )
                    move_options = ["Stay"] + [item for item in categories if item != selected_category]
                    move_to = app.selectbox(
                        "Move to",
                        move_options,
                        key=f"wiz_move_{question.canonical_id}",
                        label_visibility="collapsed",
                    )
                    if move_to != "Stay":
                        assignments = dict(app.session_state.get("wiz_category_assignments") or {})
                        assignments[question.canonical_id] = move_to
                        app.session_state["wiz_category_assignments"] = assignments
                        if not any(
                            assigned == selected_category
                            for qid, assigned in assignments.items()
                            if qid != question.canonical_id
                        ):
                            app.session_state.setdefault("wiz_empty_categories", [])
                            if selected_category not in app.session_state["wiz_empty_categories"]:
                                app.session_state["wiz_empty_categories"].append(selected_category)
                        _wizard_rerun(app)
                    if app.button("Remove", key=f"wiz_remove_{question.canonical_id}"):
                        app.session_state["wiz_pending_remove"] = question.canonical_id
                        _wizard_rerun(app)


def _render_wizard_step_filters() -> None:
    app = _require_streamlit()
    from src.ui.wizard import distinct_value_preview, eligible_filter_question_ids, question_display_text

    schema = _wizard_schema()
    dataframe = _wizard_dataframe()
    if schema is None:
        app.info("Upload files to review demographic filters.")
        return
    # Left = heading + definition, Right = Excel mock. (#6 layout)
    _lf_left, _lf_right = app.columns([1, 1])
    with _lf_left:
        app.markdown("### Local Filters")
        app.markdown(
            "<div style='font-size:13px;color:#444;line-height:1.7;min-height:120px;display:flex;flex-direction:column;justify-content:flex-start;padding-top:4px;'>"
            "Local Filters appear as drop-downs at the top of each question's sheet "
            "in the Excel output, so a reader can slice that sheet on the fly. "
            "Pick the questions you want available as slicers below.</div>",
            unsafe_allow_html=True,
        )
    with _lf_right:
        app.markdown(
            "<div style='margin:4px 0 8px;font-family:Calibri,Arial,sans-serif;"
            "border:1px solid #D4D4D4;border-radius:8px;overflow:hidden;"
            "box-shadow:0 4px 14px rgba(0,0,0,0.08);'>"
            "<div style='background:#217346;color:#fff;font-size:11px;font-weight:600;"
            "padding:6px 12px;display:flex;align-items:center;gap:6px;'>"
            "<span style='font-size:13px;'>\u25A6</span> Q5 \u2014 Brand awareness.xlsx</div>"
            "<div style='display:flex;background:#F3F3F3;border-bottom:1px solid #D4D4D4;"
            "font-size:11px;color:#666;font-weight:600;'>"
            "<div style='width:130px;padding:5px 10px;border-right:1px solid #E4E4E4;'>Filter</div>"
            "<div style='flex:1;padding:5px 10px;'>Value</div></div>"
            "<div style='display:flex;border-bottom:1px solid #EEE;font-size:12px;'>"
            "<div style='width:130px;padding:7px 10px;border-right:1px solid #EEE;color:#333;'>Region</div>"
            "<div style='flex:1;padding:7px 10px;'>"
            "<span style='display:inline-flex;align-items:center;gap:8px;background:#fff;"
            "border:1px solid #217346;border-radius:4px;padding:2px 8px;color:#217346;font-weight:600;'>"
            "APAC <span>\u25BE</span></span></div></div>"
            "<div style='display:flex;font-size:12px;'>"
            "<div style='width:130px;padding:7px 10px;border-right:1px solid #EEE;color:#333;'>Company size</div>"
            "<div style='flex:1;padding:7px 10px;'>"
            "<span style='display:inline-flex;align-items:center;gap:8px;background:#fff;"
            "border:1px solid #BBB;border-radius:4px;padding:2px 8px;color:#333;'>"
            "Enterprise <span style='color:#999;'>\u25BE</span></span></div></div>"
            "</div>",
            unsafe_allow_html=True,
        )
    selected = set(app.session_state.get("wiz_selected_demographics") or [])
    candidate_ids = [q.canonical_id for q in schema.questions if q.is_demographic or q.canonical_id in selected]
    for question_id in candidate_ids:
        question = schema.get_question(question_id)
        if question is None:
            continue
        checked = app.checkbox(
            f"{question.canonical_id} - {question_display_text(question, 72)}",
            value=question_id in selected,
            key=f"wiz_demo_{question_id}",
        )
        if checked:
            selected.add(question_id)
        else:
            selected.discard(question_id)
        app.caption(distinct_value_preview(dataframe, question))
    app.session_state["wiz_selected_demographics"] = sorted(selected)

    available = [qid for qid in eligible_filter_question_ids(schema) if qid not in selected]
    labels = {"": "Select a question"}
    labels.update({qid: f"{qid} - {question_display_text(schema.get_question(qid), 72)}" for qid in available if schema.get_question(qid)})
    choice = app.selectbox("+ Add another question as a filter", [""] + available, format_func=lambda value: labels.get(value, value), key="wiz_add_demo_choice")
    if app.button("Add filter", key="wiz_add_demo_button", disabled=(choice == "")):
        selected.add(choice)
        app.session_state["wiz_selected_demographics"] = sorted(selected)
        _wizard_rerun(app)

    app.divider()
    from src.ui.wizard import normalise_custom_filter_count

    _cf_left, _cf_right = app.columns([1, 1])
    with _cf_left:
        app.markdown("### Custom workbook filters")
        app.markdown(
            "<div style='font-size:13px;color:#444;line-height:1.7;min-height:120px;display:flex;flex-direction:column;justify-content:flex-start;padding-top:4px;'>"
            "Beyond the standard demographics, add <b>custom</b> filter slots where the "
            "reader picks <i>any</i> question of their choice and a value to filter by \u2014 "
            "applied across the whole workbook.</div>",
            unsafe_allow_html=True,
        )
        value = app.number_input(
            "Custom workbook filters",
            min_value=0,
            max_value=5,
            value=normalise_custom_filter_count(app.session_state.get("wizard_workbook_custom_filter_count", app.session_state.get("wiz_num_custom_filters"))),
            step=1,
            key="wizard_workbook_custom_filter_count",
        )
        app.session_state["wiz_num_custom_filters"] = int(value)
    with _cf_right:
        app.markdown(
            "<div style='margin:4px 0 8px;font-family:Calibri,Arial,sans-serif;"
            "border:1px solid #D4D4D4;border-radius:8px;overflow:hidden;"
            "box-shadow:0 4px 14px rgba(0,0,0,0.08);'>"
            "<div style='background:#217346;color:#fff;font-size:11px;font-weight:600;"
            "padding:6px 12px;display:flex;align-items:center;gap:6px;'>"
            "<span style='font-size:13px;'>\u25A6</span> Custom filter</div>"
            "<div style='display:flex;background:#F3F3F3;border-bottom:1px solid #D4D4D4;"
            "font-size:11px;color:#666;font-weight:600;'>"
            "<div style='width:130px;padding:5px 10px;border-right:1px solid #E4E4E4;'>Filter</div>"
            "<div style='flex:1;padding:5px 10px;'>Value</div></div>"
            # pick-a-question row with open dropdown + cursor
            "<div style='display:flex;font-size:12px;background:#FFFCF2;position:relative;'>"
            "<div style='width:130px;padding:7px 10px;border-right:1px solid #EEE;color:#9A7B00;font-weight:600;'>"
            "Pick a question</div>"
            "<div style='flex:1;padding:7px 10px;'>"
            "<span style='display:inline-flex;align-items:center;gap:8px;background:#fff;"
            "border:1.5px solid #E0B400;border-radius:4px;padding:2px 8px;color:#9A7B00;font-weight:600;'>"
            "Q3 <span style='color:#E0B400;'>\u25BE</span></span>"
            "<div style='position:absolute;left:120px;top:30px;background:#fff;border:1px solid #D4D4D4;"
            "border-radius:6px;box-shadow:0 6px 18px rgba(0,0,0,0.15);font-size:11px;z-index:5;min-width:90px;'>"
            "<div style='padding:5px 12px;color:#333;'>Q1</div>"
            "<div style='padding:5px 12px;color:#333;'>Q2</div>"
            "<div style='padding:5px 12px;background:#FFF3C4;color:#9A7B00;font-weight:700;'>Q3  \u2316</div>"
            "<div style='padding:5px 12px;color:#333;'>Q4</div></div>"
            "</div></div>"
            # value row
            "<div style='display:flex;font-size:12px;background:#FFFCF2;border-top:1px solid #F0E4B0;'>"
            "<div style='width:130px;padding:7px 10px;border-right:1px solid #EEE;color:#9A7B00;'>Value</div>"
            "<div style='flex:1;padding:7px 10px;'>"
            "<span style='display:inline-flex;align-items:center;gap:8px;background:#fff;"
            "border:1px solid #E0B400;border-radius:4px;padding:2px 8px;color:#9A7B00;font-weight:600;'>"
            "Option B <span style='color:#E0B400;'>\u25BE</span></span></div></div>"
            "</div>",
            unsafe_allow_html=True,
        )

    app.divider()
    from src.ui.wizard import normalise_per_question_filter_count

    _pq_left, _pq_right = app.columns([1, 1])
    with _pq_left:
        app.markdown("### Per-question filter rows")
        app.markdown(
            "<div style='font-size:13px;color:#444;line-height:1.7;min-height:120px;display:flex;flex-direction:column;justify-content:flex-start;padding-top:4px;'>"
            "These are filter rows shown <i>inside each individual question card</i> "
            "on the dashboard. Unlike Local Filters (which sit on every sheet), a "
            "per-question filter lets you slice just one question \u2014 handy when you "
            "want a one-off cut without changing anything else.</div>",
            unsafe_allow_html=True,
        )
        value = app.number_input(
            "Per-question filters",
            min_value=0,
            max_value=3,
            value=normalise_per_question_filter_count(app.session_state.get("wizard_per_question_filter_count", app.session_state.get("wiz_num_per_question_filters"))),
            step=1,
            key="wizard_per_question_filter_count",
        )
        app.session_state["wiz_num_per_question_filters"] = int(value)
    with _pq_right:
        app.markdown(
            "<div style='margin:4px 0 8px;font-family:Calibri,Arial,sans-serif;"
            "border:1px solid #D4D4D4;border-radius:8px;overflow:hidden;"
            "box-shadow:0 4px 14px rgba(0,0,0,0.08);'>"
            "<div style='background:#217346;color:#fff;font-size:11px;font-weight:600;"
            "padding:6px 12px;display:flex;align-items:center;gap:6px;'>"
            "<span style='font-size:13px;'>\u25A6</span> Question card \u2014 Q3</div>"
            "<div style='display:flex;background:#F3F3F3;border-bottom:1px solid #D4D4D4;"
            "font-size:11px;color:#666;font-weight:600;'>"
            "<div style='width:130px;padding:5px 10px;border-right:1px solid #E4E4E4;'>Filter</div>"
            "<div style='flex:1;padding:5px 10px;'>Value</div></div>"
            # the question this card belongs to
            "<div style='display:flex;border-bottom:1px solid #EEE;font-size:12px;'>"
            "<div style='width:130px;padding:7px 10px;border-right:1px solid #EEE;color:#333;'>Question</div>"
            "<div style='flex:1;padding:7px 10px;color:#333;font-weight:600;'>Q3 \u2014 Brand awareness</div></div>"
            # a custom per-question filter row
            "<div style='display:flex;font-size:12px;background:#FFFCF2;'>"
            "<div style='width:130px;padding:7px 10px;border-right:1px solid #EEE;color:#9A7B00;font-weight:600;'>"
            "Filter</div>"
            "<div style='flex:1;padding:7px 10px;'>"
            "<span style='display:inline-flex;align-items:center;gap:8px;background:#fff;"
            "border:1.5px solid #E0B400;border-radius:4px;padding:2px 8px;color:#9A7B00;font-weight:600;'>"
            "Region = APAC <span style='color:#E0B400;'>\u25BE</span></span></div></div>"
            "</div>"
            # only-this-question note with arrow
            "<div style='margin-top:8px;font-size:12px;color:#9A1B1B;font-weight:600;"
            "font-family:Arial,sans-serif;'>\u21B3 This filter changes only this one "
            "question \u2014 every other question stays unaffected.</div>",
            unsafe_allow_html=True,
        )


def _render_wizard_step_crosscut() -> None:
    app = _require_streamlit()
    app.markdown("### Cross-cut output format")
    app.write("When you generate cross-cuts from the dashboard, where should they live?")
    options = {"One consolidated sheet": "one_sheet", "Separate sheet per cross-cut": "separate_sheets"}
    current = app.session_state.get("wiz_crosscut_consolidation", "one_sheet")
    labels = list(options)
    index = list(options.values()).index(current) if current in options.values() else 0
    selected = app.radio("Cross-cut output format", labels, index=index, key="wiz_crosscut_choice")
    app.session_state["wiz_crosscut_consolidation"] = options[selected]
    app.caption("Setting saved ? exporter wiring coming in the next iteration.")


def _wizard_apply_overrides() -> None:
    app = _require_streamlit()
    schema = app.session_state.get("wiz_schema")
    if schema is None:
        return
    from src.ui.wizard import apply_wizard_schema_overrides, themes_from_wizard_assignments

    assignments = dict(app.session_state.get("wiz_category_assignments") or {})
    selected_demographics = list(app.session_state.get("wiz_selected_demographics") or [])
    schema_override = apply_wizard_schema_overrides(schema, assignments, selected_demographics)
    app.session_state["wiz_schema_override"] = schema_override
    app.session_state["wiz_theme_override"] = themes_from_wizard_assignments(schema_override, assignments)


def _wizard_pipeline_overrides(schema: Any, themes: dict) -> tuple[Any, dict]:
    app = _require_streamlit()
    # One-sheet mode: do NOT categorise into AI themes. Returning themes=None
    # makes the exporter fall back to a single "All Questions" sheet (+ a
    # Demographics sheet when demographic questions exist).
    if app.session_state.get("wiz_output_structure") == "one_sheet":
        return schema, None
    raw_assignments = app.session_state.get("wiz_category_assignments")
    if raw_assignments is None:
        return schema, themes
    from src.ui.wizard import apply_wizard_schema_overrides, themes_from_wizard_assignments

    assignments = dict(raw_assignments or {})
    selected_demographics = list(app.session_state.get("wiz_selected_demographics") or [])
    schema = apply_wizard_schema_overrides(
        schema,
        assignments,
        selected_demographics,
    )
    return schema, themes_from_wizard_assignments(schema, assignments)


def _wizard_rerun(app: Any) -> None:
    rerun = getattr(app, "rerun", None) or getattr(app, "experimental_rerun", None)
    if callable(rerun):
        rerun()


if __name__ == "__main__":
    main()
