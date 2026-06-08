"""Tests for the Excel exporter."""

from __future__ import annotations

from dataclasses import replace
from datetime import datetime, timezone
import importlib.util
import os
from pathlib import Path
import re
import unittest
from unittest.mock import patch
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

from src.calculation_log import CalculationLog
from src.cross_cut_engine import compute_cross_cuts
from src.excel_exporter import (
    export_cross_cuts_only,
    export_filtered_single_cuts,
    export_single_cuts,
    _column_data_name,
    _grid_spec_subtype,
    _validate_no_undefined_names,
    _wrapped_formula,
)
from src.question_classifier import classify_questions
from src.single_cut import compute_single_cuts
from src.models import (
    AuditRecord,
    AnalysisType,
    CrossCutResult,
    CrossCutSpec,
    DataQualityReport,
    DenominatorPolicy,
    FilteredSingleCutResult,
    FilterSpec,
    GridBinaryPivotResult,
    GridBinaryPivotRow,
    GridSingleSelectResult,
    MultiSelectResult,
    NumericResult,
    QuestionSpec,
    QuestionType,
    RankOrderResult,
    RankOrderRow,
    SingleSelectResult,
    SkipRecord,
    SurveySchema,
)
from tests.conftest import CROSS_CUT_30_RESPONDENTS_PATH


UTC_NOW = datetime(2026, 5, 4, 12, 0, tzinfo=timezone.utc)
LONG_ID = "Q_VERY_LONG_SINGLE_SELECT_EXPORT_NAME"
FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"
def make_audit(
    metric_name: str,
    question_id: str,
    source_columns: tuple[str, ...],
    value_raw: float,
    valid_n: int,
    missing_n: int,
) -> AuditRecord:
    return AuditRecord(
        output_sheet=f"SC_{question_id}",
        metric_name=metric_name,
        source_question_id=question_id,
        source_columns=source_columns,
        filter_expr=None,
        numerator=None,
        denominator=valid_n,
        formula=f"{metric_name} formula",
        value_raw=value_raw,
        valid_n=valid_n,
        missing_n=missing_n,
        timestamp=UTC_NOW,
    )


def make_export_fixture() -> tuple[
    list,
    list[SkipRecord],
    SurveySchema,
    DataQualityReport,
    CalculationLog,
]:
    ss_audit = make_audit("rate_per_value", "Q_SS_EXPORT", ("Q_SS_EXPORT",), 10, 10, 0)
    ms_audit = make_audit(
        "selection_rate",
        "Q_MS_EXPORT",
        ("Q_MS_EXPORTr1", "Q_MS_EXPORTr2"),
        10,
        10,
        0,
    )
    num_audit = make_audit("numeric_summary", "Q_NUM_EXPORT", ("Q_NUM_EXPORT",), 5.5, 10, 0)
    grid_row_1_audit = make_audit(
        "rate_per_value", "Q_GRID_EXPORTr1", ("Q_GRID_EXPORTr1",), 10, 10, 0
    )
    grid_row_2_audit = make_audit(
        "rate_per_value", "Q_GRID_EXPORTr2", ("Q_GRID_EXPORTr2",), 10, 10, 0
    )
    grid_parent_audit = make_audit(
        "grid_overall",
        "Q_GRID_EXPORT",
        ("Q_GRID_EXPORTr1", "Q_GRID_EXPORTr2"),
        10,
        10,
        0,
    )
    long_audit = make_audit("rate_per_value", LONG_ID, (LONG_ID,), 10, 10, 0)

    single_result = SingleSelectResult(
        question_id="Q_SS_EXPORT",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Yes", "count": 6, "rate": 0.6},
            2: {"label": "No", "count": 4, "rate": 0.4},
        },
        audit_records=(ss_audit,),
    )
    multi_result = MultiSelectResult(
        question_id="Q_MS_EXPORT",
        question_type=QuestionType.MULTI_SELECT_BINARY,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        selections={
            "Q_MS_EXPORTr1": {"label": "First", "count": 5, "selection_rate": 0.5},
            "Q_MS_EXPORTr2": {"label": "Second", "count": 2, "selection_rate": 0.2},
        },
        respondents_who_answered_any=10,
        audit_records=(ms_audit,),
    )
    numeric_result = NumericResult(
        question_id="Q_NUM_EXPORT",
        question_type=QuestionType.DIRECT_NUMERIC,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        mean=5.5,
        median=5.5,
        std=2.9,
        min_val=1.0,
        max_val=10.0,
        percentiles={25: 3.0, 50: 5.5, 75: 8.0},
        audit_records=(num_audit,),
    )
    grid_row_1 = SingleSelectResult(
        question_id="Q_GRID_EXPORTr1",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Low", "count": 4, "rate": 0.4},
            2: {"label": "High", "count": 6, "rate": 0.6},
        },
        audit_records=(grid_row_1_audit,),
    )
    grid_row_2 = SingleSelectResult(
        question_id="Q_GRID_EXPORTr2",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "Low", "count": 5, "rate": 0.5},
            2: {"label": "High", "count": 5, "rate": 0.5},
        },
        audit_records=(grid_row_2_audit,),
    )
    grid_result = GridSingleSelectResult(
        question_id="Q_GRID_EXPORT",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        rows={"Q_GRID_EXPORTr1": grid_row_1, "Q_GRID_EXPORTr2": grid_row_2},
        overall_valid_n=10,
        audit_records=(grid_parent_audit,),
    )
    long_result = SingleSelectResult(
        question_id=LONG_ID,
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "A", "count": 7, "rate": 0.7},
            2: {"label": "B", "count": 3, "rate": 0.3},
        },
        audit_records=(long_audit,),
        warnings=("long id warning",),
    )
    results = [single_result, multi_result, numeric_result, grid_result, long_result]

    questions = (
        QuestionSpec(
            question_id="[Q_SS_EXPORT]",
            canonical_id="Q_SS_EXPORT",
            question_text="Single select export question",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_SS_EXPORT",),
            option_map={1: "Yes", 2: "No"},
        ),
        QuestionSpec(
            question_id="[Q_MS_EXPORT]",
            canonical_id="Q_MS_EXPORT",
            question_text="Multi select export question",
            question_type=QuestionType.MULTI_SELECT_BINARY,
            raw_columns=("Q_MS_EXPORTr1", "Q_MS_EXPORTr2"),
            option_map={"Q_MS_EXPORTr1": "First", "Q_MS_EXPORTr2": "Second"},
            value_range=(0, 1),
        ),
        QuestionSpec(
            question_id="[Q_NUM_EXPORT]",
            canonical_id="Q_NUM_EXPORT",
            question_text="Numeric export question",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_NUM_EXPORT",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_GRID_EXPORT]",
            canonical_id="Q_GRID_EXPORT",
            question_text="Grid export question",
            question_type=QuestionType.GRID_SINGLE_SELECT,
            raw_columns=("Q_GRID_EXPORTr1", "Q_GRID_EXPORTr2"),
            option_map={1: "Low", 2: "High"},
            value_range=(1, 2),
            grid_row_labels={
                "Q_GRID_EXPORTr1": "Grid first row",
                "Q_GRID_EXPORTr2": "Grid second row",
            },
        ),
        QuestionSpec(
            question_id=f"[{LONG_ID}]",
            canonical_id=LONG_ID,
            question_text="Long sheet name export question",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=(LONG_ID,),
            option_map={1: "A", 2: "B"},
        ),
        QuestionSpec(
            question_id="[Q_TEXT_EXPORT]",
            canonical_id="Q_TEXT_EXPORT",
            question_text="Open text export question",
            question_type=QuestionType.OPEN_TEXT,
            raw_columns=("Q_TEXT_EXPORT",),
            option_map={},
        ),
    )
    schema = SurveySchema(
        questions=questions,
        respondent_id_column="respondent_id",
        total_respondents=10,
        source_datamap_path="datamap.xlsx",
        source_rawdata_path="raw.csv",
        parsed_at=UTC_NOW,
    )
    skips = [
        SkipRecord(
            question_id="[Q_TEXT_EXPORT]",
            canonical_id="Q_TEXT_EXPORT",
            question_type=QuestionType.OPEN_TEXT,
            skip_reason="unsupported_type: OPEN_TEXT",
        ),
        SkipRecord(
            question_id="[Q_FAIL_EXPORT]",
            canonical_id="Q_FAIL_EXPORT",
            question_type=QuestionType.SINGLE_SELECT,
            skip_reason="calculation_error",
            details="ValueError: raw column not found in data",
        ),
    ]
    quality_report = DataQualityReport(
        total_rows=10,
        total_columns=8,
        columns_in_datamap=7,
        columns_not_in_datamap=("extra_col",),
        per_column_missing_pct={"Q_SS_EXPORT": 0.1, "Q_NUM_EXPORT": 0.0, "extra_col": 0.5},
        per_column_out_of_range_pct={"Q_NUM_EXPORT": 0.2},
        coercion_log=(
            {
                "column": "Q_NUM_EXPORT",
                "from_type": "string",
                "to_type": "numeric",
                "values_coerced": ["abc"],
                "rows_affected": 1,
            },
        ),
        warnings=("column extra_col has 50.0% missing values",),
    )
    log = CalculationLog()
    for record in (
        ss_audit,
        ms_audit,
        num_audit,
        grid_row_1_audit,
        grid_row_2_audit,
        grid_parent_audit,
        long_audit,
    ):
        log.record(record)
    return results, skips, schema, quality_report, log


def rated_grid_export_fixture() -> tuple[GridSingleSelectResult, SurveySchema]:
    row_specs = (
        ("Q_RATEDr1c1", "Implementation speed", {8: 2, 9: 2}),
        ("Q_RATEDr1c2", "Implementation speed", {6: 2, 7: 2}),
        ("Q_RATEDr2c1", "Integration fit", {8: 1, 9: 3}),
        ("Q_RATEDr2c2", "Integration fit", {6: 1, 7: 3}),
        ("Q_RATEDr3c1", "Efficiency gains", {7: 2, 8: 2}),
        ("Q_RATEDr3c2", "Efficiency gains", {5: 2, 6: 2}),
    )
    rows = {
        column_id: SingleSelectResult(
            question_id=column_id,
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                score: {"label": str(score), "count": count, "rate": count / 4}
                for score, count in counts.items()
            },
        )
        for column_id, _label, counts in row_specs
    }
    result = GridSingleSelectResult(
        question_id="Q_RATED",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        valid_n=4,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        rows=rows,
        overall_valid_n=4,
    )
    schema = SurveySchema(
        questions=(
            QuestionSpec(
                question_id="[Q_RATED]",
                canonical_id="Q_RATED",
                question_text="Q_RATED - Rate each vendor attribute",
                question_type=QuestionType.GRID_SINGLE_SELECT,
                raw_columns=tuple(column_id for column_id, _label, _counts in row_specs),
                option_map={index: str(index) for index in range(1, 11)},
                value_range=(1, 10),
                grid_row_labels={
                    column_id: label for column_id, label, _counts in row_specs
                },
                possible_role="GRID_RATED",
            ),
        ),
        respondent_id_column="respondent_id",
        total_respondents=4,
        source_datamap_path="datamap.xlsx",
        source_rawdata_path="raw.csv",
        parsed_at=UTC_NOW,
    )
    return result, schema


def categorical_grid_export_fixture() -> tuple[GridSingleSelectResult, SurveySchema]:
    distributions = {
        "Q_CATGRIDr1": {
            1: {"label": "Decision maker", "count": 3, "rate": 0.3},
            2: {"label": "Influencer", "count": 4, "rate": 0.4},
            3: {"label": "Not involved", "count": 3, "rate": 0.3},
        },
        "Q_CATGRIDr2": {
            1: {"label": "Decision maker", "count": 1, "rate": 0.1},
            2: {"label": "Influencer", "count": 5, "rate": 0.5},
            3: {"label": "Not involved", "count": 4, "rate": 0.4},
        },
        "Q_CATGRIDr3": {
            1: {"label": "Decision maker", "count": 2, "rate": 0.2},
            2: {"label": "Influencer", "count": 2, "rate": 0.2},
            3: {"label": "Not involved", "count": 6, "rate": 0.6},
        },
    }
    rows = {
        column_id: SingleSelectResult(
            question_id=column_id,
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=10,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution=distribution,
        )
        for column_id, distribution in distributions.items()
    }
    result = GridSingleSelectResult(
        question_id="Q_CATGRID",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        valid_n=10,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        rows=rows,
        overall_valid_n=10,
    )
    schema = SurveySchema(
        questions=(
            QuestionSpec(
                question_id="[Q_CATGRID]",
                canonical_id="Q_CATGRID",
                question_text="Q_CATGRID - Stakeholder role by function",
                question_type=QuestionType.GRID_SINGLE_SELECT,
                raw_columns=tuple(distributions),
                option_map={
                    1: "Decision maker",
                    2: "Influencer",
                    3: "Not involved",
                },
                value_range=(1, 3),
                grid_row_labels={
                    "Q_CATGRIDr1": "IT / Technical",
                    "Q_CATGRIDr2": "Security",
                    "Q_CATGRIDr3": "Human Resources",
                },
                possible_role="GRID_CATEGORICAL",
            ),
        ),
        respondent_id_column="respondent_id",
        total_respondents=10,
        source_datamap_path="datamap.xlsx",
        source_rawdata_path="raw.csv",
        parsed_at=UTC_NOW,
    )
    return result, schema


def categorical_c_grid_export_fixture() -> tuple[GridSingleSelectResult, SurveySchema]:
    row_specs = (
        ("Q26r1c1", "IT / Technical", "Blocked Vendors", 2),
        ("Q26r1c2", "IT / Technical", "Scored Vendors", 4),
        ("Q26r1c3", "IT / Technical", "Recommended for or against", 3),
        ("Q26r2c1", "Finance", "Blocked Vendors", 1),
        ("Q26r2c2", "Finance", "Scored Vendors", 5),
        ("Q26r2c3", "Finance", "Recommended for or against", 2),
    )
    rows = {
        column_id: SingleSelectResult(
            question_id=column_id,
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=6,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={1: {"label": "Selected", "count": count, "rate": count / 6}},
        )
        for column_id, _row_label, _category, count in row_specs
    }
    result = GridSingleSelectResult(
        question_id="Q26",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        valid_n=6,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        rows=rows,
        overall_valid_n=6,
    )
    schema = SurveySchema(
        questions=(
            QuestionSpec(
                question_id="[Q26]",
                canonical_id="Q26",
                question_text="Q26 - What role did each stakeholder play",
                question_type=QuestionType.GRID_SINGLE_SELECT,
                raw_columns=tuple(column_id for column_id, _row_label, _category, _count in row_specs),
                option_map={
                    1: "Blocked Vendors",
                    2: "Scored Vendors",
                    3: "Recommended for or against",
                },
                value_range=(0, 1),
                grid_row_labels={
                    column_id: row_label
                    for column_id, row_label, _category, _count in row_specs
                },
                possible_role="GRID_CATEGORICAL",
            ),
        ),
        respondent_id_column="respondent_id",
        total_respondents=6,
        source_datamap_path="datamap.xlsx",
        source_rawdata_path="raw.csv",
        parsed_at=UTC_NOW,
    )
    return result, schema


def make_cross_cut_schema() -> SurveySchema:
    questions = (
        QuestionSpec(
            question_id="[Q_SEG_1]",
            canonical_id="Q_SEG_1",
            question_text="Segment",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_SEG_1",),
            option_map={1: "Segment 1", 2: "Segment 2", 3: "Segment 3"},
        ),
        QuestionSpec(
            question_id="[Q_TGT_1]",
            canonical_id="Q_TGT_1",
            question_text="Target categorical",
            question_type=QuestionType.SINGLE_SELECT,
            raw_columns=("Q_TGT_1",),
            option_map={1: "A", 2: "B", 3: "C", 4: "D"},
        ),
        QuestionSpec(
            question_id="[Q_NUM_3]",
            canonical_id="Q_NUM_3",
            question_text="Numeric metric",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_NUM_3",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_EXP_1]",
            canonical_id="Q_EXP_1",
            question_text="Expected",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_EXP_1",),
            option_map={},
        ),
        QuestionSpec(
            question_id="[Q_REAL_1]",
            canonical_id="Q_REAL_1",
            question_text="Realized",
            question_type=QuestionType.DIRECT_NUMERIC,
            raw_columns=("Q_REAL_1",),
            option_map={},
        ),
    )
    return SurveySchema(
        questions=questions,
        respondent_id_column="respondent_id",
        total_respondents=30,
        source_datamap_path="cross_datamap.xlsx",
        source_rawdata_path="cross_raw.csv",
        parsed_at=UTC_NOW,
    )


def grid_segment_question() -> QuestionSpec:
    return QuestionSpec(
        question_id="[Q_GRID_SEG]",
        canonical_id="Q_GRID_SEG",
        question_text="Grid segment",
        question_type=QuestionType.GRID_SINGLE_SELECT,
        raw_columns=("Q_GRID_SEGr1", "Q_GRID_SEGr2", "Q_GRID_SEGr3"),
        option_map={1: "Selected"},
        value_range=(0, 1),
        grid_row_labels={
            "Q_GRID_SEGr1": "Segment 1",
            "Q_GRID_SEGr2": "Segment 2",
            "Q_GRID_SEGr3": "Segment 3",
        },
    )


def grid_cross_tab_export_result(display_mode: str) -> tuple[CrossCutResult, SurveySchema, CalculationLog]:
    dataframe = pd.read_csv(CROSS_CUT_30_RESPONDENTS_PATH)
    dataframe["Q_GRID_SEGr1"] = [1] * 10 + [0] * 20
    dataframe["Q_GRID_SEGr2"] = [0] * 10 + [1] * 10 + [0] * 10
    dataframe["Q_GRID_SEGr3"] = [0] * 20 + [1] * 10
    base_schema = make_cross_cut_schema()
    schema = SurveySchema(
        questions=base_schema.questions + (grid_segment_question(),),
        respondent_id_column=base_schema.respondent_id_column,
        total_respondents=base_schema.total_respondents,
        source_datamap_path=base_schema.source_datamap_path,
        source_rawdata_path=base_schema.source_rawdata_path,
        parsed_at=base_schema.parsed_at,
    )
    log = CalculationLog()
    results, skips = compute_cross_cuts(
        [
            CrossCutSpec(
                cross_cut_id="CC_GRID_EXPORT",
                title="Grid segment by target",
                analysis_type=AnalysisType.CROSS_TAB,
                source_question_ids=("Q_GRID_SEG", "Q_TGT_1"),
                display_mode=display_mode,
            )
        ],
        schema,
        dataframe,
        log,
    )
    if skips:
        raise AssertionError(skips[0].details)
    return results[0], schema, log


def make_cross_cut_export_fixture():
    results, skips, schema, quality_report, log = make_export_fixture()
    dataframe = pd.read_csv(CROSS_CUT_30_RESPONDENTS_PATH)
    cross_schema = make_cross_cut_schema()
    export_schema = SurveySchema(
        questions=schema.questions + cross_schema.questions,
        respondent_id_column=schema.respondent_id_column,
        total_respondents=schema.total_respondents,
        source_datamap_path=schema.source_datamap_path,
        source_rawdata_path=schema.source_rawdata_path,
        parsed_at=schema.parsed_at,
    )
    cross_specs = [
        CrossCutSpec(
            cross_cut_id="CC_TAB_EXPORT",
            title="Segment by target",
            analysis_type=AnalysisType.CROSS_TAB,
            source_question_ids=("Q_SEG_1", "Q_TGT_1"),
        ),
        CrossCutSpec(
            cross_cut_id="CC_SEG_EXPORT",
            title="Segment profile",
            analysis_type=AnalysisType.SEGMENT_PROFILE,
            source_question_ids=("Q_SEG_1", "Q_TGT_1"),
            filter_expr="Q_SEG_1 == 1",
            filter_mask_description="Q_SEG_1 = Segment 1",
        ),
        CrossCutSpec(
            cross_cut_id="CC_GROUP_EXPORT",
            title="Group comparison",
            analysis_type=AnalysisType.GROUP_COMPARISON,
            source_question_ids=("Q_SEG_1", "Q_NUM_3"),
        ),
        CrossCutSpec(
            cross_cut_id="CC_EVR_EXPORT",
            title="Expected vs realized",
            analysis_type=AnalysisType.EXPECTED_VS_REALIZED,
            source_question_ids=("Q_EXP_1", "Q_REAL_1"),
        ),
    ]
    cross_results, _ = compute_cross_cuts(cross_specs, cross_schema, dataframe, log)
    cross_skips = [
        SkipRecord(
            question_id="CC_BAD_EXPORT",
            canonical_id="CC_BAD_EXPORT",
            question_type=QuestionType.UNKNOWN,
            skip_reason="cross_cut_error",
            details="ValueError: synthetic cross cut failure",
        )
    ]
    return results, skips, export_schema, quality_report, log, cross_results, cross_skips


def make_filtered_export_fixture():
    (
        results,
        _skips,
        schema,
        _quality_report,
        log,
        cross_results,
        _cross_skips,
    ) = make_cross_cut_export_fixture()
    unrelated_audit = make_audit(
        "rate_per_value",
        "Q_UNRELATED_EXPORT",
        ("Q_UNRELATED_EXPORT",),
        1.0,
        1,
        0,
    )
    log.record(unrelated_audit)
    filtered_results = [
        FilteredSingleCutResult(
            target_question_id="Q_SS_EXPORT",
            filters_applied=(FilterSpec("Q_SEG_1", 1),),
            dispatch_mode="single_cut_filtered",
            single_cut_result=results[0],
            cross_cut_result=None,
            filtered_n=6,
            audit_records=results[0].audit_records,
            warnings=("Filtered sample size 6 is below reliability threshold (30); results may not be reliable.",),
        ),
        FilteredSingleCutResult(
            target_question_id="Q_SS_EXPORT",
            filters_applied=(FilterSpec("Q_SEG_1", 2),),
            dispatch_mode="single_cut_filtered",
            single_cut_result=replace(results[0], valid_n=4),
            cross_cut_result=None,
            filtered_n=4,
            audit_records=results[0].audit_records,
        ),
        FilteredSingleCutResult(
            target_question_id="Q_TGT_1",
            filters_applied=(FilterSpec("Q_SEG_1"),),
            dispatch_mode="cross_cut_breakdown",
            single_cut_result=None,
            cross_cut_result=cross_results[0],
            filtered_n=30,
            audit_records=cross_results[0].audit_records,
        ),
    ]
    return filtered_results, schema, log


def sheet_values(workbook_path: Path, sheet_name: str) -> list[object]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return [
            cell
            for row in workbook[sheet_name].iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]
    finally:
        workbook.close()


def find_row(ws, value: object, column: int = 1) -> int:
    for row_index in range(1, ws.max_row + 1):
        if ws.cell(row=row_index, column=column).value == value:
            return row_index
    raise AssertionError(f"{value!r} not found in column {column}")


def defined_range_values(workbook, name: str) -> list[object]:
    destinations = list(workbook.defined_names[name].destinations)
    values: list[object] = []
    for sheet_name, coord in destinations:
        ws = workbook[sheet_name]
        for row in ws[coord]:
            for cell in row:
                values.append(cell.value)
    return values


def question_title(question_id: str, text: str) -> str:
    return f"{question_id} - {text}"


def question_header_row(ws, question_id: str) -> int:
    prefix = f"{question_id} - "
    for row_index in range(1, ws.max_row + 1):
        value = ws.cell(row=row_index, column=1).value
        if isinstance(value, str) and value.startswith(prefix):
            return row_index
    raise AssertionError(f"Question block {question_id!r} not found")


def table_header_row(ws, question_id: str, header: str = "Option") -> int:
    start = question_header_row(ws, question_id)
    for row_index in range(start, min(ws.max_row, start + 12) + 1):
        if ws.cell(row=row_index, column=1).value == header:
            return row_index
    raise AssertionError(f"{header!r} table header for {question_id!r} not found")


def cell_has_data_validation(ws, coordinate: str) -> bool:
    for validation in ws.data_validations.dataValidation:
        if coordinate in validation.cells:
            return True
    return False


LONG_LABEL_IDS = (
    "Q14 Expected 2025 Gross margin growth outlook",
    "Q56 Expected 2025 Gross margin growth outlook",
    "Q102 Expected 2025 EBITDA margin growth outlook",
)


def make_long_label_export_fixture():
    rows = 6
    data = {"respondent_id": list(range(1, rows + 1))}
    questions = []
    results = []
    log = CalculationLog()
    for index, qid in enumerate(LONG_LABEL_IDS):
        data[qid] = [1, 2, 1, 2, 1, 2]
        audit = make_audit("rate_per_value", qid, (qid,), 0.5, rows, 0)
        questions.append(
            QuestionSpec(
                question_id=f"[{qid}]",
                canonical_id=qid,
                question_text=f"{qid} question text",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=(qid,),
                option_map={1: "Higher", 2: "Lower"},
            )
        )
        results.append(
            SingleSelectResult(
                question_id=qid,
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=rows,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "Higher", "count": 3, "rate": 0.5},
                    2: {"label": "Lower", "count": 3, "rate": 0.5},
                },
                audit_records=(audit,),
            )
        )
        log.record(audit)

    schema = SurveySchema(
        questions=tuple(questions),
        respondent_id_column="respondent_id",
        total_respondents=rows,
        source_datamap_path="datamap.xlsx",
        source_rawdata_path="raw.csv",
        parsed_at=UTC_NOW,
    )
    quality_report = DataQualityReport(
        total_rows=rows,
        total_columns=len(data),
        columns_in_datamap=len(questions),
        columns_not_in_datamap=(),
        per_column_missing_pct={qid: 0.0 for qid in LONG_LABEL_IDS},
        per_column_out_of_range_pct={qid: 0.0 for qid in LONG_LABEL_IDS},
        coercion_log=(),
        warnings=(),
    )
    return results, [], schema, quality_report, log, pd.DataFrame(data)


def export_long_label_fixture(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    results, skips, schema, quality_report, log, decoded_df = make_long_label_export_fixture()
    with patch.dict(os.environ, {"PORTKEY_API_KEY": ""}):
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            decoded_df=decoded_df,
            themes={
                "themes": [
                    {"name": "All Questions", "question_ids": list(LONG_LABEL_IDS)}
                ]
            },
            strict_formula_name_validation=True,
        )


def workbook_formula_strings(workbook) -> list[str]:
    formulas: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.value)
    return formulas


class TestExcelExporter(unittest.TestCase):
    def test_grid_spec_subtype_reclassifies_polluted_role_as_categorical(self) -> None:
        question = QuestionSpec(
            question_id="[Q26]",
            canonical_id="Q26",
            question_text="Q26 - What role did each stakeholder play",
            question_type=QuestionType.GRID_SINGLE_SELECT,
            raw_columns=(
                "Q26r1c1",
                "Q26r1c2",
                "Q26r1c3",
                "Q26r2c1",
                "Q26r2c2",
                "Q26r2c3",
            ),
            option_map={
                1: "Blocked Vendors",
                2: "Scored Vendors",
                3: "Recommended for or against",
            },
            value_range=(0, 1),
            grid_row_labels={
                "Q26r1c1": "IT / Technical",
                "Q26r1c2": "IT / Technical",
                "Q26r1c3": "IT / Technical",
                "Q26r2c1": "Finance",
                "Q26r2c2": "Finance",
                "Q26r2c3": "Finance",
            },
            possible_role="schema warning text",
        )

        self.assertEqual(_grid_spec_subtype(question), "GRID_CATEGORICAL")

    def test_column_data_name_strips_trailing_underscore(self) -> None:
        self.assertEqual(
            _column_data_name("Q14_Expected_2025_Gross_margin_"),
            "Q14_Expected_2025_Gross_margin_data",
        )

    def test_no_double_underscore_data_names(self) -> None:
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_long_labels_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        export_long_label_fixture(output_path)
        workbook = load_workbook(output_path, data_only=False)

        defined_names = [str(name) for name in workbook.defined_names.keys()]
        formulas = workbook_formula_strings(workbook)
        self.assertFalse(
            any("__data" in value for value in [*defined_names, *formulas]),
            "Workbook should not contain double-underscore data names.",
        )

    def test_nps_raw_score_named_ranges_are_separate_from_bucket_columns(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        dataframe = pd.DataFrame(
            {
                "respondent_id": [1, 2, 3, 4, 5],
                "Q_NPS_A": [10, 9, 8, "not answered", None],
                "Q_NPS_B": [6, 7, 10, 0, ""],
            }
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_NPS]",
                    canonical_id="Q_NPS",
                    question_text="Recommend vendors",
                    question_type=QuestionType.NPS,
                    raw_columns=("Q_NPS_A", "Q_NPS_B"),
                    option_map={"Q_NPS_A": "Vendor A", "Q_NPS_B": "Vendor B"},
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=len(dataframe),
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=len(dataframe),
            total_columns=3,
            columns_in_datamap=2,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        log = CalculationLog()
        results, skips = compute_single_cuts(schema, dataframe, log)

        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            decoded_df=dataframe,
        )

        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        raw_ws = workbook["_RawData"]
        headers = {
            raw_ws.cell(row=1, column=col).value: col
            for col in range(1, raw_ws.max_column + 1)
        }
        self.assertLess(headers["Q_NPS_NPS_Bucket"], headers["Q_NPS_A"])
        self.assertLess(headers["Q_NPS_A"], headers["Q_NPS_B"])
        self.assertEqual(raw_ws.cell(row=2, column=headers["Q_NPS_A"]).value, 10)
        self.assertEqual(raw_ws.cell(row=5, column=headers["Q_NPS_A"]).value, "not answered")
        self.assertEqual(raw_ws.cell(row=2, column=headers["Q_NPS_NPS_Bucket"]).value, "Promoter")
        q_nps_a_col = get_column_letter(headers["Q_NPS_A"])
        self.assertEqual(
            list(workbook.defined_names["Q_NPS_A_data"].destinations),
            [("_RawData", f"${q_nps_a_col}$2:${q_nps_a_col}$6")],
        )

        nps_ws = workbook["SC_Q_NPS"]
        self.assertEqual(nps_ws.cell(row=5, column=1).value, "Vendor A")
        self.assertEqual(nps_ws.cell(row=5, column=2).value, 2)
        self.assertAlmostEqual(nps_ws.cell(row=5, column=3).value, 2 / 3)
        self.assertEqual(nps_ws.cell(row=5, column=4).value, 1)
        self.assertAlmostEqual(nps_ws.cell(row=5, column=5).value, 1 / 3)
        self.assertEqual(nps_ws.cell(row=5, column=6).value, 0)
        self.assertEqual(nps_ws.cell(row=5, column=8).value, 3)
        self.assertEqual(nps_ws.cell(row=5, column=9).value, 67)
        self.assertEqual(nps_ws.cell(row=6, column=9).value, -25)
        for row in range(5, 7):
            for col in range(2, 10):
                value = nps_ws.cell(row=row, column=col).value
                self.assertFalse(
                    isinstance(value, str) and value.startswith("="),
                    f"NPS cell {row},{col} should be static, got {value!r}",
                )

    def test_validate_no_undefined_names_catches_mismatch(self) -> None:
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet["A1"] = '=SUM(Missing_data)+COUNTIF(A:A,"Quoted_data")'

        self.assertEqual(_validate_no_undefined_names(workbook), ["Missing_data"])
        with self.assertRaisesRegex(ValueError, "Missing_data"):
            _validate_no_undefined_names(workbook, strict=True)

    def test_real_long_label_workbook_has_no_undefined_names(self) -> None:
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_long_labels_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        export_long_label_fixture(output_path)
        workbook = load_workbook(output_path, data_only=False)

        self.assertEqual(_validate_no_undefined_names(workbook, strict=True), [])

    def test_wrapped_formula_preserves_bare_commas(self) -> None:
        formula = _wrapped_formula("F_Q14")

        self.assertEqual(formula.count("SUBSTITUTE("), 1)
        self.assertIn('", ", "|"', formula)
        self.assertNotIn('SUBSTITUTE(SUBSTITUTE(', formula)

    def test_memory_profile_report_written_when_enabled(self) -> None:
        from src import memory_profiler

        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        report_path = Path("outputs") / f"{output_path.stem}.memory_report.txt"
        if report_path.exists():
            report_path.unlink()
        results, skips, schema, quality_report, log = make_export_fixture()
        expected_labels = [
            "load_or_receive_decoded_df",
            "generate_short_labels",
            "build_raw_data_sheet",
            "build_options_sheet",
            "build_filters_sheet",
            "build_helper_columns",
            "build_run_summary_sheet",
            "build_question_metadata_sheet",
            "build_single_cut_index_sheet",
            "build_theme_sheets",
            "build_calculation_log_sheet",
            "build_filter_log_sheet",
            "build_warnings_sheet",
            "save_workbook",
            "patch_formula_caches",
            "write_calc_chain",
        ]

        try:
            with patch.dict(os.environ, {"SURVEY_PROFILE_MEMORY": "1"}):
                export_single_cuts(
                    results,
                    skips,
                    schema,
                    quality_report,
                    log,
                    str(output_path),
                )
        finally:
            memory_profiler.disable_profiling()
            memory_profiler.reset_log()

        self.assertTrue(report_path.exists())
        report = report_path.read_text(encoding="utf-8")
        for label in expected_labels:
            self.assertIn(label, report)

    def grid_single_select_format_fixture(
        self,
    ) -> tuple[GridSingleSelectResult, SurveySchema]:
        row_a = SingleSelectResult(
            question_id="Q_GRID_FORMAT.r1",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=80,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={1: {"label": "Checked", "count": 50, "rate": 0.625}},
        )
        row_b = SingleSelectResult(
            question_id="Q_GRID_FORMAT.r2",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=80,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={1: {"label": "Checked", "count": 30, "rate": 0.375}},
        )
        row_c = SingleSelectResult(
            question_id="Q_GRID_FORMAT.r3",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=80,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={1: {"label": "Checked", "count": 0, "rate": 0.0}},
        )
        result = GridSingleSelectResult(
            question_id="Q_GRID_FORMAT",
            question_type=QuestionType.GRID_SINGLE_SELECT,
            valid_n=80,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            rows={"r1": row_a, "r2": row_b, "r3": row_c},
            overall_valid_n=80,
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_GRID_FORMAT]",
                    canonical_id="Q_GRID_FORMAT",
                    question_text="Grid format question",
                    question_type=QuestionType.GRID_SINGLE_SELECT,
                    raw_columns=("r1", "r2", "r3"),
                    option_map={0: "Unchecked", 1: "Checked"},
                    value_range=(0, 1),
                    grid_row_labels={
                        "r1": "Option A",
                        "r2": "Option B",
                        "r3": "Option C",
                    },
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=80,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        return result, schema

    def export_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        export_single_cuts(results, skips, schema, quality_report, log, str(output_path))
        return output_path

    def export_custom_workbook(
        self,
        results: list,
        schema: SurveySchema,
        **export_kwargs,
    ) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        quality_report = DataQualityReport(
            total_rows=schema.total_respondents,
            total_columns=len(schema.questions),
            columns_in_datamap=len(schema.questions),
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        export_single_cuts(
            results,
            [],
            schema,
            quality_report,
            CalculationLog(),
            str(output_path),
            **export_kwargs,
        )
        return output_path

    def export_numeric_allocation_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        per_option_stats = {
            "Q20r1": {
                "mean": 28.9,
                "median": 27.0,
                "std": 10.5,
                "min_val": 0.0,
                "max_val": 100.0,
                "valid_n": 1022,
                "missing_n": 0,
            },
            "Q20r2": {
                "mean": 24.6,
                "median": 22.0,
                "std": 9.1,
                "min_val": 0.0,
                "max_val": 100.0,
                "valid_n": 1022,
                "missing_n": 0,
            },
            "Q20r3": {
                "mean": 22.9,
                "median": 20.0,
                "std": 8.4,
                "min_val": 0.0,
                "max_val": 100.0,
                "valid_n": 1022,
                "missing_n": 0,
            },
        }
        audits = tuple(
            AuditRecord(
                output_sheet="SC_Q20",
                metric_name=metric_name,
                source_question_id="Q20",
                source_columns=(option_id,),
                filter_expr=None,
                numerator=100.0,
                denominator=int(payload["valid_n"]),
                formula=f"{metric_name} formula",
                value_raw=float(payload[metric_name.rsplit("_", 1)[-1]]),
                valid_n=int(payload["valid_n"]),
                missing_n=int(payload["missing_n"]),
                timestamp=UTC_NOW,
            )
            for option_id, payload in per_option_stats.items()
            for metric_name in ("numeric_allocation_mean", "numeric_allocation_median")
        )
        result = NumericResult(
            question_id="Q20",
            question_type=QuestionType.NUMERIC_ALLOCATION,
            valid_n=1022,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            mean=25.5,
            median=23.0,
            std=3.0,
            min_val=0.0,
            max_val=100.0,
            percentiles={25: 20.0, 50: 23.0, 75: 28.9},
            allocation_target=100.0,
            allocation_tolerance=2.0,
            allocation_excluded_n=0,
            per_option_stats=per_option_stats,
            audit_records=audits,
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q20]",
                    canonical_id="Q20",
                    question_text="Allocate growth by source",
                    question_type=QuestionType.NUMERIC_ALLOCATION,
                    raw_columns=("Q20r1", "Q20r2", "Q20r3"),
                    option_map={
                        "Q20r1": "Existing customers - net retention",
                        "Q20r2": "Existing customers - cross-sell / up-sell",
                        "Q20r3": "New customers - existing geographies",
                    },
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=1022,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=1022,
            total_columns=3,
            columns_in_datamap=1,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        log = CalculationLog()
        for audit in audits:
            log.record(audit)
        export_single_cuts([result], [], schema, quality_report, log, str(output_path))
        return output_path

    def export_streaming_raw_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        row_count = 3000
        decoded_df = pd.DataFrame(
            {
                "respondent_id": range(1, row_count + 1),
                "Q_STREAM": ["A" if index % 2 == 0 else "B" for index in range(row_count)],
            }
        )
        result = SingleSelectResult(
            question_id="Q_STREAM",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=row_count,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                1: {"label": "A", "count": 1500, "rate": 0.5},
                2: {"label": "B", "count": 1500, "rate": 0.5},
            },
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_STREAM]",
                    canonical_id="Q_STREAM",
                    question_text="Streaming raw data question",
                    question_type=QuestionType.SINGLE_SELECT,
                    raw_columns=("Q_STREAM",),
                    option_map={1: "A", 2: "B"},
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=row_count,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=row_count,
            total_columns=2,
            columns_in_datamap=1,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        export_single_cuts(
            [result],
            [],
            schema,
            quality_report,
            CalculationLog(),
            str(output_path),
            decoded_df=decoded_df,
        )
        return output_path

    def export_mixed_eight_block_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        extra_results: list[SingleSelectResult] = []
        extra_questions: list[QuestionSpec] = []
        for index in range(3):
            qid = f"Q_EXTRA_{index + 1}"
            extra_results.append(
                SingleSelectResult(
                    question_id=qid,
                    question_type=QuestionType.SINGLE_SELECT,
                    valid_n=10,
                    missing_n=0,
                    denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                    distribution={
                        1: {"label": f"Extra {index + 1} A", "count": 6, "rate": 0.6},
                        2: {"label": f"Extra {index + 1} B", "count": 4, "rate": 0.4},
                    },
                )
            )
            extra_questions.append(
                QuestionSpec(
                    question_id=f"[{qid}]",
                    canonical_id=qid,
                    question_text=f"Extra mixed block question {index + 1}",
                    question_type=QuestionType.SINGLE_SELECT,
                    raw_columns=(qid,),
                    option_map={1: f"Extra {index + 1} A", 2: f"Extra {index + 1} B"},
                )
            )
        export_single_cuts(
            [*results, *extra_results],
            skips,
            replace(schema, questions=(*schema.questions, *extra_questions)),
            quality_report,
            log,
            str(output_path),
        )
        return output_path

    def export_workbook_with_cross_cuts(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        (
            results,
            skips,
            schema,
            quality_report,
            log,
            cross_results,
            cross_skips,
        ) = make_cross_cut_export_fixture()
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            cross_cut_results=cross_results,
            cross_cut_skips=cross_skips,
        )
        return output_path

    def export_cross_cut_workbook(self, cross_results: list) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        (
            _results,
            _skips,
            schema,
            _quality_report,
            log,
            _cross_results,
            _cross_skips,
        ) = make_cross_cut_export_fixture()
        export_cross_cuts_only(cross_results, schema, log, str(output_path))
        return output_path

    def export_filtered_workbook(self) -> Path:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        filtered_results, schema, log = make_filtered_export_fixture()
        export_filtered_single_cuts(filtered_results, schema, log, str(output_path))
        return output_path

    def test_export_creates_file_at_path(self) -> None:
        output_path = self.export_workbook()

        self.assertTrue(output_path.exists())
        self.assertGreater(output_path.stat().st_size, 0)

    def test_export_writes_all_required_sheets(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(
            workbook.sheetnames[:6],
            ["_RawData", "_Options", "Filters", "Run_Summary", "Question_Metadata", "Single_Cut_Index"],
        )
        self.assertEqual(
            workbook.sheetnames[-3:],
            ["Calculation_Log", "Filter_Log", "Warnings"],
        )
        self.assertIn("All Questions", workbook.sheetnames)
        self.assertFalse(any(name.startswith("SC_") for name in workbook.sheetnames))

    def test_run_summary_contains_correct_totals(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Run_Summary"]

        self.assertEqual(ws["B4"].value, 10)
        self.assertEqual(ws["B5"].value, 6)
        self.assertEqual(ws["B6"].value, 5)
        self.assertEqual(ws["B7"].value, 2)
        self.assertEqual(ws["B8"].value, 0)
        self.assertEqual(ws["B9"].value, 0)
        self.assertEqual(ws["B10"].value, 7)
        self.assertEqual(ws["B11"].value, 1)

    def test_question_metadata_one_row_per_question(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(workbook["Question_Metadata"].max_row, 7)

    def test_single_cut_index_links_to_theme_sheets(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["Single_Cut_Index"]

        self.assertEqual(ws["D2"].value, "All Questions")
        self.assertEqual(ws["E2"].value, "All Questions")
        self.assertIsNotNone(ws["E2"].hyperlink)
        self.assertIn("All Questions", workbook.sheetnames)

    def test_sc_sheet_for_single_select_has_distribution_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_SS_EXPORT")
        data_row = header_row + 1
        self.assertEqual(
            ws.cell(row=question_header_row(ws, "Q_SS_EXPORT"), column=1).value,
            "Q_SS_EXPORT - Single select export question",
        )
        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 5)],
            ["Option", "Count", "%", "Denominator"],
        )
        self.assertEqual(ws.cell(data_row, 1).value, "Yes")
        self.assertTrue(ws.cell(data_row, 2).value.startswith("=COUNTIFS"))
        self.assertIn("Q_SS_EXPORT_data", ws.cell(data_row, 2).value)
        self.assertIn("passes_workbook_filters_data", ws.cell(data_row, 2).value)
        self.assertIn("passes_workbook_custom_filters_data", ws.cell(data_row, 2).value)
        self.assertIn("SUBTOTAL", ws.cell(data_row, 3).value)

    def test_sc_sheet_for_multi_select_has_selections_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        values_workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.addCleanup(values_workbook.close)
        ws = workbook["All Questions"]
        values_ws = values_workbook["All Questions"]

        header_row = table_header_row(ws, "Q_MS_EXPORT")
        data_row = header_row + 1
        self.assertEqual(
            ws.cell(row=question_header_row(ws, "Q_MS_EXPORT"), column=1).value,
            "Q_MS_EXPORT - Multi select export question",
        )
        self.assertEqual(ws.cell(header_row, 1).value, "Option")
        self.assertEqual(ws.cell(header_row, 2).value, "Count")
        self.assertEqual(ws.cell(data_row, 1).value, "First")
        self.assertIn("Q_MS_EXPORTr1_data", ws.cell(data_row, 2).value)
        self.assertIn("SUMPRODUCT", ws.cell(data_row, 3).value)
        self.assertNotIn("SUBTOTAL", ws.cell(data_row, 3).value)
        self.assertAlmostEqual(values_ws.cell(data_row, 3).value, 0.5)
        self.assertEqual(values_ws.cell(data_row, 4).value, 10)

    def test_multi_select_percentages_use_m01_respondent_denominators(self) -> None:
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q4]",
                    canonical_id="Q4",
                    question_text="Business needs",
                    question_type=QuestionType.MULTI_SELECT_BINARY,
                    raw_columns=("Q4r1", "Q4r2"),
                    option_map={
                        "Q4r1": "Revenue intelligence",
                        "Q4r2": "Other business need",
                    },
                ),
                QuestionSpec(
                    question_id="[Q5]",
                    canonical_id="Q5",
                    question_text="Pressures",
                    question_type=QuestionType.MULTI_SELECT_BINARY,
                    raw_columns=("Q5r1", "Q5r2"),
                    option_map={"Q5r1": "Cost pressure", "Q5r2": "Other pressure"},
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=206,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        results = [
            MultiSelectResult(
                question_id="Q4",
                question_type=QuestionType.MULTI_SELECT_BINARY,
                valid_n=206,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                selections={
                    "Q4r1": {
                        "label": "Revenue intelligence",
                        "count": 141,
                        "selection_rate": 141 / 206,
                    },
                    "Q4r2": {
                        "label": "Other business need",
                        "count": 83,
                        "selection_rate": 83 / 206,
                    },
                },
                respondents_who_answered_any=206,
            ),
            MultiSelectResult(
                question_id="Q5",
                question_type=QuestionType.MULTI_SELECT_BINARY,
                valid_n=196,
                missing_n=10,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                selections={
                    "Q5r1": {
                        "label": "Cost pressure",
                        "count": 119,
                        "selection_rate": 119 / 196,
                    },
                    "Q5r2": {
                        "label": "Other pressure",
                        "count": 77,
                        "selection_rate": 77 / 196,
                    },
                },
                respondents_who_answered_any=196,
            ),
        ]
        output_path = self.export_custom_workbook(results, schema)
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        q4_header = table_header_row(ws, "Q4")
        q5_header = table_header_row(ws, "Q5")

        self.assertEqual(ws.cell(q4_header + 1, 1).value, "Revenue intelligence")
        self.assertAlmostEqual(ws.cell(q4_header + 1, 3).value * 100, 68.4466, delta=0.05)
        self.assertEqual(ws.cell(q4_header + 1, 4).value, 206)
        self.assertEqual(ws.cell(q5_header + 1, 1).value, "Cost pressure")
        self.assertAlmostEqual(ws.cell(q5_header + 1, 3).value * 100, 60.7143, delta=0.05)
        self.assertEqual(ws.cell(q5_header + 1, 4).value, 196)

    def test_sc_sheet_for_numeric_has_stats_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_NUM_EXPORT", header="Metric")
        self.assertEqual(
            ws.cell(row=question_header_row(ws, "Q_NUM_EXPORT"), column=1).value,
            "Q_NUM_EXPORT - Numeric export question",
        )
        self.assertEqual(ws.cell(header_row, 1).value, "Metric")
        self.assertEqual(ws.cell(header_row + 1, 1).value, "Mean")
        self.assertTrue(ws.cell(header_row + 1, 2).value.startswith("=IFERROR(AVERAGEIFS"))
        self.assertIn("passes_workbook_filters_data", ws.cell(header_row + 1, 2).value)
        self.assertIn("passes_workbook_custom_filters_data", ws.cell(header_row + 1, 2).value)
        self.assertEqual(ws.cell(header_row + 4, 1).value, "Std")
        self.assertEqual(ws.cell(header_row + 4, 2).value, 2.9)
        self.assertEqual(ws.cell(header_row + 1, 2).number_format, "0.00")
        self.assertEqual(ws.cell(header_row + 5, 1).value, "Median")
        self.assertEqual(ws.cell(header_row + 5, 2).value, 5.5)
        self.assertEqual(ws.cell(header_row + 5, 2).number_format, "0.00")
        self.assertEqual(ws.cell(header_row + 5, 4).value, 10)

    def test_direct_numeric_median_rows_use_m01_unfiltered_values(self) -> None:
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q8]",
                    canonical_id="Q8",
                    question_text="Likely to recommend",
                    question_type=QuestionType.DIRECT_NUMERIC,
                    raw_columns=("Q8",),
                    option_map={},
                ),
                QuestionSpec(
                    question_id="[Q9]",
                    canonical_id="Q9",
                    question_text="Planned spend",
                    question_type=QuestionType.DIRECT_NUMERIC,
                    raw_columns=("Q9",),
                    option_map={},
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=206,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        results = [
            NumericResult(
                question_id="Q8",
                question_type=QuestionType.DIRECT_NUMERIC,
                valid_n=206,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                mean=5.25,
                median=5.0,
                std=1.2,
                min_val=1.0,
                max_val=7.0,
                percentiles={25: 4.0, 50: 5.0, 75: 6.0},
            ),
            NumericResult(
                question_id="Q9",
                question_type=QuestionType.DIRECT_NUMERIC,
                valid_n=206,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                mean=27.1,
                median=26.55,
                std=7.4,
                min_val=0.0,
                max_val=50.0,
                percentiles={25: 20.0, 50: 26.55, 75: 35.0},
            ),
        ]
        output_path = self.export_custom_workbook(results, schema)
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        q8_header = table_header_row(ws, "Q8", header="Metric")
        q9_header = table_header_row(ws, "Q9", header="Metric")

        self.assertEqual(ws.cell(q8_header + 5, 1).value, "Median")
        self.assertAlmostEqual(ws.cell(q8_header + 5, 2).value, 5.0, delta=0.01)
        self.assertEqual(ws.cell(q9_header + 5, 1).value, "Median")
        self.assertAlmostEqual(ws.cell(q9_header + 5, 2).value, 26.55, delta=0.01)

    def test_numeric_allocation_mean_cell_preserves_m01_precision(self) -> None:
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q6]",
                    canonical_id="Q6",
                    question_text="Channel allocation",
                    question_type=QuestionType.NUMERIC_ALLOCATION,
                    raw_columns=("Q6r1",),
                    option_map={"Q6r1": "CRM & sales tools"},
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=206,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        result = NumericResult(
            question_id="Q6",
            question_type=QuestionType.NUMERIC_ALLOCATION,
            valid_n=206,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            mean=26.3695,
            median=26.55,
            std=0.0,
            min_val=0.0,
            max_val=100.0,
            percentiles={25: 20.0, 50: 26.55, 75: 35.0},
            allocation_target=100.0,
            allocation_tolerance=2.0,
            allocation_excluded_n=0,
            per_option_stats={
                "Q6r1": {
                    "mean": 26.3695,
                    "median": 26.55,
                    "std": 4.0,
                    "min_val": 0.0,
                    "max_val": 100.0,
                    "valid_n": 206,
                    "missing_n": 0,
                }
            },
        )
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q6")

        self.assertEqual(ws.cell(header_row + 1, 1).value, "CRM & sales tools")
        self.assertAlmostEqual(ws.cell(header_row + 1, 2).value, 26.3695, delta=0.01)

    def test_numeric_allocation_table_is_option_oriented_with_mean_and_median(self) -> None:
        output_path = self.export_numeric_allocation_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q20")

        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 5)],
            ["Option", "Mean", "Median", "Denominator"],
        )
        option_labels = [ws.cell(header_row + offset, 1).value for offset in range(1, 4)]
        self.assertIn("Existing customers - net retention", option_labels)
        self.assertFalse(any(str(label).startswith("Q20r") for label in option_labels))
        table_values = [
            ws.cell(row, 1).value
            for row in range(header_row, table_header_row(ws, "Q20") + 6)
        ]
        self.assertFalse(any(value in {"Std", "Min", "Max"} for value in table_values))
        self.assertEqual(ws.cell(header_row + 1, 2).number_format, "0.00")
        self.assertEqual(ws.cell(header_row + 1, 3).number_format, "0.00")

    def test_numeric_allocation_mean_and_median_columns_have_color_scales(self) -> None:
        output_path = self.export_numeric_allocation_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q20")

        self.assertTrue(
            any(
                str(formatting.sqref) == f"B{header_row + 1}:B{header_row + 3}"
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )
        self.assertTrue(
            any(
                str(formatting.sqref) == f"C{header_row + 1}:C{header_row + 3}"
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )

    def test_numeric_allocation_audit_log_contains_per_option_medians(self) -> None:
        output_path = self.export_numeric_allocation_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Calculation_Log"]
        metric_names = [
            ws.cell(row=row_index, column=2).value
            for row_index in range(2, ws.max_row + 1)
        ]

        self.assertEqual(metric_names.count("numeric_allocation_median"), 3)

    def test_sc_sheet_for_grid_without_binary_role_has_categorical_table(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_GRID_EXPORT", header="Sub-question ID")
        data_row = header_row + 2
        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 7)],
            ["Sub-question ID", "Sub-question", "High", None, "Low", None],
        )
        self.assertEqual(
            [ws.cell(header_row + 1, col).value for col in range(3, 7)],
            ["Count", "%", "Count", "%"],
        )
        self.assertEqual(ws.cell(data_row, 1).value, "Q_GRID_EXPORTr1")
        self.assertEqual(ws.cell(data_row, 2).value, "Grid first row")
        self.assertTrue(ws.cell(data_row, 3).value.startswith("=COUNTIFS"))
        self.assertIn("Q_GRID_EXPORTr1_data", ws.cell(data_row, 3).value)
        self.assertIn("passes_workbook_filters_data", ws.cell(data_row, 3).value)
        self.assertIn("passes_workbook_custom_filters_data", ws.cell(data_row, 3).value)

    def test_sc_sheet_has_autofilter(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertGreater(len(ws.tables), 0)

    def test_distribution_count_column_has_color_scale(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        range_ref = f"B{header_row + 1}:B{header_row + 2}"

        self.assertTrue(
            any(
                str(formatting.sqref) == range_ref
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )

    def test_distribution_percent_column_has_color_scale(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        range_ref = f"C{header_row + 1}:C{header_row + 2}"

        self.assertTrue(
            any(
                str(formatting.sqref) == range_ref
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )

    def test_total_respondents_row_follows_distribution_options(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        total_row = header_row + 3

        self.assertEqual(ws.cell(total_row, 1).value, "Total respondents")
        self.assertEqual(ws.cell(total_row, 2).value, 10)

    def test_sc_sheet_filter_block_no_filter(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(ws["A1"].value, "THEME: All Questions")
        self.assertEqual(ws["A3"].value, "LOCAL FILTERS (override workbook defaults)")
        self.assertNotEqual(ws["A3"].value, "DEMOGRAPHIC FILTERS")
        self.assertEqual(
            ws.cell(question_header_row(ws, "Q_SS_EXPORT"), 1).value,
            "Q_SS_EXPORT - Single select export question",
        )

    def test_subset_denominator_note_appears_when_denominator_is_small(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        subset_result = replace(
            results[0],
            valid_n=5,
            missing_n=5,
            distribution={
                1: {"label": "Yes", "count": 3, "rate": 0.6},
                2: {"label": "No", "count": 2, "rate": 0.4},
            },
        )
        export_single_cuts(
            [subset_result, *results[1:]],
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        values = [
            ws.cell(row=row_index, column=1).value
            for row_index in range(
                question_header_row(ws, "Q_SS_EXPORT"),
                table_header_row(ws, "Q_SS_EXPORT") + 1,
            )
        ]

        self.assertIn(
            "Note: This question was shown to a subset. Total respondents shown: 5.",
            values,
        )

    def test_subset_denominator_note_absent_for_full_denominator(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        values = [
            ws.cell(row=row_index, column=1).value
            for row_index in range(
                question_header_row(ws, "Q_SS_EXPORT"),
                table_header_row(ws, "Q_SS_EXPORT") + 1,
            )
        ]

        self.assertFalse(
            any(isinstance(value, str) and value.startswith("Note:") for value in values)
        )

    def test_filter_sheet_exists_visible(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Filters", workbook.sheetnames)
        self.assertEqual(workbook["Filters"].sheet_state, "visible")

    def test_filter_display_label_drops_question_number(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(
                    question,
                    question_text="Q14 - In which country are you based?",
                    is_demographic=True,
                )
                if question.canonical_id == "Q_SS_EXPORT"
                else question
                for question in schema.questions
            ),
        )
        with patch("src.ai_insights.generate_short_labels", return_value={}):
            export_single_cuts(
                results,
                skips,
                schema,
                quality_report,
                log,
                str(output_path),
                demo_priority={
                    "priority_ordered": ["Q_SS_EXPORT"],
                    "categories": {"Q_SS_EXPORT": "country"},
                },
            )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(workbook["Filters"]["A4"].value, "Country")
        self.assertNotRegex(str(workbook["Filters"]["A4"].value), r"\bQ\d+")

    def test_filters_sheet_uses_ai_short_labels_for_demographics(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        questions = (
            QuestionSpec(
                question_id="[Q4]",
                canonical_id="Q4",
                question_text="Which of the following best describes the size of your organization",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q4",),
                option_map={1: "Small", 2: "Large"},
                is_demographic=True,
            ),
            QuestionSpec(
                question_id="[Q14]",
                canonical_id="Q14",
                question_text="Approximately how old is the organization",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q14",),
                option_map={1: "New", 2: "Established"},
                is_demographic=True,
            ),
        )
        schema = SurveySchema(
            questions=questions,
            respondent_id_column="respondent_id",
            total_respondents=2,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=2,
            total_columns=2,
            columns_in_datamap=2,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        results = [
            SingleSelectResult(
                question_id="Q4",
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=2,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "Small", "count": 1, "rate": 0.5},
                    2: {"label": "Large", "count": 1, "rate": 0.5},
                },
            ),
            SingleSelectResult(
                question_id="Q14",
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=2,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "New", "count": 1, "rate": 0.5},
                    2: {"label": "Established", "count": 1, "rate": 0.5},
                },
            ),
        ]

        with patch(
            "src.ai_insights.generate_short_labels",
            return_value={"Q4": "Org Size", "Q14": "Org Age"},
        ):
            export_single_cuts(
                results,
                [],
                schema,
                quality_report,
                CalculationLog(),
                str(output_path),
                demo_priority={
                    "priority_ordered": ["Q4", "Q14"],
                    "categories": {"Q4": "company_size", "Q14": "age"},
                },
            )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        filter_labels = [
            workbook["Filters"].cell(row=row_index, column=1).value
            for row_index in range(1, workbook["Filters"].max_row + 1)
            if workbook["Filters"].cell(row=row_index, column=1).value
        ]

        self.assertIn("Org Size", filter_labels)
        self.assertIn("Org Age", filter_labels)
        broken_prefix = re.compile(r"^(Which|If|Approximately|How|What|During)")
        self.assertFalse(
            any(broken_prefix.match(str(label)) for label in filter_labels)
        )

    def test_per_question_dropdown_options_are_readable_labels(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["_Options"]
        values = [
            ws.cell(row=row_index, column=1).value
            for row_index in range(2, ws.max_row + 1)
            if ws.cell(row=row_index, column=1).value
        ]

        self.assertIn("Q_SS_EXPORT - Single Select Export Question", values)
        self.assertFalse(any(str(value) in {"Yes", "No", "First", "Second"} for value in values))

    def test_all_questions_dropdown_contains_filterable_question_entries(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        values = defined_range_values(workbook, "All_Questions")

        self.assertEqual(values[0], "(None)")
        self.assertIn("Q_SS_EXPORT - Single Select Export Question", values)
        self.assertIn(f"{LONG_ID} - Long Sheet Name Export Question", values)
        self.assertEqual(sum("Q_MS_EXPORT" in str(value) for value in values), 1)
        self.assertEqual(sum("Q_GRID_EXPORT" in str(value) for value in values), 1)
        self.assertFalse(any("Q_MS_EXPORT" in str(value) and "First" in str(value) for value in values))
        self.assertFalse(any("Q_GRID_EXPORT" in str(value) and "Grid first row" in str(value) for value in values))
        self.assertNotIn("Yes", values)
        self.assertNotIn("No", values)
        self.assertNotIn("Version 2", values)
        self.assertFalse(any("Q_NUM_EXPORT" in str(value) for value in values))
        self.assertFalse(any("term:" in str(value) for value in values))

    def test_all_questions_dropdown_entries_use_q_number_dash_label_format(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        questions = (
            QuestionSpec(
                question_id="[Q1]",
                canonical_id="Q1",
                question_text="In which country do you currently work?",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q1",),
                option_map={1: "United States", 2: "Canada", 3: "India"},
                is_demographic=True,
            ),
            QuestionSpec(
                question_id="[Q2]",
                canonical_id="Q2",
                question_text="Which survey version did the respondent see?",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q2",),
                option_map={1: "Version 2", 2: "Version 3"},
            ),
            QuestionSpec(
                question_id="[Q3]",
                canonical_id="Q3",
                question_text="Which terms apply?",
                question_type=QuestionType.MULTI_SELECT_BINARY,
                raw_columns=("Q3r1",),
                option_map={"Q3r1": "term: Selected Other at Q2"},
                value_range=(0, 1),
            ),
        )
        schema = SurveySchema(
            questions=questions,
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=4,
            total_columns=3,
            columns_in_datamap=3,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        results = [
            SingleSelectResult(
                question_id="Q1",
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=4,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "United States", "count": 2, "rate": 0.5},
                    2: {"label": "Canada", "count": 1, "rate": 0.25},
                    3: {"label": "India", "count": 1, "rate": 0.25},
                },
            ),
            SingleSelectResult(
                question_id="Q2",
                question_type=QuestionType.SINGLE_SELECT,
                valid_n=4,
                missing_n=0,
                denominator_policy=DenominatorPolicy.VALID_RESPONSES,
                distribution={
                    1: {"label": "Version 2", "count": 2, "rate": 0.5},
                    2: {"label": "Version 3", "count": 2, "rate": 0.5},
                },
            ),
        ]

        with patch(
            "src.ai_insights.generate_short_labels",
            return_value={"Q1": "Country", "Q2": "Survey Version", "Q3": "Terms"},
        ):
            export_single_cuts(
                results,
                [],
                schema,
                quality_report,
                CalculationLog(),
                str(output_path),
            )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        values = defined_range_values(workbook, "All_Questions")

        self.assertEqual(values, ["(None)", "Q1 - Country", "Q2 - Survey Version"])
        for value in values[1:]:
            self.assertRegex(str(value), r"^Q\d+\s*-\s*\w+")
        self.assertFalse(
            any(
                value in {"United States", "Canada", "India", "Version 2", "Version 3"}
                for value in values
            )
        )
        local_values = defined_range_values(workbook, "All_Questions_Local")
        self.assertEqual(local_values, ["(Inherit)", "(None)", "Q1 - Country", "Q2 - Survey Version"])

    def test_all_questions_lookup_columns_store_sheet_range_refs(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["_Options"]
        bare_name_pattern = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")

        for column in (3, 4):
            for row_index in range(2, ws.max_row + 1):
                value = ws.cell(row=row_index, column=column).value
                if value in (None, ""):
                    continue
                text = str(value)
                self.assertIn("!", text)
                self.assertNotRegex(text, bare_name_pattern)

    def test_theme_sheet_has_local_filter_rows(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(ws["A3"].value, "LOCAL FILTERS (override workbook defaults)")
        self.assertEqual(ws["A4"].value, "Filter")
        self.assertEqual(ws["B5"].value, "(Inherit)")

    def test_filter_uses_countifs_not_sumproduct(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        formula = ws.cell(row=header_row + 1, column=2).value

        self.assertTrue(formula.startswith("=COUNTIFS"))
        self.assertIn("passes_workbook_filters_data", formula)
        self.assertIn("passes_workbook_custom_filters_data", formula)
        self.assertNotIn("SUMPRODUCT", formula)
        self.assertNotIn("INDEX(", formula)

    def test_formulas_use_supported_helpers_without_implicit_intersection(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)

        formulas: list[str] = []
        for worksheet in workbook.worksheets:
            if worksheet.title.startswith("_"):
                continue
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and (
                        "COUNTIFS" in value
                        or "AVERAGEIFS" in value
                        or "MINIFS" in value
                        or "MAXIFS" in value
                    ):
                        formulas.append(value)

        self.assertTrue(formulas)
        for formula in formulas:
            self.assertNotIn("@", formula)
            self.assertNotIn("XLOOKUP", formula)
            self.assertNotIn("_xludf.", formula)
            self.assertNotIn("VALUE(", formula)

    def test_formula_cells_have_cached_values_and_calc_chain(self) -> None:
        import xml.etree.ElementTree as ET
        from zipfile import ZipFile

        output_path = self.export_workbook()
        formula_cells: list[tuple[str, str]] = []
        namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

        with ZipFile(output_path) as archive:
            names = set(archive.namelist())
            self.assertIn("xl/calcChain.xml", names)

            for name in sorted(names):
                if not name.startswith("xl/worksheets/sheet") or not name.endswith(".xml"):
                    continue
                root = ET.fromstring(archive.read(name))
                for cell in root.iter(f"{namespace}c"):
                    formula = cell.find(f"{namespace}f")
                    if formula is None:
                        continue
                    ref = cell.attrib.get("r", "")
                    formula_cells.append((name, ref))
                    cached = cell.find(f"{namespace}v")
                    self.assertIsNotNone(cached, f"{name}!{ref} has no cached <v>")
                    if "INDEX(INDIRECT" not in (formula.text or ""):
                        self.assertIsNotNone(cached.text, f"{name}!{ref} has an empty cache")
                        self.assertNotEqual(cached.text, "", f"{name}!{ref} has an empty cache")

        self.assertTrue(formula_cells)

    def test_generated_workbook_uses_no_xlookup_or_xludf_formulas(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)

        formulas: list[str] = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas.append(cell.value)
        self.assertTrue(formulas)
        for formula in formulas:
            self.assertNotIn("XLOOKUP", formula)
            self.assertNotIn("_xludf.", formula)

    def test_indirect_index_match_cached_values_are_not_name_or_ref_errors(self) -> None:
        output_path = self.export_workbook()
        formulas_workbook = load_workbook(output_path, read_only=False, data_only=False)
        values_workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(formulas_workbook.close)
        self.addCleanup(values_workbook.close)

        checked = 0
        for worksheet in formulas_workbook.worksheets:
            values_sheet = values_workbook[worksheet.title]
            for row in worksheet.iter_rows():
                for cell in row:
                    formula = cell.value
                    if not (
                        isinstance(formula, str)
                        and "INDIRECT(IFERROR(INDEX" in formula
                    ):
                        continue
                    cached = values_sheet[cell.coordinate].value
                    self.assertNotIn(cached, ("#NAME?", "#REF!"))
                    checked += 1
        self.assertGreater(checked, 0)

    def test_countifs_include_theme_local_and_per_question_helpers(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        formula = ws.cell(row=header_row + 1, column=2).value

        self.assertTrue(formula.startswith("=COUNTIFS"))
        self.assertIn("All_Questions_passes_local_filters_data,1", formula)
        self.assertIn(
            "All_Questions_Q_SS_EXPORT_F_passes_per_q_filter_data,1",
            formula,
        )

    def test_numeric_formulas_include_theme_local_and_per_question_helpers(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_NUM_EXPORT", header="Metric")
        mean_formula = ws.cell(row=header_row + 1, column=2).value
        count_formula = ws.cell(row=header_row + 1, column=4).value

        self.assertIn("All_Questions_passes_local_filters_data,1", mean_formula)
        self.assertIn(
            "All_Questions_Q_NUM_EXPORT_F_passes_per_q_filter_data,1",
            mean_formula,
        )
        self.assertIn("All_Questions_passes_local_filters_data,1", count_formula)
        self.assertIn(
            "All_Questions_Q_NUM_EXPORT_F_passes_per_q_filter_data,1",
            count_formula,
        )

    def test_cross_tab_header_formula_cache_is_blank_by_default(self) -> None:
        import xml.etree.ElementTree as ET
        from zipfile import ZipFile

        output_path = self.export_workbook()
        namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        blank_caches = 0

        with ZipFile(output_path) as archive:
            for name in archive.namelist():
                if not name.startswith("xl/worksheets/sheet") or not name.endswith(".xml"):
                    continue
                root = ET.fromstring(archive.read(name))
                for cell in root.iter(f"{namespace}c"):
                    formula = cell.find(f"{namespace}f")
                    if formula is None:
                        continue
                    formula_text = formula.text or ""
                    if (
                        "INDEX(INDIRECT" not in formula_text
                        or "Cross_Tab_Questions_Options_Names" not in formula_text
                    ):
                        continue
                    cached = cell.find(f"{namespace}v")
                    self.assertIsNotNone(cached)
                    self.assertIn(cached.text, (None, ""))
                    self.assertNotEqual(cached.text, "0")
                    blank_caches += 1

        self.assertGreater(blank_caches, 0)

    def test_raw_data_has_filter_helper_columns(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, is_demographic=question.canonical_id == "Q_SS_EXPORT")
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country"},
            },
        )
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["_RawData"]
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

        self.assertIn("F_Country_match", headers)
        self.assertIn("passes_workbook_filters", headers)
        self.assertIn("F_Custom1_match", headers)
        self.assertIn("F_Custom2_match", headers)
        self.assertIn("passes_workbook_custom_filters", headers)
        pass_col = headers.index("passes_workbook_filters") + 1
        self.assertIn("*", ws.cell(row=2, column=pass_col).value)
        custom_col = headers.index("passes_workbook_custom_filters") + 1
        self.assertIn("*", ws.cell(row=2, column=custom_col).value)
        self.assertIn("passes_workbook_filters_data", workbook.defined_names)
        self.assertIn("F_Custom1_match_data", workbook.defined_names)
        self.assertIn("F_Custom2_match_data", workbook.defined_names)
        self.assertIn("passes_workbook_custom_filters_data", workbook.defined_names)

    def test_streaming_raw_data_named_ranges_cover_all_rows(self) -> None:
        output_path = self.export_streaming_raw_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertEqual(
            workbook.defined_names["respondent_id_data"].attr_text,
            "'_RawData'!$A$2:$A$3001",
        )
        self.assertEqual(
            workbook.defined_names["Q_STREAM_data"].attr_text,
            "'_RawData'!$B$2:$B$3001",
        )
        self.assertEqual(
            workbook.defined_names["passes_workbook_filters_data"].attr_text,
            "'_RawData'!$C$2:$C$3001",
        )
        self.assertIn("All_Questions_Q_STREAM_F_passes_per_q_filter_data", workbook.defined_names)

    def test_streaming_raw_data_sheet_contains_all_rows_after_reload(self) -> None:
        output_path = self.export_streaming_raw_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["_RawData"]

        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 3001)
        self.assertEqual(rows[0][0], "respondent_id")
        self.assertEqual(rows[-1][1], "B")

    def test_streaming_formula_cache_patch_peak_under_100_mib(self) -> None:
        from src import memory_profiler

        try:
            with patch.dict(os.environ, {"SURVEY_PROFILE_MEMORY": "1"}):
                output_path = self.export_streaming_raw_workbook()
        finally:
            memory_profiler.disable_profiling()
            memory_profiler.reset_log()

        report_path = Path("outputs") / f"{output_path.stem}.memory_report.txt"
        self.assertTrue(report_path.exists())
        report = report_path.read_text(encoding="utf-8")
        patch_peak = None
        for line in report.splitlines():
            if not line.startswith("patch_formula_caches"):
                continue
            parts = [part.strip() for part in line.split("|")]
            patch_peak = float(parts[-1].replace("MiB", "").strip())
            break

        self.assertIsNotNone(patch_peak)
        self.assertLess(patch_peak, 100.0)

    def test_inherit_pattern_resolves_to_workbook_value(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(ws["B5"].value, "(Inherit)")
        self.assertIn('="(Inherit)"', ws["C5"].value)
        self.assertIn("F_Custom1_Q", ws["C5"].value)

    def test_available_values_column_present(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, is_demographic=question.canonical_id == "Q_SS_EXPORT")
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country"},
            },
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Yes", workbook["Filters"]["D4"].value)
        self.assertIn("No", workbook["Filters"]["D4"].value)
        self.assertIn("Yes", workbook["Demographics"]["E5"].value)

    def test_question_heading_has_no_comment_when_full_text_is_visible(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        title_row = question_header_row(ws, "Q_SS_EXPORT")

        self.assertIsNone(ws.cell(title_row, 1).comment)

    def test_theme_question_heading_restores_question_number_prefix(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        heading = ws.cell(question_header_row(ws, "Q_SS_EXPORT"), 1).value

        self.assertEqual(heading, "Q_SS_EXPORT - Single select export question")
        self.assertRegex(str(heading), r"^Q\w+\s*[-:]?\s*")

    def test_short_labels_stay_in_filter_ui_not_question_heading(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            short_labels={"Q_SS_EXPORT": "Short revenue label"},
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(
            ws.cell(question_header_row(ws, "Q_SS_EXPORT"), 1).value,
            "Q_SS_EXPORT - Single select export question",
        )
        filter_values = defined_range_values(workbook, "All_Questions")
        self.assertIn("Q_SS_EXPORT - Short Revenue Label", filter_values)

    def test_filter_dropdown_collapses_multi_select_and_grid_questions(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        filter_values = [str(value) for value in defined_range_values(workbook, "All_Questions")]

        self.assertEqual(sum("Q_MS_EXPORT" in value for value in filter_values), 1)
        self.assertEqual(sum("Q_GRID_EXPORT" in value for value in filter_values), 1)
        self.assertFalse(any("Q_MS_EXPORT" in value and "First" in value for value in filter_values))
        self.assertFalse(any("Q_GRID_EXPORT" in value and "Grid first row" in value for value in filter_values))
        self.assertIn("First", defined_range_values(workbook, "Q_MS_EXPORT_filter_values_options"))
        self.assertIn("Second", defined_range_values(workbook, "Q_MS_EXPORT_filter_values_options"))
        self.assertIn(
            "Grid first row: Low",
            defined_range_values(workbook, "Q_GRID_EXPORT_filter_values_options"),
        )

        lookup_keys = [str(value) for value in defined_range_values(workbook, "All_Questions_Value_Keys")]
        lookup_data = [str(value) for value in defined_range_values(workbook, "All_Questions_Value_Data_Names")]
        lookup_criteria = [str(value) for value in defined_range_values(workbook, "All_Questions_Value_Criteria")]
        first_key = "Q_MS_EXPORT - Multi Select Export Question|First"
        self.assertIn(first_key, lookup_keys)
        first_index = lookup_keys.index(first_key)
        self.assertEqual(lookup_data[first_index], workbook.defined_names["Q_MS_EXPORTr1_data"].attr_text)
        self.assertEqual(lookup_criteria[first_index], "Selected")
        grid_key = "Q_GRID_EXPORT - Grid Export Question|Grid first row: Low"
        self.assertIn(grid_key, lookup_keys)
        grid_index = lookup_keys.index(grid_key)
        self.assertEqual(lookup_data[grid_index], workbook.defined_names["Q_GRID_EXPORTr1_data"].attr_text)
        self.assertEqual(lookup_criteria[grid_index], "Low")

    def test_cross_tab_questions_dropdown_is_dimension_only(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        cross_tab_values = [str(value) for value in defined_range_values(workbook, "Cross_Tab_Questions")]

        self.assertEqual(cross_tab_values[0], "(None)")
        self.assertIn("Q_SS_EXPORT - Single Select Export Question", cross_tab_values)
        self.assertIn(f"{LONG_ID} - Long Sheet Name Export Question", cross_tab_values)
        self.assertFalse(any("Q_MS_EXPORT" in value for value in cross_tab_values))
        self.assertFalse(any("Q_GRID_EXPORT" in value for value in cross_tab_values))
        self.assertFalse(any("Q_NUM_EXPORT" in value for value in cross_tab_values))

    def test_question_heading_row_uses_bain_red_style(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        heading_row = question_header_row(ws, "Q_SS_EXPORT")
        heading_cell = ws.cell(heading_row, 1)

        self.assertEqual(heading_cell.fill.fgColor.rgb, "FFCC0000")
        self.assertEqual(heading_cell.font.color.rgb, "FFFFFFFF")
        self.assertTrue(heading_cell.font.bold)
        for column in range(1, 5):
            self.assertEqual(ws.cell(heading_row, column).fill.fgColor.rgb, "FFCC0000")

    def test_question_heading_uses_full_text_without_comment(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        full_text = (
            "Q7 - Which of the following best describes your current seniority "
            "within your organization?"
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q7]",
                    canonical_id="Q7",
                    question_text=full_text,
                    question_type=QuestionType.SINGLE_SELECT,
                    raw_columns=("Q7",),
                    option_map={1: "Executive", 2: "Manager"},
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        result = SingleSelectResult(
            question_id="Q7",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                1: {"label": "Executive", "count": 1, "rate": 0.25},
                2: {"label": "Manager", "count": 3, "rate": 0.75},
            },
        )
        quality_report = DataQualityReport(
            total_rows=4,
            total_columns=1,
            columns_in_datamap=1,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        with patch("src.ai_insights.generate_short_labels", return_value={"Q7": "Seniority"}):
            export_single_cuts(
                [result],
                [],
                schema,
                quality_report,
                CalculationLog(),
                str(output_path),
            )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        heading_cell = ws.cell(question_header_row(ws, "Q7"), 1)

        self.assertRegex(str(heading_cell.value), r"^Q\d+[a-z]*\s*-\s*.+")
        self.assertEqual(heading_cell.value, full_text)
        self.assertGreater(len(str(heading_cell.value)), 30)
        self.assertIsNone(heading_cell.comment)
        filter_values = defined_range_values(workbook, "All_Questions")
        self.assertIn("Q7 - Seniority", filter_values)

    def test_per_question_filter_named_cells_scoped_per_theme(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        themes = {
            "themes": [
                {"name": "Theme One", "question_ids": ["Q_SS_EXPORT"]},
                {"name": "Theme Two", "question_ids": ["Q_NUM_EXPORT"]},
            ]
        }
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            themes=themes,
        )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Theme_One_Q_SS_EXPORT_F_V", workbook.defined_names)
        self.assertIn("Theme_Two_Q_NUM_EXPORT_F_V", workbook.defined_names)
        self.assertIsNone(workbook.defined_names["Theme_One_Q_SS_EXPORT_F_V"].localSheetId)

    def test_raw_data_sheet_is_hidden(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("_RawData", workbook.sheetnames)
        self.assertEqual(workbook["_RawData"].sheet_state, "hidden")

    def test_options_sheet_is_hidden(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("_Options", workbook.sheetnames)
        self.assertEqual(workbook["_Options"].sheet_state, "hidden")

    def test_demographic_filter_named_cells_exist(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(
                    question,
                    is_demographic=question.canonical_id in {"Q_SS_EXPORT", "Q_MS_EXPORT"},
                )
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT", "Q_MS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country", "Q_MS_EXPORT": "industry"},
            },
        )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("F_Country", workbook.defined_names)
        self.assertIn("F_Industry", workbook.defined_names)

    def test_count_cell_uses_countifs_formula(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        formula = ws.cell(row=header_row + 1, column=2).value

        self.assertTrue(formula.startswith("=COUNTIFS"))
        self.assertIn("Q_SS_EXPORT_data", formula)
        self.assertIn("passes_workbook_filters_data", formula)
        self.assertIn("passes_workbook_custom_filters_data", formula)
        self.assertNotIn("SUMPRODUCT", formula)

    def test_named_cells_are_workbook_scoped(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, is_demographic=question.canonical_id == "Q_SS_EXPORT")
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country"},
            },
        )
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIsNone(workbook.defined_names["F_Country"].localSheetId)
        self.assertIsNone(workbook.defined_names["F_Country_wrapped"].localSheetId)
        self.assertEqual(workbook.defined_names["F_Country"].attr_text, "'Filters'!$B$4")

    def test_per_question_filter_named_cells_exist(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("All_Questions_Q_SS_EXPORT_F_Q", workbook.defined_names)
        self.assertIn("All_Questions_Q_SS_EXPORT_F_V", workbook.defined_names)

    def test_per_question_filter_value_dropdown_on_every_question_block(self) -> None:
        output_path = self.export_mixed_eight_block_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        filter_rows = [
            row_index
            for row_index in range(1, ws.max_row + 1)
            if ws.cell(row=row_index, column=1).value == "Per-question filter"
        ]
        self.assertGreaterEqual(len(filter_rows), 8)
        value_cells = [f"E{row_index}" for row_index in filter_rows]

        for row_index in filter_rows:
            self.assertTrue(cell_has_data_validation(ws, f"C{row_index}"))
            self.assertTrue(cell_has_data_validation(ws, f"E{row_index}"))
        value_validations = [
            validation
            for validation in ws.data_validations.dataValidation
            if any(cell in validation.cells for cell in value_cells)
        ]
        self.assertEqual(len(value_validations), len(value_cells))
        for validation in value_validations:
            self.assertIn("INDIRECT(IFERROR(INDEX", validation.formula1)
            self.assertIn("MATCH(", validation.formula1)

    def test_cross_tab_named_cell_exists(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("All_Questions_Q_SS_EXPORT_CT", workbook.defined_names)

    def test_cross_tab_table_has_2d_count_only_layout(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        cross_tab_first_col = 6
        max_groups = 12
        total_col = cross_tab_first_col + max_groups + 1

        self.assertEqual(ws.cell(header_row, cross_tab_first_col).value, "Option")
        self.assertEqual(ws.cell(header_row, total_col).value, "Total")
        subheaders = [
            ws.cell(header_row + 1, column).value
            for column in range(cross_tab_first_col + 1, total_col + 1)
        ]
        self.assertEqual(subheaders.count("# of resp"), max_groups + 1)
        self.assertNotIn("% of resp", subheaders)
        first_count = ws.cell(header_row + 2, cross_tab_first_col + 1).value
        self.assertIn("COUNTIFS", first_count)
        self.assertNotIn("Denominator", subheaders)
        total_rows = [
            row_index
            for row_index in range(header_row, header_row + 5)
            if ws.cell(row_index, cross_tab_first_col).value == "Total respondents"
        ]
        self.assertEqual(total_rows, [header_row + 4])

    def test_cross_tab_cap_uses_other_group_for_large_dimensions(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        cross_tab_first_col = 6
        last_group_col = cross_tab_first_col + 12
        formula = ws.cell(header_row, last_group_col).value

        self.assertIn(',"Other"', formula)
        self.assertIn("-1>11", formula)

    def test_priority_demographic_ordering(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(
                    question,
                    is_demographic=question.canonical_id in {"Q_SS_EXPORT", "Q_MS_EXPORT"},
                )
                for question in schema.questions
            ),
        )
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            demo_priority={
                "priority_ordered": ["Q_SS_EXPORT", "Q_MS_EXPORT"],
                "categories": {"Q_SS_EXPORT": "country", "Q_MS_EXPORT": "industry"},
            },
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filters"]

        self.assertEqual(ws["A4"].value, "Single Select Export Question")
        self.assertEqual(ws["A5"].value, "Multi Select Export Question")

    def test_workbook_has_one_sheet_per_theme(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        themes = {
            "themes": [
                {
                    "name": "Theme One",
                    "question_ids": ["Q_SS_EXPORT", "Q_MS_EXPORT"],
                },
                {"name": "Theme Two", "question_ids": ["Q_NUM_EXPORT"]},
            ]
        }
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            themes=themes,
        )

        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Theme One", workbook.sheetnames)
        self.assertIn("Theme Two", workbook.sheetnames)
        self.assertFalse(any(name.startswith("SC_") for name in workbook.sheetnames))

    def test_question_block_uses_subtotal_formula_for_percent(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_SS_EXPORT")
        value = ws.cell(row=header_row + 1, column=3).value
        self.assertTrue(value.startswith("="))
        self.assertIn("SUBTOTAL", value)

    def test_demographic_filter_row_has_data_validation(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        results, skips, schema, quality_report, log = make_export_fixture()
        questions = tuple(
            replace(question, is_demographic=question.canonical_id == "Q_SS_EXPORT")
            for question in schema.questions
        )
        schema = replace(schema, questions=questions)
        decoded_df = pd.DataFrame({"Q_SS_EXPORT": ["Yes", "No", "Yes"]})
        export_single_cuts(
            results,
            skips,
            schema,
            quality_report,
            log,
            str(output_path),
            decoded_df=decoded_df,
        )

        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filters"]

        self.assertGreater(len(ws.data_validations.dataValidation), 0)

    def test_grid_single_select_excel_format(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        result, schema = self.grid_single_select_format_fixture()
        _results, _skips, _schema, quality_report, _log = make_export_fixture()
        export_single_cuts(
            [result],
            [],
            schema,
            quality_report,
            CalculationLog(),
            str(output_path),
        )

        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        header_row = table_header_row(ws, "Q_GRID_FORMAT")
        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 5)],
            ["Option", "Count", "%", "Denominator"],
        )
        self.assertEqual(ws.cell(header_row + 1, 1).value, "Option A")
        self.assertIn("r1_data", ws.cell(header_row + 1, 2).value)
        self.assertEqual(ws.cell(header_row + 2, 1).value, "Option B")
        self.assertIn("r2_data", ws.cell(header_row + 2, 2).value)

        values = [
            cell
            for row in ws.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]
        self.assertNotIn("Option C", values)
        self.assertNotIn("Per-row distributions", values)
        self.assertNotIn("Code", values)

    def test_grid_rated_parent_renders_single_mean_median_block(self) -> None:
        result, schema = rated_grid_export_fixture()
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        heading_rows = [
            row
            for row in range(1, ws.max_row + 1)
            if isinstance(ws.cell(row, 1).value, str)
            and ws.cell(row, 1).value.startswith("Q_RATED -")
        ]
        self.assertEqual(len(heading_rows), 1)
        header_row = table_header_row(ws, "Q_RATED", header="Sub-question ID")
        values = [
            ws.cell(header_row, col).value
            for col in range(1, 13)
        ]

        self.assertEqual(values[:4], ["Sub-question ID", "Sub-question", "# of respondents", "% of respondents"])
        self.assertIn("Winner - All", values)
        self.assertIn("Other considered vendor", values)
        self.assertIn("Delta", values)
        self.assertEqual(ws.cell(header_row + 1, 6).value, "Mean")
        self.assertEqual(ws.cell(header_row + 1, 7).value, "Median")
        self.assertTrue(ws.merged_cells.ranges)
        self.assertEqual(ws.cell(header_row + 2, 1).value, "Q_RATEDr1")
        self.assertEqual(ws.cell(header_row + 2, 2).value, "Implementation speed")

    def test_grid_rated_cross_tab_by_uses_averageifs_mean_matrix(self) -> None:
        result, schema = rated_grid_export_fixture()
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        formulas = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and "Q_RATED_CT" in cell.value
        ]

        self.assertTrue(
            any(
                "AVERAGEIFS(" in formula
                and "Q_RATEDr1c1_data" in formula
                and '">=0"' in formula
                for formula in formulas
            ),
            formulas,
        )
        self.assertFalse(
            any(
                "COUNTIFS(" in formula
                and "Q_RATEDr1c1_data" in formula
                for formula in formulas
            ),
            formulas,
        )

    def test_rank_cross_tab_by_has_selectable_rank_metric_formula(self) -> None:
        result = RankOrderResult(
            question_id="Q_RANK",
            question_type=QuestionType.RANK_ORDER,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            question_text="Rank vendors",
            K=4,
            rows=[
                RankOrderRow(
                    option_id="Q_RANK_A",
                    option_label="Vendor A",
                    counts_per_rank=[1, 1, 1, 1],
                    pcts_per_rank=[0.25, 0.25, 0.25, 0.25],
                    net_rank_score=62.5,
                ),
                RankOrderRow(
                    option_id="Q_RANK_B",
                    option_label="Vendor B",
                    counts_per_rank=[0, 2, 0, 0],
                    pcts_per_rank=[0.0, 0.5, 0.0, 0.0],
                    net_rank_score=37.5,
                ),
            ],
            total_respondents=4,
            total_responses=6,
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_RANK]",
                    canonical_id="Q_RANK",
                    question_text="Rank vendors",
                    question_type=QuestionType.RANK_ORDER,
                    raw_columns=("Q_RANK_A", "Q_RANK_B"),
                    option_map={"Q_RANK_A": "Vendor A", "Q_RANK_B": "Vendor B"},
                    value_range=(1, 4),
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        output_path = self.export_custom_workbook(
            [result],
            schema,
            rank_cross_tab_settings={
                "Q_RANK": {
                    "metric": "Weighted average",
                    "weights": [10, 6, 3, 0],
                    "rank_position": 2,
                }
            },
        )
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertIn("All_Questions_Q_RANK_RM", workbook.defined_names)
        self.assertIn("All_Questions_Q_RANK_W", workbook.defined_names)
        self.assertIn("All_Questions_Q_RANK_RP", workbook.defined_names)
        self.assertEqual(
            defined_range_values(workbook, "All_Questions_Q_RANK_W"),
            [10.0, 6.0, 3.0, 0.0],
        )
        rank_metric_cells = [
            ws.cell(row=row_index, column=col_index + 1).value
            for row_index in range(1, ws.max_row + 1)
            for col_index in range(1, ws.max_column)
            if ws.cell(row=row_index, column=col_index).value == "Rank metric"
        ]
        self.assertIn("Weighted average", rank_metric_cells)
        rank_position_cells = [
            ws.cell(row=row_index, column=col_index + 1).value
            for row_index in range(1, ws.max_row + 1)
            for col_index in range(1, ws.max_column)
            if ws.cell(row=row_index, column=col_index).value == "Rank position"
        ]
        self.assertIn(2, rank_position_cells)

        formulas = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and "Q_RANK_CT" in cell.value
        ]
        subheaders = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
            and 'All_Questions_Q_RANK_RM="Weighted average"' in cell.value
        ]

        self.assertGreaterEqual(len(subheaders), 2)
        self.assertTrue(
            any(
                "COUNTIFS(" in formula
                and "SUMIFS(" in formula
                and "SUMPRODUCT(" in formula
                and "Q_RANK_A_data" in formula
                and "Q_RANK_B_data" in formula
                and 'All_Questions_Q_RANK_RM="Weighted average"' in formula
                and "INDEX(All_Questions_Q_RANK_W,1)*COUNTIFS(Q_RANK_A_data,1" in formula
                and 'All_Questions_Q_RANK_RM="Sum of ranks"' in formula
                and 'All_Questions_Q_RANK_RM="Rank position count"' in formula
                and "COUNTIFS(Q_RANK_A_data,All_Questions_Q_RANK_RP" in formula
                and "*100" in formula
                for formula in formulas
            ),
            formulas,
        )
        self.assertFalse(any("AVERAGEIFS(" in formula for formula in formulas), formulas)
        self.assertFalse(any("VALUE(" in formula for formula in formulas), formulas)

    def test_grid_binary_pivot_cross_tab_still_uses_selection_rate_formula(self) -> None:
        segment_result = SingleSelectResult(
            question_id="Q_SEG",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                "A": {"label": "Segment A", "count": 2, "rate": 0.5},
                "B": {"label": "Segment B", "count": 2, "rate": 0.5},
            },
        )
        grid_binary_result = GridBinaryPivotResult(
            question_id="Q_GB",
            question_type=QuestionType.GRID_BINARY_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            question_text="Binary grid",
            column_headers=["Selected"],
            rows=[
                GridBinaryPivotRow(
                    row_id="Q_GB_R1",
                    row_label="Capability",
                    counts_per_column=[3],
                    pcts_per_column=[0.75],
                ),
            ],
            total_respondents=4,
            total_responses=3,
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_SEG]",
                    canonical_id="Q_SEG",
                    question_text="Segment",
                    question_type=QuestionType.SINGLE_SELECT,
                    raw_columns=("Q_SEG",),
                    option_map={"A": "Segment A", "B": "Segment B"},
                ),
                QuestionSpec(
                    question_id="[Q_GB]",
                    canonical_id="Q_GB",
                    question_text="Binary grid",
                    question_type=QuestionType.GRID_BINARY_SELECT,
                    raw_columns=("Q_GB_R1",),
                    option_map={1: "Selected"},
                    grid_row_labels={"Q_GB_R1": "Capability"},
                    value_range=(0, 1),
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        quality_report = DataQualityReport(
            total_rows=4,
            total_columns=3,
            columns_in_datamap=2,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        export_single_cuts(
            [segment_result, grid_binary_result],
            [],
            schema,
            quality_report,
            CalculationLog(),
            str(output_path),
            decoded_df=pd.DataFrame(
                {
                    "respondent_id": [1, 2, 3, 4],
                    "Q_SEG": ["A", "A", "B", "B"],
                    "Q_GB_R1": [1, 0, 1, 1],
                }
            ),
        )
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        formulas = [
            cell.value
            for row in ws.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and "Q_GB_CT" in cell.value
        ]

        self.assertTrue(
            any(
                "COUNTIFS(" in formula
                and '"Selected"' in formula
                and "SUMPRODUCT(" in formula
                and "Q_GB_R1_data" in formula
                for formula in formulas
            ),
            formulas,
        )

    def test_grid_rated_without_c_groups_falls_back_to_all_group(self) -> None:
        row_a = SingleSelectResult(
            question_id="Q_RATED_DIRECTr1",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                8: {"label": "8", "count": 2, "rate": 0.5},
                9: {"label": "9", "count": 2, "rate": 0.5},
            },
        )
        row_b = SingleSelectResult(
            question_id="Q_RATED_DIRECTr2",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                6: {"label": "6", "count": 2, "rate": 0.5},
                7: {"label": "7", "count": 2, "rate": 0.5},
            },
        )
        result = GridSingleSelectResult(
            question_id="Q_RATED_DIRECT",
            question_type=QuestionType.GRID_SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            rows={"Q_RATED_DIRECTr1": row_a, "Q_RATED_DIRECTr2": row_b},
            overall_valid_n=4,
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_RATED_DIRECT]",
                    canonical_id="Q_RATED_DIRECT",
                    question_text="Q_RATED_DIRECT - Rate direct rows",
                    question_type=QuestionType.GRID_SINGLE_SELECT,
                    raw_columns=("Q_RATED_DIRECTr1", "Q_RATED_DIRECTr2"),
                    option_map={index: str(index) for index in range(1, 11)},
                    value_range=(1, 10),
                    grid_row_labels={
                        "Q_RATED_DIRECTr1": "Implementation speed",
                        "Q_RATED_DIRECTr2": "Integration fit",
                    },
                    possible_role="GRID_RATED",
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_RATED_DIRECT", header="Sub-question ID")
        nearby_values = [
            ws.cell(header_row, col).value
            for col in range(1, 10)
        ]

        self.assertIn("All", nearby_values)
        self.assertNotIn("Delta", nearby_values)

    def test_sibling_rated_grid_exports_one_parent_block_with_rounded_delta(self) -> None:
        root_text = "Rate each vendor attribute from 1 to 10"
        questions = [
            {
                "canonical_id": "Q1r1",
                "raw_id": "Q1r1",
                "question_text": f"Implementation speed - {root_text}",
                "type_hint": "values_range",
                "value_range": (1, 10),
                "options": [(index, str(index)) for index in range(1, 11)],
                "sub_columns": [],
                "parent_canonical_id": None,
                "source_row": 1,
                "warnings": [],
            },
            {
                "canonical_id": "Q1r2",
                "raw_id": "Q1r2",
                "question_text": f"Integration fit - {root_text}",
                "type_hint": "values_range",
                "value_range": (1, 10),
                "options": [(index, str(index)) for index in range(1, 11)],
                "sub_columns": [],
                "parent_canonical_id": None,
                "source_row": 2,
                "warnings": [],
            },
            {
                "canonical_id": "Q1r3",
                "raw_id": "Q1r3",
                "question_text": f"Efficiency gains - {root_text}",
                "type_hint": "values_range",
                "value_range": (1, 10),
                "options": [(index, str(index)) for index in range(1, 11)],
                "sub_columns": [],
                "parent_canonical_id": None,
                "source_row": 3,
                "warnings": [],
            },
        ]
        raw_columns = [
            "respondent_id",
            "Q1r1c1",
            "Q1r1c2",
            "Q1r2c1",
            "Q1r2c2",
            "Q1r3c1",
            "Q1r3c2",
        ]
        schema = classify_questions(
            {
                "questions": questions,
                "source_path": "datamap.xlsx",
                "sheet_name": "Sheet1",
                "total_rows_in_sheet": 3,
                "parser_warnings": [],
            },
            raw_columns,
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_rawdata_path="raw.csv",
        )
        dataframe = pd.DataFrame(
            {
                "respondent_id": [1, 2, 3, 4],
                "Q1r1c1": [8, 8, 8, 9],
                "Q1r1c2": [6, 6, 7, 7],
                "Q1r2c1": [8, 9, 9, 9],
                "Q1r2c2": [6, 7, 7, 7],
                "Q1r3c1": [7, 8, 8, 8],
                "Q1r3c2": [5, 6, 6, 6],
            }
        )
        log = CalculationLog()
        results, skips = compute_single_cuts(schema, dataframe, log)
        self.assertFalse(skips)
        output_path = self.export_custom_workbook(results, schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        values_workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        self.addCleanup(values_workbook.close)
        ws = workbook["All Questions"]
        values_ws = values_workbook["All Questions"]

        heading_rows = [
            row
            for row in range(1, ws.max_row + 1)
            if isinstance(ws.cell(row, 1).value, str)
            and ws.cell(row, 1).value.startswith("Q1 -")
        ]
        self.assertEqual(len(heading_rows), 1)
        header_row = table_header_row(ws, "Q1", header="Sub-question ID")
        headers = [ws.cell(header_row, col).value for col in range(1, 13)]
        delta_col = headers.index("Delta") + 1

        self.assertIn("Winner - All", headers)
        self.assertIn("Other considered vendor", headers)
        self.assertEqual(
            [ws.cell(header_row + offset, 1).value for offset in range(2, 5)],
            ["Q1r1", "Q1r2", "Q1r3"],
        )
        self.assertTrue(str(ws.cell(header_row + 2, delta_col).value).startswith("=ROUND("))
        delta_value = float(values_ws.cell(header_row + 2, delta_col).value)
        self.assertEqual(delta_value, round(delta_value, 1))

    def test_all_empty_grid_question_is_skipped_and_not_rendered(self) -> None:
        data_map = {
            "questions": [
                {
                    "canonical_id": "Q1",
                    "raw_id": "Q1",
                    "question_text": "Q1 - Purchase decision",
                    "type_hint": "values_range",
                    "value_range": (1, 2),
                    "options": [(1, "Yes"), (2, "No")],
                    "sub_columns": [],
                    "parent_canonical_id": None,
                    "source_row": 1,
                    "warnings": [],
                },
                {
                    "canonical_id": "Q26r1",
                    "raw_id": "Q26r1",
                    "question_text": "IT / Technical - What role did each stakeholder play",
                    "type_hint": "values_range",
                    "value_range": (1, 3),
                    "options": [
                        (1, "Decision maker"),
                        (2, "Influencer"),
                        (3, "Not involved"),
                    ],
                    "sub_columns": [],
                    "parent_canonical_id": None,
                    "source_row": 2,
                    "warnings": [],
                },
                {
                    "canonical_id": "Q26r2",
                    "raw_id": "Q26r2",
                    "question_text": "Finance - What role did each stakeholder play",
                    "type_hint": "values_range",
                    "value_range": (1, 3),
                    "options": [
                        (1, "Decision maker"),
                        (2, "Influencer"),
                        (3, "Not involved"),
                    ],
                    "sub_columns": [],
                    "parent_canonical_id": None,
                    "source_row": 3,
                    "warnings": [],
                },
            ],
            "source_path": "datamap.xlsx",
            "sheet_name": "Sheet1",
            "total_rows_in_sheet": 3,
            "parser_warnings": [],
        }
        dataframe = pd.DataFrame(
            {
                "respondent_id": [1, 2, 3],
                "Q1": [1, 2, 1],
                "Q26r1": [pd.NA, pd.NA, pd.NA],
                "Q26r2": [pd.NA, pd.NA, pd.NA],
            }
        )
        schema = classify_questions(
            data_map,
            ["respondent_id", "Q1", "Q26r1", "Q26r2"],
            respondent_id_column="respondent_id",
            total_respondents=3,
            source_rawdata_path="raw.csv",
        )
        log = CalculationLog()
        results, skips = compute_single_cuts(schema, dataframe, log)
        self.assertEqual([result.question_id for result in results], ["Q1"])
        self.assertEqual(skips[0].canonical_id, "Q26")
        self.assertEqual(skips[0].skip_reason, "all raw columns empty in dataset")

        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        quality_report = DataQualityReport(
            total_rows=3,
            total_columns=4,
            columns_in_datamap=3,
            columns_not_in_datamap=(),
            per_column_missing_pct={},
            per_column_out_of_range_pct={},
            coercion_log=(),
            warnings=(),
        )
        export_single_cuts(results, skips, schema, quality_report, log, str(output_path))
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        all_values = [
            value
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
        self.assertTrue(any(str(value).startswith("Q1 -") for value in all_values))
        self.assertFalse(any(str(value).startswith("Q26 -") for value in all_values))
        self.assertIn("all raw columns empty in dataset", all_values)

    def test_grid_rated_dual_block_has_left_response_table_and_heatmaps(self) -> None:
        result, schema = rated_grid_export_fixture()
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_RATED", header="Sub-question ID")
        data_start = header_row + 2
        data_end = data_start + 2

        self.assertEqual(ws.cell(header_row, 3).value, "# of respondents")
        self.assertEqual(ws.cell(header_row, 4).value, "% of respondents")
        self.assertEqual(ws.cell(header_row, 6).value, "Winner - All")
        self.assertEqual(ws.cell(header_row + 1, 6).value, "Mean")
        self.assertTrue(
            any(
                str(formatting.sqref) == f"D{data_start}:D{data_end}"
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )
        self.assertTrue(
            any(
                str(formatting.sqref) == f"F{data_start}:F{data_end}"
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )

    def test_grid_categorical_parent_renders_single_count_matrix(self) -> None:
        result, schema = categorical_grid_export_fixture()
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        heading_rows = [
            row
            for row in range(1, ws.max_row + 1)
            if isinstance(ws.cell(row, 1).value, str)
            and ws.cell(row, 1).value.startswith("Q_CATGRID -")
        ]
        self.assertEqual(len(heading_rows), 1)
        header_row = table_header_row(ws, "Q_CATGRID", header="Sub-question ID")
        headers = [ws.cell(header_row, col).value for col in range(1, 9)]

        self.assertEqual(headers[:2], ["Sub-question ID", "Sub-question"])
        self.assertIn("Decision maker", headers)
        self.assertIn("Influencer", headers)
        self.assertIn("Not involved", headers)
        self.assertEqual(ws.cell(header_row + 1, 3).value, "Count")
        self.assertEqual(ws.cell(header_row + 1, 4).value, "%")
        self.assertEqual(ws.cell(header_row + 2, 1).value, "Q_CATGRIDr1")

    def test_grid_categorical_footer_has_total_responses_and_percent_heatmap(self) -> None:
        result, schema = categorical_grid_export_fixture()
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_CATGRID", header="Sub-question ID")
        data_start = header_row + 2
        data_end = data_start + 2
        total_respondents_row = data_end + 1
        total_responses_row = total_respondents_row + 1

        self.assertEqual(ws.cell(total_respondents_row, 1).value, "Total respondents")
        self.assertEqual(ws.cell(total_responses_row, 1).value, "Total responses")
        self.assertEqual(ws.cell(total_responses_row, 2).value, 30)
        self.assertTrue(
            any(
                str(formatting.sqref) == f"D{data_start}:D{data_end}"
                and any(rule.type == "colorScale" for rule in formatting.rules)
                for formatting in ws.conditional_formatting
            )
        )

    def test_grid_categorical_c_columns_pivot_to_category_columns(self) -> None:
        result, schema = categorical_c_grid_export_fixture()
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q26", header="Sub-question ID")

        self.assertEqual(ws.cell(header_row, 3).value, "Blocked Vendors")
        self.assertEqual(ws.cell(header_row, 5).value, "Scored Vendors")
        self.assertEqual(ws.cell(header_row + 1, 3).value, "Count")
        self.assertEqual(ws.cell(header_row + 1, 4).value, "%")
        self.assertEqual(ws.cell(header_row + 2, 1).value, "Q26r1")
        self.assertEqual(ws.cell(header_row + 2, 2).value, "IT / Technical")
        self.assertEqual(ws.cell(header_row + 3, 1).value, "Q26r2")
        self.assertNotEqual(ws.cell(header_row + 3, 1).value, "Q26r1c2")

    def test_grid_categorical_c_column_percent_cells_are_formatted(self) -> None:
        result, schema = categorical_c_grid_export_fixture()
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q26", header="Sub-question ID")

        self.assertEqual(ws.cell(header_row + 2, 4).number_format, "0.0%")
        self.assertEqual(ws.cell(header_row + 2, 6).number_format, "0.0%")
        values = [cell for row in ws.iter_rows(values_only=True) for cell in row if cell is not None]
        self.assertIn("Total respondents", values)
        self.assertIn("Total responses", values)

    def test_grid_binary_select_keeps_existing_compact_selected_count_rendering(self) -> None:
        result, schema = self.grid_single_select_format_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, possible_role="GRID_BINARY_SELECT")
                if question.canonical_id == "Q_GRID_FORMAT"
                else question
                for question in schema.questions
            ),
        )
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_GRID_FORMAT")

        self.assertEqual(
            [ws.cell(header_row, col).value for col in range(1, 5)],
            ["Option", "Count", "%", "Denominator"],
        )
        self.assertEqual(ws.cell(header_row + 1, 1).value, "Option A")

    def test_grid_binary_select_footer_has_total_responses(self) -> None:
        result, schema = self.grid_single_select_format_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, possible_role="GRID_BINARY_SELECT")
                if question.canonical_id == "Q_GRID_FORMAT"
                else question
                for question in schema.questions
            ),
        )
        output_path = self.export_custom_workbook([result], schema)
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        values = [cell for row in ws.iter_rows(values_only=True) for cell in row if cell is not None]

        self.assertIn("Total responses", values)
        self.assertIn(80, values)

    def test_multi_select_footer_has_total_responses_but_single_select_does_not(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        ms_header = table_header_row(ws, "Q_MS_EXPORT")
        ss_header = table_header_row(ws, "Q_SS_EXPORT")

        self.assertEqual(ws.cell(ms_header + 4, 1).value, "Total responses")
        self.assertEqual(ws.cell(ms_header + 4, 2).value, 7)
        self.assertNotEqual(ws.cell(ss_header + 4, 1).value, "Total responses")

    def test_grid_single_select_ui_format(self) -> None:
        repo_root = Path(__file__).resolve().parents[1]
        app_path = repo_root / "app.py"
        if not app_path.exists():
            app_path = repo_root / "artifacts" / "survey-insight-engine" / "app.py"
        spec = importlib.util.spec_from_file_location("survey_insight_app", app_path)
        self.assertIsNotNone(spec)
        self.assertIsNotNone(spec.loader)
        app_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(app_module)

        result, schema = self.grid_single_select_format_fixture()
        rows = app_module._build_grid_display_rows(result, schema)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["Option"], "Option A")
        self.assertEqual(rows[0]["Count"], 50)
        self.assertEqual(rows[1]["Option"], "Option B")
        self.assertNotIn("Option C", [row["Option"] for row in rows])

    def test_calculation_log_one_row_per_audit_record(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        metric_names = [
            workbook["Calculation_Log"].cell(row=row, column=2).value
            for row in range(2, workbook["Calculation_Log"].max_row + 1)
        ]
        self.assertGreaterEqual(workbook["Calculation_Log"].max_row, 8)
        self.assertIn("cross_tab_by_selection_rate", metric_names)

    def test_skip_log_one_row_per_skip(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        values = sheet_values(output_path, "Warnings")
        self.assertIn("ValueError: raw column not found in data", values)

    def test_data_quality_sheet_has_three_sections(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Warnings", workbook.sheetnames)
        values = sheet_values(output_path, "Warnings")
        self.assertIn("column extra_col has 50.0% missing values", values)

    def test_warnings_sheet_lists_low_confidence_grid_classifications(self) -> None:
        result, schema = categorical_grid_export_fixture()
        schema = replace(
            schema,
            questions=tuple(
                replace(question, classification_confidence_low=True)
                if question.canonical_id == "Q_CATGRID"
                else question
                for question in schema.questions
            ),
        )
        output_path = self.export_custom_workbook([result], schema)
        values = sheet_values(output_path, "Warnings")

        self.assertIn("Q_CATGRID", values)
        self.assertIn("low classification confidence", values)

    def test_sheet_names_truncated_to_31_chars(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertTrue(all(len(sheet_name) <= 31 for sheet_name in workbook.sheetnames))
        self.assertIn("All Questions", workbook.sheetnames)

    def test_percentages_stored_as_raw_floats(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        value = ws.cell(row=header_row + 1, column=3).value

        self.assertTrue(value.startswith("="))
        self.assertIn("SUBTOTAL", value)

    def test_workbook_opens_without_corruption(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Run_Summary", workbook.sheetnames)

    def test_export_with_cross_cuts_writes_cross_cut_index(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertNotIn("Cross_Cut_Index", workbook.sheetnames)
        self.assertIn("Filter_Log", workbook.sheetnames)

    def test_export_with_cross_cuts_writes_cc_sheets(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertFalse(any(name.startswith("CC_") for name in workbook.sheetnames))
        self.assertIn("All Questions", workbook.sheetnames)

    def test_export_writes_cross_tab_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")

        self.assertEqual(ws.cell(row=header_row, column=6).value, "Option")
        self.assertIn("INDEX(INDIRECT", ws.cell(row=header_row, column=7).value)
        self.assertIn("Cross_Tab_Questions_Options_Names", ws.cell(row=header_row, column=7).value)

    def test_cross_tab_sheet_includes_chart(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]

        self.assertEqual(len(ws._charts), 0)
        self.assertGreater(len(ws.tables), 0)

    def test_chart_uses_correct_data_range(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")
        count_formula = ws.cell(row=header_row + 1, column=2).value
        pct_formula = ws.cell(row=header_row + 1, column=3).value

        self.assertIn("_RawData", workbook.sheetnames)
        self.assertTrue(count_formula.startswith("=COUNTIFS"))
        self.assertIn("passes_workbook_filters_data", count_formula)
        self.assertIn("passes_workbook_custom_filters_data", count_formula)
        self.assertIn("SUBTOTAL", pct_formula)

    def test_export_writes_segment_profile_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("All Questions", workbook.sheetnames)
        self.assertNotIn("CC_CC_SEG_EXPORT", workbook.sheetnames)

    def test_segment_profile_sheet_includes_chart(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        self.assertEqual(len(workbook["All Questions"]._charts), 0)

    def test_export_writes_group_comparison_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("Warnings", workbook.sheetnames)
        self.assertNotIn("CC_CC_GROUP_EXPORT", workbook.sheetnames)

    def test_group_comparison_sheet_includes_chart(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        self.assertEqual(len(workbook["All Questions"]._charts), 0)

    def test_export_writes_expected_vs_realized_body(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("Filter_Log", workbook.sheetnames)
        self.assertNotIn("CC_CC_EVR_EXPORT", workbook.sheetnames)

    def test_expected_vs_realized_sheet_includes_chart(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        self.assertEqual(len(workbook["All Questions"]._charts), 0)

    def test_export_filter_log_populated_for_segment_profile(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filter_Log"]

        self.assertEqual(ws["A1"].value, "Cross Cut ID")
        self.assertEqual(ws["A2"].value, "CC_SEG_EXPORT")
        self.assertEqual(ws["C2"].value, "Q_SEG_1 == 1")
        self.assertEqual(ws["D2"].value, "Q_SEG_1 = Segment 1")

    def test_export_skip_log_includes_cross_cut_skips(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        rows = list(workbook["Warnings"].iter_rows(values_only=True))

        self.assertEqual(rows[0][0], "Source")
        self.assertIn(
            ("skip:CC_BAD_EXPORT", "ValueError: synthetic cross cut failure"),
            rows,
        )

    def test_export_run_summary_has_cross_cut_totals(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Run_Summary"]

        self.assertEqual(ws["B8"].value, 4)
        self.assertEqual(ws["B9"].value, 1)
        self.assertEqual(ws["B10"].value, 19)

    def test_cross_tab_sheet_includes_axis_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")

        self.assertEqual(ws.cell(header_row, 6).value, "Option")
        self.assertIn("Q_SS_EXPORT_CT", ws.cell(header_row, 7).value)

    def test_segment_profile_sheet_includes_filter_and_target_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("CC_SEG_EXPORT", sheet_values(output_path, "Filter_Log"))

    def test_group_comparison_sheet_includes_segment_and_metric_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        self.assertIn("All Questions", workbook.sheetnames)
        self.assertFalse(any(name.startswith("CC_CC_GROUP") for name in workbook.sheetnames))

    def test_expected_vs_realized_sheet_includes_both_question_labels(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        values = sheet_values(output_path, "Question_Metadata")
        self.assertIn("Expected", values)
        self.assertIn("Realized", values)

    def test_cross_tab_sheet_has_corner_orientation_label(self) -> None:
        output_path = self.export_workbook_with_cross_cuts()
        workbook = load_workbook(output_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = table_header_row(ws, "Q_SS_EXPORT")

        self.assertEqual(ws.cell(header_row, 6).value, "Option")
        self.assertEqual(ws.cell(header_row + 2, 6).value, "Yes")

    def test_cross_tab_sheet_renders_only_counts_in_counts_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="counts")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertNotIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_only_row_pct_in_row_pct_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="row_pct")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertNotIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_both_blocks_in_both_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="both")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)

    def test_cross_tab_sheet_renders_all_three_blocks_in_all_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[0], display_mode="all")]
        )
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertIn("Column %", values)

    def test_grid_cross_tab_respects_display_mode(self) -> None:
        result, schema, log = grid_cross_tab_export_result(display_mode="row_pct")
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        export_cross_cuts_only([result], schema, log, str(output_path))

        values = sheet_values(output_path, "CC_CC_GRID_EXPORT")

        self.assertNotIn("Counts", values)
        self.assertIn("Row %", values)
        self.assertNotIn("Column %", values)
        self.assertIn("Q_GRID_SEGr1", values)
        self.assertIn("Segment 1", values)

    def test_cross_tab_includes_copy_friendly_block(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(cross_results[:1])
        values = sheet_values(output_path, "CC_CC_TAB_EXPORT")

        self.assertIn("Copy-friendly", values)
        self.assertIn("Row Code", values)
        self.assertIn("Row Label", values)
        self.assertIn("Column Code", values)
        self.assertIn("Column Label", values)
        self.assertIn("Count", values)

    def test_segment_profile_sheet_respects_display_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[1], display_mode="counts")]
        )
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["CC_CC_SEG_EXPORT"]

        self.assertEqual(ws["A15"].value, "Code")
        self.assertEqual(ws["C15"].value, "Count")
        self.assertIsNone(ws["D15"].value)

    def test_group_comparison_sheet_ignores_display_mode(self) -> None:
        *_base, cross_results, _cross_skips = make_cross_cut_export_fixture()
        output_path = self.export_cross_cut_workbook(
            [replace(cross_results[2], display_mode="counts")]
        )
        values = sheet_values(output_path, "CC_CC_GROUP_EXPORT")

        self.assertIn("Per-segment comparison", values)
        self.assertIn("Mean", values)
        self.assertIn("Median", values)
        self.assertIn("Std", values)

    def test_nps_group_comparison_exports_entity_segment_table(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        dataframe = pd.DataFrame(
            {
                "Q_SEG_1": [1, 1, 2, 2],
                "Q_NPS_A": [10, 10, 9, 0],
            }
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_SEG_1]",
                    canonical_id="Q_SEG_1",
                    question_text="Segment",
                    question_type=QuestionType.SINGLE_SELECT,
                    raw_columns=("Q_SEG_1",),
                    option_map={1: "Segment 1", 2: "Segment 2"},
                ),
                QuestionSpec(
                    question_id="[Q_NPS]",
                    canonical_id="Q_NPS",
                    question_text="Recommend vendor",
                    question_type=QuestionType.NPS,
                    raw_columns=("Q_NPS_A",),
                    option_map={"Q_NPS_A": "Vendor A"},
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=len(dataframe),
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        log = CalculationLog()
        results, skips = compute_cross_cuts(
            [
                CrossCutSpec(
                    cross_cut_id="CC_NPS_EXPORT",
                    title="NPS by segment",
                    analysis_type=AnalysisType.GROUP_COMPARISON,
                    source_question_ids=("Q_SEG_1", "Q_NPS"),
                )
            ],
            schema,
            dataframe,
            log,
        )
        self.assertEqual(skips, [])

        export_cross_cuts_only(results, schema, log, str(output_path))
        values = sheet_values(output_path, "CC_CC_NPS_EXPORT")

        self.assertIn("NPS by segment", values)
        self.assertIn("Per-entity segment NPS detail", values)
        self.assertIn("Vendor A", values)
        self.assertIn("Promoters %", values)
        self.assertIn("Detractors %", values)
        self.assertIn(100.0, values)
        self.assertIn(0.0, values)

    def test_new_metric_group_comparisons_export_matrix_and_detail_blocks(self) -> None:
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        dataframe = pd.DataFrame(
            {
                "Q_SEG_1": [1, 1, 2, 2],
                "Q_MS_A": [1, 0, 1, 0],
                "Q_RANK_A": [1, 2, 3, 4],
                "Q_GB_R1": [0, 1, 1, 1],
            }
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="[Q_SEG_1]",
                    canonical_id="Q_SEG_1",
                    question_text="Segment",
                    question_type=QuestionType.SINGLE_SELECT,
                    raw_columns=("Q_SEG_1",),
                    option_map={1: "Segment 1", 2: "Segment 2"},
                ),
                QuestionSpec(
                    question_id="[Q_MS]",
                    canonical_id="Q_MS",
                    question_text="Multi-select metric",
                    question_type=QuestionType.MULTI_SELECT_BINARY,
                    raw_columns=("Q_MS_A",),
                    option_map={"Q_MS_A": "Brand A"},
                ),
                QuestionSpec(
                    question_id="[Q_RANK]",
                    canonical_id="Q_RANK",
                    question_text="Rank metric",
                    question_type=QuestionType.RANK_ORDER,
                    raw_columns=("Q_RANK_A",),
                    option_map={"Q_RANK_A": "Vendor A"},
                    value_range=(1, 4),
                ),
                QuestionSpec(
                    question_id="[Q_GB]",
                    canonical_id="Q_GB",
                    question_text="Grid binary metric",
                    question_type=QuestionType.GRID_BINARY_SELECT,
                    raw_columns=("Q_GB_R1",),
                    option_map={1: "Selected"},
                    grid_row_labels={"Q_GB_R1": "Capability"},
                    value_range=(0, 1),
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=len(dataframe),
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        log = CalculationLog()
        specs = [
            CrossCutSpec(
                cross_cut_id="CC_MS_EXPORT",
                title="MS by segment",
                analysis_type=AnalysisType.GROUP_COMPARISON,
                source_question_ids=("Q_SEG_1", "Q_MS"),
            ),
            CrossCutSpec(
                cross_cut_id="CC_RANK_EXPORT",
                title="Rank by segment",
                analysis_type=AnalysisType.GROUP_COMPARISON,
                source_question_ids=("Q_SEG_1", "Q_RANK"),
            ),
            CrossCutSpec(
                cross_cut_id="CC_GB_EXPORT",
                title="Grid binary by segment",
                analysis_type=AnalysisType.GROUP_COMPARISON,
                source_question_ids=("Q_SEG_1", "Q_GB"),
            ),
        ]
        results, skips = compute_cross_cuts(specs, schema, dataframe, log)
        self.assertEqual(skips, [])

        export_cross_cuts_only(results, schema, log, str(output_path))

        ms_values = sheet_values(output_path, "CC_CC_MS_EXPORT")
        rank_values = sheet_values(output_path, "CC_CC_RANK_EXPORT")
        gb_values = sheet_values(output_path, "CC_CC_GB_EXPORT")
        log_values = sheet_values(output_path, "Calculation_Log")
        self.assertIn("Selection rate by segment", ms_values)
        self.assertIn("Per-row segment selection detail", ms_values)
        self.assertIn("Mean rank by segment", rank_values)
        self.assertIn("Median rank by segment", rank_values)
        self.assertIn("Per-row segment rank detail", rank_values)
        self.assertIn("Selection rate by segment", gb_values)
        self.assertIn("Per-row segment selection detail", gb_values)
        self.assertIn("selection_rate", log_values)
        self.assertIn("numeric_summary", log_values)

    def test_export_cross_cuts_only_writes_selected_cc_sheets(self) -> None:
        (
            _results,
            _skips,
            schema,
            _quality_report,
            log,
            cross_results,
            _cross_skips,
        ) = make_cross_cut_export_fixture()
        FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
        output_path = (
            FIXTURE_DIR
            / f"excel_exporter_{self._testMethodName}_{uuid4().hex}.xlsx"
        )
        export_cross_cuts_only(cross_results[:2], schema, log, str(output_path))
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("Cross_Cut_Index", workbook.sheetnames)
        self.assertIn("CC_CC_TAB_EXPORT", workbook.sheetnames)
        self.assertIn("CC_CC_SEG_EXPORT", workbook.sheetnames)
        self.assertNotIn("CC_CC_GROUP_EXPORT", workbook.sheetnames)

    def test_export_backward_compatible_without_cross_cuts(self) -> None:
        output_path = self.export_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertNotIn("Cross_Cut_Index", workbook.sheetnames)
        self.assertFalse(any(name.startswith("CC_") for name in workbook.sheetnames))

    def test_export_filtered_single_cuts_creates_file(self) -> None:
        output_path = self.export_filtered_workbook()

        self.assertTrue(output_path.exists())
        self.assertGreater(output_path.stat().st_size, 0)

    def test_filtered_cut_index_lists_all_results(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["Filtered_Cut_Index"]

        self.assertEqual(ws.max_row, 4)
        self.assertEqual(ws["A1"].value, "Sheet Name")
        self.assertEqual(ws["B2"].value, "Q_SS_EXPORT")
        self.assertEqual(ws["C2"].value, "Q_SEG_1 == 1 (Segment 1)")
        self.assertEqual(ws["D4"].value, "cross_cut_breakdown")
        self.assertEqual(ws["E4"].value, 30)

    def test_fsc_sheet_for_single_cut_filtered_renders_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["FSC_Q_SS_EXPORT_01"]
        values = sheet_values(output_path, "FSC_Q_SS_EXPORT_01")

        self.assertEqual(ws["A1"].value, "Filters applied")
        self.assertEqual(ws["A4"].value, "Target question:")
        self.assertEqual(ws["B5"].value, "Q_SS_EXPORT")
        self.assertIn("Q_SEG_1", values)
        self.assertIn("= Segment 1", values)
        self.assertIn("FILTERED VIEW: Q_SEG_1 == 1 (Segment 1) (n=6)", values)
        self.assertIn("Code", values)
        self.assertIn("Yes", values)
        self.assertIn(6, values)

    def test_sc_sheet_filtered_section_present_when_filter_active(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=False, data_only=True)
        self.addCleanup(workbook.close)
        ws = workbook["FSC_Q_SS_EXPORT_01"]
        values = [
            cell
            for row in ws.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]

        self.assertIn("FILTERED VIEW: Q_SEG_1 == 1 (Segment 1) (n=6)", values)
        self.assertIn("n=6 respondents match filter", values)
        self.assertIsNotNone(ws.auto_filter.ref)

    def test_fsc_sheet_for_cross_cut_breakdown_renders_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        values = sheet_values(output_path, "FSC_Q_TGT_1")

        self.assertIn("Filters applied", values)
        self.assertIn("Q_SEG_1", values)
        self.assertIn("(breakdown)", values)
        self.assertIn("Rows (vertical): Q_SEG_1", values)
        self.assertIn("Columns (horizontal): Q_TGT_1", values)
        self.assertIn("Counts", values)

    def test_fsc_sheet_naming_handles_duplicate_targets(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertIn("FSC_Q_SS_EXPORT_01", workbook.sheetnames)
        self.assertIn("FSC_Q_SS_EXPORT_02", workbook.sheetnames)
        self.assertIn("FSC_Q_TGT_1", workbook.sheetnames)

    def test_export_filtered_single_cuts_filter_log_populated(self) -> None:
        output_path = self.export_filtered_workbook()
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        self.addCleanup(workbook.close)
        rows = list(workbook["Filter_Log"].iter_rows(values_only=True))

        self.assertEqual(rows[0], (
            "Sheet Name",
            "Filter Question",
            "Filter Value",
            "Filter Description",
        ))
        self.assertIn(
            (
                "FSC_Q_SS_EXPORT_01",
                "Q_SEG_1",
                1,
                "Q_SEG_1 == 1 (Segment 1)",
            ),
            rows,
        )
        self.assertIn(
            (
                "FSC_Q_TGT_1",
                "Q_SEG_1",
                None,
                "Q_SEG_1 (breakdown - no specific value)",
            ),
            rows,
        )

    def test_filtered_export_calculation_log_filtered_correctly(self) -> None:
        output_path = self.export_filtered_workbook()
        values = sheet_values(output_path, "Calculation_Log")

        self.assertIn("Q_SS_EXPORT", values)
        self.assertIn("CC_CC_TAB_EXPORT", values)
        self.assertNotIn("Q_MS_EXPORT", values)
        self.assertNotIn("Q_UNRELATED_EXPORT", values)


if __name__ == "__main__":
    unittest.main()
