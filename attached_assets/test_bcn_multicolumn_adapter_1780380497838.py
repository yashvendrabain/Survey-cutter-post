"""Tests for the BCN multi-column data-map adapter."""

from __future__ import annotations

import unittest

from openpyxl import Workbook, load_workbook

from src.adapters.bcn_multicolumn import BCN_SUB_COLUMN_RE, BcnMulticolumnAdapter
from tests.conftest import DATAMAP_FIXTURE_PATH


def _compact_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["Q1", "Employment status"])
    worksheet.append(["1", "Full-time"])
    worksheet.append(["2", "Part-time"])
    worksheet.append([None, None])
    worksheet.append(["Q2", "Industry"])
    worksheet.append(["1", "Technology"])
    worksheet.append(["2", "Retail"])
    return workbook


def _bcn_workbook(sheet_name: str = "Sheet1") -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["Question ID", "Question Text", "Type"])
    worksheet.append(["Q1", "A question", "Values: 1-2"])
    return workbook


class TestBcnMulticolumnAdapter(unittest.TestCase):
    def test_detects_bcn_fixture(self) -> None:
        workbook = load_workbook(DATAMAP_FIXTURE_PATH, read_only=True, data_only=True)
        try:
            result = BcnMulticolumnAdapter().detect(workbook)
        finally:
            workbook.close()

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("BCN-style", result.reason)

    def test_detects_explicit_question_id_header(self) -> None:
        workbook = _bcn_workbook()

        result = BcnMulticolumnAdapter().detect(workbook)

        self.assertGreaterEqual(result.confidence, 0.6)

    def test_detects_bcn_format_when_sheet_named_data_map(self) -> None:
        result = BcnMulticolumnAdapter().detect(_bcn_workbook("Data map"))

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("Data map", result.reason)

    def test_detects_bcn_format_when_sheet_named_datamap(self) -> None:
        result = BcnMulticolumnAdapter().detect(_bcn_workbook("Datamap"))

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("Datamap", result.reason)

    def test_returns_zero_when_no_datamap_sheet(self) -> None:
        workbook = _bcn_workbook("Survey Layout")

        result = BcnMulticolumnAdapter().detect(workbook)

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_detects_format_in_codebook_sheet(self) -> None:
        result = BcnMulticolumnAdapter().detect(_bcn_workbook("Codebook"))

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("Codebook", result.reason)

    def test_detects_format_in_schema_sheet(self) -> None:
        result = BcnMulticolumnAdapter().detect(_bcn_workbook("Schema"))

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("Schema", result.reason)

    def test_does_not_match_question_metadata_sheet(self) -> None:
        result = BcnMulticolumnAdapter().detect(_bcn_workbook("Question_Metadata"))

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_does_not_match_dictionary_sheet(self) -> None:
        result = BcnMulticolumnAdapter().detect(_bcn_workbook("Dictionary"))

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_does_not_match_variables_sheet(self) -> None:
        result = BcnMulticolumnAdapter().detect(_bcn_workbook("Variables"))

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_does_not_match_metadata_sheet(self) -> None:
        result = BcnMulticolumnAdapter().detect(_bcn_workbook("Metadata"))

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_raw_sub_column_pattern_boost_accepts_new_separators(self) -> None:
        workbook = _compact_workbook()
        raw_columns = [f"Q14s{i}" for i in range(1, 7)] + [
            f"Q14_{i}" for i in range(1, 7)
        ]

        result = BcnMulticolumnAdapter().detect(workbook, raw_columns)

        self.assertGreaterEqual(result.confidence, 0.5)
        self.assertIn("raw columns match", result.reason)

    def test_sub_column_regex_accepts_existing_and_new_patterns(self) -> None:
        for column in ("Q14r1", "Q14r1c2", "Q14s1", "Q14_1", "Q14r98oe"):
            with self.subTest(column=column):
                self.assertIsNotNone(BCN_SUB_COLUMN_RE.match(column))

    def test_parse_bcn_fixture_preserves_expected_question_shape(self) -> None:
        workbook = load_workbook(DATAMAP_FIXTURE_PATH, read_only=True, data_only=True)
        try:
            setattr(workbook, "_survey_source_path", str(DATAMAP_FIXTURE_PATH))
            parsed = BcnMulticolumnAdapter().parse(workbook)
        finally:
            workbook.close()

        questions = {question["canonical_id"]: question for question in parsed["questions"]}
        self.assertIn("Q3", questions)
        self.assertEqual(questions["Q3"]["options"], [(1, "Yes"), (2, "No")])
        self.assertEqual(parsed["source_path"], str(DATAMAP_FIXTURE_PATH))

    def test_low_confidence_for_compact_workbook_without_raw_bcn_columns(self) -> None:
        result = BcnMulticolumnAdapter().detect(_compact_workbook())

        self.assertLess(result.confidence, 0.3)


if __name__ == "__main__":
    unittest.main()
