"""Tests for the BCN multi-column data-map adapter."""

from __future__ import annotations

import unittest

from openpyxl import Workbook, load_workbook

from src.adapters.bcn_multicolumn import BCN_SUB_COLUMN_RE, BcnMulticolumnAdapter
from tests.conftest import DATAMAP_FIXTURE_PATH


def _compact_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["Q1", "Employment status"])
    worksheet.append(["1", "Full-time"])
    worksheet.append(["2", "Part-time"])
    worksheet.append([None, None])
    worksheet.append(["Q2", "Industry"])
    worksheet.append(["1", "Technology"])
    worksheet.append(["2", "Retail"])
    return workbook


class TestBcnMulticolumnAdapter(unittest.TestCase):
    def test_detects_bcn_fixture(self) -> None:
        workbook = load_workbook(DATAMAP_FIXTURE_PATH, read_only=True, data_only=True)
        try:
            result = BcnMulticolumnAdapter().detect(workbook)
        finally:
            workbook.close()

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("BCN-style", result.reason)

    def test_detects_explicit_question_id_header(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["Question ID", "Question Text", "Type"])
        worksheet.append(["Q1", "A question", "Values: 1-2"])

        result = BcnMulticolumnAdapter().detect(workbook)

        self.assertGreaterEqual(result.confidence, 0.6)

    def test_raw_sub_column_pattern_boost_accepts_new_separators(self) -> None:
        workbook = _compact_workbook()
        raw_columns = [f"Q14s{i}" for i in range(1, 7)] + [
            f"Q14_{i}" for i in range(1, 7)
        ]

        result = BcnMulticolumnAdapter().detect(workbook, raw_columns)

        self.assertGreaterEqual(result.confidence, 0.5)
        self.assertIn("raw columns match", result.reason)

    def test_sub_column_regex_accepts_existing_and_new_patterns(self) -> None:
        for column in ("Q14r1", "Q14r1c2", "Q14s1", "Q14_1", "Q14r98oe"):
            with self.subTest(column=column):
                self.assertIsNotNone(BCN_SUB_COLUMN_RE.match(column))

    def test_parse_bcn_fixture_preserves_expected_question_shape(self) -> None:
        workbook = load_workbook(DATAMAP_FIXTURE_PATH, read_only=True, data_only=True)
        try:
            setattr(workbook, "_survey_source_path", str(DATAMAP_FIXTURE_PATH))
            parsed = BcnMulticolumnAdapter().parse(workbook)
        finally:
            workbook.close()

        questions = {question["canonical_id"]: question for question in parsed["questions"]}
        self.assertIn("Q3", questions)
        self.assertEqual(questions["Q3"]["options"], [(1, "Yes"), (2, "No")])
        self.assertEqual(parsed["source_path"], str(DATAMAP_FIXTURE_PATH))

    def test_low_confidence_for_compact_workbook_without_raw_bcn_columns(self) -> None:
        result = BcnMulticolumnAdapter().detect(_compact_workbook())

        self.assertLess(result.confidence, 0.3)


if __name__ == "__main__":
    unittest.main()
