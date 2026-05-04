"""Fixture builders for datamap parser tests."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"
DATAMAP_FIXTURE_PATH = FIXTURE_DIR / "test_datamap_minimal.xlsx"
MISSING_SHEET_FIXTURE_PATH = FIXTURE_DIR / "missing_sheet1.xlsx"
RAW_DECODER_CSV_PATH = FIXTURE_DIR / "raw_decoder_fixture.csv"
RAW_DECODER_XLSX_PATH = FIXTURE_DIR / "raw_decoder_fixture.xlsx"
RAW_DECODER_NO_ID_CSV_PATH = FIXTURE_DIR / "raw_decoder_no_id_fixture.csv"
GOLDEN_30_RESPONDENTS_PATH = FIXTURE_DIR / "golden_30_respondents.csv"


def ensure_fixtures() -> None:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    _build_minimal_datamap(DATAMAP_FIXTURE_PATH)
    _build_missing_sheet_workbook(MISSING_SHEET_FIXTURE_PATH)
    _build_raw_decoder_fixtures()
    _build_golden_30_respondents_fixture()


def _build_minimal_datamap(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    index_sheet = workbook.create_sheet("Index")
    index_sheet.append(["[QIndex]: This sheet must be ignored"])
    index_sheet.append(["Values: 1-2"])
    index_sheet.append([None, 1, "Should not appear"])

    rows = [
        [None, None, None],
        ["  [Q3]: Are you currently in a full-time position  ", None, None],
        [" Values: 1-2 ", None, None],
        [None, 1, " Yes "],
        ["", "2", "No"],
        [None, None, None],
        [None, None, None],
        ["This is workbook junk", None, None],
        [None, None, None],
        ["Q53: Which of the following challenges...", None, None],
        ["Values: 0-1", None, None],
        [None, 0, "Unchecked"],
        [None, 1, "Checked"],
        [None, "[Q53r1]", "Knowledge gap"],
        [None, " [Q53r2] ", " Sales marketing alignment gap "],
        [None, None, None],
        ["Q15: What best describes your involvement...", None, None],
        ["Values: 1-4", None, None],
        [None, 1, "Directly involved in decision making AND budget"],
        [None, 2, "Directly involved in decision making OR budget"],
        [None, 3, "Indirectly involved"],
        [None, 4, "Not involved"],
        [None, "[Q15r1]", "Overall company strategy"],
        [None, "[Q15r2]", "Go-to-market strategy"],
        [None, "[Q15r3]", "Partner strategy"],
        [None, None, None],
        ["[Q70]: What % of your pipeline...", None, None],
        ["Values: 0-100", None, None],
        [None, None, None],
        ["Q33: Allocate 100 points...", None, None],
        ["Values: 0-999", None, None],
        [None, "[Q33r1]", "Managing pricing"],
        [None, "[Q33r2]", "Customer-led shift"],
        [None, None, None],
        [
            "[Q4r98oe]: Which of the following industries... - Other "
            "(please specify)",
            None,
            None,
        ],
        ["Open text response", None, None],
        [None, None, None],
        ["QMissingType: Header but missing type hint", None, None],
        [None, None, None],
        [None, None, None],
        ["[vQTIME_MINUTES]: Survey length in minutes", None, None],
        ["Values: -99999999999999-999999999999999", None, None],
    ]

    for row in rows:
        worksheet.append(row)

    workbook.save(path)
    workbook.close()


def _build_missing_sheet_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "IndexOnly"
    worksheet.append(["metadata"])
    workbook.save(path)
    workbook.close()


def _build_raw_decoder_fixtures() -> None:
    headers = [
        "record",
        "Q3",
        "Q53r1",
        "Q53r2",
        "Q70",
        "Q4r98oe",
        "vStatus",
        "extra_col",
    ]
    rows = [
        ["R001", "1", "1", "0", "10", "note 1", "complete", "extra"],
        ["R002", "2", "0", "1", "20", "note 2", "complete", "extra"],
        ["R003", "1", "1", "1", "101", "note 3", "complete", "extra"],
        ["R004", "NA", "0", "0", "30", "note 4", "complete", "extra"],
        ["R005", "2", "1", "0", "105", "note 5", "complete", "extra"],
        ["R006", "abc", "0", "1", "60", "", "complete", "extra"],
        ["R007", "1", "1", "0", "60", "NA", "complete", "extra"],
        ["R008", "2", "0", "1", "70", "N/A", "complete", "extra"],
        ["R009", "1", "1", "0", "80", "", "complete", "extra"],
        ["R010", "2", "0", "1", "90", "", "complete", "extra"],
        ["R011", "1", "1", "0", "100", "", "complete", "extra"],
        ["R012", "2", "0", "1", "0", "", "complete", "extra"],
        ["R013", "1", "1", "0", "25", "", "complete", "extra"],
        ["R014", "2", "0", "1", "35", "", "complete", "extra"],
        ["R015", "1", "1", "0", "45", "", "complete", "extra"],
        ["R016", "2", "0", "1", "55", "", "complete", "extra"],
        ["R017", "1", "1", "0", "65", "", "complete", "extra"],
        ["R018", "2", "0", "1", "75", "", "complete", "extra"],
        ["R019", "1", "1", "0", "85", "tail 1", "complete", "extra"],
        ["R020", "2", "0", "1", "95", "tail 2", "complete", "extra"],
    ]

    _write_csv(RAW_DECODER_CSV_PATH, headers, rows)
    _write_xlsx(RAW_DECODER_XLSX_PATH, headers, rows)

    no_id_headers = headers[1:]
    no_id_rows = [row[1:] for row in rows]
    _write_csv(RAW_DECODER_NO_ID_CSV_PATH, no_id_headers, no_id_rows)


def _write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    lines = [",".join(headers)]
    lines.extend(",".join(row) for row in rows)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    workbook = Workbook()
    index_sheet = workbook.active
    index_sheet.title = "Index"
    index_sheet.append(["Workbook metadata"])

    worksheet = workbook.create_sheet("Raw")
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _build_golden_30_respondents_fixture() -> None:
    headers = [
        "respondent_id",
        "Q_SS_1",
        "Q_SS_2",
        "Q_SS_3",
        "Q_MS_1r1",
        "Q_MS_1r2",
        "Q_MS_1r3",
        "Q_NUM_1",
        "Q_ALLOC_1r1",
        "Q_ALLOC_1r2",
        "Q_ALLOC_1r3",
    ]
    q_ss_1 = [1] * 18 + [2] * 12
    q_ss_2 = [1] * 5 + [2] * 5 + [3] * 10 + [4] * 10
    q_ss_3 = [1] * 10 + [2] * 10 + [3] * 7 + [""] * 3
    q_ms_1r1 = [1] * 12 + [0] * 18
    q_ms_1r2 = [1] * 6 + [0] * 24
    q_ms_1r3 = [0] * 30
    q_num_1 = list(range(1, 11)) * 3
    allocation_rows = (
        [[50, 30, 20] for _ in range(25)]
        + [[60, 30, 10], [33, 33, 35], [32, 33, 34]]
        + [[50, 50, 10], [30, 30, 30]]
    )

    rows = []
    for index in range(30):
        rows.append(
            [
                f"G{index + 1:03d}",
                str(q_ss_1[index]),
                str(q_ss_2[index]),
                str(q_ss_3[index]),
                str(q_ms_1r1[index]),
                str(q_ms_1r2[index]),
                str(q_ms_1r3[index]),
                str(q_num_1[index]),
                str(allocation_rows[index][0]),
                str(allocation_rows[index][1]),
                str(allocation_rows[index][2]),
            ]
        )

    _write_csv(GOLDEN_30_RESPONDENTS_PATH, headers, rows)


ensure_fixtures()
