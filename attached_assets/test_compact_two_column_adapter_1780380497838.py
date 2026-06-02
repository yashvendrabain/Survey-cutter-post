"""Tests for the compact two-column data-map adapter."""

from __future__ import annotations

import unittest

from openpyxl import Workbook, load_workbook

from src.adapters.compact_two_column import CompactTwoColumnAdapter
from tests.conftest import DATAMAP_FIXTURE_PATH


def make_compact_workbook(sheet_name: str = "Sheet1") -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["Q1", "What is your current employment status?"])
    worksheet.append([1, "Full-time"])
    worksheet.append([2, "Part-time"])
    worksheet.append(["a", "Alphabetic option"])
    worksheet.append([None, None])
    worksheet.append(["Q2", "What was your most recent industry?"])
    worksheet.append([1, "Technology"])
    worksheet.append([2, "Retail"])
    return workbook


def make_winvslag_style_compact_workbook() -> Workbook:
    workbook = Workbook()
    index = workbook.active
    index.title = "Index"
    index.append(["Contents"])
    raw = workbook.create_sheet("Raw Data")
    raw.append(["uuid", "Q1", "Q2"])
    raw.append(["r1", 1, 2])
    data_map = workbook.create_sheet("Data map")
    data_map.append(["Q1", "Employment"])
    data_map.append([1, "Full-time"])
    data_map.append([2, "Part-time"])
    return workbook


class TestCompactTwoColumnAdapter(unittest.TestCase):
    def test_detects_compact_workbook(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook())

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("compact", result.reason)

    def test_detects_compact_format_when_sheet_named_data_map(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook("Data map"))

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("Data map", result.reason)

    def test_detects_compact_format_in_winvslag2024_style_workbook(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_winvslag_style_compact_workbook())

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("Data map", result.reason)

    def test_detects_compact_format_with_codebook_sheet_name(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook("Codebook"))

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("Codebook", result.reason)

    def test_detects_compact_format_with_schema_sheet_name(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook("Schema"))

        self.assertGreaterEqual(result.confidence, 0.6)
        self.assertIn("Schema", result.reason)

    def test_does_not_match_question_metadata_sheet(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook("Question_Metadata"))

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_does_not_match_dictionary_sheet(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook("Dictionary"))

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_does_not_match_variables_sheet(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook("Variables"))

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_does_not_match_metadata_sheet(self) -> None:
        result = CompactTwoColumnAdapter().detect(make_compact_workbook("Metadata"))

        self.assertEqual(result.confidence, 0.0)
        self.assertIn("no data map sheet", result.reason)

    def test_low_confidence_for_bcn_fixture(self) -> None:
        workbook = load_workbook(DATAMAP_FIXTURE_PATH, read_only=True, data_only=True)
        try:
            result = CompactTwoColumnAdapter().detect(workbook)
        finally:
            workbook.close()

        self.assertEqual(result.confidence, 0.0)

    def test_parse_compact_workbook_preserves_codes_and_options(self) -> None:
        workbook = make_compact_workbook()
        setattr(workbook, "_survey_source_path", "compact.xlsx")

        parsed = CompactTwoColumnAdapter().parse(workbook)

        questions = {question["canonical_id"]: question for question in parsed["questions"]}
        self.assertEqual(parsed["source_path"], "compact.xlsx")
        self.assertEqual(
            questions["Q1"]["options"],
            [(1, "Full-time"), (2, "Part-time"), ("a", "Alphabetic option")],
        )
        self.assertEqual(questions["Q2"]["question_text"], "What was your most recent industry?")


if __name__ == "__main__":
    unittest.main()
