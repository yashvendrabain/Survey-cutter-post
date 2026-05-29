from __future__ import annotations

import inspect
import unittest

from openpyxl import load_workbook

from src.calculation_log import CalculationLog
from src.excel_exporter import (
    export_cross_cuts_only,
    export_filtered_single_cuts,
    export_single_cuts,
    export_winners_vs_laggards_workbook,
    write_workbook,
)
from src.models import SegmentDefinition
from tests.test_excel_exporter_winners_vs_laggards import _fixture, _raw_column_values


class TestWvlManualCohortExporter(unittest.TestCase):
    def test_exporter_entry_points_preserve_embed_input_kwargs(self) -> None:
        for function in (
            export_single_cuts,
            write_workbook,
            export_cross_cuts_only,
            export_filtered_single_cuts,
            export_winners_vs_laggards_workbook,
        ):
            with self.subTest(function=function.__name__):
                signature = inspect.signature(function)
                self.assertIn("embed_input_files", signature.parameters)
                self.assertIn("input_file_sources", signature.parameters)
                self.assertFalse(signature.parameters["embed_input_files"].default)
                self.assertIsNone(signature.parameters["input_file_sources"].default)

    def test_manual_uuid_wvl_masks_use_respondent_id_column(self) -> None:
        output_path, schema, df, results, _default_segment, log = _fixture()
        segment = SegmentDefinition(
            outcome_question_id="manual_uuid",
            segment_mode="manual_uuid",
            manual_winner_uuids=("1", "3"),
            manual_laggard_uuids=("2",),
            winner_label="Winners",
            loser_label="Laggards",
            laggard_label="Laggards",
        )
        export_winners_vs_laggards_workbook(
            output_path=output_path,
            decoded_df=df,
            schema=schema,
            single_cut_results=results,
            segment_definition=segment,
            laggard_segment_definition=segment,
            themes={"themes": [{"name": "Demographics", "question_ids": ["Q1"]}]},
            calculation_log=log,
        )
        workbook = load_workbook(output_path, data_only=False)
        self.assertEqual(
            _raw_column_values(workbook, "winners_mask_data"),
            [True, False, True, False],
        )
        self.assertEqual(
            _raw_column_values(workbook, "laggards_mask_data"),
            [False, True, False, False],
        )

    def test_manual_uuid_cohort_definition_sheet_records_lists(self) -> None:
        output_path, schema, df, results, _default_segment, _log = _fixture()
        segment = SegmentDefinition(
            outcome_question_id="manual_uuid",
            segment_mode="manual_uuid",
            manual_winner_uuids=("1", "2"),
            manual_laggard_uuids=("2", "4"),
            winner_label="Winners",
            loser_label="Laggards",
            laggard_label="Laggards",
        )
        export_winners_vs_laggards_workbook(
            output_path=output_path,
            decoded_df=df,
            schema=schema,
            single_cut_results=results,
            segment_definition=segment,
            laggard_segment_definition=segment,
            themes={"themes": [{"name": "Demographics", "question_ids": ["Q1"]}]},
            calculation_log=CalculationLog(),
            embed_input_files=False,
            input_file_sources=None,
        )
        workbook = load_workbook(output_path, data_only=False)
        sheet = workbook["_CohortDefinition"]
        self.assertEqual(sheet.sheet_state, "hidden")
        self.assertEqual(sheet["B2"].value, "manual_uuid")
        self.assertEqual(sheet["B3"].value, "(N/A for manual)")
        self.assertEqual(sheet["B7"].value, 1)
        self.assertEqual(sheet["A10"].value, "Winner uuids:")
        self.assertEqual(sheet["A11"].value, "1")
        self.assertEqual(sheet["A12"].value, "2")
        self.assertEqual(sheet["A14"].value, "Laggard uuids:")
        self.assertEqual(sheet["A15"].value, "2")
        self.assertEqual(sheet["A16"].value, "4")


if __name__ == "__main__":
    unittest.main()
