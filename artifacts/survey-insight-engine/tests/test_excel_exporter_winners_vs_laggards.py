from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.excel_exporter import export_winners_vs_laggards_workbook
from src.models import (
    DenominatorPolicy,
    GridRatedResult,
    GridRatedRow,
    QuestionSpec,
    QuestionType,
    SegmentDefinition,
    SingleSelectResult,
    SurveySchema,
)


def _fixture() -> tuple[Path, SurveySchema, pd.DataFrame, list, SegmentDefinition, CalculationLog]:
    schema = SurveySchema(
        questions=(
            QuestionSpec(
                question_id="QO",
                canonical_id="QO",
                question_text="Outcome",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("QO",),
                option_map={1: "Winner", 2: "Laggard"},
            ),
            QuestionSpec(
                question_id="QO2",
                canonical_id="QO2",
                question_text="Override outcome",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("QO2",),
                option_map={1: "Winner", 2: "Laggard"},
            ),
            QuestionSpec(
                question_id="Q1",
                canonical_id="Q1",
                question_text="Choice",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q1",),
                option_map={1: "A", 2: "B"},
            ),
            QuestionSpec(
                question_id="QG",
                canonical_id="QG",
                question_text="Rated grid",
                question_type=QuestionType.GRID_RATED,
                raw_columns=("QGr1c1", "QGr2c1"),
                option_map={},
                grid_row_labels={"QGr1c1": "Row 1", "QGr2c1": "Row 2"},
            ),
        ),
        respondent_id_column="rid",
        total_respondents=4,
        source_datamap_path="map.xlsx",
        source_rawdata_path="raw.xlsx",
        parsed_at=datetime.now(timezone.utc),
    )
    df = pd.DataFrame(
        {
            "rid": [1, 2, 3, 4],
            "QO": [1, 1, 2, 2],
            "QO2": [2, 2, 1, 1],
            "Q1": [1, 2, 1, 2],
            "QGr1c1": [8, 9, 3, 4],
            "QGr2c1": [7, 8, 2, 3],
        }
    )
    results = [
        SingleSelectResult(
            question_id="Q1",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                1: {"label": "A", "count": 2, "rate": 0.5},
                2: {"label": "B", "count": 2, "rate": 0.5},
            },
        ),
        GridRatedResult(
            question_id="QG",
            question_type=QuestionType.GRID_RATED,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            question_text="Rated grid",
            column_headers=["Score"],
            rows=(
                GridRatedRow(
                    row_id="QGr1",
                    row_label="Row 1",
                    means_per_column=[6.0],
                    valid_n_per_column=[4],
                ),
                GridRatedRow(
                    row_id="QGr2",
                    row_label="Row 2",
                    means_per_column=[5.0],
                    valid_n_per_column=[4],
                ),
            ),
            total_respondents=4,
            total_responses=8,
            show_delta=False,
        ),
    ]
    output_path = Path(tempfile.gettempdir()) / "winners_vs_laggards_test.xlsx"
    log = CalculationLog()
    segment = SegmentDefinition(
        "QO",
        "categorical",
        winner_values=(1,),
        laggard_values=(2,),
        winner_label="Winners",
        loser_label="Laggards",
        laggard_label="Laggards",
    )
    return output_path, schema, df, results, segment, log


def _build_workbook(segment: SegmentDefinition | None = None):
    output_path, schema, df, results, default_segment, log = _fixture()
    export_winners_vs_laggards_workbook(
        output_path=output_path,
        decoded_df=df,
        schema=schema,
        single_cut_results=results,
        segment_definition=segment or default_segment,
        laggard_segment_definition=segment or default_segment,
        calculation_log=log,
    )
    return output_path, load_workbook(output_path, data_only=False), log


def _data_sheet(workbook):
    return workbook["All Questions"]


class TestWinnersVsLaggardsWorkbook(unittest.TestCase):
    def test_workbook_is_produced(self) -> None:
        output_path, _workbook, _log = _build_workbook()
        self.assertTrue(output_path.exists())

    def test_run_summary_contains_labels_and_sizes(self) -> None:
        _path, workbook, _log = _build_workbook()
        values = [cell.value for row in workbook["Run_Summary"].iter_rows() for cell in row]
        self.assertIn("Winner cohort size:", values)
        self.assertIn("Laggard cohort size:", values)
        self.assertIn("Winners", values)

    def test_single_select_sheet_has_documented_columns(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        headers = [sheet.cell(row=4, column=col).value for col in range(1, 8)]
        self.assertEqual(headers, ["Option ID", "Option", "Winners count", "Winners %", "Laggards count", "Laggards %", "Delta %"])

    def test_count_percentage_and_delta_cells_are_formulas(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        self.assertTrue(str(sheet["C5"].value).startswith("="))
        self.assertTrue(str(sheet["D5"].value).startswith("=IFERROR("))
        self.assertTrue(str(sheet["G5"].value).startswith("="))
        self.assertIn("winners_mask_data", sheet["C5"].value)
        self.assertIn("laggards_mask_data", sheet["E5"].value)

    def test_grid_rated_layout_uses_formula_means(self) -> None:
        _path, workbook, _log = _build_workbook()
        sheet = _data_sheet(workbook)
        row = next(
            r
            for r in range(1, sheet.max_row + 1)
            if sheet.cell(row=r, column=1).value == "Sub-question ID"
        )
        self.assertTrue(str(sheet.cell(row=row + 1, column=3).value).startswith("=IFERROR(AVERAGEIFS"))
        self.assertTrue(str(sheet.cell(row=row + 1, column=4).value).startswith("=IFERROR(AVERAGEIFS"))
        self.assertTrue(str(sheet.cell(row=row + 1, column=5).value).startswith("="))

    def test_named_ranges_for_masks_exist(self) -> None:
        _path, workbook, _log = _build_workbook()
        self.assertIn("winners_mask_data", workbook.defined_names)
        self.assertIn("laggards_mask_data", workbook.defined_names)

    def test_override_outcome_records_laggard_source_column(self) -> None:
        segment = SegmentDefinition(
            "QO",
            "categorical",
            winner_values=(1,),
            laggard_values=(2,),
            laggard_outcome_question_id="QO2",
        )
        _path, _workbook, log = _build_workbook(segment)
        laggard_records = [
            record for record in log.all_records() if record.metric_name == "laggard_mask"
        ]
        self.assertEqual(laggard_records[-1].source_question_id, "QO2")

    def test_backward_compat_defaults_inverse_laggards(self) -> None:
        segment = SegmentDefinition("QO", "categorical", winner_values=(1,))
        _path, workbook, _log = _build_workbook(segment)
        raw = workbook["_RawData"]
        headers = [raw.cell(row=1, column=col).value for col in range(1, raw.max_column + 1)]
        lag_col = headers.index("laggards_mask_data") + 1
        self.assertEqual([raw.cell(row=row, column=lag_col).value for row in range(2, 6)], [False, False, True, True])


if __name__ == "__main__":
    unittest.main()
