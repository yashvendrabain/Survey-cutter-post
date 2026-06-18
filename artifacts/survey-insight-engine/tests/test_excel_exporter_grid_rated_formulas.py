"""Regression tests for live rated-grid formulas in exported workbooks."""

from __future__ import annotations

from datetime import datetime, timezone
import unittest
from uuid import uuid4

from openpyxl import load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.excel_exporter import export_single_cuts
from src.models import (
    DataQualityReport,
    DenominatorPolicy,
    GridRatedResult,
    GridRatedRow,
    QuestionSpec,
    QuestionType,
    SurveySchema,
)
from tests.conftest import make_temp_output_dir


UTC_NOW = datetime(2026, 5, 26, tzinfo=timezone.utc)
OUTPUT_DIR = make_temp_output_dir()


def _quality_report(dataframe: pd.DataFrame) -> DataQualityReport:
    return DataQualityReport(
        total_rows=len(dataframe),
        total_columns=len(dataframe.columns),
        columns_in_datamap=len(dataframe.columns),
        columns_not_in_datamap=(),
        per_column_missing_pct={},
        per_column_out_of_range_pct={},
        coercion_log=(),
        warnings=(),
    )


def _find_header_row(ws, header: str) -> int:
    for row_index in range(1, ws.max_row + 1):
        if ws.cell(row_index, 1).value == header:
            return row_index
    raise AssertionError(f"Header {header!r} not found")


class TestGridRatedExporterFormulas(unittest.TestCase):
    def test_grid_rated_means_delta_and_total_are_formulas(self) -> None:
        dataframe = pd.DataFrame(
            {
                "respondent_id": [1, 2, 3, 4],
                "Q30r1c1": [8, 9, 8, None],
                "Q30r1c2": [6, 7, 7, None],
                "Q30r2c1": [9, 9, None, 8],
                "Q30r2c2": [7, 6, None, 6],
            }
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="Q30",
                    canonical_id="Q30",
                    question_text="Rate each vendor attribute",
                    question_type=QuestionType.GRID_RATED,
                    raw_columns=("Q30r1c1", "Q30r1c2", "Q30r2c1", "Q30r2c2"),
                    option_map={index: str(index) for index in range(0, 11)},
                    value_range=(0, 10),
                    grid_row_labels={
                        "Q30r1c1": "Pre-purchase familiarity",
                        "Q30r1c2": "Pre-purchase familiarity",
                        "Q30r2c1": "Customer validation",
                        "Q30r2c2": "Customer validation",
                    },
                    grid_column_labels={"1": "Winner", "2": "Other vendor"},
                    possible_role="GRID_RATED",
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        result = GridRatedResult(
            question_id="Q30",
            question_type=QuestionType.GRID_RATED,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            question_text="Rate each vendor attribute",
            column_headers=["Winner", "Other vendor"],
            rows=[
                GridRatedRow("Q30r1", "Pre-purchase familiarity", [8.3, 6.7], [3, 3], 1.6),
                GridRatedRow("Q30r2", "Customer validation", [8.7, 6.3], [3, 3], 2.4),
            ],
            total_respondents=4,
            total_responses=12,
            show_delta=True,
        )

        output_path = OUTPUT_DIR / f"grid_rated_formulas_{uuid4().hex}.xlsx"
        export_single_cuts(
            [result],
            [],
            schema,
            _quality_report(dataframe),
            CalculationLog(),
            str(output_path),
            decoded_df=dataframe,
        )

        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = _find_header_row(ws, "Sub-question ID")
        data_start = header_row + 1
        data_end = data_start + 1

        formula_cells = []
        for row_index in range(data_start, data_end + 1):
            for column_index in (3, 4, 5):
                formula_cells.append(ws.cell(row_index, column_index).value)
        total_row = data_end + 1
        formula_cells.append(ws.cell(total_row, 2).value)

        for formula in formula_cells:
            self.assertTrue(str(formula).startswith("="), formula)
        mean_formula = str(ws.cell(data_start, 3).value)
        self.assertIn("AVERAGEIFS", mean_formula)
        self.assertIn("passes_workbook_filters_data", mean_formula)
        self.assertIn("passes_workbook_filters_data", str(ws.cell(total_row, 2).value))


if __name__ == "__main__":
    unittest.main()
