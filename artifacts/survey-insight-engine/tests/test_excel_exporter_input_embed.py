from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import unittest
import uuid

from openpyxl import Workbook, load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.excel_exporter import export_single_cuts
from src.models import (
    DataQualityReport,
    DenominatorPolicy,
    QuestionSpec,
    QuestionType,
    SingleSelectResult,
    SurveySchema,
)


def _write_fixture_workbook(path: Path, rows: list[list[object]], sheet_name: str = "Sheet1") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index, value=value)
    workbook.save(path)
    workbook.close()


def _fixture(tmp_dir: Path) -> tuple[Path, Path, Path, SurveySchema, pd.DataFrame, list[SingleSelectResult], DataQualityReport]:
    raw_rows = [
        ["record", "Q1", "Q2"],
        [1, 1, "alpha"],
        [2, 2, "beta"],
        [3, 1, "gamma"],
        [4, 2, "delta"],
    ]
    datamap_rows = [
        ["Variable", "Question", "Type", "Code", "Label"],
        ["Q1", "Choice", "Single select", 1, "A"],
        ["Q1", "Choice", "Single select", 2, "B"],
        ["Q2", "Text", "Open text", "", ""],
        ["", "", "", "", ""],
        ["Meta", "Source", "", "", ""],
        ["Row 7", "", "", "", ""],
        ["Row 8", "", "", "", ""],
        ["Row 9", "", "", "", ""],
        ["Row 10", "", "", "", ""],
    ]
    raw_path = tmp_dir / "raw_upload.xlsx"
    datamap_path = tmp_dir / "datamap_upload.xlsx"
    output_path = tmp_dir / "survey_analysis.xlsx"
    _write_fixture_workbook(raw_path, raw_rows, "Raw")
    _write_fixture_workbook(datamap_path, datamap_rows, "Sheet1")

    schema = SurveySchema(
        questions=(
            QuestionSpec(
                question_id="Q1",
                canonical_id="Q1",
                question_text="Choice",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("Q1",),
                option_map={1: "A", 2: "B"},
            ),
        ),
        respondent_id_column="record",
        total_respondents=4,
        source_datamap_path=str(datamap_path),
        source_rawdata_path=str(raw_path),
        parsed_at=datetime.now(timezone.utc),
    )
    decoded_df = pd.DataFrame({"record": [1, 2, 3, 4], "Q1": [1, 2, 1, 2]})
    results = [
        SingleSelectResult(
            question_id="Q1",
            question_type=QuestionType.SINGLE_SELECT,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            distribution={
                1: {"label": "A", "count": 2, "rate": 0.5},
                2: {"label": "B", "count": 2, "rate": 0.5},
            },
        )
    ]
    quality_report = DataQualityReport(
        total_rows=4,
        total_columns=2,
        columns_in_datamap=1,
        columns_not_in_datamap=(),
        per_column_missing_pct={},
        per_column_out_of_range_pct={},
        coercion_log=(),
        warnings=(),
    )
    return output_path, raw_path, datamap_path, schema, decoded_df, results, quality_report


def _export_fixture(embed_input_files: bool):
    temp_root = Path.cwd() / "outputs" / "test_input_embed"
    temp_root.mkdir(parents=True, exist_ok=True)
    tmp_dir = temp_root / f"case_{uuid.uuid4().hex}"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    output_path, raw_path, datamap_path, schema, decoded_df, results, quality_report = _fixture(tmp_dir)
    export_single_cuts(
        results=results,
        skips=[],
        schema=schema,
        quality_report=quality_report,
        log=CalculationLog(),
        output_path=str(output_path),
        themes={"themes": [{"name": "Demographics", "question_ids": ["Q1"]}]},
        decoded_df=decoded_df,
        embed_input_files=embed_input_files,
        input_file_sources={"raw_path": str(raw_path), "datamap_path": str(datamap_path), "raw_sheet": "Raw"},
    )
    return tmp_dir, output_path, load_workbook(output_path, data_only=False)


class TestExcelExporterInputEmbed(unittest.TestCase):
    def test_no_input_sheets_when_checkbox_off(self) -> None:
        _tmp, _path, workbook = _export_fixture(False)
        self.assertNotIn("Raw Data (Input)", workbook.sheetnames)
        self.assertNotIn("Data Map (Input)", workbook.sheetnames)

    def test_input_sheets_present_when_checkbox_on(self) -> None:
        _tmp, _path, workbook = _export_fixture(True)
        self.assertIn("Raw Data (Input)", workbook.sheetnames)
        self.assertIn("Data Map (Input)", workbook.sheetnames)

    def test_raw_data_input_sheet_matches_uploaded_file(self) -> None:
        _tmp, _path, workbook = _export_fixture(True)
        sheet = workbook["Raw Data (Input)"]
        values = [[sheet.cell(row=row, column=col).value for col in range(1, 4)] for row in range(1, 6)]
        self.assertEqual(
            values,
            [
                ["record", "Q1", "Q2"],
                [1, 1, "alpha"],
                [2, 2, "beta"],
                [3, 1, "gamma"],
                [4, 2, "delta"],
            ],
        )

    def test_data_map_input_sheet_matches_uploaded_file(self) -> None:
        _tmp, _path, workbook = _export_fixture(True)
        sheet = workbook["Data Map (Input)"]
        values = [[sheet.cell(row=row, column=col).value for col in range(1, 6)] for row in range(1, 11)]
        self.assertEqual(values[0], ["Variable", "Question", "Type", "Code", "Label"])
        self.assertEqual(values[1], ["Q1", "Choice", "Single select", 1, "A"])
        self.assertEqual(values[9], ["Row 10", None, None, None, None])

    def test_input_sheets_do_not_replace_underscore_rawdata(self) -> None:
        _tmp, _path, workbook = _export_fixture(True)
        self.assertIn("_RawData", workbook.sheetnames)
        self.assertIn("Raw Data (Input)", workbook.sheetnames)
        self.assertEqual(workbook["_RawData"].sheet_state, "hidden")
        self.assertEqual(workbook["Raw Data (Input)"].sheet_state, "visible")

    def test_input_sheet_position_is_after_main_content(self) -> None:
        _tmp, _path, workbook = _export_fixture(True)
        names = workbook.sheetnames
        self.assertLess(names.index("Filter_Log"), names.index("Raw Data (Input)"))
        self.assertLess(names.index("Warnings"), names.index("Raw Data (Input)"))
        self.assertLess(names.index("Raw Data (Input)"), names.index("_RawData"))
        self.assertLess(names.index("Data Map (Input)"), names.index("_RawData"))


if __name__ == "__main__":
    unittest.main()
