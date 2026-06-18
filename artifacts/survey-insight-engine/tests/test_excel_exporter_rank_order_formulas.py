"""Regression tests for live rank-order formulas in exported workbooks."""

from __future__ import annotations

from datetime import datetime, timezone
import re
import unittest
from uuid import uuid4

from openpyxl import load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.excel_exporter import export_single_cuts
from src.models import (
    DataQualityReport,
    DenominatorPolicy,
    QuestionSpec,
    QuestionType,
    RankOrderResult,
    RankOrderRow,
    SurveySchema,
)
from tests.conftest import make_temp_output_dir


UTC_NOW = datetime(2026, 5, 26, tzinfo=timezone.utc)
OUTPUT_DIR = make_temp_output_dir()


def _quality_report(dataframe: pd.DataFrame) -> DataQualityReport:
    return DataQualityReport(
        total_rows=len(dataframe),
        total_columns=len(dataframe.columns),
        columns_in_datamap=len(dataframe.columns),
        columns_not_in_datamap=(),
        per_column_missing_pct={},
        per_column_out_of_range_pct={},
        coercion_log=(),
        warnings=(),
    )


def _find_header_row(ws, header: str) -> int:
    for row_index in range(1, ws.max_row + 1):
        if ws.cell(row_index, 1).value == header:
            return row_index
    raise AssertionError(f"Header {header!r} not found")


class TestRankOrderExporterFormulas(unittest.TestCase):
    def test_rank_order_counts_percents_and_totals_are_formulas(self) -> None:
        dataframe = pd.DataFrame(
            {
                "respondent_id": [1, 2, 3, 4],
                "Q43r1": [1, 2, None, 1],
                "Q43r2": [2, 1, 1, None],
            }
        )
        schema = SurveySchema(
            questions=(
                QuestionSpec(
                    question_id="Q43",
                    canonical_id="Q43",
                    question_text="Rank the selection criteria",
                    question_type=QuestionType.RANK_ORDER,
                    raw_columns=("Q43r1", "Q43r2"),
                    option_map={"Q43r1": "Criterion A", "Q43r2": "Criterion B"},
                    value_range=(1, 2),
                ),
            ),
            respondent_id_column="respondent_id",
            total_respondents=4,
            source_datamap_path="datamap.xlsx",
            source_rawdata_path="raw.csv",
            parsed_at=UTC_NOW,
        )
        result = RankOrderResult(
            question_id="Q43",
            question_type=QuestionType.RANK_ORDER,
            valid_n=4,
            missing_n=0,
            denominator_policy=DenominatorPolicy.VALID_RESPONSES,
            question_text="Rank the selection criteria",
            K=2,
            rows=[
                RankOrderRow("Q43r1", "Criterion A", [2, 1], [0.5, 0.25]),
                RankOrderRow("Q43r2", "Criterion B", [2, 1], [0.5, 0.25]),
            ],
            total_respondents=4,
            total_responses=6,
        )

        output_path = OUTPUT_DIR / f"rank_order_formulas_{uuid4().hex}.xlsx"
        export_single_cuts(
            [result],
            [],
            schema,
            _quality_report(dataframe),
            CalculationLog(),
            str(output_path),
            decoded_df=dataframe,
        )

        workbook = load_workbook(output_path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        ws = workbook["All Questions"]
        header_row = _find_header_row(ws, "Option ID")
        data_start = header_row + 2
        data_end = data_start + 1

        formula_cells = []
        for row_index in range(data_start, data_end + 1):
            for column_index in (3, 4, 5, 6, 7):
                formula_cells.append(ws.cell(row_index, column_index).value)
        total_row = data_end + 1
        responses_row = total_row + 1
        qc_row = responses_row + 1
        formula_cells.extend(
            [
                ws.cell(total_row, 2).value,
                ws.cell(responses_row, 2).value,
                ws.cell(qc_row, 3).value,
            ]
        )

        for formula in formula_cells:
            self.assertTrue(str(formula).startswith("="), formula)
        self.assertEqual(ws.cell(header_row, 3).value, "Weighted Average")

        weighted_average = str(ws.cell(data_start, 3).value)
        # The weighted-average formula is checked structurally rather than as a
        # frozen literal so that benign exporter tweaks (row-layout shifts,
        # whitespace, formula reordering) do not break the test while a genuine
        # logic regression still does.
        self.assertTrue(
            weighted_average.startswith("=IFERROR("), weighted_average
        )
        self.assertTrue(weighted_average.endswith(",0)"), weighted_average)
        # Weighted sum indexes into the points named range, once per rank (K=2).
        for rank in (1, 2):
            self.assertIn(
                f"INDEX(All_Questions_Q43_PTS,{rank})", weighted_average
            )
            # Each rank is counted via COUNTIFS over the option data range for
            # that rank value.
            self.assertIn(
                f'COUNTIFS(Q43r1_data,"{rank}"', weighted_average
            )
        # COUNTIFS constrains on every filter named range.
        for filter_range in (
            "passes_workbook_filters_data",
            "passes_workbook_custom_filters_data",
            "All_Questions_passes_local_filters_data",
            "All_Questions_Q43_F_passes_per_q_filter_data",
        ):
            self.assertIn(filter_range, weighted_average)
        # The weighted sum is divided by the totals cell in column B; the exact
        # row is layout-dependent, so only the column anchor is pinned.
        self.assertRegex(weighted_average, r"/\$B\$\d+,0\)$")
        self.assertIn("passes_workbook_filters_data", str(ws.cell(data_start, 4).value))
        self.assertIn("passes_workbook_filters_data", str(ws.cell(data_start, 6).value))
        self.assertIn("Q43r1_data", str(ws.cell(data_start, 4).value))
        self.assertIn("Q43r1_data", str(ws.cell(data_start, 6).value))
        self.assertIn("passes_workbook_filters_data", str(ws.cell(total_row, 2).value))


if __name__ == "__main__":
    unittest.main()
