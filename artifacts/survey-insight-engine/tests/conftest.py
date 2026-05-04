"""Fixture builders for datamap parser tests."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"
DATAMAP_FIXTURE_PATH = FIXTURE_DIR / "test_datamap_minimal.xlsx"
MISSING_SHEET_FIXTURE_PATH = FIXTURE_DIR / "missing_sheet1.xlsx"


def ensure_fixtures() -> None:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    _build_minimal_datamap(DATAMAP_FIXTURE_PATH)
    _build_missing_sheet_workbook(MISSING_SHEET_FIXTURE_PATH)


def _build_minimal_datamap(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    index_sheet = workbook.create_sheet("Index")
    index_sheet.append(["[QIndex]: This sheet must be ignored"])
    index_sheet.append(["Values: 1-2"])
    index_sheet.append([None, 1, "Should not appear"])

    rows = [
        [None, None, None],
        ["  [Q3]: Are you currently in a full-time position  ", None, None],
        [" Values: 1-2 ", None, None],
        [None, 1, " Yes "],
        ["", "2", "No"],
        [None, None, None],
        [None, None, None],
        ["This is workbook junk", None, None],
        [None, None, None],
        ["Q53: Which of the following challenges...", None, None],
        ["Values: 0-1", None, None],
        [None, 0, "Unchecked"],
        [None, 1, "Checked"],
        [None, "[Q53r1]", "Knowledge gap"],
        [None, " [Q53r2] ", " Sales marketing alignment gap "],
        [None, None, None],
        ["Q15: What best describes your involvement...", None, None],
        ["Values: 1-4", None, None],
        [None, 1, "Directly involved in decision making AND budget"],
        [None, 2, "Directly involved in decision making OR budget"],
        [None, 3, "Indirectly involved"],
        [None, 4, "Not involved"],
        [None, "[Q15r1]", "Overall company strategy"],
        [None, "[Q15r2]", "Go-to-market strategy"],
        [None, "[Q15r3]", "Partner strategy"],
        [None, None, None],
        ["[Q70]: What % of your pipeline...", None, None],
        ["Values: 0-100", None, None],
        [None, None, None],
        ["Q33: Allocate 100 points...", None, None],
        ["Values: 0-999", None, None],
        [None, "[Q33r1]", "Managing pricing"],
        [None, "[Q33r2]", "Customer-led shift"],
        [None, None, None],
        [
            "[Q4r98oe]: Which of the following industries... - Other "
            "(please specify)",
            None,
            None,
        ],
        ["Open text response", None, None],
        [None, None, None],
        ["QMissingType: Header but missing type hint", None, None],
        [None, None, None],
        [None, None, None],
        ["[vQTIME_MINUTES]: Survey length in minutes", None, None],
        ["Values: -99999999999999-999999999999999", None, None],
    ]

    for row in rows:
        worksheet.append(row)

    workbook.save(path)
    workbook.close()


def _build_missing_sheet_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "IndexOnly"
    worksheet.append(["metadata"])
    workbook.save(path)
    workbook.close()


ensure_fixtures()
