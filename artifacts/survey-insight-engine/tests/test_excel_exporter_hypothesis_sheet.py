from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
import uuid
import unittest

from openpyxl import load_workbook
import pandas as pd

from src.calculation_log import CalculationLog
from src.excel_exporter import export_single_cuts
from src.hypothesis_validator import validate_hypothesis
from src.models import (
    DataQualityReport,
    DenominatorPolicy,
    HypothesisSpec,
    QuestionSpec,
    QuestionType,
    SingleSelectResult,
    SurveySchema,
)


def _scratch_xlsx(name: str) -> Path:
    root = Path.cwd() / "outputs" / "test_tmp"
    root.mkdir(parents=True, exist_ok=True)
    return root / f"{name}_{uuid.uuid4().hex}.xlsx"


def _schema() -> SurveySchema:
    return SurveySchema(
        questions=(
            QuestionSpec(
                question_id="QY",
                canonical_id="QY",
                question_text="Outcome numeric",
                question_type=QuestionType.DIRECT_NUMERIC,
                raw_columns=("QY",),
                option_map={},
            ),
            QuestionSpec(
                question_id="QX",
                canonical_id="QX",
                question_text="Predictor numeric",
                question_type=QuestionType.DIRECT_NUMERIC,
                raw_columns=("QX",),
                option_map={},
            ),
            QuestionSpec(
                question_id="QCAT",
                canonical_id="QCAT",
                question_text="Category",
                question_type=QuestionType.SINGLE_SELECT,
                raw_columns=("QCAT",),
                option_map={1: "A", 2: "B"},
            ),
        ),
        respondent_id_column="respondent_id",
        total_respondents=80,
        source_datamap_path="datamap.xlsx",
        source_rawdata_path="raw.xlsx",
        parsed_at=datetime.now(timezone.utc),
    )


def _quality_report() -> DataQualityReport:
    return DataQualityReport(
        total_rows=80,
        total_columns=4,
        columns_in_datamap=3,
        columns_not_in_datamap=(),
        per_column_missing_pct={},
        per_column_out_of_range_pct={},
        coercion_log=(),
        warnings=(),
    )


def _single_result() -> SingleSelectResult:
    return SingleSelectResult(
        question_id="QCAT",
        question_type=QuestionType.SINGLE_SELECT,
        valid_n=80,
        missing_n=0,
        denominator_policy=DenominatorPolicy.VALID_RESPONSES,
        distribution={
            1: {"label": "A", "count": 40, "rate": 0.5},
            2: {"label": "B", "count": 40, "rate": 0.5},
        },
    )


def _write_workbook(include_hypothesis: bool = True) -> Path:
    schema = _schema()
    log = CalculationLog()
    hypothesis_results = []
    if include_hypothesis:
        df = pd.DataFrame(
            {
                "respondent_id": range(80),
                "QY": range(80),
                "QX": range(80),
                "QCAT": [1, 2] * 40,
            }
        )
        hypothesis_results.append(
            validate_hypothesis(
                HypothesisSpec("QY tracks QX", "QY", "QX", "correlated_positive"),
                df,
                schema,
                log,
                cross_cut_lookup={("QY", "QX"): "Run_Summary"},
            )
        )
    output_path = _scratch_xlsx("hypothesis")
    export_single_cuts(
        [_single_result()],
        [],
        schema,
        _quality_report(),
        log,
        str(output_path),
        decoded_df=pd.DataFrame(
            {
                "respondent_id": range(80),
                "QY": range(80),
                "QX": range(80),
                "QCAT": [1, 2] * 40,
            }
        ),
        hypothesis_results=hypothesis_results,
    )
    return output_path


class TestExcelExporterHypothesisSheet(unittest.TestCase):
    def test_hypothesis_check_sheet_written_when_results_non_empty(self) -> None:
        workbook = load_workbook(_write_workbook(True), data_only=False)
        try:
            self.assertIn("Hypothesis_Check", workbook.sheetnames)
        finally:
            workbook.close()

    def test_hypothesis_check_sheet_omitted_when_empty(self) -> None:
        workbook = load_workbook(_write_workbook(False), data_only=False)
        try:
            self.assertNotIn("Hypothesis_Check", workbook.sheetnames)
        finally:
            workbook.close()

    def test_effect_size_cell_is_formula_reference_to_calculation_log(self) -> None:
        workbook = load_workbook(_write_workbook(True), data_only=False)
        try:
            value = workbook["Hypothesis_Check"]["I2"].value
            self.assertIsInstance(value, str)
            self.assertTrue(value.startswith("="))
            self.assertIn("Calculation_Log", value)
        finally:
            workbook.close()

    def test_p_value_cell_is_formula_reference_to_calculation_log(self) -> None:
        workbook = load_workbook(_write_workbook(True), data_only=False)
        try:
            value = workbook["Hypothesis_Check"]["K2"].value
            self.assertIsInstance(value, str)
            self.assertTrue(value.startswith("="))
            self.assertIn("Calculation_Log", value)
        finally:
            workbook.close()

    def test_related_cross_cut_hyperlink_points_to_existing_sheet(self) -> None:
        workbook = load_workbook(_write_workbook(True), data_only=False)
        try:
            cell = workbook["Hypothesis_Check"]["Q2"]
            self.assertEqual(cell.value, "Run_Summary")
            self.assertIsNotNone(cell.hyperlink)
            self.assertIn("Run_Summary", cell.hyperlink.target)
            self.assertIn("Run_Summary", workbook.sheetnames)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
